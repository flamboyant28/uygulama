import streamlit as st
import pandas as pd
import math
import io
import os
import plotly.graph_objects as go
import folium
from streamlit_folium import st_folium
from math import radians, sin, cos, sqrt, atan2
from datetime import time, timedelta, datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Araç Yol Bilgisayarı", page_icon="🚗", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.page-title { font-size:2rem; font-weight:700; color:#1F4E78; margin-bottom:4px; }
.page-sub   { color:#888; font-size:0.9rem; margin-bottom:20px; }
.status-ok   { background:linear-gradient(135deg,#375623,#70AD47); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(55,86,35,.3); }
.status-fail { background:linear-gradient(135deg,#7B0000,#C00000); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(192,0,0,.3); }
.status-icon { font-size:2rem; }
.status-detail { font-weight:400; font-size:0.85rem; opacity:.85; margin-top:2px; }
.section-header { font-size:1rem; font-weight:700; color:#1F4E78;
    border-left:4px solid #1F4E78; padding-left:10px; margin:24px 0 10px; }
.route-strip { background:#EBF3FB; border-radius:10px; padding:12px 16px; margin:8px 0 10px;
               border-left:4px solid #1F4E78; }
.route-strip-title { font-size:0.72rem; font-weight:700; color:#1F4E78;
    text-transform:uppercase; letter-spacing:.6px; margin-bottom:8px; }
.route-seg { display:flex; align-items:center; gap:6px; font-size:0.85rem; color:#333; margin:4px 0; }
.route-seg-km { margin-left:auto; font-weight:700; color:#1F4E78; background:white;
    padding:2px 8px; border-radius:20px; font-size:0.78rem; border:1px solid #C8DDF0; }
.route-total { border-top:1px dashed #C8DDF0; margin-top:8px; padding-top:8px;
    font-weight:700; font-size:0.9rem; color:#1F4E78; display:flex; justify-content:space-between; }
.etki-box { background:#FFF8E7; border:1px solid #FFE082; border-radius:8px;
    padding:10px 14px; font-size:0.82rem; color:#5D4037; margin-top:6px; }
[data-testid="metric-container"] { background:white; border:1px solid #E8EDF2;
    border-radius:12px; padding:14px 18px; box-shadow:0 2px 8px rgba(0,0,0,.05); }
[data-testid="metric-container"] > label { font-size:0.72rem !important; color:#888 !important;
    font-weight:600 !important; text-transform:uppercase; letter-spacing:.5px; }
[data-testid="stMetricValue"] > div { font-size:1.4rem !important; font-weight:700 !important; color:#1F4E78 !important; }
[data-testid="stMetricDelta"] > div { font-size:0.75rem !important; color:#999 !important; }
[data-testid="stMetricDeltaIcon"] { display:none !important; }
[data-testid="stSidebar"] { background:#F8FAFC; border-right:1px solid #E2E8F0; }
[data-testid="stSidebar"] label { font-size:0.8rem !important; font-weight:600 !important; color:#555 !important; }
</style>
""", unsafe_allow_html=True)

# ── Veri ─────────────────────────────────────────────────────────────────────
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ilce_data import ILCELER
IL_LIST = sorted(ILCELER.keys())

def ilce_listesi(il): return sorted(ILCELER.get(il, {}).keys())
def koordinat(il, ilce): return ILCELER[il][ilce]

ROAD_FACTOR = 1.30

def haversine_km(c1, c2):
    lat1,lon1 = c1; lat2,lon2 = c2
    R = 6371
    dlat,dlon = radians(lat2-lat1), radians(lon2-lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    return round(R*2*atan2(sqrt(a),sqrt(1-a))*ROAD_FACTOR)

def fmt_sure(h_float):
    h = int(h_float); m = int(round((h_float-h)*60))
    if m == 60: h+=1; m=0
    return f"{h}s {m:02d}dk" if h else f"{m}dk"

# ── Session state ─────────────────────────────────────────────────────────────
if "duraks"  not in st.session_state:
    st.session_state.duraks = [("İstanbul","Kadıköy"),("Ankara","Çankaya")]
if "km_ov"   not in st.session_state:
    st.session_state.km_ov = [0]
if "fiyatlar" not in st.session_state:
    st.session_state.fiyatlar = {"benzin":64.88,"dizel":55.0,"lpg":25.0,"elektrik":12.0,"kaynak":"","ts":None,"guncelleme":""}

def _sync():
    n = len(st.session_state.duraks)-1
    while len(st.session_state.km_ov) < n: st.session_state.km_ov.append(0)
    while len(st.session_state.km_ov) > n: st.session_state.km_ov.pop()
_sync()

# ── Yakıt fiyatlarını JSON'dan oku ───────────────────────────────────────────
def fiyat_json_oku():
    """fiyatlar.json dosyasından fiyatları okur."""
    try:
        import json as _json
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fiyatlar.json")
        with open(json_path, encoding="utf-8") as f:
            data = _json.load(f)
        st.session_state.fiyatlar["benzin"]    = float(data.get("benzin",   64.88))
        st.session_state.fiyatlar["dizel"]     = float(data.get("dizel",    55.00))
        st.session_state.fiyatlar["lpg"]       = float(data.get("lpg",      25.00))
        st.session_state.fiyatlar["elektrik"]  = float(data.get("elektrik", 12.00))
        st.session_state.fiyatlar["kaynak"]    = data.get("kaynak", "")
        st.session_state.fiyatlar["guncelleme"]= data.get("guncelleme", "")
        st.session_state.fiyatlar["ts"]        = data.get("guncelleme", "")
        return True
    except FileNotFoundError:
        return False
    except Exception as e:
        st.sidebar.warning(f"fiyatlar.json okunamadı: {e}")
        return False

# Uygulama ilk açıldığında JSON'dan otomatik yükle
if "json_yuklendi" not in st.session_state:
    fiyat_json_oku()
    st.session_state.json_yuklendi = True

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Güzergah ──────────────────────────────────────────────────────────────
    st.markdown("### 📍 Güzergah")
    to_remove = None
    n_d = len(st.session_state.duraks)
    for i in range(n_d):
        il, ilce = st.session_state.duraks[i]
        lbl = "🚩 Kalkış" if i==0 else ("🏁 Varış" if i==n_d-1 else f"📍 {i}. Durak")
        c1,c2 = st.columns([5,1])
        with c1:
            new_il   = st.selectbox(lbl, IL_LIST,
                index=IL_LIST.index(il) if il in IL_LIST else 0, key=f"il_{i}")
            ilceler  = ilce_listesi(new_il)
            safe_ilce= ilce if (new_il==il and ilce in ilceler) else ilceler[0]
            new_ilce = st.selectbox("", ilceler,
                index=ilceler.index(safe_ilce), key=f"ilce_{i}", label_visibility="collapsed")
            st.session_state.duraks[i] = (new_il, new_ilce)
        with c2:
            st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)
            if n_d > 2 and st.button("✕", key=f"del_{i}"):
                to_remove = i
        if i < n_d-1:
            a_il,a_ilce = st.session_state.duraks[i]
            b_il,b_ilce = st.session_state.duraks[i+1]
            auto_km = haversine_km(koordinat(a_il,a_ilce), koordinat(b_il,b_ilce))
            ov = st.number_input(f"↕ Segment km (0 = tahmini {auto_km} km)",
                value=int(st.session_state.km_ov[i]), min_value=0, step=1,
                key=f"km_ov_{i}", help="Google Maps'teki gerçek km'yi gir, 0 = otomatik")
            st.session_state.km_ov[i] = ov

    if to_remove is not None:
        st.session_state.duraks.pop(to_remove); _sync(); st.rerun()

    ca, cb = st.columns(2)
    with ca:
        if st.button("＋ Durak", use_container_width=True):
            st.session_state.duraks.append(("Ankara", ilce_listesi("Ankara")[0]))
            _sync(); st.rerun()
    with cb:
        if st.button("↺ Sıfırla", use_container_width=True):
            st.session_state.duraks=[("İstanbul","Kadıköy"),("Ankara","Çankaya")]
            st.session_state.km_ov=[0]; st.rerun()

    segments = []
    for i in range(len(st.session_state.duraks)-1):
        a_il,a_ilce = st.session_state.duraks[i]
        b_il,b_ilce = st.session_state.duraks[i+1]
        ov = st.session_state.km_ov[i]
        km = ov if ov>0 else haversine_km(koordinat(a_il,a_ilce), koordinat(b_il,b_ilce))
        src = "📍" if ov>0 else "📐"
        segments.append((f"{a_il}/{a_ilce}", f"{b_il}/{b_ilce}", km, src))
    total_yol = sum(s[2] for s in segments)

    if segments:
        segs_html = "".join(
            f'<div class="route-seg">{src} {a} → {b}<span class="route-seg-km">{km} km</span></div>'
            for a,b,km,src in segments)
        st.markdown(f"""<div class="route-strip">
            <div class="route-strip-title">Güzergah Özeti</div>
            {segs_html}
            <div class="route-total"><span>Toplam</span><span>{total_yol} km</span></div>
        </div>""", unsafe_allow_html=True)
        st.caption("📍 Manuel  ·  📐 Tahmini (haversine ×1.30)")
    yol = max(total_yol, 1)

    # ── Mola ──────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### ⏱ Mola Ayarları")
    mola_araligi = st.number_input("Kaç saatte bir mola", value=2, min_value=1, max_value=6, step=1)
    mola_suresi  = st.number_input("Mola süresi (dk)",    value=15, min_value=5, max_value=120, step=5)

    # ── Araç ──────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🚗 Araç Bilgileri")
    hiz      = st.number_input("Ortalama Hız (km/s)", value=90, min_value=1, max_value=300)
    depo     = st.number_input("Depo Kapasitesi (L)",  value=53.0, min_value=1.0, step=0.5, format="%.1f")
    tuketim  = st.number_input("100km Tüketim (L)",    value=7.4,  min_value=0.1, step=0.1, format="%.1f")
    cikis    = st.time_input("Çıkış Saati", value=time(8,0))
    kisi     = st.number_input("Kişi Sayısı", value=1, min_value=1, max_value=20)
    aylik_km = st.number_input("Aylık Ortalama km", value=2000, min_value=0, step=100)

    # ── Hava & Yük Etkisi ─────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🌡 Hava & Yük Etkisi")
    sicaklik = st.slider("Hava Sıcaklığı (°C)", -20, 45, 20)
    yuk_kg   = st.slider("Araç Yükü / Bagaj (kg)", 0, 500, 0, step=25)

    if    sicaklik <  0: hava_kat = 1.15; hava_acik = "+15% (çok soğuk)"
    elif  sicaklik < 10: hava_kat = 1.08; hava_acik = "+8% (soğuk)"
    elif  sicaklik < 25: hava_kat = 1.00; hava_acik = "Etki yok (ideal)"
    elif  sicaklik < 35: hava_kat = 1.05; hava_acik = "+5% (sıcak)"
    else:                hava_kat = 1.08; hava_acik = "+8% (çok sıcak)"

    yuk_ek       = yuk_kg / 100 * 0.3        # her 100 kg → +0.3 L/100km
    tuketim_d    = tuketim * hava_kat + yuk_ek
    duzeltme_pct = (tuketim_d / tuketim - 1) * 100

    st.markdown(f"""<div class="etki-box">
    🌡 Hava: {hava_acik}<br>
    ⚖ Yük: +{yuk_ek:.1f} L/100km<br>
    <b>Düzeltilmiş tüketim: {tuketim:.1f} → {tuketim_d:.1f} L/100km ({duzeltme_pct:+.0f}%)</b>
    </div>""", unsafe_allow_html=True)

    # ── Yakıt Fiyatları ───────────────────────────────────────────────────────
    st.divider()
    st.markdown("### ⛽ Yakıt Fiyatları")

    col_j1, col_j2 = st.columns(2)
    with col_j1:
        if st.button("📂 JSON'dan Yükle", use_container_width=True,
                     help="fiyatlar.json dosyasındaki fiyatları uygular"):
            if fiyat_json_oku():
                st.success("Yüklendi!")
                st.rerun()
            else:
                st.error("fiyatlar.json bulunamadı")
    with col_j2:
        if st.session_state.fiyatlar.get("guncelleme"):
            kaynak = st.session_state.fiyatlar.get("kaynak","")
            st.caption(f"📅 {st.session_state.fiyatlar['guncelleme']}" +
                       (f"\n📌 {kaynak}" if kaynak else ""))

    fiyat  = st.number_input("Benzin (TL/L)",  value=st.session_state.fiyatlar["benzin"],
                              min_value=0.01, step=0.5, format="%.2f", key="inp_benzin")
    st.session_state.fiyatlar["benzin"] = fiyat

    dizel_fiyat   = st.number_input("Dizel (TL/L)",    value=st.session_state.fiyatlar["dizel"],
                                     min_value=0.01, step=0.5, format="%.2f", key="inp_dizel")
    st.session_state.fiyatlar["dizel"] = dizel_fiyat

    lpg_fiyat     = st.number_input("LPG (TL/L)",      value=st.session_state.fiyatlar["lpg"],
                                     min_value=0.01, step=0.5, format="%.2f", key="inp_lpg")
    st.session_state.fiyatlar["lpg"] = lpg_fiyat

    dizel_tuketim = st.number_input("Dizel 100km (L)", value=6.5,  min_value=0.1, step=0.1, format="%.1f")
    lpg_tuketim   = st.number_input("LPG 100km (L)",   value=11.0, min_value=0.1, step=0.1, format="%.1f")

    st.divider()
    st.markdown("### ⚡ Elektrikli Araç (EV)")
    ev_fiyat   = st.number_input("Elektrik (TL/kWh)", value=st.session_state.fiyatlar.get("elektrik", 12.0), min_value=0.01, step=0.5, format="%.2f", key="inp_elektrik")
    st.session_state.fiyatlar["elektrik"] = ev_fiyat
    ev_tuketim = st.number_input("EV 100km (kWh)",    value=20.0,  min_value=0.1,  step=0.1, format="%.1f")
    ev_batarya = st.number_input("Batarya (kWh)",     value=75.0,  min_value=1.0,  step=1.0, format="%.1f")

# ── Hesaplamalar ─────────────────────────────────────────────────────────────
# Düzeltilmiş tüketim kullan
tuk = tuketim_d  # hava + yük etkili değer

km_yakit         = tuk / 100
km_maliyet       = km_yakit * fiyat
yol_yakiti       = yol * km_yakit
yol_maliyeti     = yol_yakiti * fiyat
tam_depo_menzil  = (depo / tuk) * 100
tam_depo_maliyet = depo * fiyat
km100_maliyet    = tuk * fiyat
gerekli_depo     = math.ceil(yol / tam_depo_menzil)
depoda_kalan     = depo - yol_yakiti
gidis_donus      = yol_maliyeti * 2
kisi_basi        = gidis_donus / kisi

sure_saat   = yol / hiz
mola_sayisi = int(sure_saat / mola_araligi)
mola_dakika = mola_sayisi * mola_suresi
molali_sure = sure_saat + mola_dakika / 60
varis_dt    = datetime.combine(datetime.today(), cikis) + timedelta(hours=molali_sure)
varis_str   = varis_dt.strftime("%H:%M")
sure_str    = fmt_sure(molali_sure)
surus_str   = fmt_sure(sure_saat)
mola_str    = f"{mola_dakika}dk" + (f" ({mola_sayisi}×{mola_suresi}dk)" if mola_sayisi>0 else " (mola yok)")

aylik_gider   = aylik_km * km_maliyet
yillik_gider  = aylik_gider * 12
yakit_yeterli = yol <= tam_depo_menzil

def hesapla(km, fp, tuk_):
    m = km * tuk_ / 100; return m * fp, m

bnz_m, bnz_y = hesapla(yol, fiyat, tuk)
diz_m, diz_y = hesapla(yol, dizel_fiyat, dizel_tuketim * (tuk/tuketim))
lpg_m, lpg_y = hesapla(yol, lpg_fiyat,   lpg_tuketim   * (tuk/tuketim))
ev_enerji    = yol * ev_tuketim / 100
ev_m         = ev_enerji * ev_fiyat
ev_menzil    = (ev_batarya / ev_tuketim) * 100 if ev_tuketim > 0 else 0
ev_tas       = bnz_m - ev_m
ev_aylik_tas = (aylik_km * ev_tas / yol) if yol > 0 else 0
ev_yillik_tas= ev_aylik_tas * 12
ev_oran      = bnz_m / ev_m if ev_m > 0 else 0

def seg_sure_h(km):
    h = km / hiz
    return h + int(h / mola_araligi) * mola_suresi / 60

# Segment saatleri
def build_seg_rows():
    cur = datetime.combine(datetime.today(), cikis)
    rows = []
    for a,b,km,src in segments:
        dep = cur.strftime("%H:%M")
        sh  = seg_sure_h(km)
        arr_dt = cur + timedelta(hours=sh)
        arr = arr_dt.strftime("%H:%M")
        mol_dk = int((km/hiz)/mola_araligi) * mola_suresi
        rows.append({"Segment":f"{a} → {b}", "Mesafe":f"{km} km ({src})",
                     "Çıkış":dep, "Varış":arr,
                     "Sürüş":fmt_sure(km/hiz), "Mola":f"{mol_dk}dk" if mol_dk else "—",
                     "Toplam":fmt_sure(sh), "Yakıt Mal.":f"{km*km_maliyet:,.0f} ₺"})
        cur = arr_dt
    return rows

# Yakıt dolum noktaları (hangi duraklar)
def yakit_durak_indeksleri():
    idxs = []; kum = 0; sonraki = tam_depo_menzil
    for i,(a,b,km,src) in enumerate(segments):
        if kum + km > sonraki:
            idxs.append(i); sonraki += tam_depo_menzil
        kum += km
    return idxs

# ── Sayfa ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="page-title">🚗 Araç Yol Bilgisayarı</div>', unsafe_allow_html=True)
rl = " → ".join(f"{il}/{ilce}" for il,ilce in st.session_state.duraks)
st.markdown(f'<div class="page-sub">📍 {rl} &nbsp;·&nbsp; {total_yol} km</div>', unsafe_allow_html=True)

# Durum
if yakit_yeterli:
    st.markdown(f"""<div class="status-ok"><div class="status-icon">✅</div>
      <div><div>YAKIT YETERLİ</div>
      <div class="status-detail">Menzil {tam_depo_menzil:.0f} km · Yol {yol} km · Depoda {depoda_kalan:.1f} L kalır · Tüketim {tuk:.1f} L/100km</div>
      </div></div>""", unsafe_allow_html=True)
else:
    st.markdown(f"""<div class="status-fail"><div class="status-icon">⛽</div>
      <div><div>YAKIT YETMEZ — {gerekli_depo} depo gerekli</div>
      <div class="status-detail">Menzil {tam_depo_menzil:.0f} km · Yol {yol} km · {abs(depoda_kalan):.1f} L eksik · Tüketim {tuk:.1f} L/100km</div>
      </div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Ana metrikler (tablo dışında — her tab'da görünür)
r1c1,r1c2,r1c3,r1c4 = st.columns(4)
r1c1.metric("Yol Maliyeti",    f"{yol_maliyeti:,.0f} ₺",  f"{km_maliyet:.2f} ₺/km", delta_color="off")
r1c2.metric("Gidiş-Dönüş",    f"{gidis_donus:,.0f} ₺")
r1c3.metric("Kişi Başı (G/D)", f"{kisi_basi:,.0f} ₺",    f"{kisi} kişi",            delta_color="off")
r1c4.metric("Tahmini Varış",   varis_str, f"Çıkış {cikis.strftime('%H:%M')} · {sure_str}", delta_color="off")
st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# Tabs
tab_ozet, tab_harita, tab_rapor = st.tabs(["📊 Özet", "🗺 Harita", "📄 Rapor"])

# ═══════════════════════ ÖZET ═════════════════════════════════════════════════
with tab_ozet:
    r2c1,r2c2,r2c3,r2c4 = st.columns(4)
    r2c1.metric("Tam Depo Menzili",   f"{tam_depo_menzil:.0f} km")
    r2c2.metric("Tam Depo Maliyeti",  f"{tam_depo_maliyet:,.0f} ₺")
    r2c3.metric("Aylık Yakıt Gideri", f"{aylik_gider:,.0f} ₺", f"Yıllık {yillik_gider:,.0f} ₺", delta_color="off")
    r2c4.metric("100km Maliyeti",     f"{km100_maliyet:,.0f} ₺", f"Tüketim {tuk:.1f} L", delta_color="off")

    st.markdown('<div class="section-header">⏱ Süre Bilgileri</div>', unsafe_allow_html=True)
    t1,t2,t3,t4 = st.columns(4)
    t1.metric("Sürüş Süresi",  surus_str, f"{yol} km / {hiz} km/s", delta_color="off")
    t2.metric("Mola Süresi",   mola_str,  f"Her {mola_araligi}s'de {mola_suresi}dk", delta_color="off")
    t3.metric("Toplam Süre",   sure_str,  "Sürüş + mola", delta_color="off")
    t4.metric("Varış",         varis_str, f"Çıkış: {cikis.strftime('%H:%M')}", delta_color="off")

    if len(segments) > 1:
        st.markdown('<div class="section-header">📍 Güzergah Detayı</div>', unsafe_allow_html=True)
        seg_df = pd.DataFrame(build_seg_rows())
        st.dataframe(seg_df, hide_index=True, use_container_width=True)

    st.markdown('<div class="section-header">⛽ Yakıt Tipi Karşılaştırması</div>', unsafe_allow_html=True)

    def fmt_tas(v):
        if v == 0: return "—"
        return f"{'+'if v>0 else ''}{v:,.0f} ₺"

    dizel_tuk_d = dizel_tuketim * (tuk/tuketim)
    lpg_tuk_d   = lpg_tuketim   * (tuk/tuketim)

    comp = pd.DataFrame([
        {"Yakıt":"Benzin",           "Fiyat":f"{fiyat:.2f} ₺/L",      "100km":f"{tuk:.1f} L",          "Yol Yakıtı":f"{bnz_y:.1f} L",   "Yol Mal.":f"{bnz_m:,.0f} ₺",  "G/D":f"{bnz_m*2:,.0f} ₺",  "Tasarruf":"—"},
        {"Yakıt":"Dizel",            "Fiyat":f"{dizel_fiyat:.2f} ₺/L", "100km":f"{dizel_tuk_d:.1f} L",  "Yol Yakıtı":f"{diz_y:.1f} L",   "Yol Mal.":f"{diz_m:,.0f} ₺",  "G/D":f"{diz_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-diz_m)},
        {"Yakıt":"LPG",              "Fiyat":f"{lpg_fiyat:.2f} ₺/L",   "100km":f"{lpg_tuk_d:.1f} L",    "Yol Yakıtı":f"{lpg_y:.1f} L",   "Yol Mal.":f"{lpg_m:,.0f} ₺",  "G/D":f"{lpg_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-lpg_m)},
        {"Yakıt":"⚡ EV",            "Fiyat":f"{ev_fiyat:.2f} ₺/kWh",  "100km":f"{ev_tuketim:.1f} kWh", "Yol Yakıtı":f"{ev_enerji:.1f} kWh","Yol Mal.":f"{ev_m:,.0f} ₺", "G/D":f"{ev_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-ev_m)},
    ])
    st.dataframe(comp, hide_index=True, use_container_width=True)

    st.markdown('<div class="section-header">📊 Maliyet Grafiği</div>', unsafe_allow_html=True)
    fig = go.Figure(go.Bar(
        x=["Benzin","Dizel","LPG","⚡ EV"], y=[bnz_m,diz_m,lpg_m,ev_m],
        marker_color=["#ED7D31","#2E75B6","#70AD47","#7030A0"],
        text=[f"{v:,.0f} ₺" for v in [bnz_m,diz_m,lpg_m,ev_m]],
        textposition="outside", textfont=dict(size=13,family="Inter"),
    ))
    fig.add_hline(y=bnz_m, line_dash="dot", line_color="#ED7D31",
                  annotation_text=f"Benzin: {bnz_m:,.0f} ₺",
                  annotation_position="top right",
                  annotation_font=dict(color="#ED7D31",size=11))
    fig.update_layout(plot_bgcolor="white", paper_bgcolor="white",
        yaxis=dict(title=dict(text="Yol Maliyeti (₺)",font=dict(size=12)), gridcolor="#F0F0F0"),
        xaxis=dict(tickfont=dict(size=13)), margin=dict(t=20,b=20,l=20,r=20), height=300, showlegend=False)
    st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="section-header">⚡ EV Tasarruf Analizi</div>', unsafe_allow_html=True)
    e1,e2,e3,e4 = st.columns(4)
    e1.metric("Bu Seyahat",    f"{'+'if ev_tas>=0 else ''}{ev_tas:,.0f} ₺",         "Benzin'e kıyasla",  delta_color="off")
    e2.metric("Aylık Tasarruf",f"{'+'if ev_aylik_tas>=0 else ''}{ev_aylik_tas:,.0f} ₺", f"{aylik_km:,} km/ay",delta_color="off")
    e3.metric("Yıllık Tasarruf",f"{'+'if ev_yillik_tas>=0 else ''}{ev_yillik_tas:,.0f} ₺")
    e4.metric("Benzin/EV Oran",f"{ev_oran:.2f}×","EV ucuz" if ev_oran>1 else "EV pahalı",delta_color="off")

# ═══════════════════════ HARİTA ═══════════════════════════════════════════════
with tab_harita:
    stop_coords = [koordinat(il,ilce) for il,ilce in st.session_state.duraks]
    center_lat  = sum(c[0] for c in stop_coords) / len(stop_coords)
    center_lon  = sum(c[1] for c in stop_coords) / len(stop_coords)

    m = folium.Map(location=[center_lat,center_lon], zoom_start=6,
                   tiles="CartoDB positron")

    # Güzergah çizgisi
    folium.PolyLine(stop_coords, weight=3, color="#1F4E78", opacity=0.8,
                    dash_array=None).add_to(m)

    # Durak marker'ları
    seg_rows = build_seg_rows() if len(segments) > 0 else []
    for i,(il,ilce) in enumerate(st.session_state.duraks):
        lat, lon = koordinat(il, ilce)
        if i == 0:
            color, icon_name, label = "green",  "play",      "Kalkış"
        elif i == len(st.session_state.duraks)-1:
            color, icon_name, label = "red",    "stop",      "Varış"
        else:
            color, icon_name, label = "blue",   "info-sign", f"{i}. Durak"

        # Çıkış/varış saatini bul
        if i < len(seg_rows):
            zaman = f"Çıkış: {seg_rows[i]['Çıkış']}"
        elif seg_rows:
            zaman = f"Varış: {seg_rows[-1]['Varış']}"
        else:
            zaman = ""

        popup_html = f"""
        <b>{label}</b><br>
        {il} / {ilce}<br>
        {zaman}
        """
        folium.Marker([lat,lon],
            popup=folium.Popup(popup_html, max_width=180),
            tooltip=f"{label}: {il}/{ilce}",
            icon=folium.Icon(color=color, icon=icon_name)
        ).add_to(m)

    # Yakıt dolum noktaları
    yakit_idxler = yakit_durak_indeksleri()
    for idx in yakit_idxler:
        # Bu segmentin başındaki durak
        durak_idx = idx
        if durak_idx < len(st.session_state.duraks):
            il, ilce = st.session_state.duraks[durak_idx]
            lat, lon = koordinat(il, ilce)
            kum_km   = sum(s[2] for s in segments[:idx])
            folium.Marker([lat,lon],
                popup=folium.Popup(
                    f"<b>⛽ Yakıt Durağı</b><br>{il}/{ilce}<br>Yaklaşık {kum_km} km'de", max_width=160),
                tooltip=f"⛽ Yakıt al: {il}/{ilce}",
                icon=folium.Icon(color="orange", icon="tint", prefix="fa")
            ).add_to(m)

    col_map, col_info = st.columns([3, 1])
    with col_map:
        st_folium(m, height=520, use_container_width=True)
    with col_info:
        st.markdown("**Harita Göstergesi**")
        st.markdown("🟢 Kalkış noktası")
        st.markdown("🔵 Ara duraklar")
        st.markdown("🔴 Varış noktası")
        st.markdown("🟠 Yakıt dolum noktası")
        st.divider()
        st.markdown(f"**Toplam:** {total_yol} km")
        st.markdown(f"**Süre:** {sure_str}")
        st.markdown(f"**Varış:** {varis_str}")
        if yakit_idxler:
            st.divider()
            st.markdown(f"**Yakıt durağı:** {len(yakit_idxler)} kez")
            st.markdown(f"Menzil: {tam_depo_menzil:.0f} km")
        else:
            st.divider()
            st.markdown("✅ Tek depoda gidilir")
            st.markdown(f"Kalacak: {depoda_kalan:.1f} L")

# ═══════════════════════ RAPOR ════════════════════════════════════════════════
with tab_rapor:
    st.markdown('<div class="section-header">📄 Rapor Oluştur</div>', unsafe_allow_html=True)
    st.markdown("Seyahat planını PDF veya Excel olarak indir.")

    tarih = datetime.now().strftime("%d.%m.%Y")
    rota  = " → ".join(f"{il}/{ilce}" for il,ilce in st.session_state.duraks)

    # ── PDF ───────────────────────────────────────────────────────────────────
    def olustur_pdf():
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        stiller = getSampleStyleSheet()
        baslik_stili  = ParagraphStyle("B", parent=stiller["Heading1"],
            fontSize=16, textColor=colors.HexColor("#1F4E78"), spaceAfter=4)
        alt_stili     = ParagraphStyle("A", parent=stiller["Normal"],
            fontSize=9, textColor=colors.grey, spaceAfter=12)
        bolum_stili   = ParagraphStyle("S", parent=stiller["Heading2"],
            fontSize=11, textColor=colors.HexColor("#1F4E78"), spaceBefore=14, spaceAfter=6)

        def tablo(veriler, col_w=None):
            t = Table(veriler, colWidths=col_w)
            t.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F0F4F8")]),
                ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
                ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",   (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ])); return t

        story = [
            Paragraph("🚗 Araç Yol Bilgisayarı — Seyahat Raporu", baslik_stili),
            Paragraph(f"{rota}  ·  {total_yol} km  ·  {tarih}", alt_stili),
            Paragraph("Özet Bilgiler", bolum_stili),
            tablo([
                ["Yol","Gidiş-Dönüş","Kişi Başı","Toplam Süre","Varış"],
                [f"{yol} km", f"{gidis_donus:,.0f} ₺", f"{kisi_basi:,.0f} ₺", sure_str, varis_str],
            ], [3*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]),
            Spacer(1,0.3*cm),
            tablo([
                ["Menzil","Tüketim (düz.)","Yakıt Fiyatı","Aylık Gider","Yıllık Gider"],
                [f"{tam_depo_menzil:.0f} km", f"{tuk:.1f} L/100km",
                 f"{fiyat:.2f} ₺/L", f"{aylik_gider:,.0f} ₺", f"{yillik_gider:,.0f} ₺"],
            ], [3.5*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]),
        ]

        if len(segments) > 1:
            story.append(Paragraph("Güzergah Detayı", bolum_stili))
            rows = build_seg_rows()
            tablo_v = [["Segment","Mesafe","Çıkış","Varış","Süre","Yakıt"]]
            for r in rows:
                tablo_v.append([r["Segment"], r["Mesafe"], r["Çıkış"], r["Varış"],
                                r["Toplam"], r["Yakıt Mal."]])
            story.append(tablo([tablo_v[0]] + tablo_v[1:],
                [6.5*cm,2.5*cm,1.8*cm,1.8*cm,2*cm,2.4*cm]))

        story.append(Paragraph("Yakıt Karşılaştırması", bolum_stili))
        story.append(tablo([
            ["Yakıt","Fiyat","Yol Maliyeti","G/D Toplam","Tasarruf"],
            ["Benzin",  f"{fiyat:.2f} ₺/L",      f"{bnz_m:,.0f} ₺", f"{bnz_m*2:,.0f} ₺", "—"],
            ["Dizel",   f"{dizel_fiyat:.2f} ₺/L", f"{diz_m:,.0f} ₺", f"{diz_m*2:,.0f} ₺", fmt_tas(bnz_m-diz_m)],
            ["LPG",     f"{lpg_fiyat:.2f} ₺/L",   f"{lpg_m:,.0f} ₺", f"{lpg_m*2:,.0f} ₺", fmt_tas(bnz_m-lpg_m)],
            ["EV",      f"{ev_fiyat:.2f} ₺/kWh",  f"{ev_m:,.0f} ₺",  f"{ev_m*2:,.0f} ₺",  fmt_tas(bnz_m-ev_m)],
        ], [3.5*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]))

        if duzeltme_pct != 0:
            story.append(Spacer(1,0.3*cm))
            story.append(Paragraph(
                f"ℹ Tüketim {sicaklik}°C hava ve {yuk_kg}kg yük için düzeltilmiştir "
                f"({tuketim:.1f} → {tuk:.1f} L/100km, {duzeltme_pct:+.0f}%).", alt_stili))

        doc.build(story)
        buf.seek(0); return buf

    # ── Excel ─────────────────────────────────────────────────────────────────
    def olustur_excel():
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Seyahat Raporu"
        DARK  = "1F4E78"; WHT = "FFFFFF"; LBLU = "DEEAF1"; YELL = "FFF2CC"
        def hd(cell, val, bg=DARK, fc=WHT, bold=True, sz=11):
            cell.value = val
            cell.font  = Font(bold=bold, color=fc, size=sz, name="Calibri")
            cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=s,right=s,top=s,bottom=s)
        def vl(cell, val, bg=LBLU, bold=False):
            cell.value = val
            cell.font  = Font(bold=bold, color="000000", size=10, name="Calibri")
            cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=s,right=s,top=s,bottom=s)

        r = 1
        ws.merge_cells(f"A{r}:G{r}")
        hd(ws[f"A{r}"], f"ARAÇ YOL BİLGİSAYARI — {tarih}", sz=13)
        ws.row_dimensions[r].height = 24; r += 1
        ws.merge_cells(f"A{r}:G{r}")
        hd(ws[f"A{r}"], rota, bg="2E4057", sz=10)
        ws.row_dimensions[r].height = 18; r += 2

        hdrs = ["Yol","Gidiş-Dönüş","Kişi Başı","Toplam Süre","Menzil","Tüketim","Varış"]
        vals = [f"{yol} km", f"{gidis_donus:,.0f} ₺", f"{kisi_basi:,.0f} ₺",
                sure_str, f"{tam_depo_menzil:.0f} km", f"{tuk:.1f} L/100km", varis_str]
        for j,(h,v) in enumerate(zip(hdrs,vals),1):
            hd(ws.cell(r,j), h); vl(ws.cell(r+1,j), v)
        ws.row_dimensions[r].height=18; ws.row_dimensions[r+1].height=18; r+=3

        if len(segments)>1:
            seg_hdrs=["Segment","Mesafe","Çıkış","Varış","Sürüş","Mola","Toplam","Yakıt"]
            for j,h in enumerate(seg_hdrs,1): hd(ws.cell(r,j),h)
            ws.row_dimensions[r].height=18; r+=1
            for row in build_seg_rows():
                for j,k in enumerate(["Segment","Mesafe","Çıkış","Varış","Sürüş","Mola","Toplam","Yakıt Mal."],1):
                    bg="FFFFFF" if r%2==0 else "F5F8FC"
                    vl(ws.cell(r,j), row.get(k,""), bg=bg)
                ws.row_dimensions[r].height=16; r+=1
            r+=1

        yakit_hdrs=["Yakıt","Fiyat","Yol Maliyeti","G/D Toplam","Tasarruf"]
        for j,h in enumerate(yakit_hdrs,1): hd(ws.cell(r,j),h)
        ws.row_dimensions[r].height=18; r+=1
        for yakdata in [
            ("Benzin",  f"{fiyat:.2f}₺/L",      f"{bnz_m:,.0f}₺", f"{bnz_m*2:,.0f}₺","—"),
            ("Dizel",   f"{dizel_fiyat:.2f}₺/L", f"{diz_m:,.0f}₺", f"{diz_m*2:,.0f}₺",fmt_tas(bnz_m-diz_m)),
            ("LPG",     f"{lpg_fiyat:.2f}₺/L",   f"{lpg_m:,.0f}₺", f"{lpg_m*2:,.0f}₺",fmt_tas(bnz_m-lpg_m)),
            ("⚡ EV",   f"{ev_fiyat:.2f}₺/kWh",  f"{ev_m:,.0f}₺",  f"{ev_m*2:,.0f}₺", fmt_tas(bnz_m-ev_m)),
        ]:
            bg = "FFFFFF" if r%2==0 else "F5F8FC"
            for j,v in enumerate(yakdata,1): vl(ws.cell(r,j),v,bg=bg)
            ws.row_dimensions[r].height=16; r+=1

        for col in ws.columns:
            col_letter = None
            for cell in col:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
            mx = max((len(str(c.value or "")) for c in col if hasattr(c, "value")), default=10)
            ws.column_dimensions[col_letter].width = min(mx + 3, 30)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    # İndirme butonları
    dc1, dc2, dc3 = st.columns([1,1,2])
    with dc1:
        try:
            pdf_buf = olustur_pdf()
            st.download_button("📥 PDF İndir", data=pdf_buf,
                file_name=f"seyahat_{tarih.replace('.','')}.pdf",
                mime="application/pdf", use_container_width=True)
        except Exception as e:
            st.error(f"PDF hatası: {e}")
    with dc2:
        try:
            xl_buf = olustur_excel()
            st.download_button("📥 Excel İndir", data=xl_buf,
                file_name=f"seyahat_{tarih.replace('.','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.error(f"Excel hatası: {e}")

    st.divider()
    st.markdown("**Rapor içeriği:**")
    st.markdown("- Özet bilgiler (yol, maliyet, süre, varış)\n"
                "- Güzergah detay tablosu (çıkış/varış saatleri)\n"
                "- Yakıt tipi karşılaştırması (benzin, dizel, LPG, EV)\n"
                "- Hava/yük düzeltme notu")
    st.divider()
    st.caption("⛽ Yakıt fiyatlarını güncellemek için fiyatlar.json dosyasını düzenleyin.")
    st.caption("📐 Mesafeler koordinat tabanlı tahminidir (haversine × 1.30 yol katsayısı). Gerçek mesafe için navigasyon uygulaması kullanın.")
    st.caption("© Yolculuk-Yakıt-Zaman-Masraf Hesaplayıcı | Streamlit + Python | 2026 Enes Özkan")
