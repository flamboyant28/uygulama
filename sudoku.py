"""
Sudoku Çözücü & Üretici
Bağımlılıklar: reportlab, openpyxl
pip install reportlab openpyxl
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import random, copy, time

# ── Renkler ───────────────────────────────────────────────────────────────────
BG       = "#1a1a2e"
BG2      = "#16213e"
BG3      = "#0f3460"
ACCENT   = "#e94560"
ACCENT2  = "#0f9b8e"
TEXT     = "#eaeaea"
TEXT2    = "#a0a0b0"
GIVEN    = "#ffffff"
USER     = "#64b5f6"
ERROR    = "#ef5350"
SUCCESS  = "#66bb6a"
HINT     = "#ffd54f"
BORDER   = "#2a2a4a"

CELL_SIZE = 58
GRID_PAD  = 24

# ── Sudoku Motoru ──────────────────────────────────────────────────────────────
class SudokuEngine:
    @staticmethod
    def is_valid(board, row, col, num):
        # Satır
        if num in board[row]: return False
        # Sütun
        if num in [board[r][col] for r in range(9)]: return False
        # 3x3 kutu
        br, bc = (row // 3) * 3, (col // 3) * 3
        for r in range(br, br+3):
            for c in range(bc, bc+3):
                if board[r][c] == num: return False
        return True

    @staticmethod
    def solve(board, find_all=False):
        """Backtracking çözücü. find_all=True ise birden fazla çözüm arar."""
        solutions = []
        def _solve(b):
            if find_all and len(solutions) > 1:
                return
            for row in range(9):
                for col in range(9):
                    if b[row][col] == 0:
                        nums = list(range(1, 10))
                        random.shuffle(nums)
                        for num in nums:
                            if SudokuEngine.is_valid(b, row, col, num):
                                b[row][col] = num
                                _solve(b)
                                if not find_all and solutions:
                                    return
                                b[row][col] = 0
                        return
            solutions.append(copy.deepcopy(b))
        _solve(board)
        return solutions

    @staticmethod
    def has_unique_solution(board):
        sols = SudokuEngine.solve(copy.deepcopy(board), find_all=True)
        return len(sols) == 1

    @classmethod
    def generate(cls, difficulty="orta"):
        """Geçerli ve benzersiz çözümlü sudoku üret"""
        # Boş tahtayı doldur
        board = [[0]*9 for _ in range(9)]
        sols  = cls.solve(board)
        full  = sols[0]

        # Zorluk → kaç hücre kaldırılacak
        remove_counts = {"kolay": 36, "orta": 46, "zor": 54, "uzman": 58}
        to_remove = remove_counts.get(difficulty, 46)

        puzzle = copy.deepcopy(full)
        cells  = list(range(81))
        random.shuffle(cells)
        removed = 0

        for idx in cells:
            if removed >= to_remove:
                break
            r, c   = divmod(idx, 9)
            backup = puzzle[r][c]
            puzzle[r][c] = 0
            if cls.has_unique_solution(puzzle):
                removed += 1
            else:
                puzzle[r][c] = backup

        return puzzle, full

    @staticmethod
    def get_errors(board, solution):
        errors = set()
        for r in range(9):
            for c in range(9):
                if board[r][c] != 0 and board[r][c] != solution[r][c]:
                    errors.add((r, c))
        return errors

    @staticmethod
    def count_filled(board):
        return sum(1 for r in range(9) for c in range(9) if board[r][c] != 0)

# ── PDF Export ─────────────────────────────────────────────────────────────────
def export_pdf(puzzles_data, filename, per_page=1, show_solution=False):
    """
    puzzles_data: list of (puzzle, solution, difficulty) tuples
    per_page: 1 veya 4
    """
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    W, H = A4  # 595 x 842 pt

    c = rl_canvas.Canvas(filename, pagesize=A4)

    def draw_sudoku(c, board, x0, y0, size, title=""):
        cell = size / 9

        # Başlık
        if title:
            c.setFont("Helvetica-Bold", 11)
            c.setFillColor(colors.HexColor("#333333"))
            c.drawString(x0, y0 + size + 14, title)

        # Arka plan
        c.setFillColor(colors.white)
        c.rect(x0, y0, size, size, fill=1, stroke=0)

        # Hücre değerleri
        for row in range(9):
            for col in range(9):
                val = board[row][col]
                if val != 0:
                    cx = x0 + col * cell + cell/2
                    cy = y0 + (8 - row) * cell + cell/2 - 5
                    c.setFont("Helvetica-Bold", int(cell * 0.45))
                    c.setFillColor(colors.HexColor("#1a1a2e"))
                    c.drawCentredString(cx, cy, str(val))

        # İnce çizgiler (hücre sınırları)
        c.setStrokeColor(colors.HexColor("#aaaaaa"))
        c.setLineWidth(0.5)
        for i in range(10):
            if i % 3 != 0:
                c.line(x0 + i*cell, y0, x0 + i*cell, y0 + size)
                c.line(x0, y0 + i*cell, x0 + size, y0 + i*cell)

        # Kalın çizgiler (3x3 kutu sınırları)
        c.setStrokeColor(colors.HexColor("#1a1a2e"))
        c.setLineWidth(2.5)
        for i in range(0, 10, 3):
            c.line(x0 + i*cell, y0, x0 + i*cell, y0 + size)
            c.line(x0, y0 + i*cell, x0 + size, y0 + i*cell)

    if per_page == 1:
        for puzzle, solution, difficulty in puzzles_data:
            # Bulmaca sayfası
            margin = 72
            size   = W - 2 * margin

            c.setFont("Helvetica-Bold", 18)
            c.setFillColor(colors.HexColor("#1a1a2e"))
            c.drawCentredString(W/2, H - 50, "SUDOKU")
            c.setFont("Helvetica", 11)
            c.setFillColor(colors.HexColor("#666666"))
            c.drawCentredString(W/2, H - 68, f"Zorluk: {difficulty.capitalize()}")

            y0 = (H - size) / 2
            draw_sudoku(c, puzzle, margin, y0, size)

            # Alt bilgi
            filled = SudokuEngine.count_filled(puzzle)
            c.setFont("Helvetica", 9)
            c.setFillColor(colors.HexColor("#999999"))
            c.drawCentredString(W/2, 30, f"Doldurulacak: {81 - filled} hücre")
            c.showPage()

            # Çözüm sayfası
            if show_solution:
                c.setFont("Helvetica-Bold", 18)
                c.setFillColor(colors.HexColor("#1a1a2e"))
                c.drawCentredString(W/2, H - 50, "ÇÖZÜM")
                c.setFont("Helvetica", 11)
                c.setFillColor(colors.HexColor("#666666"))
                c.drawCentredString(W/2, H - 68, f"Zorluk: {difficulty.capitalize()}")
                draw_sudoku(c, solution, margin, y0, size)
                c.showPage()

    elif per_page == 4:
        # Her sayfaya 4 bulmaca (2x2)
        for page_start in range(0, len(puzzles_data), 4):
            batch = puzzles_data[page_start:page_start+4]

            # Başlık
            c.setFont("Helvetica-Bold", 14)
            c.setFillColor(colors.HexColor("#1a1a2e"))
            c.drawCentredString(W/2, H - 35, "SUDOKU")

            margin  = 30
            gap     = 20
            size    = (W - 2*margin - gap) / 2
            positions = [
                (margin,            H/2 + gap/2),
                (margin + size + gap, H/2 + gap/2),
                (margin,            H/2 - size - gap/2),
                (margin + size + gap, H/2 - size - gap/2),
            ]

            for i, (puzzle, solution, difficulty) in enumerate(batch):
                x0, y0 = positions[i]
                draw_sudoku(c, puzzle, x0, y0, size,
                            title=f"#{page_start+i+1} — {difficulty.capitalize()}")

            c.showPage()

            # Çözüm sayfası
            if show_solution:
                c.setFont("Helvetica-Bold", 14)
                c.setFillColor(colors.HexColor("#1a1a2e"))
                c.drawCentredString(W/2, H - 35, "ÇÖZÜMLER")

                for i, (puzzle, solution, difficulty) in enumerate(batch):
                    x0, y0 = positions[i]
                    draw_sudoku(c, solution, x0, y0, size,
                                title=f"#{page_start+i+1} Çözüm")
                c.showPage()

    c.save()

# ── Excel Export ───────────────────────────────────────────────────────────────
def export_excel(puzzles_data, filename, show_solution=False):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thick = Side(style="medium", color="1a1a2e")
    thin  = Side(style="thin",   color="aaaaaa")

    def get_border(row, col):
        top    = thick if row % 3 == 0 else thin
        left   = thick if col % 3 == 0 else thin
        bottom = thick if row == 8 else (thick if (row+1) % 3 == 0 else thin)
        right  = thick if col == 8 else (thick if (col+1) % 3 == 0 else thin)
        return Border(top=top, left=left, bottom=bottom, right=right)

    def write_board(ws, board, given_board, start_row, start_col, col_w=4.0, row_h=22):
        for r in range(9):
            for c in range(9):
                cell = ws.cell(row=start_row+r, column=start_col+c)
                val  = board[r][c]
                cell.value     = val if val != 0 else None
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = get_border(r, c)
                if given_board and given_board[r][c] != 0:
                    cell.font = Font(name="Calibri", size=11, bold=True,  color="1a1a2e")
                    cell.fill = PatternFill("solid", fgColor="EFEFEF")
                elif val != 0:
                    cell.font = Font(name="Calibri", size=11, bold=False, color="1f6feb")
                else:
                    cell.font = Font(name="Calibri", size=11)
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
        for c in range(9):
            ws.column_dimensions[get_column_letter(start_col+c)].width = col_w
        for r in range(9):
            ws.row_dimensions[start_row+r].height = row_h

    # 2 sütun × 3 satır = 6 bulmaca/sayfa
    # Sütun başlangıçları: 1 ve 11 (9 sütun + 1 boşluk)
    # Satır başlangıçları: 3, 14, 25 (1 başlık + 1 satır boşluk aralıklı)
    PER_PAGE    = 6
    COL_OFFSETS = [1, 11]
    ROW_OFFSETS = [3, 14, 25]

    # slot_idx → (row_grp, col_grp)
    POSITIONS = [(r, c) for r in range(3) for c in range(2)]  # 6 slot

    def make_sheet(ws, batch, page_start, is_solution=False):
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 20

        # Sayfa başlığı
        ws.merge_cells("A1:S1")
        t = ws.cell(row=1, column=1,
                    value="ÇÖZÜMLER" if is_solution else
                    f"SUDOKU — {batch[0][2].upper()}")
        t.font      = Font(name="Calibri", size=13, bold=True,
                           color="0f9b8e" if is_solution else "1a1a2e")
        t.alignment = Alignment(horizontal="center")

        for slot_idx, (puzzle, solution, difficulty) in enumerate(batch):
            ri, ci    = POSITIONS[slot_idx]
            start_col = COL_OFFSETS[ci]
            start_row = ROW_OFFSETS[ri]
            num       = page_start + slot_idx + 1
            board     = solution if is_solution else puzzle
            given     = puzzle   if is_solution else None

            # Bulmaca başlığı
            hdr_row = start_row - 1
            ws.merge_cells(start_row=hdr_row, start_column=start_col,
                           end_row=hdr_row,   end_column=start_col+8)
            hdr = ws.cell(row=hdr_row, column=start_col,
                          value=f"#{num} {'Çözüm' if is_solution else difficulty.capitalize()}")
            hdr.font      = Font(name="Calibri", size=9, bold=True,
                                 color="0f9b8e" if is_solution else "555555")
            hdr.alignment = Alignment(horizontal="left")
            ws.row_dimensions[hdr_row].height = 14

            write_board(ws, board, given, start_row, start_col,
                        col_w=4.0, row_h=22)

        # Yazdırma — dikey A4, tek sayfaya sığdır
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left   = 0.35
        ws.page_margins.right  = 0.35
        ws.page_margins.top    = 0.4
        ws.page_margins.bottom = 0.4

    for page_idx, page_start in enumerate(range(0, len(puzzles_data), PER_PAGE)):
        batch = puzzles_data[page_start:page_start+PER_PAGE]
        name  = f"Sayfa {page_idx+1}" if len(puzzles_data) > PER_PAGE else "Bulmacalar"
        make_sheet(wb.create_sheet(title=name), batch, page_start, is_solution=False)

        if show_solution:
            sol_name = (f"Çözümler {page_idx+1}"
                        if len(puzzles_data) > PER_PAGE else "Çözümler")
            make_sheet(wb.create_sheet(title=sol_name), batch, page_start, is_solution=True)

    wb.save(filename)

# ── Ana Uygulama ──────────────────────────────────────────────────────────────
class SudokuApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sudoku")
        self.configure(bg=BG)
        self.resizable(False, False)

        self.board    = [[0]*9 for _ in range(9)]
        self.given    = [[False]*9 for _ in range(9)]
        self.solution = [[0]*9 for _ in range(9)]
        self.cells    = {}       # (r,c) → Entry widget
        self.selected = None
        self.timer_running = False
        self.start_time    = 0
        self.elapsed       = 0

        self._setup_style()
        self._build_ui()
        self._new_game("orta")

    def _setup_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TCombobox", fieldbackground=BG2, background=BG2,
                        foreground=TEXT, selectbackground=BG3)
        style.configure("Accent.TButton", background=ACCENT, foreground="white",
                        font=("Consolas",10,"bold"), borderwidth=0, focuscolor="none")
        style.configure("Green.TButton",  background=ACCENT2, foreground="white",
                        font=("Consolas",10,"bold"), borderwidth=0, focuscolor="none")
        style.configure("Ghost.TButton",  background=BG3, foreground=TEXT,
                        font=("Consolas",10), borderwidth=0, focuscolor="none")
        style.map("Accent.TButton", background=[("active","#c73652")])
        style.map("Green.TButton",  background=[("active","#0d8070")])
        style.map("Ghost.TButton",  background=[("active","#1a4a80")])

    def _build_ui(self):
        # ── Üst bar ────────────────────────────────────────────────────────
        topbar = tk.Frame(self, bg=BG2, pady=12)
        topbar.pack(fill="x")

        tk.Label(topbar, text="SUDOKU", font=("Consolas",20,"bold"),
                 bg=BG2, fg=TEXT).pack(side="left", padx=20)

        # Timer
        self._timer_lbl = tk.Label(topbar, text="00:00", font=("Consolas",16,"bold"),
                                    bg=BG2, fg=ACCENT2)
        self._timer_lbl.pack(side="right", padx=20)
        tk.Label(topbar, text="⏱", font=("Consolas",14), bg=BG2, fg=TEXT2).pack(
            side="right", padx=(0,4))

        # ── Kontrol paneli ─────────────────────────────────────────────────
        ctrl = tk.Frame(self, bg=BG, pady=10)
        ctrl.pack(fill="x", padx=GRID_PAD)

        tk.Label(ctrl, text="Zorluk:", font=("Consolas",10),
                 bg=BG, fg=TEXT2).pack(side="left")

        self._diff_var = tk.StringVar(value="orta")
        diff_cb = ttk.Combobox(ctrl, textvariable=self._diff_var,
                               values=["kolay","orta","zor","uzman"],
                               state="readonly", width=8,
                               font=("Consolas",10))
        diff_cb.pack(side="left", padx=(6,16))

        ttk.Button(ctrl, text="🎲 Yeni Oyun", style="Accent.TButton",
                   command=self._on_new_game).pack(side="left", padx=4, ipady=4)
        ttk.Button(ctrl, text="✓ Çöz", style="Green.TButton",
                   command=self._solve_current).pack(side="left", padx=4, ipady=4)
        ttk.Button(ctrl, text="💡 İpucu", style="Ghost.TButton",
                   command=self._hint).pack(side="left", padx=4, ipady=4)
        ttk.Button(ctrl, text="🗑 Temizle", style="Ghost.TButton",
                   command=self._clear_user).pack(side="left", padx=4, ipady=4)
        ttk.Button(ctrl, text="✔ Kontrol", style="Ghost.TButton",
                   command=self._check).pack(side="left", padx=4, ipady=4)

        # ── Grid ───────────────────────────────────────────────────────────
        grid_frame = tk.Frame(self, bg="#1a1a2e", padx=GRID_PAD, pady=8)
        grid_frame.pack()

        canvas_size = CELL_SIZE * 9 + 4
        self._canvas = tk.Canvas(grid_frame, width=canvas_size, height=canvas_size,
                                  bg=BG2, highlightthickness=0)
        self._canvas.pack()
        self._canvas.bind("<Button-1>", self._on_canvas_click)
        self.bind("<Key>", self._on_key)

        # ── Numpad ─────────────────────────────────────────────────────────
        numpad = tk.Frame(self, bg=BG, pady=8)
        numpad.pack()

        self._num_btns     = {}  # numara → buton widget
        self._remain_lbls  = {}  # numara → kalan sayı label

        for n in range(1, 10):
            btn = tk.Button(numpad, text=str(n), font=("Consolas",14,"bold"),
                            bg=BG3, fg=TEXT, relief="flat", bd=0,
                            width=3, height=1, activebackground=ACCENT,
                            activeforeground="white",
                            command=lambda v=n: self._input_number(v))
            btn.grid(row=0, column=n-1, padx=3, pady=(4,0), ipady=4)
            self._num_btns[n] = btn

            # Kalan sayı etiketi
            lbl = tk.Label(numpad, text="", font=("Consolas",8,"bold"),
                           bg=BG, fg=ACCENT2, width=3)
            lbl.grid(row=1, column=n-1, padx=3, pady=(0,4))
            self._remain_lbls[n] = lbl

        # Sil butonu
        tk.Button(numpad, text="⌫", font=("Consolas",14), bg=BG3, fg=ACCENT,
                  relief="flat", bd=0, width=3, height=1,
                  activebackground="#3a1a1a", activeforeground=ERROR,
                  command=lambda: self._input_number(0)).grid(
                  row=0, column=9, padx=(8,3), pady=4, ipady=4)

        # ── Export paneli ──────────────────────────────────────────────────
        exp_frame = tk.Frame(self, bg=BG2, pady=10)
        exp_frame.pack(fill="x")

        tk.Label(exp_frame, text="Dışa Aktar:", font=("Consolas",10),
                 bg=BG2, fg=TEXT2).pack(side="left", padx=16)

        self._export_count = tk.IntVar(value=1)
        tk.Label(exp_frame, text="Bulmaca sayısı:", font=("Consolas",9),
                 bg=BG2, fg=TEXT2).pack(side="left")
        tk.Spinbox(exp_frame, from_=1, to=200, textvariable=self._export_count,
                   width=4, font=("Consolas",10), bg=BG3, fg=TEXT,
                   buttonbackground=BG3, relief="flat").pack(side="left", padx=(4,12))

        self._per_page = tk.IntVar(value=1)
        tk.Label(exp_frame, text="Sayfa başına:", font=("Consolas",9),
                 bg=BG2, fg=TEXT2).pack(side="left")
        for v, lbl in [(1,"1"), (4,"4")]:
            tk.Radiobutton(exp_frame, text=lbl, variable=self._per_page, value=v,
                           font=("Consolas",10), bg=BG2, fg=TEXT,
                           activebackground=BG2, selectcolor=BG3).pack(side="left", padx=4)

        self._show_sol = tk.BooleanVar(value=True)
        tk.Checkbutton(exp_frame, text="Çözüm dahil", variable=self._show_sol,
                       font=("Consolas",9), bg=BG2, fg=TEXT2,
                       activebackground=BG2, selectcolor=BG3).pack(side="left", padx=12)

        ttk.Button(exp_frame, text="📄 PDF", style="Ghost.TButton",
                   command=self._export_pdf).pack(side="left", padx=4, ipady=4)
        ttk.Button(exp_frame, text="📊 Excel", style="Ghost.TButton",
                   command=self._export_excel).pack(side="left", padx=4, ipady=4)

        # ── Durum çubuğu ───────────────────────────────────────────────────
        self._status = tk.Label(self, text="", font=("Consolas",10),
                                 bg=BG, fg=TEXT2, pady=6)
        self._status.pack()

    def _update_remain(self):
        """Her sayının tahtada kaç kez daha konulabileceğini güncelle"""
        # Tahtadaki mevcut sayıları say (hatalı olanlar dahil)
        counts = {n: 0 for n in range(1, 10)}
        for r in range(9):
            for c in range(9):
                v = self.board[r][c]
                if v != 0:
                    counts[v] += 1

        for n in range(1, 10):
            remaining = 9 - counts[n]
            lbl = self._remain_lbls[n]
            btn = self._num_btns[n]
            if remaining <= 0:
                # Tamamlandı — buton solar, sayı gizlenir
                lbl.config(text="✓", fg=SUCCESS)
                btn.config(bg="#1a3a2a", fg="#3a6a4a")
            elif remaining <= 2:
                # Az kaldı — sarı uyarı
                lbl.config(text=str(remaining), fg=HINT)
                btn.config(bg=BG3, fg=TEXT)
            else:
                lbl.config(text=str(remaining), fg=ACCENT2)
                btn.config(bg=BG3, fg=TEXT)

    # ── Çizim ──────────────────────────────────────────────────────────────
    def _draw_board(self):
        self._canvas.delete("all")
        size = CELL_SIZE * 9

        errors = SudokuEngine.get_errors(self.board, self.solution) if any(
            self.solution[r][c] for r in range(9) for c in range(9)) else set()

        for row in range(9):
            for col in range(9):
                x0 = col * CELL_SIZE
                y0 = row * CELL_SIZE
                x1 = x0 + CELL_SIZE
                y1 = y0 + CELL_SIZE

                # Hücre arka planı
                if (row, col) == self.selected:
                    fill = "#2a3a6a"
                elif self.selected and (
                    self.selected[0] == row or self.selected[1] == col or
                    (row//3 == self.selected[0]//3 and col//3 == self.selected[1]//3)):
                    fill = "#1e2a4a"
                else:
                    fill = BG2

                self._canvas.create_rectangle(x0, y0, x1, y1,
                                               fill=fill, outline="")

                val = self.board[row][col]
                if val != 0:
                    if (row, col) in errors:
                        color = ERROR
                    elif self.given[row][col]:
                        color = GIVEN
                    else:
                        color = USER
                    self._canvas.create_text(
                        x0 + CELL_SIZE//2, y0 + CELL_SIZE//2,
                        text=str(val),
                        font=("Consolas", 22, "bold" if self.given[row][col] else "normal"),
                        fill=color
                    )

        # İnce çizgiler
        for i in range(10):
            w = 3 if i % 3 == 0 else 1
            c = TEXT if i % 3 == 0 else BORDER
            self._canvas.create_line(i*CELL_SIZE, 0, i*CELL_SIZE, size, fill=c, width=w)
            self._canvas.create_line(0, i*CELL_SIZE, size, i*CELL_SIZE, fill=c, width=w)

        # Kalan sayıları güncelle
        if hasattr(self, '_remain_lbls'):
            self._update_remain()

    # ── Oyun kontrolü ──────────────────────────────────────────────────────
    def _new_game(self, difficulty=None):
        if difficulty is None:
            difficulty = self._diff_var.get()
        self._set_status("Bulmaca üretiliyor...", TEXT2)
        self.update()

        puzzle, solution = SudokuEngine.generate(difficulty)
        self.board    = copy.deepcopy(puzzle)
        self.solution = solution
        self.given    = [[puzzle[r][c] != 0 for c in range(9)] for r in range(9)]
        self.selected = None

        self._draw_board()
        self._start_timer()
        filled = SudokuEngine.count_filled(puzzle)
        self._set_status(f"Yeni oyun — {difficulty.capitalize()} · {81-filled} hücre boş", ACCENT2)

    def _on_new_game(self):
        self._new_game(self._diff_var.get())

    def _on_canvas_click(self, event):
        col = event.x // CELL_SIZE
        row = event.y // CELL_SIZE
        if 0 <= row < 9 and 0 <= col < 9:
            self.selected = (row, col)
            self._draw_board()

    def _on_key(self, event):
        if not self.selected: return
        row, col = self.selected

        if event.keysym in ("Up","Down","Left","Right"):
            dr = {"Up":-1,"Down":1}.get(event.keysym, 0)
            dc = {"Left":-1,"Right":1}.get(event.keysym, 0)
            nr, nc = max(0,min(8,row+dr)), max(0,min(8,col+dc))
            self.selected = (nr, nc)
            self._draw_board()
        elif event.char in "123456789":
            self._input_number(int(event.char))
        elif event.keysym in ("BackSpace","Delete","0"):
            self._input_number(0)

    def _input_number(self, num):
        if not self.selected: return
        row, col = self.selected
        if self.given[row][col]: return
        self.board[row][col] = num
        self._draw_board()
        self._check_complete()

    def _solve_current(self):
        self.board = copy.deepcopy(self.solution)
        self._draw_board()
        self._stop_timer()
        self._set_status("Çözüm gösterildi!", ACCENT2)

    def _hint(self):
        """Rastgele bir boş hücreyi doldur"""
        empties = [(r, c) for r in range(9) for c in range(9)
                   if self.board[r][c] == 0]
        if not empties:
            self._set_status("Tüm hücreler dolu!", SUCCESS)
            return
        row, col = random.choice(empties)
        self.board[row][col] = self.solution[row][col]
        self.selected = (row, col)
        self._draw_board()
        self._check_complete()
        self._set_status(f"İpucu: {row+1}. satır, {col+1}. sütun → {self.solution[row][col]}", HINT)

    def _clear_user(self):
        for r in range(9):
            for c in range(9):
                if not self.given[r][c]:
                    self.board[r][c] = 0
        self._draw_board()
        self._set_status("Kullanıcı girişleri temizlendi.", TEXT2)

    def _check(self):
        errors = SudokuEngine.get_errors(self.board, self.solution)
        self._draw_board()
        if not errors:
            filled = SudokuEngine.count_filled(self.board)
            if filled == 81:
                self._set_status("🎉 Mükemmel! Sudoku tamamlandı!", SUCCESS)
            else:
                self._set_status("✓ Şu ana kadar hata yok!", SUCCESS)
        else:
            self._set_status(f"⚠ {len(errors)} hatalı hücre var (kırmızı ile gösterildi)", ERROR)

    def _check_complete(self):
        if all(self.board[r][c] != 0 for r in range(9) for c in range(9)):
            errors = SudokuEngine.get_errors(self.board, self.solution)
            if not errors:
                self._stop_timer()
                elapsed = self.elapsed
                m, s = divmod(int(elapsed), 60)
                self._set_status(f"🎉 Tebrikler! {m:02d}:{s:02d} sürede tamamladın!", SUCCESS)

    # ── Timer ──────────────────────────────────────────────────────────────
    def _start_timer(self):
        self._stop_timer()
        self.start_time    = time.time()
        self.elapsed       = 0
        self.timer_running = True
        self._tick()

    def _stop_timer(self):
        self.timer_running = False

    def _tick(self):
        if not self.timer_running: return
        self.elapsed = time.time() - self.start_time
        m, s = divmod(int(self.elapsed), 60)
        self._timer_lbl.config(text=f"{m:02d}:{s:02d}")
        self.after(1000, self._tick)

    # ── Status ─────────────────────────────────────────────────────────────
    def _set_status(self, msg, color=TEXT2):
        self._status.config(text=msg, fg=color)

    # ── Export ─────────────────────────────────────────────────────────────
    def _generate_puzzles(self):
        count      = self._export_count.get()
        difficulty = self._diff_var.get()
        self._set_status(f"{count} bulmaca üretiliyor...", TEXT2)
        self.update()

        puzzles_data = []
        for i in range(count):
            p, s = SudokuEngine.generate(difficulty)
            puzzles_data.append((p, s, difficulty))
            self._set_status(f"{i+1}/{count} üretildi...", ACCENT2)
            self.update()
        return puzzles_data

    def _export_pdf(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF dosyası","*.pdf")],
            title="PDF olarak kaydet"
        )
        if not filename: return
        try:
            puzzles_data = self._generate_puzzles()
            export_pdf(puzzles_data, filename,
                       per_page=self._per_page.get(),
                       show_solution=self._show_sol.get())
            self._set_status(f"✓ PDF kaydedildi: {filename}", SUCCESS)
            import os; os.startfile(filename)
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def _export_excel(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel dosyası","*.xlsx")],
            title="Excel olarak kaydet"
        )
        if not filename: return
        try:
            puzzles_data = self._generate_puzzles()
            export_excel(puzzles_data, filename,
                         show_solution=self._show_sol.get())
            self._set_status(f"✓ Excel kaydedildi: {filename}", SUCCESS)
            import os; os.startfile(filename)
        except Exception as e:
            messagebox.showerror("Hata", str(e))


if __name__ == "__main__":
    app = SudokuApp()
    app.mainloop()