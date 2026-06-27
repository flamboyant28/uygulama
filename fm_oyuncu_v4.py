import streamlit as st
import random
import math
import textwrap
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# =========================================================
# ÜLKE PROFİLLERİ
# =========================================================

COUNTRY_PROFILES = {
    "Türkiye"         : {"tech" : 1.35, "phys" : 1.30},
    "Brezilya"        : {"tech" : 1.25, "phys" : 0.90},
    "Arjantin"        : {"tech" : 1.20, "phys" : 0.95},
    "Uruguay"         : {"tech" : 1.10, "phys" : 1.05},
    "Kolombiya"       : {"tech" : 1.15, "phys" : 1.00},
    "Şili"            : {"tech" : 1.05, "phys" : 1.00},
    "Peru"            : {"tech" : 1.05, "phys" : 0.95},
    "Ekvador"         : {"tech" : 1.00, "phys" : 1.10},
    "İspanya"         : {"tech" : 1.15, "phys" : 0.90},
    "Portekiz"        : {"tech" : 1.20, "phys" : 0.95},
    "Fransa"          : {"tech" : 1.05, "phys" : 1.15},
    "İtalya"          : {"tech" : 1.10, "phys" : 1.00},
    "Belçika"         : {"tech" : 1.05, "phys" : 1.05},
    "Hollanda"        : {"tech" : 1.10, "phys" : 1.00},
    "İsviçre"         : {"tech" : 1.00, "phys" : 1.05},
    "Almanya"         : {"tech" : 0.95, "phys" : 1.20},
    "İngiltere"       : {"tech" : 0.95, "phys" : 1.15},
    "Danimarka"       : {"tech" : 0.95, "phys" : 1.20},
    "İsveç"           : {"tech" : 0.90, "phys" : 1.25},
    "Norveç"          : {"tech" : 0.90, "phys" : 1.30},
    "Finlandiya"      : {"tech" : 0.90, "phys" : 1.20},
    "Sırbistan"       : {"tech" : 0.90, "phys" : 1.20},
    "Hırvatistan"     : {"tech" : 1.05, "phys" : 1.05},
    "Bosna-Hersek"    : {"tech" : 1.00, "phys" : 1.10},
    "Polonya"         : {"tech" : 0.95, "phys" : 1.15},
    "Macaristan"      : {"tech" : 0.95, "phys" : 1.10},
    "Çekya"           : {"tech" : 1.00, "phys" : 1.10},
    "Romanya"         : {"tech" : 0.95, "phys" : 1.10},
    "Ukrayna"         : {"tech" : 0.90, "phys" : 1.20},
    "Rusya"           : {"tech" : 0.90, "phys" : 1.25},
    "Nijerya"         : {"tech" : 0.85, "phys" : 1.30},
    "Gana"            : {"tech" : 0.90, "phys" : 1.25},
    "Fildişi Sahili"  : {"tech" : 0.90, "phys" : 1.30},
    "Senegal"         : {"tech" : 0.95, "phys" : 1.25},
    "Fas"             : {"tech" : 1.00, "phys" : 1.10},
    "Cezayir"         : {"tech" : 1.00, "phys" : 1.05},
    "Mısır"           : {"tech" : 0.95, "phys" : 1.05},
    "Japonya"         : {"tech" : 1.05, "phys" : 0.95},
    "Güney Kore"      : {"tech" : 1.00, "phys" : 1.00},
    "İran"            : {"tech" : 0.95, "phys" : 1.10},
    "Suudi Arabistan" : {"tech" : 0.90, "phys" : 1.05},
    "Avustralya"      : {"tech" : 0.90, "phys" : 1.20},
    "ABD"             : {"tech" : 0.85, "phys" : 1.25},
    "Kanada"          : {"tech" : 0.85, "phys" : 1.20},
    "Meksika"         : {"tech" : 1.05, "phys" : 0.95},

}

COUNTRY_PA_BONUS = {
    "Brezilya": 8, "Arjantin": 6, "Fransa": 6,
    "Almanya": 5, "Portekiz": 6, "İspanya": 5,
    "Türkiye": 6,
}

# =========================================================
# BAYRAK EMOJİLERİ
# =========================================================

COUNTRY_FLAG = {
    "Türkiye": "🇹🇷", "Brezilya": "🇧🇷", "Arjantin": "🇦🇷", "Uruguay": "🇺🇾",
    "Kolombiya": "🇨🇴", "Şili": "🇨🇱", "Peru": "🇵🇪", "Ekvador": "🇪🇨",
    "İspanya": "🇪🇸", "Portekiz": "🇵🇹", "Fransa": "🇫🇷", "İtalya": "🇮🇹",
    "Belçika": "🇧🇪", "Hollanda": "🇳🇱", "İsviçre": "🇨🇭", "Almanya": "🇩🇪",
    "İngiltere": "🏴󠁧󠁢󠁥󠁮󠁧󠁿", "Danimarka": "🇩🇰", "İsveç": "🇸🇪", "Norveç": "🇳🇴",
    "Finlandiya": "🇫🇮", "Sırbistan": "🇷🇸", "Hırvatistan": "🇭🇷", "Bosna-Hersek": "🇧🇦",
    "Polonya": "🇵🇱", "Macaristan": "🇭🇺", "Çekya": "🇨🇿", "Romanya": "🇷🇴",
    "Ukrayna": "🇺🇦", "Rusya": "🇷🇺", "Nijerya": "🇳🇬", "Gana": "🇬🇭",
    "Fildişi Sahili": "🇨🇮", "Senegal": "🇸🇳", "Fas": "🇲🇦", "Cezayir": "🇩🇿",
    "Mısır": "🇪🇬", "Japonya": "🇯🇵", "Güney Kore": "🇰🇷", "İran": "🇮🇷",
    "Suudi Arabistan": "🇸🇦", "Avustralya": "🇦🇺", "ABD": "🇺🇸", "Kanada": "🇨🇦",
    "Meksika": "🇲🇽",
}

# =========================================================
# İSİM VERİTABANI
# =========================================================

NAMES = {

    # ═══════════════════════════════════════════════════════
    # TÜRK İYE
    # ═══════════════════════════════════════════════════════
    "Türkiye": {
        "first": [
            "Arda", "Burak", "Hakan", "Kerem", "Cengiz", "Serdar", "Ozan", "Yusuf",
            "Okay", "Mert", "Ferdi", "Taylan", "Barış", "Dorukhan", "İlkay", "Emre",
            "Berk", "Uğur", "Salih", "Kaan", "Metehan", "Yunus", "Berkay", "Halil",
            "Lamine", "Enes", "Furkan", "Umut", "Tunahan", "Sefa", "Müzeyyen", "Atakan",
            "Cenk", "Gökhan", "Volkan", "Semih", "Nuri", "Hamit", "Olcan", "Tuncay",
            "Bülent", "Rüştü", "Alpay", "Yıldıray", "Muzaffer", "Erhan", "Tolga", "Recep",
            "Miraç", "Mete", "Ercan", "Yılmaz", "Abdullah", "Aykut", "Polat", "Naci",
            "Ahmet", "Mehmet", "Mustafa", "Ali", "Hasan", "Hüseyin", "İbrahim", "İsmail",
            "Osman", "Ramazan", "Ömer", "Bekir", "Bilal", "Harun", "Eren", "Batuhan",
            "Berke", "Emir", "Arif", "Onur", "Sinan", "Levent", "Selim", "Kadir",
            "Adem", "Samet", "Talha", "Yiğit", "Alperen", "Doğukan", "Kürşat", "Sarp",
            "Efe", "Çağrı", "Çağatay", "Deniz", "Can", "Caner", "Cihan", "Hüsnü",
            "Zafer", "Orhan", "Şükrü", "Necati", "Vedat", "Tamer", "Erol", "Fikret",
            "Koray", "Ufuk", "Serkan", "İlker", "Eray", "Erdi", "Erkan", "Ergin",
            "Fatih", "Yasin", "Rıdvan", "Şaban", "Haluk", "Metin", "Taner", "Coşkun",
            "Alp", "Alper", "Altan", "Anıl", "Aras", "Arel", "Ata", "Atilla",
            "Aybars", "Aydın", "Aytaç", "Baha", "Bahadır", "Baki", "Baran", "Bartu",
            "Batur", "Bedir", "Bedrettin", "Berat", "Bora", "Buğra", "Burhan", "Celal",
            "Cem", "Cemal", "Cemil", "Cihaner", "Coşar", "Cüneyt", "Davut", "Demirhan",
            "Derviş", "Doğan", "Doğaner", "Dursun", "Emin", "Emrah", "Ender", "Engin",
            "Eraycan", "Ercüment", "Erdal", "Erdem", "Erenay", "Erkut", "Ersin", "Ertuğrul",
            "Esat", "Ethem", "Fahri", "Faruk", "Ferhat", "Fırat", "Galip", "Gani",
            "Giray", "Gökay", "Gökberk", "Göksel", "Görkem", "Haldun", "Halit", "Hamza",
            "Hazar", "Hikmet", "Hüseyin", "İdris", "İlhan", "İlyas", "İrfan", "İshak",
            "Kaan", "Kadirhan", "Kamil", "Kasım", "Kemal", "Kenan", "Levent", "Lokman",
            "Mahmut", "Mazhar", "Mehmetcan", "Melih", "Memet", "Mesut", "Mithat", "Murat",
            "Mustafa", "Nail", "Nazım", "Necip", "Nejat", "Nihat", "Oğulcan", "Oğuz",
            "Oğuzhan", "Onat", "Orkun", "Orçun", "Orkun", "Özgür", "Özcan", "Özer",
            "Ragıp", "Raşit", "Resul", "Sabit", "Sadık", "Saffet", "Sait", "Salim",
            "Sami", "Sedat", "Selçuk", "Serhat", "Seyit", "Suat", "Süleyman", "Şener",
            "Şevket", "Tacettin", "Talip", "Tarık", "Tayfun", "Temel", "Teoman", "Tufan",
            "Turgay", "Turhan", "Uğurcan", "Ulvi", "Ümit", "Vahit", "Veli", "Yahya",
            "Yakup", "Yavuzhan", "Yılmaz", "Yunus Emre", "Zafer", "Zeki",

            # İKİ İSİMLİLER

            "Mehmet Ali", "Mehmet Can", "Mehmet Emin", "Mehmet Akif", "Mehmet Fatih",
            "Ahmet Can", "Ahmet Emre", "Ahmet Yasin", "Ahmet Furkan", "Ahmet Kerem",
            "Mustafa Can", "Mustafa Emre", "Mustafa Kemal", "Mustafa Burak", "Ali Can",
            "Ali Emre", "Ali Kerem", "Ali Furkan", "Hasan Hüseyin", "Hasan Can",
            "Hasan Basri", "Hüseyin Can", "Hüseyin Emre", "İbrahim Halil", "İbrahim Can",
            "İbrahim Ethem", "Yusuf Emre", "Yusuf Can", "Yusuf Kerem", "Ömer Faruk",
            "Ömer Can", "Ömer Halis", "Abdullah Can", "Abdullah Emre", "Emirhan",
            "Emirhan Can", "Emirhan Ali", "Furkan Can", "Furkan Emre", "Furkan Ali",
            "Enes Can", "Enes Emre", "Enes Ali", "Burak Can", "Burak Emre",
            "Onur Can", "Onur Emre", "Kerem Can", "Kerem Ali", "Oğuzhan Can",
            "Oğuzhan Emre", "Batuhan Can", "Batuhan Emre", "Talha Can", "Talha Emre",
            "Samet Can", "Samet Emre", "Yasin Can", "Yasin Emre", "Halil İbrahim",
            "Halil Can", "Recep Can", "Bilal Can", "Bilal Emre", "Harun Can",
            "Harun Emre", "Salih Can", "Salih Emre", "Kaan Can", "Kaan Emre",
            "Mert Can", "Mert Emre", "Efe Can", "Efe Emre", "Arda Can",
            "Arda Emre", "Hasan Basri", "Mehmet Talha", "Mehmet Salih", "Mehmet Hüseyin",
            "Mehmet Burak", "Mehmet Enes", "Mehmet Onur", "Mehmet Samet", "Mehmet Serkan",
            "Mehmet Kaan", "Mehmet Arda", "Ahmet Talha", "Ahmet Salih", "Ahmet Onur",
            "Ahmet Samet", "Ahmet Kaan", "Ahmet Arda", "Ahmet Batuhan", "Ahmet Oğuzhan",
            "Ahmet Serhat", "Ahmet Berk", "Mustafa Talha", "Mustafa Salih", "Mustafa Onur",
            "Mustafa Samet", "Mustafa Kaan", "Mustafa Arda", "Mustafa Efe", "Mustafa Oğuz",
            "Mustafa Serhat", "Mustafa Berk", "Ali Talha", "Ali Salih", "Ali Onur",
            "Ali Samet", "Ali Kaan", "Ali Arda", "Ali Efe", "Ali Oğuz",
            "Ali Serhat", "Ali Berk", "Hasan Ali", "Hasan Emre", "Hasan Furkan",
            "Hasan Talha", "Hasan Oğuz", "Hasan Kaan", "Hasan Berk", "Hasan Efe",
            "Hasan Arda", "Hüseyin Ali", "Hüseyin Furkan", "Hüseyin Talha", "Hüseyin Oğuz",
            "Hüseyin Kaan", "Hüseyin Berk", "Hüseyin Efe", "Hüseyin Arda", "İbrahim Talha",
            "İbrahim Salih", "İbrahim Onur", "İbrahim Samet", "İbrahim Kaan", "İbrahim Arda",
            "İbrahim Berk", "İbrahim Efe", "Yusuf Talha", "Yusuf Salih", "Yusuf Onur",
            "Yusuf Samet", "Yusuf Kaan", "Yusuf Arda", "Yusuf Berk", "Yusuf Efe",
            "Ömer Talha", "Ömer Salih", "Ömer Onur", "Ömer Samet", "Ömer Kaan",
            "Ömer Arda", "Ömer Berk", "Ömer Efe", "Abdullah Talha", "Abdullah Salih",
            "Abdullah Onur", "Abdullah Samet", "Abdullah Kaan", "Abdullah Arda", "Abdullah Berk",
            "Abdullah Efe", "Furkan Talha", "Furkan Salih", "Furkan Onur", "Furkan Samet",
            "Furkan Kaan", "Furkan Arda", "Furkan Berk", "Furkan Efe", "Enes Talha",
            "Enes Salih", "Enes Onur", "Enes Samet", "Enes Kaan", "Enes Arda",
            "Enes Berk", "Enes Efe", "Burak Talha", "Burak Salih", "Burak Onur",
            "Burak Samet", "Burak Kaan", "Burak Arda", "Burak Berk", "Burak Efe",
            "Onur Talha", "Onur Salih", "Onur Samet", "Onur Kaan", "Onur Arda",
            "Onur Berk", "Onur Efe", "Kerem Talha", "Kerem Salih", "Kerem Onur",
            "Kerem Samet", "Kerem Kaan", "Kerem Arda", "Kerem Berk", "Kerem Efe",
            "Oğuzhan Talha", "Oğuzhan Salih", "Oğuzhan Onur", "Oğuzhan Samet", "Oğuzhan Kaan",
            "Oğuzhan Arda", "Oğuzhan Berk", "Oğuzhan Efe", "Batuhan Talha", "Batuhan Salih",
            "Batuhan Onur", "Batuhan Samet", "Batuhan Kaan", "Batuhan Arda", "Batuhan Berk",
            "Batuhan Efe", "Talha Emre", "Talha Ali", "Talha Kaan", "Talha Arda",
            "Talha Berk", "Talha Efe", "Samet Ali", "Samet Kaan", "Samet Arda",
            "Samet Berk", "Samet Efe", "Yasin Ali", "Yasin Kaan", "Yasin Arda",
            "Yasin Berk", "Yasin Efe", "Halil Can", "Halil Emre", "Halil Talha",
            "Halil Arda", "Halil Berk", "Recep Can", "Recep Emre", "Recep Talha",
            "Recep Arda", "Recep Berk", "Bilal Can", "Bilal Emre", "Bilal Talha",
            "Bilal Arda", "Bilal Berk", "Harun Can", "Harun Emre", "Harun Talha",
            "Harun Arda", "Harun Berk", "Salih Can", "Salih Emre", "Salih Talha",
            "Salih Arda", "Salih Berk", "Kaan Emre", "Kaan Ali", "Kaan Talha",
            "Kaan Arda", "Kaan Berk", "Kaan Efe", "Mert Emre", "Mert Ali",
            "Mert Talha", "Mert Arda", "Mert Berk", "Mert Efe", "Efe Emre",
            "Efe Ali", "Efe Talha", "Efe Arda", "Efe Berk", "Arda Emre",
            "Arda Ali", "Arda Talha", "Arda Berk", "Arda Efe"
        ],
        "last": [
            "Turan", "Yılmaz", "Çalhanoğlu", "Demir", "Ünder", "Dursun", "Kabak",
            "Yazıcı", "Okay", "Müldür", "Kadıoğlu", "Korkmaz", "Alper", "Atan",
            "Güler", "Kılınç", "Aktürkoğlu", "Arslan", "Özcan", "Doğan", "Şen",
            "Yüksek", "Demiral", "İbrahim", "Yıldız", "Karaman", "Kesgin", "Kaya",
            "Çelik", "Öztürk", "Ayhan", "Taşdemir", "Kaplan", "Çetin", "Aydın",
            "Polat", "Erdoğan", "Aslan", "Kurt", "Koç", "Şahin", "Yıldırım", "Güneş",
            "Bulut", "Cihan", "Akpınar", "Sönmez", "Uysal", "Toprak", "Bıyık",
            "Acar", "Aksoy", "Aktaş", "Altun", "Avcı", "Ateş", "Başar", "Bozkurt",
            "Dağ", "Durmaz", "Ekinci", "Elmas", "Gök", "Gökçe", "Gündüz", "Güven",
            "Işık", "İnce", "Karaca", "Karadağ", "Karakoç", "Kara", "Kesen", "Korkut",
            "Köse", "Kurtuluş", "Mertoğlu", "Odabaşı", "Özdemir", "Özkan", "Özsoy",
            "Sağlam", "Sezer", "Sarı", "Taş", "Tekin", "Tosun", "Uçar", "Ünal",
            "Varol", "Yavuz", "Yurt", "Zengin", "Çakır", "Çakmak", "Çevik", "Çiçek",
            "Özçelik", "Özkan", "Özarslan", "Duman", "Kılıç", "Bayraktar", "Pektaş",
            "Abacı", "Açıkgöz", "Adıgüzel", "Ağaoğlu", "Akbulut", "Akçay", "Akdeniz",
            "Akgül", "Akın", "Akkaya", "Akkurt", "Akman", "Akpınar", "Aksu", "Akyüz",
            "Albayrak", "Alkan", "Altay", "Altındağ", "Altıntaş", "Arıcı", "Arıkan",
            "Arslan", "Arslaner", "Aslanoğlu", "Atalay", "Atmaca", "Avşar", "Ayan",
            "Aydemir", "Aydoğdu", "Aykaç", "Aykut", "Ayoğlu", "Ayyıldız", "Balcı",
            "Baltacı", "Barut", "Başaran", "Başoğlu", "Baydar", "Bayram", "Baysal",
            "Bektaş", "Beşer", "Bilen", "Bilgin", "Bostancı", "Boz", "Bozdoğan",
            "Boztepe", "Budak", "Bulut", "Büyükkaya", "Candan", "Canoğlu", "Cebeci",
            "Cevahir", "Coşar", "Coşkuner", "Çalışkan", "Çapan", "Çavuş", "Çiftçi",
            "Çolak", "Çubukçu", "Dağlı", "Demirci", "Demirtaş", "Denizci", "Dereli",
            "Dikmen", "Dinç", "Doğru", "Durak", "Duru", "Efeoğlu", "Ekşi", "Elibol",
            "Erarslan", "Eraslan", "Erbaş", "Erdoğan", "Erduran", "Ergün", "Ersoy",
            "Ertürk", "Esen", "Esenyurt", "Eski", "Gedik", "Genç", "Gencer", "Giray",
            "Göçer", "Gökmen", "Görgülü", "Güçlü", "Güler", "Gültekin", "Gümüş",
            "Gündoğdu", "Güneş", "Güngör", "Gür", "Gürbüz", "Gürkan", "Gürsoy",
            "Hacıoğlu", "Harmancı", "İldiz", "Irmak", "İmamoğlu", "İpek", "İşler",
            "Kahraman", "Kalender", "Kalkan", "Kamacı", "Karabulut", "Karagöz",
            "Karakaş", "Karataş", "Kargın", "Kavak", "Kavaklı", "Kayaalp", "Kayhan",
            "Keser", "Kılıçaslan", "Kır", "Kıran", "Kireçci", "Koçak", "Konak",
            "Korkut", "Köksal", "Köseoğlu", "Kuru", "Kuş", "Kutlu", "Kuyucu",
            "Maden", "Mavioğlu", "Mercan", "Mutlu", "Nalbant", "Oflaz", "Okur",
            "Onaran", "Öcal", "Önder", "Özbay", "Özbek", "Özdaş", "Özdoğan", "Özgen",
            "Özmen", "Özpolat", "Öztuna", "Saçan", "Sancak", "Saygın", "Seçkin",
            "Sefer", "Selçuk", "Sert", "Sevim", "Sönmez", "Sucu", "Sümer",
            "Şanlı", "Şimşek", "Şirin", "Tan", "Tanrıverdi", "Taşkın", "Tatar",
            "Tekbaş", "Terzi", "Tetik", "Tiryaki", "Tunç", "Turanlı", "Türker",
            "Ural", "Uslu", "Uysal", "Uzun", "Üçok", "Ülgen", "Ünsal", "Üstün",
            "Yalçın", "Yaman", "Yardımcı", "Yazgan", "Yıldız", "Yıldızhan",
            "Yorulmaz", "Yurtsever", "Zorlu"
        ],
    },

    # ═══════════════════════════════════════════════════════
    # GÜNEY AMERİKA
    # ═══════════════════════════════════════════════════════
    "Brezilya": {
        "first": [
            "Gabriel", "Lucas", "Bruno", "Rodrygo", "Vinícius", "Endrick", "Rafael",
            "Marquinhos", "Casemiro", "Raphinha", "Richarlison", "Everton", "Matheus",
            "Danilo", "Alex", "Felipe", "Diego", "Neymar", "Thiago", "Fred", "Fabinho",
            "Roberto", "Ronaldo", "Ronaldinho", "Kaká", "Denilson", "Adriano", "Robinho",
            "Douglas", "Willian", "David", "Anderson", "Fernandinho", "Gilberto", "Hulk",
            "Anderson", "Militão", "Paquetá", "Rodriguinho", "Claudinho", "Artur",
        ],
        "last": [
            "Silva", "Santos", "Pereira", "Oliveira", "Costa", "Rodrigues", "Alves",
            "Ferreira", "Gomes", "Araújo", "Moura", "Militão", "Cunha", "Telles",
            "Barbosa", "Junior", "Firmino", "Paquetá", "Coutinho", "Souza",
            "Nascimento", "Carvalho", "Lima", "Mendes", "Ribeiro", "Cavalcanti",
            "Andrade", "Freitas", "Cardoso", "Marques", "Lopes", "Nunes", "Batista",
            "Campos", "Xavier", "Monteiro", "Azevedo", "Correia", "Vieira", "Ramos",
        ],
    },
    "Arjantin": {
        "first": [
            "Lionel", "Julián", "Rodrigo", "Lautaro", "Paulo", "Alexis", "Nicolás",
            "Cristian", "Leandro", "Marcos", "Exequiel", "Ángel", "Giovani", "Enzo",
            "Thiago", "Facundo", "Ramiro", "Agustín", "Germán", "Federico",
            "Sergio", "Javier", "Pablo", "Carlos", "Mauro", "Erik", "Lucas", "Manuel",
            "Nahuel", "Alejandro", "Lisandro", "Valentín", "Matías", "Iván", "Diego",
            "Roberto", "Miguel", "Ezequiel", "Maximiliano", "Santiago", "Gonzalo",
        ],
        "last": [
            "Messi", "Álvarez", "De Paul", "Martínez", "Dybala", "Mac Allister",
            "Fernández", "Acuña", "Otamendi", "Molina", "Palacios", "Correa",
            "Lo Celso", "Romero", "Paredes", "Di María", "Tagliafico", "Agüero",
            "Kun", "Tevez", "Higuaín", "Maradona", "Simeone", "Batistuta", "Crespo",
            "Riquelme", "Aimar", "Veron", "Saviola", "Zanetti", "Ayala", "Samuel",
            "Díaz", "Gómez", "Rodríguez", "García", "López", "Torres", "Herrera",
        ],
    },
    "Uruguay": {
        "first": [
            "Luis", "Diego", "Edinson", "Rodrigo", "Darwin", "Facundo", "Federico",
            "Sebastián", "Giorgian", "Martín", "Maxi", "Ronald", "Brian", "Nahitan",
            "José", "Pablo", "Diego", "Guillermo", "Álvaro", "Fernando", "Diego",
        ],
        "last": [
            "Suárez", "Forlán", "Cavani", "Bentancur", "Núñez", "Pellistri", "Valverde",
            "Coates", "De Arrascaeta", "Vecino", "Gómez", "Ugarte", "Nández", "Araújo",
            "García", "Rodríguez", "Alonso", "Perdomo", "Olivera", "Muslera", "Giménez",
        ],
    },
    "Kolombiya": {
        "first": [
            "James", "Radamel", "Juan", "Luis", "Carlos", "David", "Cucho", "Jhon", "Miguel",
            "Daniel", "Sebastián", "Yerry", "Davinson", "Wílmar", "Stefan", "Rafael",
            "Alfredo", "Fredy", "Teófilo", "Giovanni", "Víctor", "Camilo", "Kevin",
        ],
        "last": [
            "Rodríguez", "Falcao", "Cuadrado", "Díaz", "Muriel", "Córdoba", "Mina",
            "Sánchez", "Barrios", "Medina", "Mojica", "Santos", "Ospina", "Martínez",
            "Cardona", "Quintero", "Zapata", "Murillo", "Cuellar", "Borja", "Machado",
        ],
    },
    "Şili": {
        "first": [
            "Alexis", "Arturo", "Gary", "Charles", "Claudio", "Mauricio", "Erick", "Ben",
            "Sebastián", "Pablo", "Eduardo", "Jorge", "Christian", "Felipe", "Ivan",
        ],
        "last": [
            "Sánchez", "Vidal", "Medel", "Aránguiz", "Bravo", "Isla", "Pulgar", "Brereton",
            "Valdivia", "Beausejour", "Fuenzalida", "Valdés", "Fernández", "Morales", "Vargas",
        ],
    },
    "Meksika": {
        "first": [
            "Hirving", "Raúl", "Guillermo", "Carlos", "Andres", "Miguel", "Diego", "Javier",
            "Héctor", "Roberto", "Jorge", "Luis", "Edson", "Santiago", "Johan", "Alexis",
        ],
        "last": [
            "Lozano", "Jiménez", "Ochoa", "Guardado", "Moreno", "Lainez", "Herrera",
            "Márquez", "Vela", "Chicharito", "Corona", "Araujo", "Sánchez", "Pineda", "Álvarez",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # BATI & GÜNEY AVRUPA
    # ═══════════════════════════════════════════════════════
    "İspanya": {
        "first": [
            "Pedri", "Gavi", "Ansu", "Ferran", "Dani", "Marcos", "Rodri", "Alejandro",
            "Carlos", "Álvaro", "Unai", "Fabián", "José", "Pablo", "Mikel", "Yerlan",
            "Bryan", "Nico", "Lamine", "Marc", "Andrés", "Xavi", "David", "Iker",
            "Fernando", "Cesc", "Juan", "Raúl", "Fernando", "Carles", "Sergio",
            "Jordi", "Nacho", "Aymeric", "Eric", "Kepa", "Pau", "Álex", "Daniel",
        ],
        "last": [
            "González", "Fati", "Torres", "Olmo", "Carvajal", "Llorente", "Simón",
            "Ruiz", "Vivian", "Morata", "Laporte", "Gil", "Gaya", "Merino",
            "Zubimendi", "Cubarsi", "Yamal", "Grimaldo", "Iniesta", "Xavi",
            "Villa", "Puyol", "Piqué", "Casillas", "Ramos", "Alba", "Busquets",
            "Pedri", "Bellerin", "Navas", "Aspas", "Canales", "Gayà", "Parejo",
        ],
    },
    "Portekiz": {
        "first": [
            "Cristiano", "Bruno", "João", "Diogo", "Rúben", "Rafael", "Gonçalo",
            "Bernardo", "Nuno", "Vitinha", "Matheus", "Otávio", "Pepe",
            "Ricardo", "Renato", "Sérgio", "Domingos", "José", "André", "Hugo",
            "Luís", "Nani", "Deco", "Simão", "Hélder", "Figo", "Eusébio", "Rui",
        ],
        "last": [
            "Ronaldo", "Fernandes", "Félix", "Jota", "Dias", "Leão", "Ramos",
            "Silva", "Mendes", "Sanches", "Nunes", "Cancelo", "Moutinho", "Pepe",
            "Horta", "Guerreiro", "Trincão", "Palhinha", "Semedo", "Vieirinha",
            "Coentrao", "Veloso", "Postiga", "Figo", "Costa", "Soares", "Carvalho",
        ],
    },
    "Fransa": {
        "first": [
            "Kylian", "Ousmane", "Antoine", "Aurélien", "Theo", "Adrien", "Jules",
            "Benjamin", "Eduardo", "Marcus", "William", "Randal", "Christopher",
            "Jonathan", "Kingsley", "Axel", "Dayot", "Ibrahima", "Matteo", "Youssouf",
            "Zinedine", "Thierry", "Patrick", "Lilian", "Laurent", "Claude", "Marcel",
            "Franck", "Samir", "Nicolas", "Hugo", "Raphaël", "Steve", "Ferland", "Wesley",
        ],
        "last": [
            "Mbappé", "Dembélé", "Griezmann", "Tchouaméni", "Hernández", "Rabiot",
            "Koundé", "Pavard", "Camavinga", "Thuram", "Saliba", "Konaté", "Giroud",
            "Kanté", "Maignan", "Upamecano", "Zaire-Emery", "Barcola", "Guendouzi",
            "Zidane", "Henry", "Vieira", "Thuram", "Blanc", "Desailly", "Ribéry",
            "Benzema", "Nasri", "Evra", "Sagna", "Lloris", "Varane", "Pogba", "Mendy",
        ],
    },
    "İtalya": {
        "first": [
            "Federico", "Nicolo", "Sandro", "Giovanni", "Marco", "Gianluigi", "Alessandro",
            "Matteo", "Davide", "Giacomo", "Lorenzo", "Manuel", "Bryan", "Moise",
            "Luca", "Wilfried", "Gianluca", "Riccardo", "Roberto", "Mateo",
            "Francesco", "Andrea", "Giorgio", "Daniele", "Gennaro", "Filippo", "Fabio",
            "Simone", "Ciro", "Stephan", "Jorginho", "Emerson", "Leonardo", "Sassuolo",
        ],
        "last": [
            "Chiesa", "Barella", "Tonali", "Donnarumma", "Bastoni", "Verratti", "Dimarco",
            "Cristante", "Scamacca", "Bonaventura", "Pellegrini", "Locatelli", "Pessina",
            "Kean", "Zaccagni", "Gnonto", "Raspadori", "Calafiori", "Frattesi", "Retegui",
            "Totti", "Pirlo", "Buffon", "Cannavaro", "Maldini", "Nesta", "Del Piero",
            "Inzaghi", "Vieri", "Baggio", "Zambrotta", "Gattuso", "De Rossi", "Chiellini",
        ],
    },
    "Hollanda": {
        "first": [
            "Virgil", "Frenkie", "Memphis", "Davy", "Denzel", "Ryan", "Wout", "Donyell",
            "Teun", "Xavi", "Cody", "Nathan", "Jurriën", "Justin", "Tijjani", "Marten",
            "Georginio", "Quinten", "Jordan", "Lutsharel", "Johan", "Arjen", "Wesley",
            "Robin", "Rafael", "Clarence", "Patrick", "Nigel", "Dirk", "Jan", "Mark",
        ],
        "last": [
            "van Dijk", "de Jong", "Depay", "Klaassen", "Dumfries", "Gravenberch",
            "Weghorst", "Malen", "Koopmeiners", "Simons", "Gakpo", "Aké", "Timber",
            "Bijlow", "de Roon", "Wijnaldum", "Geertruida", "Cruyff", "Robben",
            "Sneijder", "van Persie", "van Nistelrooy", "Seedorf", "Davids", "Bergkamp",
            "van der Sar", "Stam", "van Bommel", "Zenden", "Kuyt", "Arjen", "Blind",
        ],
    },
    "Belçika": {
        "first": [
            "Kevin", "Eden", "Romelu", "Axel", "Yannick", "Dries", "Jan", "Toby",
            "Thomas", "Leandro", "Charles", "Youri", "Adnan", "Michy", "Jeremy",
            "Divock", "Nacer", "Marouane", "Mousa", "Christian", "Loïs", "Lander",
        ],
        "last": [
            "De Bruyne", "Hazard", "Lukaku", "Witsel", "Carrasco", "Mertens", "Vertonghen",
            "Alderweireld", "Vermaelen", "Trossard", "De Ketelaere", "Tielemans", "Januzaj",
            "Batshuayi", "Dendoncker", "Origi", "Chadli", "Fellaini", "Dembélé", "Benteke",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # KUZEY AVRUPA
    # ═══════════════════════════════════════════════════════
    "Almanya": {
        "first": [
            "Jamal", "Florian", "Kai", "Leroy", "Thomas", "Joshua", "Leon", "Niklas",
            "Ilkay", "Antonio", "Serge", "Robin", "Marco", "Timo", "Matthijs",
            "Nico", "Maximilian", "Julian", "Christopher", "Jonas", "Lukas",
            "Bastian", "Mario", "Miroslav", "Oliver", "Michael", "Philipp", "Per",
            "Sami", "Manuel", "Toni", "Mario", "André", "Emre", "Boateng", "Khedira",
        ],
        "last": [
            "Musiala", "Wirtz", "Havertz", "Sané", "Müller", "Kimmich", "Goretzka",
            "Süle", "Gündoğan", "Rüdiger", "Gnabry", "Koch", "Reus", "Werner",
            "de Ligt", "Schlotterbeck", "Arnold", "Draxler", "Nkunku", "Hofmann",
            "Schweinsteiger", "Götze", "Klose", "Podolski", "Lahm", "Mertesacker",
            "Hummels", "Neuer", "Kroos", "Bender", "Boateng", "Khedira", "Özil",
        ],
    },
    "İngiltere": {
        "first": [
            "Harry", "Jude", "Phil", "Marcus", "Bukayo", "Declan", "Jordan", "Raheem",
            "Mason", "Jack", "Kyle", "Trent", "Luke", "Reece", "Conor", "Kieran",
            "John", "James", "Kalvin", "Saka", "Wayne", "Steven", "Frank", "Ashley",
            "Rio", "John", "Peter", "Michael", "Owen", "Gerrard", "Lampard", "Terry",
            "Cole", "Scholes", "Giggs", "Beckham", "Shearer", "Lineker", "Robbie",
        ],
        "last": [
            "Kane", "Bellingham", "Foden", "Rashford", "Saka", "Rice", "Henderson",
            "Sterling", "Mount", "Grealish", "Walker", "Alexander-Arnold", "Shaw",
            "James", "Gallagher", "Trippier", "Stones", "Maddison", "Phillips", "Gordon",
            "Rooney", "Gerrard", "Lampard", "Ferdinand", "Terry", "Cole", "Scholes",
            "Beckham", "Shearer", "Owen", "Heskey", "Smith", "Johnson", "Carragher",
        ],
    },
    "Danimarka": {
        "first": [
            "Christian", "Pierre-Emile", "Kasper", "Simon", "Jannik", "Joakim",
            "Andreas", "Mathias", "Mikkel", "Thomas", "Rasmus", "Victor", "Joachim",
            "Yussuf", "Martin", "Henrik", "Peter", "Brian", "Nicklas", "Daniel",
        ],
        "last": [
            "Eriksen", "Højbjerg", "Schmeichel", "Kjær", "Vestergaard", "Maehle",
            "Christensen", "Lindstrøm", "Damsgaard", "Delaney", "Hojlund", "Poulsen",
            "Skov", "Sand", "Tomasson", "Rommedahl", "Priske", "Gronkjaer", "Jensen",
        ],
    },
    "İsveç": {
        "first": [
            "Zlatan", "Alexander", "Emil", "Jordan", "Viktor", "Ludwig", "Dejan",
            "Mikael", "Henrik", "Fredrik", "Marcus", "Samuel", "Kristoffer", "Sebastian",
            "Robin", "Pontus", "Mattias", "Oscar", "Anthony", "Seb",
        ],
        "last": [
            "Ibrahimović", "Isak", "Forsberg", "Larsson", "Nilsson", "Ekdal", "Kallman",
            "Ljungberg", "Svensson", "Gustafsson", "Lindelöf", "Olsson", "Danielson",
            "Claesson", "Augustinsson", "Bengtsson", "Toivonen", "Berg", "Zengin",
        ],
    },
    "Norveç": {
        "first": [
            "Erling", "Martin", "Alexander", "Joshua", "Veton", "Kristian", "Ola",
            "Mohamed", "Mathias", "Fredrik", "Sander", "Tarik", "Leo", "Håvard", "Mats",
        ],
        "last": [
            "Haaland", "Ødegaard", "Sørloth", "King", "Berisha", "Thorstvedt", "Solskjær",
            "Elyounoussi", "Ajer", "Strand Larsen", "Berge", "Henriksen", "Østigård", "Normann",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # DOĞU AVRUPA & BALKAN
    # ═══════════════════════════════════════════════════════
    "Sırbistan": {
        "first": [
            "Aleksandar", "Dušan", "Sergej", "Nemanja", "Nikola", "Filip", "Andrija",
            "Luka", "Stefan", "Marko", "Vladimir", "Milan", "Ivan", "Darko", "Dejan",
        ],
        "last": [
            "Mitrović", "Vlahović", "Milinković-Savić", "Matić", "Gudelj", "Kostić",
            "Pavlović", "Jović", "Tadić", "Živković", "Maksimović", "Nastasić", "Lazović",
        ],
    },
    "Hırvatistan": {
        "first": [
            "Luka", "Ivan", "Mateo", "Ivan", "Marcelo", "Sime", "Josip", "Ante",
            "Mario", "Marko", "Nikola", "Domagoj", "Vedran", "Eduardo", "Ognjen",
        ],
        "last": [
            "Modrić", "Rakitić", "Kovačić", "Perišić", "Brozović", "Vrsaljko", "Vlašić",
            "Mandzukić", "Vida", "Corluka", "Subašić", "Kramarić", "Orsić", "Gvardiol",
        ],
    },
    "Polonya": {
        "first": [
            "Robert", "Piotr", "Wojciech", "Arkadiusz", "Kamil", "Bartosz", "Jakub",
            "Grzegorz", "Karol", "Maciej", "Sebastian", "Łukasz", "Damian", "Tomasz", "Rafał",
        ],
        "last": [
            "Lewandowski", "Zieliński", "Szczęsny", "Milik", "Grosicki", "Bednarek",
            "Moder", "Krychowiak", "Linetty", "Glik", "Piątek", "Zalewski", "Frankowski",
        ],
    },
    "Ukrayna": {
        "first": [
            "Andriy", "Ruslan", "Oleksandr", "Mykhailo", "Artem", "Taras", "Viktor",
            "Olexandr", "Yevhen", "Serhiy", "Andriy", "Roman", "Vasyl", "Ivan", "Dmytro",
        ],
        "last": [
            "Shevchenko", "Malinovskyi", "Zinchenko", "Mudryk", "Dovbyk", "Stepanenko",
            "Shaparenko", "Rebrov", "Luzhny", "Voronin", "Konoplyanka", "Yarmolenko",
        ],
    },
    "Rusya": {
        "first": [
            "Artem", "Aleksandr", "Fyodor", "Andrey", "Denis", "Anton", "Daler",
            "Roman", "Yuri", "Igor", "Dmitri", "Sergei", "Vasili", "Nikita", "Pavel",
        ],
        "last": [
            "Dzyuba", "Golovin", "Smolov", "Miranchuk", "Zhirkov", "Arshavin",
            "Pavlyuchenko", "Kerzhakov", "Berezutski", "Ignashevich", "Malafeev",
        ],
    },
    "Romanya": {
        "first": [
            "Ianis", "Draguș", "Florin", "Denis", "Alexandru", "Răzvan", "Ciprian",
            "Gheorghe", "Cosmin", "Adrian", "Cristian", "Bogdan", "Valentin", "Daniel",
        ],
        "last": [
            "Hagi", "Drăguș", "Tătărușanu", "Alibec", "Chipciu", "Marin", "Nedelcearu",
            "Stanciu", "Moruțan", "Ioniță", "Maxim", "Keșeru", "Pușcaș", "Rus",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # AFRİKA
    # ═══════════════════════════════════════════════════════
    "Nijerya": {
        "first": [
            "Victor", "Alex", "Kelechi", "Wilfred", "Odion", "Samuel", "Emmanuel", "Ola",
            "Taiwo", "Calvin", "Chidera", "Terem", "Frank", "Cyriel", "Bright", "Henry",
            "Ahmed", "Sunday", "David", "Tyronne", "Nwankwo", "Jay-Jay", "Yakubu",
            "Mikel", "Obafemi", "Austin", "Wilson", "Peter", "Obinna", "Sunday",
        ],
        "last": [
            "Osimhen", "Iwobi", "Iheanacho", "Ndidi", "Ighalo", "Chukwueze", "Dennis",
            "Awoniyi", "Bassey", "Ejuke", "Moffi", "Onyeka", "Ugbo", "Eze", "Musa",
            "Kanu", "Okocha", "Martins", "Mikel", "Martins", "Lawal", "Yekini", "Amokachi",
        ],
    },
    "Gana": {
        "first": [
            "Thomas", "Jordan", "Mohammed", "André", "Daniel", "Inaki", "Tariq",
            "Baba", "Jonathan", "Mubarak", "Caleb", "Edwin", "Kudus", "Joseph", "Kasim",
            "Asamoah", "Michael", "Sulley", "Kevin", "Anthony", "Quincy", "Abedi",
        ],
        "last": [
            "Partey", "Ayew", "Kudus", "Djiku", "Mensah", "Williams", "Lamptey",
            "Rahman", "Asante", "Wakaso", "Kyereh", "Agyemang-Badu", "Acquah",
            "Gyan", "Essien", "Muntari", "Boateng", "Appiah", "Pele", "Kuffour",
        ],
    },
    "Fildişi Sahili": {
        "first": [
            "Didier", "Franck", "Wilfried", "Serge", "Nicolas", "Sebastien", "Eric",
            "Salomon", "Jean-Daniel", "Max-Alain", "Ghislain", "Ibrahim", "Kouassi",
            "Yaya", "Kolo", "Gervinho", "Romaric", "Siaka", "Emmanuel", "Jonathan",
        ],
        "last": [
            "Drogba", "Kessié", "Zaha", "Aurier", "Pépé", "Haller", "Gradel",
            "Kalou", "Bamba", "Gradel", "Diallo", "Touré", "Touré", "Gervinho",
            "Romaric", "Akpala", "Ndri", "Zokora", "Doumbia", "Dindane",
        ],
    },
    "Senegal": {
        "first": [
            "Sadio", "Kalidou", "Idrissa", "Ismaila", "Cheikhou", "Bamba", "Nampalys",
            "Formose", "Pape", "Abdou", "Nicolas", "Moussa", "Alfred", "Lamine",
            "Habib", "Krepin", "Saliou", "Fodé", "Henri", "Mamadou", "Kara",
        ],
        "last": [
            "Mané", "Koulibaly", "Gueye", "Sarr", "Kouyaté", "Diallo", "Mendy",
            "Jackson", "Diatta", "Gomis", "Camara", "Diaw", "Sabaly", "Ciss",
            "Diouf", "Ndiaye", "Ndoye", "Sane", "Wagué", "Ba", "Ndoye",
        ],
    },
    "Fas": {
        "first": [
            "Achraf", "Hakim", "Youssef", "Sofyan", "Jawad", "Nayef", "Noussair",
            "Amine", "Selim", "Azzedine", "Abde", "Ilias", "Zakaria", "Brahim",
            "Munir", "Fouad", "Yassine", "Reda", "Mehdi", "Nabil", "Mbark",
        ],
        "last": [
            "Hakimi", "Ziyech", "En-Nesyri", "Amrabat", "El Yamiq", "Aguerd", "Saiss",
            "Mazraoui", "Harit", "Benrahma", "Ounahi", "Ezzalzouli", "Chair", "Aboukhlal",
            "El Kaabi", "Boussoufa", "Chamakh", "Benatia", "Belhanda", "Labyad", "Larousi",
        ],
    },
    "Cezayir": {
        "first": [
            "Riyad", "Islam", "Andy", "Ismaël", "Youcef", "Sofiane", "Adlene", "Nabil",
            "Rachid", "Mehdi", "Hicham", "Bilal", "Baghdad", "Aissa", "Faouzi",
        ],
        "last": [
            "Mahrez", "Slimani", "Delort", "Bennacer", "Atal", "Feghouli", "Guedioura",
            "Ghilas", "Brahimi", "Lacen", "Boudebouz", "Ghoulam", "Bounedjah", "Mandi",
        ],
    },
    "Mısır": {
        "first": [
            "Mohamed", "Ahmed", "Omar", "Mahmoud", "Amr", "Karim", "Ramadan",
            "Trezeguet", "Mostafa", "Akram", "Marwan", "Zizo", "Hamdi", "Saad",
        ],
        "last": [
            "Salah", "Elneny", "El-Shenawy", "Hegazi", "Hamdy", "Ashraf", "Fathy",
            "Trezeguet", "Mohamed", "Kahraba", "Nesim", "Ramadan", "Warda", "Elmohamady",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # ASYA
    # ═══════════════════════════════════════════════════════
    "Japonya": {
        "first": [
            "Takumi", "Daichi", "Hiroki", "Ao", "Wataru", "Ritsu", "Takehiro", "Junya",
            "Kaoru", "Maya", "Ko", "Shogo", "Keito", "Soma", "Ayase", "Shuichi",
            "Shinji", "Keisuke", "Shinji", "Yuya", "Makoto", "Yoshinori", "Genki", "Reo",
        ],
        "last": [
            "Minamino", "Kamada", "Ito", "Tanaka", "Endo", "Doan", "Tomiyasu",
            "Mitoma", "Yoshida", "Itakura", "Ueda", "Nakamura", "Ogawa", "Gonda",
            "Kagawa", "Honda", "Okazaki", "Nagatomo", "Kubo", "Hasebe", "Morishige",
        ],
    },
    "Güney Kore": {
        "first": [
            "Son", "Lee", "Kim", "Park", "Hwang", "Cho", "Na", "Kwon", "Jung",
            "Chang", "Heung", "Min", "Jae", "Sung", "Woo", "Hyun", "Ji", "Young",
        ],
        "last": [
            "Heung-Min", "Jae-Sung", "In-Beom", "Seung-Ho", "Hee-Chan", "Sang-Ho",
            "Min-Jae", "Young-Gwon", "Woo-Young", "Tae-Hwan", "Kang-In", "Ju-Chan",
        ],
    },
    "İran": {
        "first": [
            "Sardar", "Mehdi", "Ali", "Alireza", "Saman", "Karim", "Roozbeh",
            "Morteza", "Saeid", "Masoud", "Mehdi", "Ashkan", "Ahmad", "Hassan",
        ],
        "last": [
            "Azmoun", "Taremi", "Gholizadeh", "Beiranvand", "Aghaei", "Ansarifard",
            "Hajsafi", "Pouraliganji", "Rezaeian", "Mohammadi", "Jalali", "Shojaei",
        ],
    },
    "Suudi Arabistan": {
        "first": [
            "Salem", "Firas", "Mohammed", "Saleh", "Ahmed", "Abdullah", "Nasser",
            "Yasser", "Sami", "Sultan", "Turki", "Hattan", "Nawaf", "Ali",
        ],
        "last": [
            "Al-Dawsari", "Al-Buraikan", "Al-Owais", "Al-Shahrani", "Al-Malki",
            "Al-Faraj", "Al-Qasim", "Al-Abed", "Al-Ghannam", "Al-Hazazi", "Al-Yami",
        ],
    },

    # ═══════════════════════════════════════════════════════
    # KUZEY AMERİKA & AVUSTRALYA
    # ═══════════════════════════════════════════════════════
    "ABD": {
        "first": [
            "Christian", "Tyler", "Weston", "Gio", "Josh", "Ricardo", "Reggie",
            "Sergiño", "DeAndre", "Zack", "Malik", "Antonee", "Jordan", "Brenden", "John",
        ],
        "last": [
            "Pulisic", "Adams", "McKennie", "Reyna", "Sargent", "Pepi", "Cannon",
            "Dest", "Yedlin", "Steffen", "Tillman", "Robinson", "Morris", "Aaronson", "Brooks",
        ],
    },
    "Kanada": {
        "first": [
            "Alphonso", "Jonathan", "Cyle", "Jonathan", "Tajon", "Liam", "Ismaël", "Milan",
            "Sam", "Richie", "Scott", "David", "Alistair", "Atiba", "Junior",
        ],
        "last": [
            "Davies", "David", "Larin", "Osorio", "Buchanan", "Miller", "Koné", "Borjan",
            "Adekugbe", "Laryea", "Kennedy", "Hoilett", "Johnston", "Hutchinson", "Hoilett",
        ],
    },
    "Avustralya": {
        "first": [
            "Mathew", "Jackson", "Mitchell", "Martin", "Awer", "Aaron", "Craig",
            "Ryan", "Maty", "Brad", "Riley", "Nestory", "Milos", "Jason", "Garang",
        ],
        "last": [
            "Leckie", "Irvine", "Duke", "Boyle", "Mabil", "Mooy", "Goodwin",
            "Ryan", "Ryan", "Smith", "McGree", "Juric", "Degenek", "Davidson", "Kuol",
        ],
    },
}

# Varsayılan (diğer ülkeler)
NAMES["Varsayılan"] = {
    "first": [
        "Alex", "Marco", "Lucas", "David", "Carlos", "Ivan", "Stefan", "Nikola",
        "Milan", "Pablo", "Diego", "Sergio", "João", "Luca", "Felix", "Max",
        "Kevin", "Patrick", "Thomas", "André", "Nicolas", "Antoine", "Baptiste", "Hugo",
    ],
    "last": [
        "Silva", "Garcia", "Müller", "Rossi", "Kovač", "Santos", "Fischer", "Martínez",
        "López", "Romano", "Costa", "Fernández", "Weber", "Becker", "Schulz", "Johansson",
        "Petrov", "Novak", "Horváth", "Popescu", "Ionescu", "Bogdanov", "Vlasov", "Radu",
    ],
}

NAMES["Peru"] = {
    "first": ["Paolo", "Jefferson", "Gianluca", "André", "Christian", "Luis", "Pedro",
              "Raúl", "Alex", "Josepmir", "Sergio", "Andy", "Alexander", "Yoshimar", "Renato"],
    "last": ["Guerrero", "Farfán", "Lapadula", "Carrillo", "Cueva", "Advíncula", "Abram",
             "García", "Aquino", "Ballón", "Peña", "Polo", "Valera", "Yotún", "Tapia"],
}
NAMES["Ekvador"] = {
    "first": ["Moisés", "Enner", "Jhegson", "Ángel", "Félix", "Antonio", "Renato",
              "Xavier", "Robert", "Luis", "William", "Gonzalo", "Djorkaeff", "Jorge", "Pervis"],
    "last": ["Caicedo", "Valencia", "Méndez", "Mena", "Torres", "Valencia", "Ibarra",
             "Arreaga", "Arboleda", "Chala", "Rodríguez", "Plata", "Pineida", "Cifuentes", "Estupiñán"],
}
NAMES["İsviçre"] = {
    "first": ["Granit", "Xherdan", "Haris", "Ricardo", "Fabian", "Steven", "Silvan",
              "Remo", "Edimilson", "Ruben", "Manuel", "Philipp", "Yvon", "Denis", "Breel"],
    "last": ["Xhaka", "Shaqiri", "Seferović", "Rodríguez", "Schär", "Zuber", "Widmer",
             "Freuler", "Fernandes", "Vargas", "Akanji", "Senderos", "Mvogo", "Zakaria", "Embolo"],
}
NAMES["Finlandiya"] = {
    "first": ["Teemu", "Joel", "Juhani", "Robert", "Fredrik", "Petteri", "Pyry", "Rasmus",
              "Joni", "Lukas", "Jasse", "Onni", "Miro", "Riku", "Ilmari"],
    "last": ["Pukki", "Pohjanpalo", "Lod", "Taylor", "Jensen", "Forss", "Nissilä",
             "Schuller", "Toivio", "Hradecky", "Uronen", "Raitala", "Pentz", "Kamara", "Hamalainen"],
}
NAMES["Bosna-Hersek"] = {
    "first": ["Edin", "Miralem", "Vedad", "Emir", "Rade", "Sead", "Haris", "Asmir",
              "Senad", "Ermin", "Armin", "Zlatan", "Darko", "Nermin", "Muhamed"],
    "last": ["Džeko", "Pjanić", "Ibišević", "Spahić", "Krunic", "Kolašinac", "Dijakovic",
             "Begović", "Lulic", "Zec", "Hadziahmetovic", "Duljević", "Šunjić", "Mujakić", "Tahirović"],
}
NAMES["Macaristan"] = {
    "first": ["Dominik", "Roland", "Attila", "Ádám", "Loïc", "Barnabás", "Kevin",
              "Ákos", "László", "Zsolt", "Gábor", "Balázs", "Mihály", "Péter", "Tamás"],
    "last": ["Szoboszlai", "Sallai", "Fiola", "Szalai", "Nego", "Varga", "Csoboth",
             "Schafer", "Orbán", "Gulácsi", "Lovrencsics", "Botka", "Ferenczi", "Gazdag", "Könyves"],
}
NAMES["Çekya"] = {
    "first": ["Patrik", "Vladimír", "Ondřej", "Tomáš", "Jan", "Adam", "Lukáš",
              "Jakub", "Antonín", "Michal", "Matěj", "Pavel", "Ladislav", "Radoslav", "Jiří"],
    "last": ["Schick", "Coufal", "Čelůstka", "Souček", "Bořil", "Hložek", "Jankto",
             "Kuchta", "Barák", "Jurásek", "Jurečka", "Sadílek", "Kral", "Hlozek", "Lingr"],
}


def generate_name(country: str) -> str:
    pool = NAMES.get(country, NAMES["Varsayılan"])
    return f"{random.choice(pool['first'])} {random.choice(pool['last'])}"


# =========================================================
# MEVKİ TANIMI
# =========================================================

ALL_POSITIONS = [
    "ST", "KF (Sol)", "KF (Sağ)", "OOS",
    "KANAT (Sol)", "KANAT (Sağ)", "OS", "DM",
    "DOS", "KB (Sol)", "KB (Sağ)", "D (Sol)", "D (Sağ)", "KL",
]

POSITION_BASE = {
    "ST": "ST", "KF (Sol)": "KF", "KF (Sağ)": "KF",
    "OOS": "OOS", "KANAT (Sol)": "KANAT", "KANAT (Sağ)": "KANAT",
    "OS": "OS", "DM": "DM", "DOS": "DOS",
    "KB (Sol)": "KB", "KB (Sağ)": "KB",
    "D (Sol)": "D", "D (Sağ)": "D", "KL": "KL",
}

LEFT_POSITIONS = {"KF (Sol)", "KANAT (Sol)", "KB (Sol)", "D (Sol)"}
RIGHT_POSITIONS = {"KF (Sağ)", "KANAT (Sağ)", "KB (Sağ)", "D (Sağ)"}

# =========================================================
# ATTRIBUTE LİSTELERİ
# =========================================================

TECHNICAL = [
    "Bitiricilik", "Dripling", "İlk Kontrol", "Kafa Vuruşu", "Markaj",
    "Orta Yapma", "Pas", "Teknik", "Top Kapma", "Uzaktan Şut",
    "Korner", "Penaltı Kullanma", "Serbest Vuruş Kullanma", "Uzun Taç",
]
MENTAL = [
    "Agresiflik", "Cesaret", "Çalışkanlık", "Karar Alma", "Kararlılık",
    "Konsantrasyon", "Liderlik", "Önsezi", "Özel Yetenek", "Pozisyon Alma",
    "Soğukkanlılık", "Takım Oyunu", "Topsuz Alan", "Vizyon",
]
PHYSICAL = [
    "Çeviklik", "Dayanıklılık", "Denge", "Güç",
    "Hız", "Hızlanma", "Vücut Zindeliği", "Zıplama",
]
GOALKEEPER = [
    "Ani Çıkış Eğilimi", "Birebir", "Bölge Hakimiyeti", "Degaj",
    "Eksantriklik", "Elle Kontrol", "Elle Oyun Başlatma",
    "Hava Topları", "İletişim", "İlk Kontrol (K)", "Pas (K)",
    "Refleksler", "Yumrukla Uzaklaştırma",
]
HIDDEN = [
    "Uyum", "Hırs", "Tartışma", "Aidiyet Duygusu",
    "Baskıya Dayanıklılık", "Profesyonellik",
    "Sportmenlik", "Huy", "Çok Yönlülük",
    "Çirkeflik", "Önemli Maçlar", "Sakatlanma Eğilimi", "Süreklilik",
]

GK_ATTRS_SET = set(GOALKEEPER)
HIDDEN_SET = set(HIDDEN)

SCOUT_GRADES = [(170, "A"), (150, "B"), (130, "C"), (110, "D"), (0, "E")]
STAR_CA_THRESHOLD = 150

# =========================================================
# BASKINDAYAK
# =========================================================

FOOT_WEIGHTS = {
    "KF (Sol)"    : {"Sol" : 0.20, "Sağ" : 0.70, "Her İkisi" : 0.10},  # ters ayaklı içe keser
    "KF (Sağ)"    : {"Sol" : 0.70, "Sağ" : 0.20, "Her İkisi" : 0.10},
    "KANAT (Sol)" : {"Sol" : 0.55, "Sağ" : 0.35, "Her İkisi" : 0.10},
    "KANAT (Sağ)" : {"Sol" : 0.35, "Sağ" : 0.55, "Her İkisi" : 0.10},
    "KB (Sol)"    : {"Sol" : 0.60, "Sağ" : 0.30, "Her İkisi" : 0.10},
    "KB (Sağ)"    : {"Sol" : 0.25, "Sağ" : 0.65, "Her İkisi" : 0.10},
    "D (Sol)"     : {"Sol" : 0.60, "Sağ" : 0.30, "Her İkisi" : 0.10},
    "D (Sağ)"     : {"Sol" : 0.25, "Sağ" : 0.65, "Her İkisi" : 0.10},

}
DEFAULT_FOOT =      {"Sol": 0.30, "Sağ": 0.60, "Her İkisi": 0.10}


def generate_foot(position: str) -> str:
    w = FOOT_WEIGHTS.get(position, DEFAULT_FOOT)
    return random.choices(list(w.keys()), weights=list(w.values()))[0]


# ── ZAYIF AYAK GÜCÜ ─────────────────────────────────────────────────────
# FMScout CA ağırlıkları (0-10 → ×0.4 = 0-4 ölçeği)
# Zayıf ayak tüm pozisyonlarda CA'ya katkı yapar — özellikle hücumda kritik!
# Zayıf ayak CA ağırlıkları — FM editor'dan doğrulanan GERÇEK değerler
# (FM değeri ÷ 2.5 = bizim 0-4 skalası)
WEAK_FOOT_CA_WEIGHT = {
    "ST"    : 3.0,  # FM=7.5  ✅ DOĞRULANDI (Haaland)
    "KF"    : 2.2,  # FM=5.5  ✅ DOĞRULANDI (Saka, Vinicius)
    "OOS"   : 2.8,  # FM=7.0  ✅ DOĞRULANDI (Dybala)
    "KANAT" : 2.4,  # FM=6.0  ✅ DOĞRULANDI (Arnold) ← eskiden 2.2!
    "OS"    : 2.0,  # FM=5.0  ✅ DOĞRULANDI (Kanté)
    "DM"    : 2.0,  # FM=5.0  ✅ DOĞRULANDI (Casemiro)
    "KB"    : 1.6,  # FM=4.0  ✅ DOĞRULANDI (Gosens)
    "D"     : 1.6,  # FM=4.0  ✅ DOĞRULANDI (Robertson, Trippier)
    "DOS"   : 1.8,  # FM=4.5  ✅ DOĞRULANDI (Van Dijk) ← eskiden 0.4!
    "KL"    : 1.2,  # FM=3.0  ✅ DOĞRULANDI (Alisson, Ederson)

}

# FM'de zayıf ayak kategorileri (1-20 skalası)
WEAK_FOOT_LABELS = {
    (1, 4)   : "Çok Zayıf",
    (5, 8)   : "Zayıf",
    (9, 12)  : "Orta",
    (13, 16) : "Güçlü",
    (17, 20) : "Çok Güçlü",

}


def weak_foot_label(val: int) -> str:
    for (lo, hi), label in WEAK_FOOT_LABELS.items():
        if lo <= val <= hi:
            return label
    return "Orta"


def generate_dominant_foot_value() -> int:
    """
    Baskın ayak gücü (1-20). Genellikle çok yüksek (15-20).
    FM'de çoğu oyuncunun baskın ayağı 15-20 arasında.
    """
    weights = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 5, 14, 25, 30, 25]
    # index 0-19 → değer 1-20
    import random
    val = random.choices(range(1, 21), weights=weights)[0]
    return val


def generate_weak_foot(dominant_foot: str, position: str) -> int:
    """
    Baskın ayağa ve mevkiye göre zayıf ayak gücü üretir (1-20).

    Her İkisi → zayıf ayak yüksek (iyi iki ayaklı)
    Tek ayak   → mevkiye göre değişken, genelde düşük
    Hücum mevkileri → bazen ters ayakla tehlikeli (daha yüksek)
    """
    base = POSITION_BASE.get(position, position)

    if dominant_foot == "Her İkisi":
        # İki ayaklı: zayıf ayak da güçlü
        return random.randint(14, 19)

    # Mevkiye göre dağılım
    if base in ("ST", "KF", "OOS"):
        # Hücumcular için iki ayaklı olmak çok değerli → biraz daha yüksek
        weights = [5, 25, 35, 25, 8, 2]  # 1-3, 4-6, 7-9, 10-12, 13-15, 16-18
        ranges = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15), (16, 18)]
    elif base in ("KANAT",):
        # Kanat için çapraz ayak önemli
        weights = [8, 30, 35, 20, 6, 1]
        ranges = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15), (16, 18)]
    elif base in ("DOS", "D", "KB", "DM"):
        # Defans için çok kritik değil
        weights = [15, 35, 30, 15, 4, 1]
        ranges = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15), (16, 18)]
    else:
        weights = [10, 30, 35, 18, 6, 1]
        ranges = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15), (16, 18)]

    lo, hi = random.choices(ranges, weights=weights)[0]
    return random.randint(lo, hi)


# =========================================================
# BOY / KİLO
# =========================================================

HEIGHT_BY_BASE = {
    "KL": (185, 200), "DOS": (182, 200), "D": (176, 188),
    "KB": (172, 184), "DM": (177, 188), "OS": (173, 184),
    "OOS": (170, 181), "KANAT": (170, 180), "KF": (170, 180),
    "ST": (177, 192),
}


def generate_height_weight(position: str, phys_attrs: dict) -> tuple:
    base = POSITION_BASE[position]
    lo, hi = HEIGHT_BY_BASE.get(base, (172, 185))
    strength_factor = (phys_attrs.get("Güç", 10) + phys_attrs.get("Zıplama", 10)) / 2
    bias = int((strength_factor - 10) * 0.4)
    height = max(165, min(200, random.randint(lo, hi) + bias))
    # Kilo: boy - 100 ± rastgele küçük fark + güç bonus
    weight = int((height - 100) + random.randint(-3, 5) + (phys_attrs.get("Güç", 10) - 10) * 0.5)
    weight = max(60, min(97, weight))
    return height, weight


# =========================================================
# TRAITS & ROLLER
# =========================================================

TRAITS_BY_POSITION = {
    "ST"          : ["Bitiriciliği Sever", "Ceza Sahasına Geç Koşu", "Topu Saklama"],
    "KF"          : ["Çaprazdan Girer", "İçe Keser", "Hızlı Koşar"],
    "KF (Sol)"    : ["İçe Keser (Sol)", "Sağ Ayağıyla Çapraz Şut", "Hızlı Koşar"],
    "KF (Sağ)"    : ["İçe Keser (Sağ)", "Sol Ayağıyla Çapraz Şut", "Hızlı Koşar"],
    "OOS"         : ["Kilit Pas Dener", "Tempo Belirler", "Derine İner"],
    "KANAT"       : ["Kanat Koşusu Yapar", "Orta Açar"],
    "KANAT (Sol)" : ["Sol Kanattan Orta Açar", "İçe Keser", "Kanat Koşusu Yapar"],
    "KANAT (Sağ)" : ["Sağ Kanattan Orta Açar", "İçe Keser", "Kanat Koşusu Yapar"],
    "OS"          : ["İki Yönlü Oynar", "Topla İleri Çıkar"],
    "DM"          : ["Takımı Topla Besler", "Önde Durur"],
    "DOS"         : ["Havada Güçlü", "İlk Topa Yetişir"],
    "KB"          : ["Üst Üste Çakılır", "Kanat Yardımı Sever"],
    "KB (Sol)"    : ["Sol Kanat Yardımı Yapar", "Hücuma Katılır"],
    "KB (Sağ)"    : ["Sağ Kanat Yardımı Yapar", "Hücuma Katılır"],
    "D"           : ["Basit Oynar", "Sert Müdahaleden Kaçınmaz"],
    "D (Sol)"     : ["Basit Oynar", "Sol Kanat Desteği Verir"],
    "D (Sağ)"     : ["Basit Oynar", "Sağ Kanat Desteği Verir"],
    "KL"          : ["Uzun Degaj Atar", "Kalesinden Çıkmayı Sever"],

}

ROLE_POOLS = {
    "D"           : ["Standart Bek (S)", "Standart Bek (D)", "Standart Bek (H)", "Çakılı Bek (S)", "İki Yönlü Bek (D)",
                     "İki Yönlü Bek (H)", "Sahte Bek (D)", "Sahte Bek (H)"],
    "D (Sol)"     : ["Standart Sol Bek (S)", "Standart Sol Bek (D)", "Standart Sol Bek (H)", "Çakılı Sol Bek (S)",
                     "İki Yönlü Sol Bek (D)", "İki Yönlü Sol Bek (H)", "Sahte Sol Bek (D)", "Sahte Sol Bek (H)"],
    "D (Sağ)"     : ["Standart Sağ Bek (S)", "Standart Sağ Bek (D)", "Standart Sağ Bek (H)", "Çakılı Sağ Bek (S)",
                     "İki Yönlü Sağ Bek (D)", "İki Yönlü Sağ Bek (H)", "Sahte Sağ Bek (D)", "Sahte Sağ Bek (H)"],
    "KB"          : ["Kanat Bek (S)", "Kanat Bek (D)", "Kanat Bek (H)", "Standart Bek (D)", "İki Yönlü Bek (H)"], "KB (Sol)"                                                                                    : ["Sol Kanat Bek (S)", "Sol Kanat Bek (D)", "Sol Kanat Bek (H)", "Standart Sol Bek (D)", "İki Yönlü Sol Bek (H)"],
    "KB (Sağ)"    : ["Sağ Kanat Bek (S)", "Sağ Kanat Bek (D)", "Sağ Kanat Bek (H)", "Standart Sağ Bek (D)", "İki Yönlü Sağ Bek (H)"],
    "DOS"         : ["Standart Stoper (S)", "Standart Stoper (K)", "Pasör Stoper (S)", "Çakılı Stoper (S)", "Libero (D)", "Libero (H)"],
    "DM"          : ["Ön Libero (S)", "Defansif Orta Saha (S)", "Defansif Orta Saha (D)", "Regista (D)",
                     "Serbest Defansif (D)", "Defansif Oyun Kurucu (D)"],
    "OS"          : ["Savaşçı Orta Saha (S)", "Savaşçı Orta Saha (D)", "Gezgin Oyun Kurucu (D)", "İki Yönlü Orta Saha (D)",
                     "Dinamo (D)", "Mezzala (D)", "Mezzala (H)"],
    "KANAT"       : ["Defansif Kanat (S)", "Çalışkan Kanat (D)", "Kanat Oyun Kurucu (H)", "Ters Ayaklı Kanat (H)","Kanat Oyuncusu (H)"],
    "KANAT (Sol)" : ["Sol Kanat (S)", "Sol Çalışkan Kanat (D)", "Sol Kanat Oyun Kurucu (H)", "Sol Ters Ayaklı Kanat (H)",
                     "Sol Kanat Oyuncusu (H)"],
    "KANAT (Sağ)" : ["Sağ Kanat (S)", "Sağ Çalışkan Kanat (D)", "Sağ Kanat Oyun Kurucu (H)", "Sağ Ters Ayaklı Kanat (H)",
                     "Sağ Kanat Oyuncusu (H)"],
    "OOS"         : ["On Numara (H)", "Ofansif Orta Saha (D)", "Ofansif Orta Saha (H)", "Ofansif Oyun Kurucu (H)",
                     "Raumdeuter (H)", "Enganche (D)"],
    "KF"          : ["Kanat Forvet (D)", "Kanat Forvet (H)"],
    "KF (Sol)"    : ["Sol Kanat Forvet (D)", "Sol Kanat Forvet (H)", "Sol İçe Kesen Kanat (H)", "Sol İçe Kesen Kanat (D)"],
    "KF (Sağ)"    : ["Sağ Kanat Forvet (D)", "Sağ Kanat Forvet (H)", "Sağ İçe Kesen Kanat (H)", "Sağ İçe Kesen Kanat (D)"],
    "ST"          : ["Gizli Forvet (H)", "Fırsatçı Golcü (H)", "Yaratıcı Forvet (H)", "Sahte Forvet (D)", "Komple Forvet (H)",
                     "Pivot Santrafor (H)", "Çalışkan Forvet (S)", "Çalışkan Forvet (H)"],
    "KL"          : ["Kaleci (S)", "Libero Kaleci (S)", "Libero Kaleci (D)", "Libero Kaleci (H)"],

}

# =========================================================
# İKİNCİL MEVKİ SİSTEMİ
# =========================================================

# Her birincil mevki için FM mantığına göre olası ikincil mevkiler
SECONDARY_POSITION_POOL = {
    "ST"          : ["KF (Sol)", "KF (Sağ)", "OOS"],
    "KF (Sol)"    : ["ST", "KANAT (Sol)", "OOS", "KF (Sağ)"],
    "KF (Sağ)"    : ["ST", "KANAT (Sağ)", "OOS", "KF (Sol)"],
    "OOS"         : ["OS", "KF (Sol)", "KF (Sağ)"],
    "KANAT (Sol)" : ["KF (Sol)", "KB (Sol)", "OS"],
    "KANAT (Sağ)" : ["KF (Sağ)", "KB (Sağ)", "OS"],
    "OS"          : ["DM", "OOS", "KANAT (Sol)", "KANAT (Sağ)"],
    "DM"          : ["OS", "DOS"],
    "DOS"         : ["DM", "D (Sol)", "D (Sağ)"],
    "KB (Sol)"    : ["D (Sol)", "KANAT (Sol)", "OS"],
    "KB (Sağ)"    : ["D (Sağ)", "KANAT (Sağ)", "OS"],
    "D (Sol)"     : ["KB (Sol)", "DOS"],
    "D (Sağ)"     : ["KB (Sağ)", "DOS"],
    "KL"          : [],  # Kalecinin ikincil mevkisi olmaz

}

# FM yeterlilik seviyeleri (yüksekten düşüğe)
SECONDARY_LEVELS = ["Doğal Mevkii", "Başarılı", "Yetkin", "Zayıf"]

# Seviye renk kodları — kart üzerinde badge rengi
SECONDARY_LEVEL_COLORS = {
    "Doğal Mevkii" : "#2ecc71",  # yeşil   — neredeyse birincil gibi
    "Başarılı"     : "#f1c40f",  # sarı    — iyi ama birincil değil
    "Yetkin"       : "#e67e22",  # turuncu — orta seviye
    "Zayıf"        : "#e74c3c",  # kırmızı — zayıf yeterlilik

}


def generate_secondary_position(primary_pos: str):
    """
    İkincil mevki üretir. (mevki, seviye) tuple'ı veya None döndürür.

    AYARLANABILIRLER:
    ─────────────────────────────────────────────────────
    SECONDARY_PROB → ikincil mevki çıkma olasılığı
      0.40 = %40  ← şu an aktif
      0.70 = %70  (daha sık görmek istersen)
      1.00 = her oyuncuda
      0.00 = hiçbir oyuncuda

    level_weights → [Natural, Accomplished, Competent, Unconvincing]
      şu an [10, 30, 40, 20]:
        Doğal Mevkii      ~ %10  (nadir)
        Başarılı ~ %30  (orta)
        Yetkin    ~ %40  (en sık)
        Zayıf ~ %20  (nadir)
    ─────────────────────────────────────────────────────
    """
    # ── OLASILIĞI BURADAN DEĞİŞTİR ───────────────────────────────
    SECONDARY_PROB = float(st.session_state.get('cfg_sec_prob', 0.40))  # Sidebar'dan
    # ─────────────────────────────────────────────────────────────

    pool = SECONDARY_POSITION_POOL.get(primary_pos, [])
    if not pool:
        return None  # Kaleci veya tanımsız mevki

    if random.random() > SECONDARY_PROB:
        return None  # Bu oyuncuda 2. mevki yok

    sec_pos = random.choice(pool)

    # ── SEVİYE AĞIRLIKLARINI BURADAN DEĞİŞTİR ────────────────────
    # Sıra: [Natural, Accomplished, Competent, Unconvincing]
    level_weights = [10, 30, 40, 20]
    # ─────────────────────────────────────────────────────────────
    level = random.choices(SECONDARY_LEVELS, weights=level_weights)[0]
    return (sec_pos, level)


# =========================================================
# ATTRIBUTE AĞIRLIKLARI
# =========================================================

# =========================================================
# ATTRIBUTE AĞIRLIKLARI — FMScout TAM TABLO
# Kaynak: FM Scout CA hesaplama metodolojisi
#
# KRİTİK DEĞİŞİKLİK: Artık TÜM sıfır-olmayan attr'lar dahil.
# Eski versiyon yalnızca ~13 "key attr" kullanıyordu → CA şişiriliyordu.
# Şimdi FM gibi ~30 attr kullanılıyor → FM CA'ya ±3-5 yakınlık.
#
# FM'de CA'ya SIFIR katkı yapan attr'lar (hiç eklenmedi):
#   Agresiflik, Liderlik, Kararlılık, Özel Yetenek, Vücut Zindeliği
# =========================================================

ATTRIBUTE_WEIGHTS = {

    # ── KALECİ (GK) — GERÇEK FM (ALİSSON = EDERSON — AYNI!) ──────
    "KL": {
        "Karar Alma"         : 4.0,  # FM=10 ← EN YÜKSEK
        "Çeviklik"           : 3.2,  # FM=8
        "Refleksler"         : 3.2,  # FM=8
        "Elle Kontrol"       : 3.2,  # FM=8
        "Konsantrasyon"      : 2.4,  # FM=6
        "Cesaret"            : 2.4,  # FM=6
        "Hızlanma"           : 2.4,  # FM=6
        "Bölge Hakimiyeti"   : 2.4,  # FM=6
        "Hava Topları"       : 2.4,  # FM=6
        "Pozisyon Alma"      : 2.0,  # FM=5
        "Degaj"              : 2.0,  # FM=5
        "İletişim"           : 2.0,  # FM=5
        "Güç"                : 1.6,  # FM=4
        "Birebir"            : 1.6,  # FM=4
        "Hız"                : 1.2,  # FM=3
        "Önsezi"             : 1.2,  # FM=3
        "Elle Oyun Başlatma" : 1.2,  # FM=3
        "Pas (K)"            : 1.2,  # FM=3  (GK'nın pasosu)
        "Soğukkanlılık"      : 0.8,  # FM=2
        "Denge"              : 0.8,  # FM=2
        "Liderlik"           : 0.8,  # FM=2
        "Takım Oyunu"        : 0.8,  # FM=2
        "Zıplama"            : 0.4,  # FM=1
        "Dayanıklılık"       : 0.4,  # FM=1
        "Çalışkanlık"        : 0.4,  # FM=1
        "Teknik"             : 0.4,  # FM=1
        "İlk Kontrol (K)"    : 0.4,  # FM=1
        "Vizyon"             : 0.4,  # FM=1
        "Kafa Vuruşu"        : 0.4,  # FM=1

    },

    # ── STOPER (DC) — GERÇEK FM VERİSİ (VAN DİJK) ────────────────
    "DOS": {
        "Karar Alma"             : 4.0,  # FM=10 ← EN YÜKSEK
        "Pozisyon Alma"          : 3.2,  # FM=8
        "Markaj"                 : 3.2,  # FM=8
        "Çeviklik"               : 2.4,  # FM=6
        "Zıplama"                : 2.4,  # FM=6
        "Güç"                    : 2.4,  # FM=6
        "Hızlanma"               : 2.4,  # FM=6
        "Hız"                    : 2.0,  # FM=5
        "Önsezi"                 : 2.0,  # FM=5
        "Top Kapma"              : 2.0,  # FM=5
        "Kafa Vuruşu"            : 2.0,  # FM=5
        "Konsantrasyon"          : 1.6,  # FM=4
        "Dayanıklılık"           : 1.2,  # FM=3
        "Soğukkanlılık"          : 0.8,  # FM=2
        "Cesaret"                : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Liderlik"               : 0.8,  # FM=2
        "Çalışkanlık"            : 0.8,  # FM=2
        "İlk Kontrol"            : 0.8,  # FM=2
        "Pas"                    : 0.8,  # FM=2
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Takım Oyunu"            : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Teknik"                 : 0.4,  # FM=1
        "Vizyon"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Topsuz Alan"            : 0.4,  # FM=1
        "Uzaktan Şut"            : 0.4,  # FM=1
        "Bitiricilik"            : 0.4,  # FM=1
        "Dripling"               : 0.4,  # FM=1
        "Orta Yapma"             : 0.4,  # FM=1

    },

    # ── BEK (DRL) — GERÇEK FM (ROBERTSON = TRİPPİER — AYNI!) ─────
    "D": {
        "Hızlanma"               : 3.2,  # FM=8 ← EN YÜKSEK
        "Dayanıklılık"           : 2.8,  # FM=7
        "Hız"                    : 2.4,  # FM=6
        "Çeviklik"               : 2.0,  # FM=5
        "Karar Alma"             : 2.0,  # FM=5
        "Güç"                    : 1.6,  # FM=4
        "Konsantrasyon"          : 1.2,  # FM=3
        "Teknik"                 : 1.2,  # FM=3
        "İlk Kontrol"            : 1.2,  # FM=3
        "Pozisyon Alma"          : 1.2,  # FM=3
        "Önsezi"                 : 1.2,  # FM=3
        "Top Kapma"              : 1.2,  # FM=3
        "Pas"                    : 1.2,  # FM=3
        "Orta Yapma"             : 1.2,  # FM=3
        "Soğukkanlılık"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Çalışkanlık"            : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Vizyon"                 : 0.8,  # FM=2
        "Topsuz Alan"            : 0.8,  # FM=2
        "Markaj"                 : 0.8,  # FM=2
        "Dripling"               : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Uzaktan Şut"            : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1
        "Bitiricilik"            : 0.4,  # FM=1

    },

    # ── KANAT BEK (WBRL) — GERÇEK FM (GOSENS) ────────────────────
    "KB": {
        "Hızlanma"               : 2.8,  # FM=7
        "Karar Alma"             : 2.8,  # FM=7
        "Çeviklik"               : 2.4,  # FM=6
        "Dayanıklılık"           : 2.4,  # FM=6
        "Hız"                    : 2.0,  # FM=5
        "Konsantrasyon"          : 1.6,  # FM=4
        "Güç"                    : 1.6,  # FM=4
        "Pozisyon Alma"          : 1.6,  # FM=4
        "Top Kapma"              : 1.6,  # FM=4
        "İlk Kontrol"            : 1.2,  # FM=3
        "Önsezi"                 : 1.2,  # FM=3
        "Markaj"                 : 1.2,  # FM=3
        "Soğukkanlılık"          : 0.8,  # FM=2
        "Cesaret"                : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Zıplama"                : 0.8,  # FM=2
        "Çalışkanlık"            : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Teknik"                 : 0.8,  # FM=2
        "Vizyon"                 : 0.8,  # FM=2
        "Pas"                    : 0.8,  # FM=2
        "Kafa Vuruşu"            : 0.8,  # FM=2
        "Orta Yapma"             : 0.8,  # FM=2
        "Liderlik"               : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Topsuz Alan"            : 0.4,  # FM=1
        "Uzaktan Şut"            : 0.4,  # FM=1
        "Bitiricilik"            : 0.4,  # FM=1
        "Dripling"               : 0.4,  # FM=1

    },

    # ── DEFANSİF OS (DM) — GERÇEK FM (CASEMIRO) ──────────────────
    # NOT: FM'de DM ve MC için AYNI ağırlık profili kullanılıyor!
    "DM": {
        "Karar Alma"             : 3.2,  # FM=8 ← EN YÜKSEK
        "Top Kapma"              : 2.8,  # FM=7
        "Çeviklik"               : 2.4,  # FM=6
        "Hızlanma"               : 2.4,  # FM=6
        "Güç"                    : 2.0,  # FM=5
        "Pozisyon Alma"          : 2.0,  # FM=5
        "Önsezi"                 : 2.0,  # FM=5
        "Hız"                    : 1.6,  # FM=4
        "Dayanıklılık"           : 1.6,  # FM=4
        "Çalışkanlık"            : 1.6,  # FM=4
        "İlk Kontrol"            : 1.6,  # FM=4
        "Vizyon"                 : 1.6,  # FM=4
        "Pas"                    : 1.6,  # FM=4
        "Konsantrasyon"          : 1.2,  # FM=3
        "Teknik"                 : 1.2,  # FM=3
        "Markaj"                 : 1.2,  # FM=3
        "Uzaktan Şut"            : 1.2,  # FM=3
        "Soğukkanlılık"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Bitiricilik"            : 0.8,  # FM=2
        "Dripling"               : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Topsuz Alan"            : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1
        "Orta Yapma"             : 0.4,  # FM=1

    },

    # ── MERKEZ OS (MC) — GERÇEK FM (KANTÉ) ───────────────────────
    # NOT: FM'de DM ve MC için AYNI ağırlık profili — Casemiro = Kanté!
    "OS": {
        "Karar Alma"             : 3.2,  # FM=8 ← EN YÜKSEK
        "Top Kapma"              : 2.8,  # FM=7
        "Çeviklik"               : 2.4,  # FM=6
        "Hızlanma"               : 2.4,  # FM=6
        "Güç"                    : 2.0,  # FM=5
        "Pozisyon Alma"          : 2.0,  # FM=5
        "Önsezi"                 : 2.0,  # FM=5
        "Hız"                    : 1.6,  # FM=4
        "Dayanıklılık"           : 1.6,  # FM=4
        "Çalışkanlık"            : 1.6,  # FM=4
        "İlk Kontrol"            : 1.6,  # FM=4
        "Vizyon"                 : 1.6,  # FM=4
        "Pas"                    : 1.6,  # FM=4
        "Konsantrasyon"          : 1.2,  # FM=3
        "Teknik"                 : 1.2,  # FM=3
        "Markaj"                 : 1.2,  # FM=3
        "Uzaktan Şut"            : 1.2,  # FM=3
        "Soğukkanlılık"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Bitiricilik"            : 0.8,  # FM=2
        "Dripling"               : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Topsuz Alan"            : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1
        "Orta Yapma"             : 0.4,  # FM=1

    },

    # ── KANAT (MRL) — GERÇEK FM (ARNOLD) ────────────────────────
    # NOT: FM bu pozisyonu "playmaker" ağırlıklı değerlendiriyor!
    # Arnold gibi oyuncular DRL değil MRL olarak CA hesaplanıyor.
    # Hızlanma değil Karar Alma en yüksek — sürpriz!
    "KANAT": {
        "Karar Alma"             : 2.8,  # FM=7 ← EN YÜKSEK (beklenmedik!)
        "Çeviklik"               : 2.4,  # FM=6
        "Dayanıklılık"           : 2.4,  # FM=6
        "Hızlanma"               : 2.4,  # FM=6
        "İlk Kontrol"            : 2.4,  # FM=6
        "Vizyon"                 : 2.4,  # FM=6
        "Pas"                    : 2.4,  # FM=6
        "Hız"                    : 2.0,  # FM=5
        "Güç"                    : 1.6,  # FM=4
        "Teknik"                 : 1.6,  # FM=4
        "Soğukkanlılık"          : 1.2,  # FM=3
        "Çalışkanlık"            : 1.2,  # FM=3
        "Pozisyon Alma"          : 1.2,  # FM=3
        "Önsezi"                 : 1.2,  # FM=3
        "Top Kapma"              : 1.2,  # FM=3
        "Topsuz Alan"            : 1.2,  # FM=3
        "Markaj"                 : 1.2,  # FM=3
        "Uzaktan Şut"            : 1.2,  # FM=3
        "Konsantrasyon"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Bitiricilik"            : 0.8,  # FM=2
        "Dripling"               : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1
        "Orta Yapma"             : 0.4,  # FM=1

    },

    # ── KANAT FORVET (AMRL) — GERÇEK FM (SAKA=VINİCİUS) ──────────
    # NOT: FM'de KF sol ve KF sağ AYNI ağırlık profili!
    "KF": {
        "Hız"                    : 4.0,  # FM=10 ← MAKSIMUM
        "Hızlanma"               : 4.0,  # FM=10 ← MAKSIMUM
        "Dayanıklılık"           : 2.8,  # FM=7
        "Çeviklik"               : 2.4,  # FM=6
        "İlk Kontrol"            : 2.0,  # FM=5
        "Karar Alma"             : 2.0,  # FM=5
        "Dripling"               : 2.0,  # FM=5
        "Orta Yapma"             : 2.0,  # FM=5
        "Teknik"                 : 1.6,  # FM=4
        "Soğukkanlılık"          : 1.2,  # FM=3
        "Güç"                    : 1.2,  # FM=3
        "Çalışkanlık"            : 1.2,  # FM=3
        "Önsezi"                 : 1.2,  # FM=3
        "Vizyon"                 : 1.2,  # FM=3
        "Konsantrasyon"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Top Kapma"              : 0.8,  # FM=2
        "Pas"                    : 0.8,  # FM=2
        "Topsuz Alan"            : 0.8,  # FM=2
        "Uzaktan Şut"            : 0.8,  # FM=2
        "Bitiricilik"            : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Pozisyon Alma"          : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Markaj"                 : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1

    },

    # ── FORVET ARKASI (AMC) — GERÇEK FM (DYBALA) ─────────────────
    "OOS": {
        "Hızlanma"               : 3.6,  # FM=9 ← EN YÜKSEK
        "Hız"                    : 2.8,  # FM=7
        "Çeviklik"               : 2.4,  # FM=6
        "Dayanıklılık"           : 2.4,  # FM=6
        "Karar Alma"             : 2.4,  # FM=6
        "Vizyon"                 : 2.4,  # FM=6
        "Teknik"                 : 2.0,  # FM=5
        "İlk Kontrol"            : 2.0,  # FM=5
        "Pas"                    : 1.6,  # FM=4
        "Soğukkanlılık"          : 1.2,  # FM=3
        "Güç"                    : 1.2,  # FM=3
        "Çalışkanlık"            : 1.2,  # FM=3
        "Önsezi"                 : 1.2,  # FM=3
        "Topsuz Alan"            : 1.2,  # FM=3
        "Uzaktan Şut"            : 1.2,  # FM=3
        "Bitiricilik"            : 1.2,  # FM=3
        "Dripling"               : 1.2,  # FM=3
        "Konsantrasyon"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Takım Oyunu"            : 0.8,  # FM=2
        "Pozisyon Alma"          : 0.8,  # FM=2
        "Top Kapma"              : 0.8,  # FM=2
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Zıplama"                : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Markaj"                 : 0.4,  # FM=1
        "Kafa Vuruşu"            : 0.4,  # FM=1
        "Orta Yapma"             : 0.4,  # FM=1

    },

    # ── SANTRAFOR (SC) — GERÇEK FM VERİSİ (HAALAND) ──────────────
    "ST": {
        # Teknik
        "Hızlanma"               : 4.0,  # FM=10 ← EN YÜKSEK
        "Bitiricilik"            : 3.2,  # FM=8
        "Hız"                    : 2.8,  # FM=7
        "Soğukkanlılık"          : 2.4,  # FM=6
        "Çeviklik"               : 2.4,  # FM=6
        "Dayanıklılık"           : 2.4,  # FM=6
        "Güç"                    : 2.4,  # FM=6
        "İlk Kontrol"            : 2.4,  # FM=6
        "Topsuz Alan"            : 2.4,  # FM=6
        "Kafa Vuruşu"            : 2.4,  # FM=6
        "Zıplama"                : 2.0,  # FM=5
        "Karar Alma"             : 2.0,  # FM=5  (FMScout'tan farklı : eskiden 2.4)
        "Önsezi"                 : 2.0,  # FM=5
        "Dripling"               : 2.0,  # FM=5
        "Teknik"                 : 1.6,  # FM=4
        "Konsantrasyon"          : 0.8,  # FM=2
        "Denge"                  : 0.8,  # FM=2
        "Çalışkanlık"            : 0.8,  # FM=2
        "Pozisyon Alma"          : 0.8,  # FM=2
        "Vizyon"                 : 0.8,  # FM=2
        "Pas"                    : 0.8,  # FM=2  (eskiden 0.4!)
        "Uzaktan Şut"            : 0.8,  # FM=2  (eskiden 0.4!)
        "Orta Yapma"             : 0.8,  # FM=2  (eskiden 0!)
        "Cesaret"                : 0.4,  # FM=1
        "Liderlik"               : 0.4,  # FM=1
        "Serbest Vuruş Kullanma" : 0.4,  # FM=1
        "Uzun Taç"               : 0.4,  # FM=1
        "Takım Oyunu"            : 0.4,  # FM=1
        "Korner"                 : 0.4,  # FM=1
        "Top Kapma"              : 0.4,  # FM=1
        "Penaltı Kullanma"       : 0.4,  # FM=1
        "Markaj"                 : 0.4,  # FM=1

    },
}


# =========================================================
# MEVKİ ARKETİPLERİ — POSITION_ATTR_MEANS override'ları
# Her arketip; sadece farklılaşan attr'ların mean değerlerini içerir.
# Listelenmeyen attr'lar POSITION_ATTR_MEANS'taki varsayılan değeri alır.
# =========================================================

ARCHETYPES = {

    # ── SANTRAFOR (ST) ────────────────────────────────────
    "ST": {
        "🎲 Rastgele": None,

        "🎯 Fırsatçı Golcü": {
            # Lewandowski, Müller tipi — ceza sahası yırtıcısı
            "Bitiricilik": 17, "Topsuz Alan": 16, "Soğukkanlılık": 16,
            "Penaltı Kullanma": 14, "Pozisyon Alma": 15, "Önsezi": 14,
            "Hız": 11, "Hızlanma": 11, "Güç": 10, "Kafa Vuruşu": 12,
            "Pas": 7, "Dripling": 9, "Teknik": 9,
        },

        "💪 Fiziksel Pres (Osimhen)": {
            # Hız + güç + agresif pressing
            "Hız": 17, "Hızlanma": 17, "Güç": 15, "Dayanıklılık": 15,
            "Çalışkanlık": 15, "Agresiflik": 14, "Cesaret": 14,
            "Topsuz Alan": 15, "Bitiricilik": 13,
            "Teknik": 8, "Pas": 6, "Dripling": 10, "Kafa Vuruşu": 13,
        },

        "⭐ Komple Forvet (Benzema)": {
            # Her şey dengeli — zayıf nokta yok
            "Bitiricilik": 14, "Dripling": 13, "İlk Kontrol": 14,
            "Teknik": 13, "Pas": 12, "Vizyon": 12,
            "Topsuz Alan": 14, "Soğukkanlılık": 14,
            "Hız": 13, "Hızlanma": 13, "Güç": 12,
            "Karar Alma": 14, "Önsezi": 13,
        },

        "🎭 Sahte Forvet (Firmino)": {
            # Oyun kurma, asist, pressing — gol ikincil
            "Pas": 14, "Vizyon": 14, "Takım Oyunu": 15,
            "Çalışkanlık": 15, "Topsuz Alan": 13, "İlk Kontrol": 13,
            "Dripling": 12, "Agresiflik": 13, "Markaj": 11,
            "Bitiricilik": 9, "Güç": 8, "Kafa Vuruşu": 7,
            "Hız": 12, "Hızlanma": 12,
            "_caps": {"Markaj": 12, "Agresiflik": 14},
        },

        "🏰 Target Man (Giroud)": {
            # Hava hakimiyeti, sırt dönük oyun, asist
            "Kafa Vuruşu": 17, "Güç": 16, "Zıplama": 16,
            "İlk Kontrol": 13, "Pas": 12, "Topsuz Alan": 14,
            "Cesaret": 14, "Karar Alma": 13,
            "Hız": 7, "Hızlanma": 7, "Dripling": 7,
            "Bitiricilik": 13, "Soğukkanlılık": 13,
        },

        "⚡ Hız Canavarı (Mbappé)": {
            # Hız + dripling + bitiricilik — gücü düşük
            "Hız": 19, "Hızlanma": 19, "Çeviklik": 16, "Denge": 14,
            "Dripling": 15, "Bitiricilik": 15, "Topsuz Alan": 14,
            "İlk Kontrol": 13, "Soğukkanlılık": 13,
            "Güç": 8, "Kafa Vuruşu": 7, "Pas": 8,
        },

        "🎨 Teknik Golcü (Del Piero)": {
            # Teknik + uzaktan şut + serbest vuruş
            "Teknik": 16, "Dripling": 15, "Uzaktan Şut": 15,
            "Serbest Vuruş Kullanma": 15, "İlk Kontrol": 15,
            "Soğukkanlılık": 14, "Bitiricilik": 14, "Vizyon": 12,
            "Hız": 11, "Hızlanma": 11, "Güç": 8, "Kafa Vuruşu": 8,
        },

            
    "🎪 Trequartista (Cantona/Ibra)": {
        # Serbest ruh — sahaya iner, yaratır, pozisyon tutmaz
        "Vizyon": 16, "Teknik": 15, "İlk Kontrol": 15, "Dripling": 14,
        "Soğukkanlılık": 16, "Karar Alma": 14, "Özel Yetenek": 15,
        "Güç": 15, "Bitiricilik": 13,
        "Çalışkanlık": 7, "Pozisyon Alma": 6, "Hız": 10,
    },

    "🔫 Penaltı Alıcısı (Suarez/Neymar)": {
        # Küçük, çevik, faul kazanan — faul çıkarımı yüksek
        "Çeviklik": 17, "Denge": 16, "Dripling": 16, "Hızlanma": 16,
        "Hız": 15, "Bitiricilik": 14, "Soğukkanlılık": 14,
        "Agresiflik": 13, "İlk Kontrol": 14,
        "Güç": 7, "Kafa Vuruşu": 6, "Zıplama": 8,
    },

    "🌙 Gece Yarısı Golcüsü": {
        # 90 dakika ortalıkta yoktur, son anda çıkar gol atar
        "Soğukkanlılık": 19, "Bitiricilik": 17, "Önsezi": 16,
        "Topsuz Alan": 15, "Karar Alma": 15, "Pozisyon Alma": 14,
        "Hız": 11, "Hızlanma": 11, "Güç": 11,
        "Çalışkanlık": 7, "Dayanıklılık": 8, "Teknik": 9,
    },

    "👑 Kaptan Santrafor (Totti/Shearer)": {
            # Her takımda bir tane — sahada lider, soyunma odasında otorite
            "Bitiricilik": 15, "İlk Kontrol": 14, "Kafa Vuruşu": 14,
            "Pas": 13, "Soğukkanlılık": 16, "Topsuz Alan": 15,
            "Liderlik": 17, "Karar Alma": 15, "Önsezi": 14,
            "Cesaret": 15, "Takım Oyunu": 15, "Kararlılık": 16,
            "Hız": 11, "Güç": 14, "Dayanıklılık": 15,
        },

    "🏃 Pressing Canavarı (Vardy)": {
            # Yüksek tempo, hız, çalışkanlık, topsuz alan
            # Markaj: pressing ST rakip stoper'ı markajlar
            "Hız": 17, "Hızlanma": 17, "Dayanıklılık": 16,
            "Çalışkanlık": 16, "Topsuz Alan": 16, "Agresiflik": 14,
            "Markaj": 13, "Bitiricilik": 13, "Soğukkanlılık": 12,
            "Teknik": 8, "Pas": 7, "Kafa Vuruşu": 9, "Güç": 10,
            "_caps": {"Markaj": 14, "Agresiflik": 16},
        },
    },

    # ── KANAT FORVET (KF) ─────────────────────────────────
    "KF": {
        "🎲 Rastgele": None,

        "⚡ Hız Kanadı (Saka/Vinicius)": {
            "Hız": 17, "Hızlanma": 17, "Çeviklik": 16,
            "Dripling": 16, "İlk Kontrol": 14, "Topsuz Alan": 13,
            "Bitiricilik": 12, "Orta Yapma": 10,
            "Güç": 8, "Kafa Vuruşu": 6,
        },

        "🔄 İçe Kesen Golcü (Robben/Salah)": {
            "Dripling": 16, "Uzaktan Şut": 15, "Bitiricilik": 14,
            "Hız": 15, "Hızlanma": 15, "Soğukkanlılık": 13,
            "İlk Kontrol": 14, "Teknik": 14,
            "Orta Yapma": 6, "Kafa Vuruşu": 6,
        },

            
    "🎩 Playmaker Kanat (Silva/Iniesta)": {
        # Kanat mevkiinde oynayan oyun kurucu — dar + pas odaklı
        "Pas": 15, "Vizyon": 15, "İlk Kontrol": 15, "Teknik": 15,
        "Dripling": 14, "Çeviklik": 15, "Denge": 15, "Soğukkanlılık": 14,
        "Karar Alma": 14, "Takım Oyunu": 14,
        "Hız": 11, "Güç": 8, "Orta Yapma": 9,
    },

    "🔋 Yorulmaz Kanat (Milner/Valencia)": {
        # Teknik zayıf ama 90 dakika aynı tempoda koşar
        "Dayanıklılık": 18, "Çalışkanlık": 17, "Vücut Zindeliği": 17,
        "Hız": 15, "Hızlanma": 15, "Orta Yapma": 13,
        "Takım Oyunu": 15, "Kararlılık": 15,
        "Teknik": 9, "Dripling": 10, "Bitiricilik": 9,
    },

    "👑 Kaptan Kanat (Giggs/Figo)": {
            "Dripling": 14, "Hız": 14, "Hızlanma": 14,
            "Pas": 14, "Vizyon": 14, "İlk Kontrol": 14,
            "Liderlik": 17, "Karar Alma": 15, "Cesaret": 14,
            "Kararlılık": 16, "Takım Oyunu": 15,
            "Dayanıklılık": 15, "Vücut Zindeliği": 15,
        },

    "🎯 Orta Açan Kanat (Trippier/Trent)": {
            "Orta Yapma": 16, "Pas": 15, "Serbest Vuruş Kullanma": 14,
            "Korner": 15, "Vizyon": 13, "İlk Kontrol": 13,
            "Hız": 12, "Hızlanma": 12, "Dripling": 11,
            "Bitiricilik": 8, "Güç": 9,
        },

        "💪 Çalışkan Kanat (Sterling)": {
            "Dayanıklılık": 16, "Çalışkanlık": 16, "Hız": 15,
            "Hızlanma": 15, "Dripling": 14, "Topsuz Alan": 14,
            "Bitiricilik": 12, "İlk Kontrol": 13,
            "Kafa Vuruşu": 7, "Güç": 9,
        },
    },

    # ── FORVET ARKASI (OOS) ───────────────────────────────
    "OOS": {
        "🎲 Rastgele": None,

        "🎨 Yaratıcı Oyun Kurucu (De Bruyne)": {
            "Pas": 17, "Vizyon": 17, "Uzaktan Şut": 15,
            "İlk Kontrol": 15, "Teknik": 15, "Karar Alma": 15,
            "Serbest Vuruş Kullanma": 14, "Önsezi": 14,
            "Hız": 12, "Güç": 9,
        },

        "🎯 Golcü 10 (Dybala/Özil)": {
            "Dripling": 15, "Teknik": 16, "İlk Kontrol": 15,
            "Bitiricilik": 13, "Soğukkanlılık": 14, "Vizyon": 14,
            "Uzaktan Şut": 13, "Pas": 14,
            "Güç": 7, "Kafa Vuruşu": 6, "Dayanıklılık": 9,
        },

            
    "🦎 Kameleon (Müller)": {
        # Mevki belirsiz — ne ST ne OOS, ikisi de değil ikisi de
        # Topsuz alan + önsezi ile her yerde gol tehlikesi
        "Topsuz Alan": 18, "Önsezi": 17, "Karar Alma": 16,
        "Pas": 13, "Takım Oyunu": 16, "Bitiricilik": 14,
        "Soğukkanlılık": 15, "Pozisyon Alma": 14,
        "Hız": 13, "Güç": 12, "Teknik": 10,
        "Dripling": 9, "Kafa Vuruşu": 12,
    },

    "👑 Kaptan Forvet Arkası (Totti/Zidane)": {
            "Pas": 16, "Vizyon": 16, "İlk Kontrol": 15,
            "Teknik": 14, "Dripling": 13, "Soğukkanlılık": 16,
            "Liderlik": 17, "Karar Alma": 16, "Kararlılık": 15,
            "Cesaret": 14, "Takım Oyunu": 15,
            "Hız": 11, "Güç": 11,
        },

    "🤝 İkinci Forvet (Griezmann)": {
            "Topsuz Alan": 16, "Bitiricilik": 14, "Soğukkanlılık": 14,
            "Dripling": 13, "Hız": 14, "Hızlanma": 14,
            "Çalışkanlık": 14, "Takım Oyunu": 13,
            "Pas": 10, "Vizyon": 11,
        },
    },

    # ── STOPER (DOS) ──────────────────────────────────────
    "DOS": {
        "🎲 Rastgele": None,

        "🏰 Sert Adam (Van Dijk)": {
            "Markaj": 16, "Top Kapma": 16, "Güç": 17, "Zıplama": 16,
            "Kafa Vuruşu": 16, "Cesaret": 15, "Pozisyon Alma": 15,
            "Hız": 11, "Pas": 10, "Dripling": 5,
        },

        "🎨 Pasör Stoper (Beckenbauer/Baresi)": {
            "Pas": 14, "Vizyon": 13, "İlk Kontrol": 13, "Teknik": 12,
            "Karar Alma": 15, "Pozisyon Alma": 15, "Önsezi": 13,
            "Markaj": 14, "Top Kapma": 14, "Güç": 13,
            "Hız": 10,
        },

            
    "🦅 Libero (Beckenbauer/Baresi)": {
        # Topla çıkan, sweeper — neredeyse orta saha gibi pas yapar
        "Pas": 15, "Vizyon": 13, "Karar Alma": 15, "İlk Kontrol": 13,
        "Teknik": 12, "Pozisyon Alma": 16, "Konsantrasyon": 15,
        "Markaj": 14, "Top Kapma": 13, "Güç": 13,
        "Hız": 12, "Hızlanma": 12,
    },

    "👊 Kirli İşçi Stoper (Vidic/Terry)": {
        # Markaj + agresiflik + cesaret maksimum — topla değil savaşla önler
        "Markaj": 17, "Agresiflik": 16, "Cesaret": 17,
        "Güç": 16, "Kafa Vuruşu": 16, "Zıplama": 15,
        "Top Kapma": 15, "Kararlılık": 16, "Konsantrasyon": 14,
        "Pas": 7, "Teknik": 6, "Hız": 10,
        "_caps": {"Markaj": 18, "Agresiflik": 17},
    },

    "👑 Kaptan Stoper (Maldini/Cannavaro)": {
            "Markaj": 16, "Top Kapma": 15, "Pozisyon Alma": 16,
            "Güç": 14, "Kafa Vuruşu": 15, "Zıplama": 15,
            "Liderlik": 17, "Cesaret": 17, "Kararlılık": 16,
            "Karar Alma": 16, "Soğukkanlılık": 16, "Konsantrasyon": 15,
            "Hız": 12, "Dayanıklılık": 15,
        },

    "⚡ Hızlı Stoper (Koulibaly)": {
            "Hız": 15, "Hızlanma": 15, "Çeviklik": 13,
            "Markaj": 15, "Top Kapma": 15, "Güç": 15,
            "Pozisyon Alma": 14, "Kafa Vuruşu": 14,
            "Pas": 9, "Teknik": 8,
        },
    },

    # ── DEFANSİF ORTA SAHA (DM) ───────────────────────────
    "DM": {
        "🎲 Rastgele": None,

            
    "⚙️ Motor Oyuncu (Kanté)": {
        # Sadece koşar, top çalar, dağıtır — teknik bekleme
        "Hız": 16, "Hızlanma": 16, "Dayanıklılık": 18,
        "Top Kapma": 17, "Çalışkanlık": 17, "Pozisyon Alma": 15,
        "Agresiflik": 14, "Güç": 13, "Kararlılık": 15,
        "Pas": 11, "Vizyon": 8, "Teknik": 9,
        "_caps": {"Agresiflik": 16},
    },

    "👑 Kaptan DM (Mascherano/Busquets)": {
            "Top Kapma": 15, "Pozisyon Alma": 15, "Karar Alma": 16,
            "Pas": 14, "Çalışkanlık": 15, "Dayanıklılık": 15,
            "Liderlik": 17, "Cesaret": 15, "Kararlılık": 16,
            "Soğukkanlılık": 15, "Takım Oyunu": 16,
            "Hız": 10, "Güç": 13,
        },

    "💥 Destroyer (Kanté/Casemiro)": {
            "Top Kapma": 17, "Agresiflik": 16, "Çalışkanlık": 16,
            "Pozisyon Alma": 16, "Dayanıklılık": 15, "Karar Alma": 14,
            "Güç": 14, "Markaj": 13,
            "Pas": 10, "Vizyon": 8, "Dripling": 8,
        },

        "🎨 Regista (Pirlo/Busquets)": {
            "Pas": 16, "Vizyon": 16, "Teknik": 15, "İlk Kontrol": 15,
            "Karar Alma": 16, "Soğukkanlılık": 15, "Önsezi": 14,
            "Top Kapma": 13, "Pozisyon Alma": 14,
            "Hız": 8, "Hızlanma": 8, "Güç": 9,
        },
    },

    # ── MERKEZ ORTA SAHA (OS) ─────────────────────────────
    "OS": {
        "🎲 Rastgele": None,

        "⚙️ Box-to-Box (Pogba/Vidal)": {
            "Dayanıklılık": 16, "Güç": 15, "Hız": 13, "Hızlanma": 13,
            "Pas": 14, "Uzaktan Şut": 13, "Top Kapma": 14,
            "Çalışkanlık": 15, "Karar Alma": 14,
        },

        "🎨 Oyun Kurucu (Iniesta/Modric)": {
            "Teknik": 16, "İlk Kontrol": 16, "Pas": 15, "Vizyon": 15,
            "Dripling": 15, "Karar Alma": 15, "Soğukkanlılık": 14,
            "Çeviklik": 14, "Denge": 14,
            "Güç": 8, "Kafa Vuruşu": 7,
        },

            
    "🎯 Set Piece Uzmanı (Pirlo/Beckham)": {
        # Durağan topun efendisi — serbest vuruş + korner + pas
        "Serbest Vuruş Kullanma": 17, "Korner": 16, "Pas": 16,
        "Uzaktan Şut": 15, "Teknik": 15, "Vizyon": 15,
        "İlk Kontrol": 14, "Soğukkanlılık": 15, "Karar Alma": 14,
        "Hız": 8, "Hızlanma": 8, "Dayanıklılık": 11, "Güç": 9,
    },

    "🧠 Derin Playmaker (Xavi/Kroos)": {
        # Her şeyi ilk temasta çözer, asla koşmaz ama top hep onda
        "Pas": 18, "Vizyon": 17, "İlk Kontrol": 16, "Teknik": 16,
        "Soğukkanlılık": 17, "Karar Alma": 17, "Konsantrasyon": 15,
        "Takım Oyunu": 15, "Pozisyon Alma": 14,
        "Hız": 8, "Hızlanma": 8, "Güç": 8, "Dayanıklılık": 12,
    },

    "⚡ Geç Hücumcu OS (Lampard 2/Scholes)": {
        # Savunma + orta + hücum — ceza sahasına girer, gol atar
        "Uzaktan Şut": 16, "Bitiricilik": 14, "Pas": 14,
        "Dayanıklılık": 15, "Çalışkanlık": 14, "Top Kapma": 13,
        "Güç": 13, "Hız": 12, "Hızlanma": 12,
        "Soğukkanlılık": 14, "Karar Alma": 14,
    },

    "👑 Kaptan OS (Gerrard/Lampard)": {
            "Pas": 15, "Uzaktan Şut": 14, "Dayanıklılık": 16,
            "Çalışkanlık": 15, "Top Kapma": 13, "Bitiricilik": 13,
            "Liderlik": 17, "Cesaret": 16, "Kararlılık": 17,
            "Karar Alma": 15, "Soğukkanlılık": 15, "Takım Oyunu": 14,
            "Güç": 14, "Hız": 12,
        },

    "💪 Savaşçı OS (Gattuso)": {
            "Agresiflik": 16, "Çalışkanlık": 16, "Dayanıklılık": 16,
            "Top Kapma": 15, "Güç": 15, "Markaj": 13,
            "Karar Alma": 13, "Cesaret": 15,
            "Pas": 10, "Vizyon": 8, "Teknik": 9,
        },

        "🎯 Gol Atan OS (Lampard/Gerrard)": {
            "Uzaktan Şut": 16, "Bitiricilik": 14, "Soğukkanlılık": 14,
            "Pas": 14, "Karar Alma": 14, "Dayanıklılık": 15,
            "Güç": 14, "Hız": 12, "Hızlanma": 12,
            "Top Kapma": 11, "Vizyon": 13,
        },
    },

    # ── KANAT (KANAT Sol/Sağ) ─────────────────────────────
    "KANAT": {
        "🎲 Rastgele": None,

        "⚡ Hız Kanadı (Adama/Leao)": {
            "Hız": 18, "Hızlanma": 18, "Çeviklik": 15, "Denge": 13,
            "Dripling": 15, "İlk Kontrol": 13, "Dayanıklılık": 14,
            "Orta Yapma": 11, "Bitiricilik": 10,
            "Güç": 9, "Pas": 8, "Teknik": 9,
        },

        "🎨 Teknik Kanat (Iniesta/Silva)": {
            "Teknik": 16, "Dripling": 16, "İlk Kontrol": 15,
            "Vizyon": 14, "Pas": 14, "Çeviklik": 15, "Denge": 15,
            "Hız": 12, "Hızlanma": 12,
            "Güç": 7, "Orta Yapma": 10,
        },

        "🔄 Geniş Kanat (Beckham/Navas)": {
            "Orta Yapma": 17, "Korner": 15, "Serbest Vuruş Kullanma": 14,
            "Pas": 14, "Çalışkanlık": 15, "Dayanıklılık": 15,
            "Hız": 13, "Hızlanma": 13, "İlk Kontrol": 13,
            "Bitiricilik": 8, "Güç": 9,
        },

        "💪 Çalışkan Kanat (Mane/Sancho)": {
            "Dayanıklılık": 16, "Çalışkanlık": 16, "Agresiflik": 12,
            "Hız": 15, "Hızlanma": 15, "Dripling": 14,
            "Bitiricilik": 12, "Topsuz Alan": 13,
            "Top Kapma": 10, "Güç": 10,
        },

        "🎯 İçe Kesen Kanat (Robben/Mkhitaryan)": {
            "Dripling": 16, "Uzaktan Şut": 15, "Bitiricilik": 13,
            "Teknik": 15, "İlk Kontrol": 14, "Hız": 14, "Hızlanma": 14,
            "Çeviklik": 15, "Soğukkanlılık": 13,
            "Orta Yapma": 6, "Güç": 8,
        },
    },

    # ── KANAT BEK (KB Sol/Sağ) ────────────────────────────
    "KB": {
        "🎲 Rastgele": None,

        "⚡ Hücum Bekı (Trent/Alba)": {
            "Orta Yapma": 16, "Pas": 15, "Hız": 16, "Hızlanma": 16,
            "İlk Kontrol": 14, "Çalışkanlık": 15, "Dayanıklılık": 15,
            "Korner": 13, "Vizyon": 13,
            "Markaj": 11, "Top Kapma": 11, "Güç": 10,
        },

        "🛡️ Defans Odaklı KB (Azpilicueta)": {
            "Markaj": 15, "Top Kapma": 15, "Pozisyon Alma": 15,
            "Güç": 14, "Dayanıklılık": 15, "Çalışkanlık": 15,
            "Karar Alma": 14, "Konsantrasyon": 14,
            "Orta Yapma": 9, "Pas": 10, "Hız": 12,
        },

            
    "🚂 Hücum Treni KB (Alaba/Blind)": {
        # Kanat bek ama orta sahada gibi oynuyor — pas + vizyon
        "Pas": 15, "Vizyon": 14, "İlk Kontrol": 14,
        "Orta Yapma": 14, "Hız": 14, "Hızlanma": 14,
        "Markaj": 12, "Top Kapma": 12, "Dayanıklılık": 15,
        "Karar Alma": 14, "Pozisyon Alma": 13,
    },

    "👑 Kaptan Kanat Bek (Alba/Lahm)": {
            "Hız": 13, "Orta Yapma": 13, "Markaj": 13, "Pas": 13,
            "Çalışkanlık": 15, "Dayanıklılık": 15, "Pozisyon Alma": 14,
            "Liderlik": 17, "Cesaret": 15, "Kararlılık": 16,
            "Karar Alma": 15, "Soğukkanlılık": 15, "Takım Oyunu": 16,
        },

    "🔄 Dengeli KB (Carvajal/Cancelo)": {
            "Hız": 14, "Hızlanma": 14, "Orta Yapma": 14,
            "Markaj": 13, "Top Kapma": 13, "Pas": 13,
            "Çalışkanlık": 14, "Dayanıklılık": 14,
            "Karar Alma": 13, "Pozisyon Alma": 13,
        },
    },

    # ── BAĞLAYAN BEK / TAM BEK (D Sol/Sağ) ──────────────
    "D": {
        "🎲 Rastgele": None,

        "🛡️ Klasik Bek (Cafu/Makelele)": {
            "Markaj": 15, "Top Kapma": 15, "Güç": 14,
            "Pozisyon Alma": 15, "Konsantrasyon": 14, "Karar Alma": 14,
            "Dayanıklılık": 14, "Çalışkanlık": 14,
            "Hız": 11, "Orta Yapma": 8, "Pas": 9,
        },

        "⚡ Modern Tam Bek (Hakimi/Davies)": {
            "Hız": 17, "Hızlanma": 17, "Çeviklik": 14,
            "Orta Yapma": 14, "Dripling": 13, "Dayanıklılık": 15,
            "Markaj": 13, "Top Kapma": 12, "Çalışkanlık": 15,
            "Güç": 11, "Pas": 12,
        },

            
    "🚀 Overlapping Bek (Roberto Carlos)": {
        # Hücuma katılan bek — sol/sağ fark etmez, sürekli üste çıkar
        "Hız": 17, "Hızlanma": 17, "Dayanıklılık": 16,
        "Orta Yapma": 14, "Uzaktan Şut": 13, "Çalışkanlık": 15,
        "Markaj": 12, "Top Kapma": 12, "Güç": 13,
        "Pas": 12, "Çeviklik": 14,
    },

    "👑 Kaptan Tam Bek (Cafu/Roberto Carlos)": {
            "Hız": 13, "Markaj": 13, "Top Kapma": 13,
            "Çalışkanlık": 16, "Dayanıklılık": 16, "Pozisyon Alma": 14,
            "Liderlik": 17, "Cesaret": 15, "Kararlılık": 16,
            "Karar Alma": 15, "Soğukkanlılık": 15, "Takım Oyunu": 16,
            "Güç": 13, "Hızlanma": 13,
        },

    "🔄 Pas Yapan Bek (Alaba/Trent sol)": {
            "Pas": 15, "Vizyon": 13, "İlk Kontrol": 13,
            "Karar Alma": 14, "Konsantrasyon": 13,
            "Markaj": 13, "Top Kapma": 13, "Pozisyon Alma": 14,
            "Hız": 12, "Orta Yapma": 12,
        },
    },

    # ── KALECI (KL) ───────────────────────────────────────
    "KL": {
        "🎲 Rastgele": None,

        "🧤 Klasik Kaleci (Buffon/Casillas)": {
            # Refleksler ve pozisyon alma ön planda
            "Refleksler": 17, "Elle Kontrol": 15, "Konumlama": 16,
            "Çıkış": 13, "Top Tutma": 15, "İletişim": 14,
            "Hava Hakimiyeti": 13, "Uzağa Atış": 11,
            "Güç": 13, "Zıplama": 14,
        },

        "⚡ Sweeper Kaleci (Neuer/Alisson)": {
            # Ayak topu + çıkış + refleksler
            "Refleksler": 16, "Çıkış": 17, "Konumlama": 15,
            "Top Tutma": 14, "İletişim": 15, "Hava Hakimiyeti": 15,
            "Uzağa Atış": 15, "Elle Kontrol": 14,
            "Hız": 13, "Hızlanma": 13, "Güç": 14,
            "İlk Kontrol": 14, "Pas": 13,
        },

        "🎯 Penaltı Uzmanı (Courtois/Oblak)": {
            "Refleksler": 18, "Elle Kontrol": 16, "Konumlama": 17,
            "Top Tutma": 16, "Soğukkanlılık": 16, "Konsantrasyon": 15,
            "Çıkış": 12, "Hava Hakimiyeti": 14,
            "Güç": 14, "Zıplama": 15,
        },

        
    "👣 Ayak Topu Kalecisi (Ederson/Alisson)": {
        # Kaleci ama 11. adam — pas + ilk kontrol çok yüksek
        "Refleksler": 15, "Konumlama": 15, "Top Tutma": 14,
        "Çıkış": 14, "Hava Hakimiyeti": 13, "İletişim": 14,
        "Uzağa Atış": 16,
        "Pas": 16, "İlk Kontrol": 14, "Teknik": 13,
        "Hız": 13, "Hızlanma": 13,
    },

    "📐 Pozisyon Kalecisi (Buffon/Schmeichel)": {
        # Refleks değil pozisyonla kurtarır — büyük, dominant
        "Konumlama": 18, "Konsantrasyon": 17, "Refleksler": 16,
        "Top Tutma": 16, "Hava Hakimiyeti": 17, "İletişim": 15,
        "Güç": 16, "Zıplama": 17,
        "Çıkış": 9, "Uzağa Atış": 11,
    },

    "💪 Lider Kaleci (Casillas/De Gea)": {
            "Refleksler": 17, "İletişim": 17, "Liderlik": 16,
            "Elle Kontrol": 15, "Konumlama": 16, "Top Tutma": 15,
            "Soğukkanlılık": 15, "Konsantrasyon": 15,
            "Güç": 13, "Zıplama": 14,
        },
    },

    # ── ST EK ARKETİPLER ──────────────────────────────────
    # (Mevcut ST arketiplerinin yanına ekstra 2 tane)
    # Bu blok ARCHETYPES["ST"]'e eklenecek,
    # aşağıda ARCHETYPES_EXTRA olarak tutulup birleştirilecek

}

# ST'ye ek arketipler — ARCHETYPES["ST"]'e birleştir
ARCHETYPES["ST"].update({
    "🦁 Fiziksel Dominant (Lukaku/Ibra)": {
        "Güç": 18, "Zıplama": 16, "Kafa Vuruşu": 16, "Dayanıklılık": 15,
        "Çeviklik": 12, "Bitiricilik": 14, "Topsuz Alan": 13,
        "İlk Kontrol": 12, "Güç": 18,
        "Hız": 11, "Hızlanma": 11, "Pas": 7, "Dripling": 9,
    },
    "🎭 Hareket Santrafor (Aguero/Benzema 2)": {
        "Topsuz Alan": 17, "Önsezi": 16, "Hız": 15, "Hızlanma": 15,
        "Soğukkanlılık": 15, "Bitiricilik": 15, "Pozisyon Alma": 15,
        "Çeviklik": 14, "Karar Alma": 14,
        "Güç": 9, "Kafa Vuruşu": 8, "Pas": 9,
    },

    "🔄 Pivot Santrafor (Drogba/Cavani)": {
        # Sırt dönük oyun — top tutundurma, asist, ikinci dalga hazırlama
        # Target Man'dan farkı: havada değil yerde dominant, teknik daha önemli
        "Güç": 17, "Denge": 16, "İlk Kontrol": 16,
        "Pas": 14, "Karar Alma": 15, "Teknik": 13,
        "Çeviklik": 13, "Topsuz Alan": 14,
        "Bitiricilik": 12, "Soğukkanlılık": 13,
        "Hız": 9, "Hızlanma": 9, "Kafa Vuruşu": 11,
    },

    "⭐ Kilitleyen Santrafor (Hakan Şükür)": {
        # FM2006 orijinal verisi — Galatasaray, 33 yaş
        # İMZA: Zıplama 19, Çalışkanlık 17, Takım Oyunu 16,
        #        İlk Kontrol 16, Dayanıklılık 17 — Finishing sadece 14!
        # Gol makinesi değil, takım için koşan havada dominant santrafor
        # Pozisyon Alma 6 — klasik "doğru yerde olmayan ama işe yarayan" tip
        "Bitiricilik": 14, "İlk Kontrol": 15, "Kafa Vuruşu": 16,
        "Pas": 15, "Teknik": 11, "Uzaktan Şut": 10,
        "Soğukkanlılık": 15, "Vizyon": 13, "Karar Alma": 12,
        "Takım Oyunu": 15, "Çalışkanlık": 16, "Önsezi": 11,
        "Liderlik": 14, "Topsuz Alan": 14, "Pozisyon Alma": 6,
        "Hız": 13, "Hızlanma": 11, "Çeviklik": 13, "Denge": 10,
        "Zıplama": 18, "Dayanıklılık": 16, "Vücut Zindeliği": 16,
        "Güç": 14, "Dripling": 9,
    },
})

# Desteklenen mevkiler için arketip pool'ları
ARCHETYPE_BASE_MAP = {
    "ST"          : "ST",
    "KF (Sol)"    : "KF",   "KF (Sağ)"    : "KF",
    "OOS"         : "OOS",
    "KANAT (Sol)" : "KANAT","KANAT (Sağ)" : "KANAT",
    "OS"          : "OS",   "DM"          : "DM",
    "DOS"         : "DOS",
    "KB (Sol)"    : "KB",   "KB (Sağ)"    : "KB",
    "D (Sol)"     : "D",    "D (Sağ)"     : "D",
    "KL"          : "KL",
}

# =========================================================
# POZİSYON ATTR ORTALAMALARI
# =========================================================

POSITION_ATTR_MEANS = {
    "ST": {
        "Bitiricilik"     : 13, "Dripling"        : 11, "İlk Kontrol"           : 11, "Kafa Vuruşu" : 10, "Uzaktan Şut" : 10,
        "Teknik"          : 9, "Pas"              : 8, "Orta Yapma"             : 3, "Top Kapma"    : 4, "Markaj"       :  3,
        "Korner"          : 3, "Penaltı Kullanma" : 9, "Serbest Vuruş Kullanma" : 6, "Uzun Taç"     : 3, "Topsuz Alan"  : 13,
        "Soğukkanlılık"   : 12, "Pozisyon Alma"   : 11, "Karar Alma"            : 11, "Önsezi"      : 11, "Çalışkanlık" :  9,
        "Vizyon"          : 7, "Konsantrasyon"    : 9, "Takım Oyunu"            : 8, "Agresiflik"   : 7, "Cesaret"      :  9,
        "Liderlik"        : 9, "Kararlılık"       : 9, "Özel Yetenek"           : 8, "Hız"          : 12, "Hızlanma"    : 12,
        "Çeviklik"        : 10, "Denge"           : 9, "Dayanıklılık"           : 10, "Güç"         : 10, "Zıplama"     : 11,
        "Vücut Zindeliği" : 10,

    },
    "KF": {
        "Bitiricilik"     : 11, "Dripling"        : 13, "İlk Kontrol"           : 12, "Kafa Vuruşu" : 7, "Uzaktan Şut"  : 9,
        "Teknik"          : 12, "Pas"             : 9, "Orta Yapma"             : 7, "Top Kapma"    : 4, "Markaj"       : 3,
        "Korner"          : 4, "Penaltı Kullanma" : 7, "Serbest Vuruş Kullanma" : 6, "Uzun Taç"     : 3, "Topsuz Alan"  : 12,
        "Soğukkanlılık"   : 11, "Pozisyon Alma"   : 10, "Karar Alma"            : 10, "Önsezi"      : 11, "Çalışkanlık" : 10,
        "Vizyon"          : 9, "Konsantrasyon"    : 9, "Takım Oyunu"            : 9,"Agresiflik"    : 7, "Cesaret"      : 8,
        "Liderlik"        : 9, "Kararlılık"       : 9, "Özel Yetenek"           : 10, "Hız"         : 13, "Hızlanma"    : 13,
        "Çeviklik"        : 12, "Denge"           : 11, "Dayanıklılık"          : 10, "Güç"         : 8, "Zıplama"      : 8,
        "Vücut Zindeliği" : 10,

    },
    "OOS": {
        "Bitiricilik"     : 9, "Dripling"          : 11, "İlk Kontrol"           : 13, "Kafa Vuruşu" : 6, "Uzaktan Şut"  : 8,
        "Teknik"          : 13, "Pas"              : 13, "Orta Yapma"            : 6, "Top Kapma"    : 5, "Markaj"       : 4,
        "Korner"          : 10, "Penaltı Kullanma" : 7, "Serbest Vuruş Kullanma" : 12, "Uzun Taç"    : 3, "Topsuz Alan"  : 11,
        "Soğukkanlılık"   : 12, "Pozisyon Alma"    : 10, "Karar Alma"            : 12, "Önsezi"      : 12, "Çalışkanlık" : 9,
        "Vizyon"          : 13, "Konsantrasyon"    : 10, "Takım Oyunu"           : 10, "Agresiflik"  : 6, "Cesaret"      : 7,
        "Liderlik"        : 10, "Kararlılık"       : 9, "Özel Yetenek"           : 12, "Hız"         : 10, "Hızlanma"    : 11,
        "Çeviklik"        : 11, "Denge"            : 10, "Dayanıklılık"          : 9, "Güç"          : 7, "Zıplama"      : 6,
        "Vücut Zindeliği" : 9,

    },
    "KANAT": {
        "Bitiricilik"     : 7, "Dripling"         : 12, "İlk Kontrol"           : 11, "Kafa Vuruşu" : 6, "Uzaktan Şut"  : 7,
        "Teknik"          : 11, "Pas"             : 10, "Orta Yapma"            : 13, "Top Kapma"   : 5, "Markaj"       : 5,
        "Korner"          : 9, "Penaltı Kullanma" : 5, "Serbest Vuruş Kullanma" : 6, "Uzun Taç"     : 3, "Topsuz Alan"  : 11,
        "Soğukkanlılık"   : 9, "Pozisyon Alma"    : 9, "Karar Alma"             : 10, "Önsezi"      : 10, "Çalışkanlık" : 11,
        "Vizyon"          : 10, "Konsantrasyon"   : 9, "Takım Oyunu"            : 10, "Agresiflik"  : 7, "Cesaret"      : 7,
        "Liderlik"        : 9, "Kararlılık"       : 8, "Özel Yetenek"           : 8, "Hız"          : 13, "Hızlanma"    : 13,
        "Çeviklik"        : 11, "Denge"           : 10, "Dayanıklılık"          : 11, "Güç"         : 7, "Zıplama"      : 7,
        "Vücut Zindeliği" : 10,

    },
    "OS": {
        "Bitiricilik"     : 7, "Dripling"          : 9, "İlk Kontrol"            : 12, "Kafa Vuruşu" : 7, "Uzaktan Şut"  : 7,
        "Teknik"          : 11, "Pas"              : 13, "Orta Yapma"            : 8, "Top Kapma"    : 9, "Markaj"       : 7,
        "Korner"          : 11, "Penaltı Kullanma" : 5, "Serbest Vuruş Kullanma" : 10, "Uzun Taç"    : 5, "Topsuz Alan"  : 9,
        "Soğukkanlılık"   : 11, "Pozisyon Alma"    : 12, "Karar Alma"            : 12, "Önsezi"      : 11, "Çalışkanlık" : 12,
        "Vizyon"          : 12, "Konsantrasyon"    : 11, "Takım Oyunu"           : 12, "Agresiflik"  : 9, "Cesaret"      : 8,
        "Liderlik"        : 9, "Kararlılık"        : 11, "Özel Yetenek"          : 9, "Hız"          : 9, "Hızlanma"     : 9,
        "Çeviklik"        : 9, "Denge"             : 10, "Dayanıklılık"          : 12, "Güç"         : 10, "Zıplama"     : 7,
        "Vücut Zindeliği" : 11,

    },
    "DM": {
        "Bitiricilik"     : 5, "Dripling"         : 7, "İlk Kontrol"            : 10, "Kafa Vuruşu" : 8, "Uzaktan Şut"  : 5,
        "Teknik"          : 9, "Pas"              : 11, "Orta Yapma"            : 6, "Top Kapma"    : 12, "Markaj"      : 10,
        "Korner"          : 4, "Penaltı Kullanma" : 4, "Serbest Vuruş Kullanma" : 5, "Uzun Taç"     : 5, "Topsuz Alan"  : 6,
        "Soğukkanlılık"   : 11, "Pozisyon Alma"   : 13, "Karar Alma"            : 12, "Önsezi"      : 11, "Çalışkanlık" : 12,
        "Vizyon"          : 9, "Konsantrasyon"    : 12, "Takım Oyunu"           : 12, "Agresiflik"  : 11, "Cesaret"     : 10,
        "Liderlik"        : 8, "Kararlılık"       : 11, "Özel Yetenek"          : 7, "Hız"          : 8, "Hızlanma"     : 8,
        "Çeviklik"        : 9, "Denge"            : 10, "Dayanıklılık"          : 12, "Güç"         : 11, "Zıplama"     : 8,
        "Vücut Zindeliği" : 11,

    },
    "DOS": {
        "Bitiricilik"     : 4, "Dripling"         : 4, "İlk Kontrol"            : 9, "Kafa Vuruşu" : 12, "Uzaktan Şut" : 4,
        "Teknik"          : 8, "Pas"              : 10, "Orta Yapma"            : 4, "Top Kapma"   : 13, "Markaj"      : 13,
        "Korner"          : 3, "Penaltı Kullanma" : 4, "Serbest Vuruş Kullanma" : 4, "Uzun Taç"    : 6, "Topsuz Alan"  : 4,
        "Soğukkanlılık"   : 11, "Pozisyon Alma"   : 13, "Karar Alma"            : 12, "Önsezi"     : 10, "Çalışkanlık" : 10,
        "Vizyon"          : 7, "Konsantrasyon"    : 12, "Takım Oyunu"           : 10, "Agresiflik" : 11, "Cesaret"     : 12,
        "Liderlik"        : 9, "Kararlılık"       : 12, "Özel Yetenek"          : 6, "Hız"         : 9, "Hızlanma"     : 9,
        "Çeviklik"        : 8, "Denge"            : 10, "Dayanıklılık"          : 11, "Güç"        : 13, "Zıplama"     : 12,
        "Vücut Zindeliği" : 11,

    },
    "D": {
        "Bitiricilik"     : 4, "Dripling"         : 7, "İlk Kontrol"            : 9, "Kafa Vuruşu" : 10, "Uzaktan Şut" : 4,
        "Teknik"          : 8, "Pas"              : 9, "Orta Yapma"             : 9, "Top Kapma"   : 12, "Markaj"      : 12,
        "Korner"          : 4, "Penaltı Kullanma" : 4, "Serbest Vuruş Kullanma" : 4, "Uzun Taç"    : 5, "Topsuz Alan"  : 5,
        "Soğukkanlılık"   : 11, "Pozisyon Alma"   : 12, "Karar Alma"            : 11, "Önsezi"     : 9, "Çalışkanlık"  : 10,
        "Vizyon"          : 6, "Konsantrasyon"    : 11, "Takım Oyunu"           : 10, "Agresiflik" : 10, "Cesaret"     : 11,
        "Liderlik"        : 9, "Kararlılık"       : 11, "Özel Yetenek"          : 5, "Hız"         : 9, "Hızlanma"     : 9,
        "Çeviklik"        : 9, "Denge"            : 10, "Dayanıklılık"          : 11, "Güç"        : 12, "Zıplama"     : 9,
        "Vücut Zindeliği" : 11,

    },
    "KB": {
        "Bitiricilik"     : 5, "Dripling"         : 9, "İlk Kontrol"            : 10, "Kafa Vuruşu" : 7, "Uzaktan Şut" : 5,
        "Teknik"          : 9, "Pas"              : 10, "Orta Yapma"            : 12, "Top Kapma"   : 11, "Markaj"     : 10,
        "Korner"          : 7, "Penaltı Kullanma" : 4, "Serbest Vuruş Kullanma" : 4, "Uzun Taç"     : 5, "Topsuz Alan" : 7,
        "Soğukkanlılık"   : 10, "Pozisyon Alma"   : 10, "Karar Alma"            : 10, "Önsezi"      : 9, "Çalışkanlık" : 11,
        "Vizyon"          : 8, "Konsantrasyon"    : 9, "Takım Oyunu"            : 11, "Agresiflik"  : 9, "Cesaret"     : 9,
        "Liderlik"        : 9, "Kararlılık"       : 10, "Özel Yetenek"          : 6, "Hız"          : 12, "Hızlanma"   : 12,
        "Çeviklik"        : 11, "Denge"           : 10, "Dayanıklılık"          : 12, "Güç"         : 9, "Zıplama"     : 8,
        "Vücut Zindeliği" : 11,

    },
    "KL": {
        "Bitiricilik"     : 2, "Dripling"         : 2, "İlk Kontrol"            : 7, "Kafa Vuruşu" : 4, "Uzaktan Şut"  : 2,
        "Teknik"          : 4, "Pas"              : 9, "Orta Yapma"             : 2, "Top Kapma"   : 2, "Markaj"       : 3,
        "Korner"          : 2, "Penaltı Kullanma" : 4, "Serbest Vuruş Kullanma" : 2, "Uzun Taç"    : 9, "Topsuz Alan"  : 3,
        "Soğukkanlılık"   : 12, "Pozisyon Alma"   : 4, "Karar Alma"             : 12, "Önsezi"     : 11, "Çalışkanlık" : 9,
        "Vizyon"          : 8, "Konsantrasyon"    : 12, "Takım Oyunu"           : 9, "Agresiflik"  : 6, "Cesaret"      : 11,
        "Liderlik"        : 10, "Kararlılık"      : 11, "Özel Yetenek"          : 7, "Hız"         : 5, "Hızlanma"     : 5,
        "Çeviklik"        : 11, "Denge"           : 10, "Dayanıklılık"          : 8, "Güç"         : 9, "Zıplama"      : 11,
        "Vücut Zindeliği" : 10,

    },
}

for _sol, _sag, _base in [("D (Sol)", "D (Sağ)", "D"), ("KB (Sol)", "KB (Sağ)", "KB"),
                          ("KANAT (Sol)", "KANAT (Sağ)", "KANAT"), ("KF (Sol)", "KF (Sağ)", "KF")]:
    POSITION_ATTR_MEANS[_sol] = dict(POSITION_ATTR_MEANS[_base])
    POSITION_ATTR_MEANS[_sag] = dict(POSITION_ATTR_MEANS[_base])

for _sol in ["D (Sol)", "KB (Sol)", "KANAT (Sol)", "KF (Sol)"]:
    m = POSITION_ATTR_MEANS[_sol]
    m["Orta Yapma"] = min(20, int(m.get("Orta Yapma", 8) * 1.10))
    m["Hızlanma"]   = min(20, int(m.get("Hızlanma", 10) * 1.05))

for _sag in ["D (Sağ)", "KB (Sağ)", "KANAT (Sağ)", "KF (Sağ)"]:
    m = POSITION_ATTR_MEANS[_sag]
    m["Top Kapma"] = min(20, int(m.get("Top Kapma", 10) * 1.05))
    m["Markaj"]    = min(20, int(m.get("Markaj", 8) * 1.05))

# =========================================================
# POZİSYON SINIRLARI
# =========================================================

ATTACK_ATTRS = {"Bitiricilik", "Uzaktan Şut", "Topsuz Alan", "Soğukkanlılık", "Dripling", "Penaltı Kullanma",
                "Serbest Vuruş Kullanma", "Korner"}
DEFENSE_ATTRS = {"Top Kapma", "Markaj", "Agresiflik"}

# Kafa Vuruşu hem defans hem hücum için kullanılır.
# Mevkiye göre ayrı üst sınır uygulanır.
HEADING_CAP = {
    "KL"  : 8,
    "DOS" : 20, "D"        : 18, "D (Sol)"    : 18, "D (Sağ)"    : 18,
    "KB"  : 15, "KB (Sol)" : 15, "KB (Sağ)"   : 15,
    "DM"  : 14, "OS"       : 13,
    "OOS" : 10, "KANAT"    : 9, "KANAT (Sol)" : 9, "KANAT (Sağ)" : 9,
    "KF"  : 8, "KF (Sol)"  : 8, "KF (Sağ)"    : 8,
    "ST"  : 17,

}

POSITION_ATTACK_CAP = {
    "KL" : 5, "DOS"       : 7, "D"         : 8, "D (Sol)"      : 8, "D (Sağ)"      : 8,
    "KB" : 10, "KB (Sol)" : 10, "KB (Sağ)" : 10, "DM"          : 10,
    "OS" : 14, "OOS"      : 19, "KANAT"    : 17, "KANAT (Sol)" : 17, "KANAT (Sağ)" : 17,
    "KF" : 18, "KF (Sol)" : 18, "KF (Sağ)" : 18, "ST"          : 20,

}
POSITION_DEFENSE_CAP = {
    "KL" : 9, "DOS"       : 20, "D"        : 19, "D (Sol)"    : 19, "D (Sağ)"    : 19,
    "KB" : 17, "KB (Sol)" : 17, "KB (Sağ)" : 17, "DM"         : 18,
    "OS" : 13, "OOS"      : 9, "KANAT"     : 9, "KANAT (Sol)" : 9, "KANAT (Sağ)" : 9,
    "KF" : 7, "KF (Sol)"  : 7, "KF (Sağ)"  : 7, "ST"          : 6,

}
PRESET_ATTR_CAP = {"Average": 14, "Wonderkid": 17, "Star": 18, "Superstar": 20}

GK_ATTR_MEANS = {
    "Refleksler"       : 13, "Elle Kontrol" : 12, "Birebir"          : 11, "Hava Topları"      : 12,
    "Bölge Hakimiyeti" : 11, "İletişim"     : 10, "Degaj"            : 9, "Elle Oyun Başlatma" : 9,
    "İlk Kontrol (K)"  : 10, "Pas (K)"      : 9, "Ani Çıkış Eğilimi" : 7, "Eksantriklik"       : 6, "Yumrukla Uzaklaştırma" : 8,

}


# =========================================================
# CA / PA MATEMATİĞİ
# =========================================================

def _ca_formula(avg: float, attrs: dict = None) -> int:
    """
    FMScout metodolojisi:
    - 6 → CA ~1, 16 → CA ~199  (doğrusal bölge)
    - 17+ olan attribute'lar CA'ya orantısız katkı yapar (geliştirmesi daha pahalı)
    """
    if avg <= 5.0: return 1
    base = int(min(200, (avg - 5.0) ** 1.2 * 12))

    # 17+ bonus: FM'de yüksek attr'ları geliştirmek daha pahalı
    # dolayısıyla aynı CA'ya sahip iki oyuncudan 17+ attr'ı olan daha değerli
    if attrs:
        bonus = sum((v - 16) ** 2 * 0.9 for v in attrs.values() if v >= 17)
        base = int(min(200, base + bonus))

    return base


def _avg_from_ca(ca):
    if ca <= 1: return 5.0
    return (ca / 12.0) ** (1.0 / 1.2) + 5.0


_REFERENCE_AVG = _avg_from_ca(110)


def _level_mult(target_ca):
    return _avg_from_ca(target_ca) / _REFERENCE_AVG


def get_ca_pa(preset, age, country):
    cb = COUNTRY_PA_BONUS.get(country, 0)
    age_prog = min(1.0, (age - 15) / 20.0)

    if preset == "Average":
        ca = random.randint(int(65 + age_prog * 70), min(int(82 + age_prog * 55), 140))
        ca = max(45, min(140, ca))
        pa = min(140, ca + random.randint(0, max(3, int(18 - age_prog * 15))))
    elif preset == "Wonderkid":
        pa = min(200, random.randint(165, 200) + cb)
        wkp = min(1.0, (age - 15) / 12.0)
        ca = random.randint(int(45 + wkp * 95), int(70 + wkp * 85))
        ca = max(25, min(ca, pa - 28))
    elif preset == "Star":
        pa = min(200, random.randint(150, 180) + cb // 2)
        sp = min(1.0, (age - 15) / 15.0)
        ca_lo = int(92 + sp * 72)
        ca = random.randint(ca_lo, max(ca_lo, min(int(118 + sp * 58), pa)))
    elif preset == "Superstar":
        pa = min(200, random.randint(180, 200) + cb // 3)
        sp = min(1.0, (age - 15) / 15.0)
        ca_lo = int(122 + sp * 63)
        ca = random.randint(ca_lo, max(ca_lo, min(int(152 + sp * 48), pa)))
    else:
        ca, pa = 100, 120

    pa = min(200, pa)
    ca = max(1, min(ca, pa))
    return ca, pa


# =========================================================
# ATTRIBUTE ÜRETİMİ
# =========================================================

def age_max_attr(age):
    return min(20, int(8 + age * 0.6))


def physical_age_mult(age):
    if age >= 33: return 0.67
    if age >= 30: return 0.82
    if age >= 28: return 0.93
    return 1.0


def _roll(mean, std, max_val):
    return max(1, min(max_val, min(20, int(random.gauss(mean, std)))))


def generate_all_attributes(position, age, preset, country, target_ca,
                             archetype_means=None, cap_overrides=None):
    # Arketip varsa base means üzerine override uygula — _caps key'ini ayır
    base_means = dict(POSITION_ATTR_MEANS[position])
    if archetype_means:
        _clean = {k: v for k, v in archetype_means.items() if not k.startswith("_")}
        base_means.update(_clean)
    means = base_means
    profile = COUNTRY_PROFILES[country]
    lmult   = _level_mult(target_ca)
    phys_m  = physical_age_mult(age)
    max_a   = age_max_attr(age)
    a_cap   = PRESET_ATTR_CAP[preset]

    if preset == "Wonderkid" and age < 18:
        a_cap = min(a_cap, 16)

    tech, mental, phys, gk, hidden = {}, {}, {}, {}, {}

    for attr in TECHNICAL:
        m = means.get(attr, 7) * lmult * profile["tech"]
        val = _roll(m, 2.0, min(max_a, a_cap))
        if attr in ATTACK_ATTRS:
            val = min(val, (cap_overrides or {}).get(attr, POSITION_ATTACK_CAP.get(position, 20)))
        if attr in DEFENSE_ATTRS:
            val = min(val, (cap_overrides or {}).get(attr, POSITION_DEFENSE_CAP.get(position, 20)))
        if attr == "Kafa Vuruşu":
            val = min(val, (cap_overrides or {}).get(attr, HEADING_CAP.get(position, 20)))
        tech[attr] = val

    for attr in MENTAL:
        m = means.get(attr, 9) * lmult
        if country == "Türkiye": m *= 1.10
        mental[attr] = _roll(m, 2.0, min(max_a, a_cap))

    for attr in PHYSICAL:
        m = means.get(attr, 9) * lmult * profile["phys"] * phys_m
        phys[attr] = _roll(m, 1.8, min(max_a, a_cap))

    base = POSITION_BASE[position]
    if base == "KL":
        for attr in GOALKEEPER:
            _gk_m = means.get(attr, GK_ATTR_MEANS.get(attr, 9)) * lmult
            gk[attr] = _roll(_gk_m, 2.0, min(max_a, a_cap))
    else:
        for attr in GOALKEEPER:
            gk[attr] = random.randint(1, min(5, max_a))

    hidden = _generate_hidden(preset)

    return tech, mental, phys, gk, hidden


# =========================================================
# GİZLİ ATTRIBUTE ÜRETİMİ (PRESET BAZLI)
# =========================================================

# Her preset için hangi hidden attr'lar yüksek olabilir
_HIDDEN_BOOSTS = {
    "Superstar": {
        # Güçlü yanlar — yüksek gelme ihtimali yüksek
        "Profesyonellik"       : (16, 20),
        "Hırs"                 : (15, 20),
        "Önemli Maçlar"        : (15, 20),
        "Baskıya Dayanıklılık" : (14, 19),
        "Kararlılık"           : (14, 19),  # mental'de de var ama hidden'da ayrı
        "Süreklilik"           : (16, 20),

    },
    "Star": {
        "Profesyonellik"       : (14, 19),
        "Hırs"                 : (13, 18),
        "Önemli Maçlar"        : (12, 18),
        "Baskıya Dayanıklılık" : (12, 17),
        "Süreklilik"           : (13, 17),

    },
    "Wonderkid": {
        # Wonderkid'lerde hırs çok yüksek, ama profesyonellik değişken
        "Hırs"                 : (16, 20),
        "Önemli Maçlar"        : (13, 19),
        "Baskıya Dayanıklılık" : (11, 18),
        "Profesyonellik"       : (9, 18),  # geniş aralık — bazısı disiplinli değil

    },
    "Average": {},  # tamamen rastgele
}

# Negatif attr'lar — her preset'te yüksek gelebilir (ama düşük olasılıkla)
_NEGATIVE_HIDDEN = {
    "Sakatlanma Eğilimi" : 0.15,  # %15 ihtimalle yüksek (14-20)
    "Tartışma"           : 0.20,
    "Çirkeflik"          : 0.18,
    "Huy"                : 0.15,  # düşük = kötü huy

}


def _generate_hidden(preset: str) -> dict:
    """
    Preset'e göre gizli attribute üretir.
    - Superstar / Wonderkid → bazı attr'lar belirgin yüksek
    - Negatif attr'lar her preset'te nadiren yüksek gelebilir
    - Kaç attr'ın boost alacağı da rastgele (5-6 civarı)
    """
    boosts = _HIDDEN_BOOSTS.get(preset, {})
    result = {}

    for attr in HIDDEN:
        if attr in boosts and random.random() < 0.75:
            # %75 ihtimalle boost aralığından üret
            lo, hi = boosts[attr]
            result[attr] = random.randint(lo, hi)
        elif attr in _NEGATIVE_HIDDEN and random.random() < _NEGATIVE_HIDDEN[attr]:
            # Negatif attr — nadiren yüksek
            result[attr] = random.randint(13, 20)
        else:
            # Normal dağılım (tüm presetler için)
            result[attr] = _roll(10, 2.8, 20)

    return result


# =========================================================
# KİŞİLİK TİPİ TESPİTİ
# =========================================================

PERSONALITY_PROFILES = [
    # ── POZİTİF KARAKTERLERs ──────────────────────────────────────
    {
        "name": "⭐ Lider",
        "desc": "Sahada ve soyunma odasında söz sahibi. Zor anlarda takımını toplar, genç oyunculara yol gösterir.",
        "color": "#FFD700",
        "conditions": {
            "Liderlik": 14, "Baskıya Dayanıklılık": 13, "Önemli Maçlar": 13, "Süreklilik": 15,
        },
        "negative": {},
    },
    {
        "name": "🔥 Hırslı Profesyonel",
        "desc": "Antrenman delisi. Son antrenmana kadar sahada kalan, potansiyelini sonuna kadar sıkıştırmak için her şeyi göze alan tip.",
        "color": "#e67e22",
        "conditions": {
            "Hırs": 15, "Profesyonellik": 14,
        },
        "negative": {},
    },
    {
        "name": "🤝 Takım Oyuncusu",
        "desc": "Bireysel istatistiklerden çok takımın başarısını önemser. Gol atmak değil, gol attırmak mutlu eder onu.",
        "color": "#2ecc71",
        "conditions": {
            "Sportmenlik": 13, "Aidiyet Duygusu": 13, "Süreklilik": 14,
        },
        "negative": {"Tartışma": 12, "Çirkeflik": 12},
    },
    {
        "name": "💎 Büyük Maç Oyuncusu",
        "desc": "Turnuva finalleri, derbiler, şampiyonluk maçları... Rakam ne kadar büyük olursa performansı o kadar artar.",
        "color": "#9b59b6",
        "conditions": {
            "Önemli Maçlar": 15, "Baskıya Dayanıklılık": 13, "Süreklilik": 14,
        },
        "negative": {},
    },
    {
        "name": "🧠 Çalışkan & Disiplinli",
        "desc": "Parlak yeteneği olmayabilir ama her antrenmanda yüzde yüz verir. Teknik direktörün güvendiği, soyunma odasının saygı duyduğu isim.",
        "color": "#3498db",
        "conditions": {
            "Profesyonellik": 15, "Çok Yönlülük": 12,
        },
        "negative": {"Tartışma": 13},
    },
    {
        "name": "🧊 Buzdan Sinirler",
        "desc": "Penaltı vuruşunda nabzı düşen, maç sonu baskısında en berrak kafayı koruyan isim. Kriz anı onun sahnesi.",
        "color": "#00bcd4",
        "conditions": {
            "Baskıya Dayanıklılık": 16, "Önemli Maçlar": 14,
        },
        "negative": {},
    },
    {
        "name": "🦁 Savaşçı",
        "desc": "Her topu son nefesiyle koşar, her müdahaleyi beden bütünlüğünü riske ederek yapar. İndirimli oynamayı bilmez.",
        "color": "#ff6b35",
        "conditions": {
            "Huy": 16, "Baskıya Dayanıklılık": 13,
        },
        "negative": {"Çirkeflik": 14},
    },
    {
        "name": "📖 Sakin Profesyonel",
        "desc": "Ne başarı şımartır ne başarısızlık yıkar. Sessiz sedasız, yıllarca aynı kalitede oynayan istikrar abidesi.",
        "color": "#27ae60",
        "conditions": {
            "Profesyonellik": 14, "Süreklilik": 16, "Sportmenlik": 13,
        },
        "negative": {"Tartışma": 11, "Hırs": 12},
    },
    {
        "name": "🌟 Tecrübeli Usta",
        "desc": "Yıllar onu olgunlaştırdı. Genç oyunculara mentor olan, soyunma odasını bir arada tutan, deneyimiyle değer yaratan isim.",
        "color": "#f39c12",
        "conditions": {
            "Liderlik": 13, "Süreklilik": 15, "Profesyonellik": 13, "Aidiyet Duygusu": 14,
        },
        "negative": {},
    },
    {
        "name": "🎯 Mükemmeliyetçi",
        "desc": "En sert eleştirmeni kendisi. İyi oynasa bile maç sonrası analize dalar, eksikleri not alır, bir sonraki maça daha iyi hazırlanır.",
        "color": "#8e44ad",
        "conditions": {
            "Hırs": 15, "Profesyonellik": 16,
        },
        "negative": {"Tartışma": 10},
    },
    # ── KARISIK / NÖTR KARAKTERLERs ──────────────────────────────
    {
        "name": "⚡ Karizmatik İsyankâr",
        "desc": "Sahada görkemli, soyunma odasında çalkantılı. Yönetmesi güç ama sağladığı fark tartışılmaz. Teknik direktörünü tüketir, taraftarı büyüler.",
        "color": "#e74c3c",
        "conditions": {
            "Hırs": 14, "Tartışma": 13,
        },
        "negative": {"Sportmenlik": 12},
    },
    {
        "name": "🎭 Medya Yıldızı",
        "desc": "Kameralar açılınca performansı artar. Sosyal medya fenomeni, marka değeri yüksek; ama saha dışı gürültü bazen odağını dağıtır.",
        "color": "#e91e63",
        "conditions": {
            "Hırs": 14, "Çirkeflik": 13,
        },
        "negative": {"Profesyonellik": 13, "Aidiyet Duygusu": 11},
    },
    {
        "name": "💤 Tembel Deha",
        "desc": "Sahaya girdiğinde 'neden bu kadar geç keşfettik' dedirtir; ama antrenmanda o yeteneğin yarısı kadar efor sarf eder.",
        "color": "#607d8b",
        "conditions": {
            "Çok Yönlülük": 14,
        },
        "negative": {"Profesyonellik": 10, "Hırs": 11},
    },
    {
        "name": "💔 Motivasyon Yoksunu",
        "desc": "Ateşi çoktan söndü. Sözleşmeyi doldurma modunda; maç maç varlık gösterir ama asla o eski seviyeye ulaşamaz.",
        "color": "#546e7a",
        "conditions": {},
        "negative": {"Hırs": 8, "Profesyonellik": 10, "Süreklilik": 10},
    },
    # ── SORUNLU KARAKTERLERs ─────────────────────────────────────
    {
        "name": "😤 Sorunlu Karakter",
        "desc": "Teknik direktörün kabusu. Soyunma odasında huzursuzluk yaratır, kliklere yol açar, kulübü yıpratır.",
        "color": "#c0392b",
        "conditions": {
            "Tartışma": 15, "Çirkeflik": 14,
        },
        "negative": {},
    },
    {
        "name": "💣 Zehirli Unsur",
        "desc": "Tek başına soyunma odasının havasını zehirleyebilir. Yönetim transfer penceresinde bir an olsun rahatlar.",
        "color": "#b71c1c",
        "conditions": {
            "Tartışma": 16, "Çirkeflik": 15, "Huy": 15,
        },
        "negative": {"Sportmenlik": 10, "Aidiyet Duygusu": 9},
    },
    # ── ÖZEL DURUMLAR ─────────────────────────────────────────────
    {
        "name": "🏥 Sakatlanmaya Meyilli",
        "desc": "Yeteneği tartışılmaz ama sağlığı sürekli sorun çıkarıyor. Her 10 maçta bir antrenörü tedirgin eden yeni bir teşhis.",
        "color": "#7f8c8d",
        "conditions": {
            "Sakatlanma Eğilimi": 15,
        },
        "negative": {},
    },
    {
        "name": "🌍 Uyumsuz Gezgin",
        "desc": "Yeni bir şehre, yeni bir dile, yeni bir kültüre alışmak için zamana ihtiyacı var. Yabancı ligde başlangıç döneminde zorlanır.",
        "color": "#95a5a6",
        "conditions": {},
        "negative": {"Uyum": 8, "Aidiyet Duygusu": 8},
    },
    {
        "name": "📋 Standart Profesyonel",
        "desc": "Belirgin bir karakter özelliği öne çıkmıyor. Tahmin edilebilir, dengeli, sürprizsiz bir profil.",
        "color": "#aaaaaa",
        "conditions": {},
        "negative": {},
        "_fallback": True,
    },
]


def detect_personality(hidden: dict) -> dict:
    """
    Hidden attr değerlerine göre kişilik profili döndürür.
    Birden fazla profil eşleşebilir — en güçlü olanı seçilir.
    """
    scores = []

    for profile in PERSONALITY_PROFILES:
        if profile.get("_fallback"):
            continue  # Fallback profili eşleşme döngüsüne girmesin

        score = 0
        matched = True

        # Pozitif koşullar
        for attr, threshold in profile["conditions"].items():
            val = hidden.get(attr, 10)
            if val >= threshold:
                score += (val - threshold) * 2
            else:
                matched = False
                break

        if not matched:
            continue

        # Negatif koşullar (düşük olması beklenen)
        for attr, threshold in profile.get("negative", {}).items():
            val = hidden.get(attr, 10)
            if val >= threshold:
                matched = False
                break

        if matched:
            scores.append((profile, score))

    if not scores:
        # Hiçbiri eşleşmediyse fallback profili bul
        for p in PERSONALITY_PROFILES:
            if p.get("_fallback"):
                return p
        return {"name": "📋 Standart Profesyonel", "desc": "Dengeli profil.", "color": "#aaaaaa"}

    # En yüksek skorlu profili döndür
    best = sorted(scores, key=lambda x: -x[1])[0][0]
    return best


def personality_card_html(profile: dict, hidden: dict) -> str:
    """Kişilik kartı HTML'i."""
    color = profile["color"]
    name = profile["name"]
    desc = profile["desc"]

    # Öne çıkan hidden attr'lar (13+)
    highlights = [(k, v) for k, v in hidden.items() if v >= 14]
    highlights.sort(key=lambda x: -x[1])

    badges = ""
    for attr, val in highlights[:5]:
        shade = "#2ecc71" if val >= 17 else "#f1c40f" if val >= 14 else "#aaa"
        badges += f"<span style='background:#21262d;border-radius:10px;padding:2px 8px;font-size:11px;color:{shade};margin:2px;display:inline-block'>{attr} {val}</span>"

    return f"""
    <div style='
        background:linear-gradient(135deg,#0d1117,#161b22);
        border:1px solid {color}44;
        border-left:3px solid {color};
        border-radius:10px;padding:14px 18px;margin:8px 0;
    '>
        <div style='font-size:16px;font-weight:800;color:{color};margin-bottom:4px'>{name}</div>
        <div style='font-size:12px;color:#8b949e;margin-bottom:10px;font-style:italic'>{desc}</div>
        <div style='line-height:2'>{badges}</div>
    </div>"""


def calculate_ca(attrs, position, weak_foot: int = 4):
    """
    FM'e yakın CA hesabı.

    FM'de zayıf ayak CA'ya mevkiye göre ağırlıklı katkı yapar.
    FMScout araştırması: SC için 7.5/10, AMRL 7/10, DOS 1/10 vb.

    Soft cap: 16+ attr'lar için azalan getiri (FM gerçeği).
    """
    base = POSITION_BASE[position]
    weights = ATTRIBUTE_WEIGHTS.get(base, {})
    total = weight_sum = 0.0
    key_attrs = {}

    _SOFT_CAP_THRESHOLD = 16
    _SOFT_CAP_FACTOR = 0.30

    for attr, w in weights.items():
        if attr in HIDDEN_SET: continue
        if base != "KL" and attr in GK_ATTRS_SET: continue
        v = attrs.get(attr, 1)
        if v > _SOFT_CAP_THRESHOLD:
            v = _SOFT_CAP_THRESHOLD + (v - _SOFT_CAP_THRESHOLD) * _SOFT_CAP_FACTOR
        total += v * w
        weight_sum += w
        key_attrs[attr] = v

    # Zayıf ayak CA'ya dahil et
    wf_weight = WEAK_FOOT_CA_WEIGHT.get(base, 1.6)
    wf_v = weak_foot
    if wf_v > _SOFT_CAP_THRESHOLD:
        wf_v = _SOFT_CAP_THRESHOLD + (wf_v - _SOFT_CAP_THRESHOLD) * _SOFT_CAP_FACTOR
    total += wf_v * wf_weight
    weight_sum += wf_weight

    if weight_sum == 0: return 1
    return _ca_formula(total / weight_sum, key_attrs)


# =========================================================
# EN İYİ ROL TESPİTİ
# =========================================================

ROLE_KEY_ATTRS = {
    "Golcü"           : {"Bitiricilik" : 4, "Topsuz Alan"   : 3, "Soğukkanlılık"    : 3, "Hızlanma"     : 2},
    "Oyun Kurucu"     : {"Pas"         : 4, "Vizyon"        : 4, "Teknik"           : 3, "Karar Alma"   : 3},
    "Hız Kanadı"      : {"Hız"         : 4, "Hızlanma"      : 4, "Dripling"         : 3, "Topsuz Alan"  : 2},
    "Orta Açan Kanat" : {"Orta Yapma"  : 4, "Hız"           : 3, "Hızlanma"         : 3, "Çalışkanlık"  : 2},
    "Defansif OS"     : {"Top Kapma"   : 4, "Pozisyon Alma" : 4, "Karar Alma"       : 3, "Dayanıklılık" : 2},
    "Atak Stoper"     : {"Markaj"      : 4, "Top Kapma"     : 4, "Güç"              : 3, "Kafa Vuruşu"  : 3},
    "İkili Kanat Bek" : {"Hız"         : 4, "Orta Yapma"    : 3, "Dayanıklılık"     : 3, "Dripling"     : 2},
    "Kaleci"          : {"Refleksler"  : 4, "Elle Kontrol"  : 4, "Bölge Hakimiyeti" : 3},

}


def detect_best_role(all_attrs):
    scores = {}
    for role, weights in ROLE_KEY_ATTRS.items():
        score = sum(all_attrs.get(attr, 1) * w for attr, w in weights.items())
        total = sum(weights.values())
        scores[role] = score / total
    return sorted(scores.items(), key=lambda x: -x[1])[:3]


# =========================================================
# TRANSFER DEĞERİ
# =========================================================

def calculate_transfer_value(ca, pa, age):
    """FM benzeri transfer değeri hesabı (€)."""
    peak_age = 26
    age_factor = max(0.3, 1.0 - abs(age - peak_age) * 0.04)
    potential_bonus = (pa - ca) / 200 * 0.5 + 0.5
    _tv_katsayi = int(st.session_state.get('cfg_tv_m', 300)) * 1_000_000
    base = (ca / 200) ** 2.2 * _tv_katsayi
    value = base * age_factor * potential_bonus

    if value >= 80_000_000:
        return f"{value / 1_000_000:.0f} Milyon €"
    elif value >= 10_000_000:
        return f"{value / 1_000_000:.1f} Milyon €"
    elif value >= 1_000_000:
        return f"{value / 1_000_000:.2f} Milyon €"
    elif value >= 100_000:
        return f"{value / 1_000:.0f}Bin €"
    else:
        return f"€{value:,.0f}"


# =========================================================
# YARDIMCI FONKSİYONLAR
# =========================================================

def scout_grade(pa):
    for limit, grade in SCOUT_GRADES:
        if pa >= limit: return grade
    return "E"


def years_to_star(age, ca, pa):
    if ca >= STAR_CA_THRESHOLD: return 0
    remaining = max(1, 26 - age)
    yearly = (pa - ca) / remaining
    if yearly <= 0: return None
    return max(1, int((STAR_CA_THRESHOLD - ca) / yearly))


def generate_traits(position):
    pool = TRAITS_BY_POSITION.get(position) or TRAITS_BY_POSITION.get(POSITION_BASE[position], [])
    if not pool: return []
    return random.sample(pool, random.randint(1, min(3, len(pool))))


def calculate_effective_pa(pa, professionalism, preset):
    base = {"Average": 0.72, "Wonderkid": 0.90, "Star": 0.95, "Superstar": 0.97}.get(preset, 0.75)
    pf = max(-0.2, min(0.3, (professionalism - 10) / 20))
    reach = max(0.60, min(0.98, base + pf + random.uniform(-0.04, 0.04)))
    return int(pa * reach)


def simulate_growth(age, ca, pa, professionalism=15, position="ST"):
    years, current = [], ca
    base = POSITION_BASE.get(position, position)
    for year in range(age, 44):
        af = 1.30 if year <= 20 else 1.00 if year <= 24 else 0.80 if year <= 27 else 0.50 if year <= 30 else 0.20 if year <= 35 else 0.05
        if base in ("OOS", "OS") and year >= 22: af *= 1.15
        gap_r = max(0.0, (pa - current) / max(pa, 1))
        pf = max(0.25, gap_r ** 0.65)
        delta = int(10 * af * pf * (professionalism / 15))
        if delta < 1 and year < 28 and current < pa: delta = 1
        if year >= 30 and current > pa * 0.9: delta = -random.choice([0, 1])
        if year >= 35: delta = -random.randint(1, 3)
        if year >= 40: delta = -random.randint(2, 5)
        current = max(1, min(pa, current + delta))
        years.append((year + 1, current))
    return years


def ca_bar(ca, pa):
    pct_ca = ca / 200 * 100
    pct_pa = pa / 200 * 100
    return f"""
    <div style='margin:6px 0'>
      <div style='display:flex;justify-content:space-between;font-size:11px;color:#aaa;margin-bottom:3px'>
        <span>CA {ca}</span><span>PA {pa}</span>
      </div>
      <div style='background:#1e1e2e;border-radius:6px;height:10px;position:relative;overflow:hidden'>
        <div style='position:absolute;left:0;top:0;height:100%;width:{pct_pa:.1f}%;background:rgba(255,165,0,0.25);border-radius:6px'></div>
        <div style='position:absolute;left:0;top:0;height:100%;width:{pct_ca:.1f}%;background:linear-gradient(90deg,#2ecc71,#27ae60);border-radius:6px'></div>
      </div>
    </div>"""


# =========================================================
# RADAR CHART — FM STİLİ (8 KATEGORİ)
# =========================================================

RADAR_CATEGORIES = {
    "Savunma"  : ["Top Kapma", "Markaj", "Pozisyon Alma"],
    "Fiziksel" : ["Dayanıklılık", "Denge", "Vücut Zindeliği"],
    "Hız"      : ["Hız", "Hızlanma", "Çeviklik"],
    "Vizyon"   : ["Vizyon", "Karar Alma", "Önsezi", "Konsantrasyon"],
    "Hücum"    : ["Bitiricilik", "Uzaktan Şut", "Topsuz Alan", "Soğukkanlılık", "Penaltı Kullanma"],
    "Teknik"   : ["Dripling", "İlk Kontrol", "Teknik", "Orta Yapma", "Pas", "Serbest Vuruş Kullanma"],
    "Hava"     : ["Kafa Vuruşu", "Zıplama", "Güç"],
    "Zihinsel" : ["Çalışkanlık", "Cesaret", "Liderlik", "Takım Oyunu"],

}

RADAR_CATEGORIES_GK = {
    "Şut Karşılama" : ["Refleksler", "Elle Kontrol", "Birebir"],
    "Fiziksel"      : ["Dayanıklılık", "Güç", "Denge"],
    "Hız"           : ["Hız", "Hızlanma", "Çeviklik"],
    "Zihinsel"      : ["Karar Alma", "Konsantrasyon", "Önsezi"],
    "İletişim"      : ["İletişim", "Bölge Hakimiyeti"],
    "Eksantriklik"  : ["Eksantriklik", "Ani Çıkış Eğilimi"],
    "Hava"          : ["Hava Topları", "Zıplama"],
    "Dağıtım"       : ["Degaj", "Elle Oyun Başlatma", "Pas (K)"],

}


def radar_chart(all_attrs, position):
    base = POSITION_BASE[position]
    cats = RADAR_CATEGORIES_GK if base == "KL" else RADAR_CATEGORIES

    cat_names = list(cats.keys())
    cat_vals = []
    for cat, attrs in cats.items():
        vals = [all_attrs.get(a, 1) for a in attrs if a in all_attrs]
        cat_vals.append(round(sum(vals) / len(vals)) if vals else 1)

    N = len(cat_names)
    # 12'den saat yönünde açılar (kartezyen: π/2 = yukarı, CW = azalan)
    angles_rad = [np.pi / 2 - i * 2 * np.pi / N for i in range(N)]

    # Değerleri 0-1 arasına normalize et (20 maksimum)
    vals_norm = [v / 20.0 for v in cat_vals]

    fig, ax = plt.subplots(figsize=(3.8, 3.8))
    fig.patch.set_facecolor("#0d1117")
    ax.set_facecolor("#0d1117")
    ax.set_xlim(-1.45, 1.45)
    ax.set_ylim(-1.45, 1.45)
    ax.set_aspect("equal")
    ax.axis("off")

    # Grid halkaları
    for r_frac in [0.25, 0.5, 0.75, 1.0]:
        cx = [r_frac * np.cos(a) for a in np.linspace(0, 2 * np.pi, 120)]
        cy = [r_frac * np.sin(a) for a in np.linspace(0, 2 * np.pi, 120)]
        ax.plot(cx, cy, color="#1e2235", linewidth=0.8, zorder=1)

    # Eksen çizgileri
    for a in angles_rad:
        ax.plot([0, np.cos(a)], [0, np.sin(a)], color="#1e2235", linewidth=0.8, zorder=1)

    # Polygon köşeleri
    px = [v * np.cos(a) for v, a in zip(vals_norm, angles_rad)]
    py = [v * np.sin(a) for v, a in zip(vals_norm, angles_rad)]
    # Kapat
    px_closed = px + [px[0]]
    py_closed = py + [py[0]]

    ax.fill(px_closed, py_closed, color="#2ecc71", alpha=0.18, zorder=2)
    ax.plot(px_closed, py_closed, color="#2ecc71", linewidth=2.2, zorder=3)
    ax.scatter(px, py, color="#2ecc71", s=28, zorder=4)

    # Etiketler
    for a, name, val in zip(angles_rad, cat_names, cat_vals):
        lx = 1.18 * np.cos(a)
        ly = 1.18 * np.sin(a)

        ha = "center"
        if lx > 0.15:
            ha = "left"
        elif lx < -0.15:
            ha = "right"
        va = "center"
        if ly > 0.15:
            va = "bottom"
        elif ly < -0.15:
            va = "top"

        # Değer (yeşil, büyük)
        ax.text(lx, ly, str(val),
                ha=ha, va=va, fontsize=9.5, fontweight="bold",
                color="#2ecc71", zorder=5)
        # Kategori adı (gri, küçük) — biraz daha dışarıda
        ax.text(lx * 1.22, ly * 1.22, name,
                ha=ha, va=va, fontsize=7,
                color="#8b949e", zorder=5)

    plt.tight_layout(pad=0.1)
    return fig

    plt.tight_layout(pad=0.3)
    return fig


# =========================================================
# KATEGORİ ORTALAMA GÖSTERGESİ
# =========================================================

def category_bar(label, attrs_dict, color):
    avg = sum(attrs_dict.values()) / max(len(attrs_dict), 1)
    pct = avg / 20 * 100
    return f"""
    <div style='margin:5px 0'>
      <div style='display:flex;justify-content:space-between;font-size:11px;color:#ccc;margin-bottom:2px'>
        <span>{label}</span><span style='color:{color};font-weight:bold'>{avg:.1f}</span>
      </div>
      <div style='background:#1e1e2e;border-radius:4px;height:7px'>
        <div style='width:{pct:.1f}%;height:100%;background:{color};border-radius:4px;opacity:0.85'></div>
      </div>
    </div>"""


# =========================================================
# ATTRIBUTE TABLOSU
# =========================================================

def render_attributes_colored(title, attrs):
    st.markdown(f"### {title}")
    html = "<table style='width:100%;border-collapse:collapse;'>"
    for k, v in attrs.items():
        if v <= 4:
            color = "#9e9e9e"
        elif v <= 9:
            color = "#e0e0e0"
        elif v <= 13:
            color = "#f1c40f"
        elif v <= 16:
            color = "#2ecc71"
        else:
            color = "#3498db"
        html += f"<tr><td style='padding:2px 6px;font-size:0.78rem'>{k}</td><td style='padding:2px 6px;text-align:right;font-weight:bold;color:{color};font-size:0.78rem'>{v}</td></tr>"
    html += "</table>"
    st.markdown(html, unsafe_allow_html=True)


def build_player_card(name, flag, country, age, position, preset, grade,
                      badge_color, grade_color, ca, pa,
                      wf_color, weak_foot, wf_label,
                      height, weight, foot_icon, foot_label, dominant_foot_val,
                      value, role, tech, mental, phys, gk=None, secondary_pos=None):
    """Profil kartı HTML — modül seviyesinde, sıfır indent sorunu."""
    return (
            "<div style='"
            "background:linear-gradient(135deg,#0d1117 0%,#161b22 60%,#1a1f2e 100%);"
            "border:1px solid #30363d;border-radius:14px;padding:20px 24px;"
            "font-family:Segoe UI,sans-serif;margin-bottom:16px;"
            "box-shadow:0 4px 20px rgba(0,0,0,0.5)'>"

            # Üst satır: preset + scout
            "<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:14px'>"
            f"<span style='background:{badge_color};color:#000;font-size:11px;font-weight:800;"
            f"padding:3px 10px;border-radius:20px;letter-spacing:0.5px'>{preset.upper()}</span>"
            f"<span style='font-size:22px;font-weight:900;color:{grade_color}'>{grade}</span>"
            "</div>"

            # İsim + Ülke
            f"<div style='font-size:22px;font-weight:800;color:#f0f0f0;margin-bottom:2px'>{name}</div>"
            f"<div style='font-size:13px;color:#8b949e;margin-bottom:14px'>"
            f"{flag} {country} &nbsp;·&nbsp; {age} yaş &nbsp;·&nbsp; {position}"
            + (f" &nbsp;<span style='background:#21262d;border-radius:6px;padding:1px 8px;"
               f"font-size:11px;font-weight:700;color:#9b59b6'>"
               f"🎭 {gp.get('archetype','')}</span>"
               if gp.get("archetype") else "")
            + (f" &nbsp;<span style='background:#21262d;border-radius:6px;padding:1px 8px;"
               f"font-size:11px;font-weight:700;color:{SECONDARY_LEVEL_COLORS.get(secondary_pos[1], '#aaa')}'>"
               f"{secondary_pos[0]} <span style='opacity:0.7'>({secondary_pos[1]})</span></span>"
               if secondary_pos else "")
            + "</div>"

            # CA/PA bar
            + ca_bar(ca, pa) +

            # Zayıf ayak
            f"<div style='display:flex;align-items:center;gap:10px;margin:8px 0;padding:6px 10px;"
            f"background:#1a1f2e;border-radius:8px;border-left:3px solid {wf_color}'>"
            f"<span style='font-size:11px;color:#8b949e'>ZAYIF AYAK</span>"
            f"<div style='flex:1;background:#0d1117;border-radius:4px;height:8px;overflow:hidden'>"
            f"<div style='width:{int(weak_foot / 20 * 100)}%;height:100%;background:{wf_color};border-radius:4px'></div>"
            f"</div>"
            f"<span style='font-size:12px;font-weight:800;color:{wf_color}'>{weak_foot}</span>"
            f"<span style='font-size:11px;color:#8b949e'>{wf_label}</span>"
            "</div>"

            # Biyometrik grid
            "<div style='display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:14px 0'>"
            "<div style='background:#21262d;border-radius:8px;padding:8px;text-align:center'>"
            "<div style='font-size:10px;color:#8b949e;margin-bottom:2px'>BOY</div>"
            f"<div style='font-size:15px;font-weight:700;color:#f0f0f0'>{height} cm</div></div>"
            "<div style='background:#21262d;border-radius:8px;padding:8px;text-align:center'>"
            "<div style='font-size:10px;color:#8b949e;margin-bottom:2px'>KİLO</div>"
            f"<div style='font-size:15px;font-weight:700;color:#f0f0f0'>{weight} kg</div></div>"
            "<div style='background:#21262d;border-radius:8px;padding:8px;text-align:center'>"
            "<div style='font-size:10px;color:#8b949e;margin-bottom:2px'>BASKINDAYAK</div>"
            f"<div style='font-size:12px;font-weight:700;color:#f0f0f0'>{foot_icon} {foot_label}</div>"
            f"<div style='font-size:11px;color:#3498db;font-weight:800;margin-top:2px'>{dominant_foot_val}</div></div>"
            "<div style='background:#21262d;border-radius:8px;padding:8px;text-align:center'>"
            "<div style='font-size:10px;color:#8b949e;margin-bottom:2px'>DEĞER</div>"
            f"<div style='font-size:13px;font-weight:700;color:#2ecc71'>{value}</div></div>"
            "</div>"

            # Rol
            "<div style='font-size:12px;color:#8b949e;margin-bottom:4px'>ROL</div>"
            f"<div style='font-size:13px;color:#58a6ff;font-weight:600;margin-bottom:12px'>{role}</div>"

            # Kategori barları
            "<div style='font-size:12px;color:#8b949e;margin-bottom:6px'>KATEGORİ</div>"
            + category_bar("⚙️ Teknik", tech, "#58a6ff")
            + category_bar("🧠 Zihinsel", mental, "#bc8cff")
            + category_bar("💪 Fiziksel", phys, "#2ecc71")
            + (category_bar("🧤 Kaleci", gk, "#e67e22") if gk and position == "KL" else "")
            + "</div>"
    )


# =========================================================
# STREAMLIT UI
# =========================================================

st.set_page_config("FM Oyuncu Üretici", layout="wide")
st.title("⚽ Football Manager Oyuncu Üretici")

# =========================================================
# ⚙️ SIDEBAR AYARLAR PANELİ
# =========================================================
with st.sidebar:
    st.markdown("## ⚙️ Ayarlar")
    st.caption("Bu parametreler yeni üretimler için geçerlidir.")

    with st.expander("🔀 İkincil Mevki", expanded=False):
        _sec_prob = st.slider("Çıkma Olasılığı", 0.0, 1.0,
                              float(st.session_state.get("cfg_sec_prob", 0.40)), 0.05,
                              key="cfg_sec_prob")
        st.caption(f"Şu an: %{int(_sec_prob*100)}")
        st.info("📍 Etki: **Oyuncu Üret** sekmesindeki tüm yeni üretimler", icon="ℹ️")

    with st.expander("💰 Transfer Değeri", expanded=False):
        _tv_m = st.select_slider("Katsayı (Milyon €)",
                                 options=[100,150,200,250,300,350,400,450,500],
                                 value=int(st.session_state.get("cfg_tv_m", 300)),
                                 key="cfg_tv_m")
        st.caption(f"CA=200, 26 yaş → maks {_tv_m} Milyon €")
        st.info("📍 Etki: **tüm sekmelerdeki** transfer değeri hesabı", icon="ℹ️")

    with st.expander("🏋️ Kariyer & Yaş", expanded=False):
        _ret_age = st.slider("Emeklilik Yaşı", 38, 50,
                             int(st.session_state.get("cfg_ret_age", 46)),
                             key="cfg_ret_age")
        _max_games = st.slider("Sezon Max Maç", 40, 80,
                               int(st.session_state.get("cfg_max_games", 70)),
                               key="cfg_max_games")
        st.caption(f"Emeklilik: {_ret_age} yaş | Max maç: {_max_games}")
        st.info("📍 Etki: **Kariyer Modu** (Simülasyon sekmesi)", icon="ℹ️")

    with st.expander("🎓 Akademi Varsayılanları", expanded=False):
        _ak_years = st.slider("Simülasyon Yılı", 3, 8,
                              int(st.session_state.get("cfg_ak_years", 5)),
                              key="cfg_ak_years")
        _ak_count = st.slider("Akademi Oyuncu Sayısı", 6, 16,
                              int(st.session_state.get("cfg_ak_count", 10)),
                              key="cfg_ak_count")
        st.info("📍 Etki: **Akademi** sekmesi", icon="ℹ️")

    with st.expander("🏆 Mini Lig Varsayılanları", expanded=False):
        _lig_teams = st.slider("Takım Sayısı", 4, 24,
                               int(st.session_state.get("cfg_lig_teams", 8)),
                               key="cfg_lig_teams")
        st.info("📍 Etki: **Mini Lig** sekmesi", icon="ℹ️")

    st.divider()
    st.markdown("**Aktif Ayarlar**")
    st.markdown(
        f"🔀 2. mevki **%{int(_sec_prob*100)}** &nbsp;·&nbsp; "
        f"💰 **{_tv_m}M €** &nbsp;·&nbsp; "
        f"⚰️ **{_ret_age} yaş** &nbsp;·&nbsp; "
        f"⚽ **{_max_games} maç**"
    )

if "position" not in st.session_state:
    st.session_state.position = "ST"
if "gen_player" not in st.session_state:
    st.session_state.gen_player = None
if "player_pool" not in st.session_state:
    st.session_state.player_pool = []   # Oyuncu havuzu

# ── ANA SEKMELER ────────────────────────────────────────
tab1, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11 = st.tabs([
    "🎲 Oyuncu Üret",
    "🧮 Manuel CA",
    "🏟️ Kadro Üretici",
    "📅 Simülasyon",
    "📖 Analitik",
    "🎓 Akademi",
    "🏆 Mini Lig",
    "💼 Transfer & Sözleşme",
    "📦 Oyuncu Havuzu",
    "🔍 FM Analiz",
])

with tab1:
    # ── KONTROLLER ──────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        age = st.slider("Yaş", 15, 45, 22)
    with c2:
        country = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()))
    with c3:
        preset = st.selectbox("Oyuncu Profili", ["Average", "Wonderkid", "Star", "Superstar"],
                              help="Average: max 140 CA/PA\nWonderkid: PA 165-200, düşük CA\nStar: 150-180\nSuperstar: 180-200")

    preset_desc = {
        "Average"   : "🟦 Sıradan oyuncu — CA/PA max 140, attr max 14",
        "Wonderkid" : "🟨 Gelişim potansiyeli yüksek — PA 165-200, CA yaşa bağlı düşük",
        "Star"      : "🟧 Yıldız oyuncu — CA/PA 150-180, attr max 18",
        "Superstar" : "🔴 Süper yıldız — CA/PA 180-200, attr max 20",

    }
    st.info(preset_desc[preset])

    # ── MEVKİ SEÇİMİ ────────────────────────────────────────
    st.subheader("📍 Mevki Seç")
    ROW1 = ["ST", "KF (Sol)", "KF (Sağ)", "OOS", "KANAT (Sol)", "KANAT (Sağ)", "OS"]
    ROW2 = ["DM", "DOS", "KB (Sol)", "KB (Sağ)", "D (Sol)", "D (Sağ)", "KL"]

    cols1 = st.columns(len(ROW1))
    for i, p in enumerate(ROW1):
        lbl = f"✅ {p}" if st.session_state.position == p else p
        if cols1[i].button(lbl, key=f"btn_{p}", use_container_width=True):
            st.session_state.position = p

    cols2 = st.columns(len(ROW2))
    for i, p in enumerate(ROW2):
        lbl = f"✅ {p}" if st.session_state.position == p else p
        if cols2[i].button(lbl, key=f"btn_{p}", use_container_width=True):
            st.session_state.position = p

    position = st.session_state.position

    # ── ARKETİP SEÇİCİ ──────────────────────────────────────
    _arch_base = ARCHETYPE_BASE_MAP.get(position)
    _selected_archetype = None
    _arch_means = None
    if _arch_base and _arch_base in ARCHETYPES:
        _arch_options = list(ARCHETYPES[_arch_base].keys())
        _arch_labels  = {
            "ST"   : "Santrafor Tipi",
            "KF"   : "Kanat Forvet Tipi",
            "OOS"  : "Forvet Arkası Tipi",
            "KANAT": "Kanat Tipi",
            "OS"   : "Orta Saha Tipi",
            "DM"   : "Defansif OS Tipi",
            "DOS"  : "Stoper Tipi",
            "KB"   : "Kanat Bek Tipi",
            "D"    : "Tam Bek Tipi",
            "KL"   : "Kaleci Tipi",
        }.get(_arch_base, "Arketip")
        _arch_col1, _arch_col2 = st.columns([1, 3])
        with _arch_col1:
            _selected_archetype = st.selectbox(
                f"🎭 {_arch_labels}",
                _arch_options,
                key="selected_archetype",
                help="Rastgele seçersen her üretimde farklı tip çıkabilir"
            )
        with _arch_col2:
            if _selected_archetype and _selected_archetype != "🎲 Rastgele":
                _arch_means = ARCHETYPES[_arch_base][_selected_archetype]
                # Öne çıkan özellikler
                _top_attrs = sorted(
                    [(k,v) for k,v in _arch_means.items() if isinstance(v, (int,float)) and v >= 14],
                    key=lambda x: -x[1]
                )[:6]
                _badges = ""
                for _a, _v in _top_attrs:
                    _c = "#3498db" if _v>=17 else "#2ecc71" if _v>=15 else "#f1c40f"
                    _badges += (f"<span style='background:#21262d;border-radius:8px;"
                                f"padding:2px 8px;font-size:11px;color:{_c};"
                                f"margin:2px;display:inline-block'>{_a} {_v}</span>")
                if _badges:
                    st.markdown(
                        f"<div style='padding:6px 0'>**Öne çıkan:** {_badges}</div>",
                        unsafe_allow_html=True
                    )
            elif _selected_archetype == "🎲 Rastgele":
                st.info("Her üretimde rastgele bir arketip seçilir.", icon="🎲")

    # ── ÜRETİM ──────────────────────────────────────────────
    if st.button("🎲 Oyuncu Üret", type="primary"):

        # Temel üretim
        role = random.choice(ROLE_POOLS.get(position, ROLE_POOLS[POSITION_BASE[position]]))
        target_ca, pa = get_ca_pa(preset, age, country)
        # Rastgele arketip seçimi
        _final_arch_means = None
        _arch_base_gen = ARCHETYPE_BASE_MAP.get(position)
        _selected = st.session_state.get("selected_archetype", "🎲 Rastgele")
        if _arch_base_gen and _arch_base_gen in ARCHETYPES:
            if _selected == "🎲 Rastgele":
                # Rastgele arketip seç (Rastgele seçeneği hariç)
                _non_random = [k for k in ARCHETYPES[_arch_base_gen] if k != "🎲 Rastgele"]
                if _non_random:
                    _rand_arch = random.choice(_non_random)
                    _final_arch_means = ARCHETYPES[_arch_base_gen][_rand_arch]
                    st.session_state["_last_archetype"] = _rand_arch
            elif _selected != "🎲 Rastgele":
                _final_arch_means = ARCHETYPES[_arch_base_gen].get(_selected)
                st.session_state["_last_archetype"] = _selected

        _arch_caps = None
        if _final_arch_means and "_caps" in _final_arch_means:
            _arch_caps = _final_arch_means["_caps"]
        tech, mental, phys, gk, hidden = generate_all_attributes(
            position, age, preset, country, target_ca,
            archetype_means=_final_arch_means,
            cap_overrides=_arch_caps
        )
        base_pos = POSITION_BASE[position]
        all_attrs = {**tech, **mental, **phys, **gk} if base_pos == "KL" else {**tech, **mental, **phys}

        foot = generate_foot(position)
        weak_foot = generate_weak_foot(foot, position)
        dominant_foot_val = generate_dominant_foot_value()
        ca = min(calculate_ca(all_attrs, position, weak_foot), pa)
        name = generate_name(country)
        height, weight = generate_height_weight(position, phys)
        value = calculate_transfer_value(ca, pa, age)
        flag = COUNTRY_FLAG.get(country, "🏳️")
        best_roles = detect_best_role(all_attrs)
        traits = generate_traits(position)
        secondary_pos = generate_secondary_position(position)
        personality = detect_personality(hidden)

        # Session state'e kaydet — attr tablosu düzenlenebilsin
        st.session_state.gen_player = {
            "name": name, "flag": flag, "country": country, "age": age,
            "position": position, "preset": preset, "pa": pa,
            "archetype": st.session_state.get("_last_archetype", None),
            "foot": foot, "weak_foot": weak_foot, "dominant_foot_val": dominant_foot_val,
            "height": height, "weight": weight, "role": role,
            "tech": dict(tech), "mental": dict(mental), "phys": dict(phys),
            "gk": dict(gk), "hidden": dict(hidden),
            "traits": traits, "personality": personality,
            "secondary_pos": secondary_pos,
        }
        # edit_ widget keylerini sıfırla → yeni oyuncu değerleri görünsün
        for _k in list(st.session_state.keys()):
            if _k.startswith("edit_"):
                del st.session_state[_k]
        # Widget başlangıç değerlerini önceden ayarla
        for _a, _v in tech.items():   st.session_state[f"edit_tech_{_a}"] = _v
        for _a, _v in mental.items(): st.session_state[f"edit_ment_{_a}"] = _v
        for _a, _v in phys.items():   st.session_state[f"edit_phys_{_a}"] = _v
        for _a, _v in gk.items():     st.session_state[f"edit_gk_{_a}"] = _v
        for _a, _v in hidden.items(): st.session_state[f"edit_hid_{_a}"] = _v
        st.rerun()

    # ── Üretilmiş oyuncu görüntüle (session_state'ten) ──────────
    if st.session_state.gen_player:
        gp                = st.session_state.gen_player
        name              = gp["name"];
        flag              = gp["flag"];
        country_g         = gp["country"]
        age_g             = gp["age"];
        position_g        = gp["position"]
        preset_g          = gp["preset"];
        pa                = gp["pa"]
        foot              = gp["foot"];
        weak_foot         = gp["weak_foot"]
        dominant_foot_val = gp["dominant_foot_val"]
        height            = gp["height"];
        weight            = gp["weight"];
        role              = gp["role"]
        traits            = gp["traits"]
        secondary_pos     = gp.get("secondary_pos", None)



        # ── Widget'lardan güncel attr değerlerini oku ──────────────
        # Her değişiklikte tüm bileşenler bu güncel değerlerle yeniden çizilir
        def _read_attrs(src_dict, prefix):
            return {a: st.session_state.get(f"edit_{prefix}_{a}", v)
                    for a, v in src_dict.items()}


        tech   = _read_attrs(gp["tech"], "tech")
        mental = _read_attrs(gp["mental"], "ment")
        phys   = _read_attrs(gp["phys"], "phys")
        gk     = _read_attrs(gp["gk"], "gk")
        hidden = _read_attrs(gp["hidden"], "hid")


        # Tüm türetilen değerleri güncel attr'larla yeniden hesapla
        base_pos = POSITION_BASE[position_g]
        all_attrs = {**tech, **mental, **phys, **gk} if base_pos == "KL" else {**tech, **mental, **phys}
        ca = min(calculate_ca(all_attrs, position_g, weak_foot), pa)
        value = calculate_transfer_value(ca, pa, age_g)
        best_roles = detect_best_role(all_attrs)
        personality = detect_personality(hidden)  # hidden değişince kişilik de değişsin

        # Orijinal CA (ilk üretildiğindeki) — banner için
        orig_ca = gp.get("_orig_ca", ca)
        if "_orig_ca" not in gp:
            st.session_state.gen_player["_orig_ca"] = ca
            orig_ca = ca

        badge_color = {
            "Average": "#3498db", "Wonderkid": "#f1c40f", "Star": "#e67e22", "Superstar": "#e74c3c"
        }.get(preset_g, "#3498db")

        foot_icon = {"Sol": "🦶", "Sağ": "🦶", "Her İkisi": "⚡"}.get(foot, "🦶")
        foot_label = {"Sol": "Sol Ayak", "Sağ": "Sağ Ayak", "Her İkisi": "Her İkisi"}.get(foot, foot)
        wf_label = weak_foot_label(weak_foot)
        wf_color = ("#e74c3c" if weak_foot <= 4 else "#e67e22" if weak_foot <= 8
        else "#f1c40f" if weak_foot <= 12 else "#2ecc71" if weak_foot <= 16
        else "#3498db")

        grade = scout_grade(pa)
        grade_color = {"A": "#2ecc71", "B": "#27ae60", "C": "#f1c40f", "D": "#e67e22", "E": "#e74c3c"}.get(grade,
                                                                                                           "#aaa")

        card_html = build_player_card(
            name=name, flag=flag, country=country_g, age=age_g, position=position_g,
            preset=preset_g, grade=grade, badge_color=badge_color, grade_color=grade_color,
            ca=ca, pa=pa,
            wf_color=wf_color, weak_foot=weak_foot, wf_label=wf_label,
            height=height, weight=weight,
            foot_icon=foot_icon, foot_label=foot_label, dominant_foot_val=dominant_foot_val,
            value=value, role=role,
            tech=tech, mental=mental, phys=phys, gk=gk,
            secondary_pos=secondary_pos,
        )
        st.markdown(card_html, unsafe_allow_html=True)

        # ── Havuza Ekle butonu ─────────────────────────────
        _pool_ids = [p["_id"] for p in st.session_state.player_pool]
        _cur_id   = f"{name}_{age_g}_{ca}_{position_g}"
        _in_pool  = _cur_id in _pool_ids
        _hv_col1, _hv_col2, _hv_col3 = st.columns([1, 1, 4])
        with _hv_col1:
            if _in_pool:
                st.success("✅ Havuzda", icon="📦")
            else:
                if st.button("📦 Havuza Ekle", key="add_pool"):
                    pool_entry = {
                        "_id"        : _cur_id,
                        "_idx"       : len(st.session_state.player_pool),
                        "name"       : name,
                        "flag"       : flag,
                        "position"   : position_g,
                        "age"        : age_g,
                        "country"    : country_g,
                        "preset"     : preset_g,
                        "ca"         : ca,
                        "pa"         : pa,
                        "personality": personality,
                        "height"     : height,
                        "weight"     : weight,
                        "foot"       : foot,
                        "weak_foot"  : weak_foot,
                        "favori"     : False,
                        "tech"       : dict(tech),
                        "mental"     : dict(mental),
                        "phys"       : dict(phys),
                        "gk"         : dict(gk),
                        "hidden"     : dict(hidden),
                    }
                    st.session_state.player_pool.append(pool_entry)
                    st.rerun()
        with _hv_col2:
            st.caption(f"📦 Havuzda {len(st.session_state.player_pool)} oyuncu")

        # ══════════════════════════════════════════════════════
        # ANA LAYOUT
        # ══════════════════════════════════════════════════════
        col_left, col_mid, col_right = st.columns([1.1, 1, 2.2])

        with col_left:
            # ── Radar Chart ─────────────────────────────────
            st.subheader("📡 Radar")
            fig_r = radar_chart(all_attrs, position_g)
            st.pyplot(fig_r, use_container_width=True)
            plt.close(fig_r)

            # ── En İyi Roller ───────────────────────────────
            st.subheader("🎯 En Uygun Roller")
            for i, (role_name, score) in enumerate(best_roles):
                medal = ["🥇", "🥈", "🥉"][i]
                bar_w = int(score / 20 * 100)
                st.markdown(
                    f"""<div style='margin:4px 0'>
                    <div style='font-size:12px;color:#ccc'>{medal} {role_name}</div>
                    <div style='background:#1e1e2e;border-radius:4px;height:6px;margin-top:2px'>
                      <div style='width:{bar_w}%;height:100%;background:#58a6ff;border-radius:4px'></div>
                    </div></div>""",
                    unsafe_allow_html=True
                )

            # ── Kişilik ─────────────────────────────────────
            st.subheader("🧬 Kişilik")
            st.markdown(personality_card_html(personality, hidden), unsafe_allow_html=True)

            # ── PPM / Trait ─────────────────────────────────
            if traits:
                st.subheader("💡 PPM")
                for t in traits:
                    st.markdown(
                        f"<span style='background:#21262d;border-radius:12px;padding:3px 10px;font-size:12px;color:#ccc'>{t}</span>",
                        unsafe_allow_html=True)

        with col_mid:
            # ── Kariyer & Gelişim ───────────────────────────
            st.subheader("📊 Kariyer")

            col_ca, col_pa = st.columns(2)
            col_ca.metric("CA", ca)
            col_pa.metric("PA", pa)
            col_ca.metric("Gap", pa - ca)
            col_pa.metric("Scout", grade)

            years = years_to_star(age, ca, pa)
            if years is None:
                st.warning("⚠️ Yıldız ihtimali düşük")
            elif years == 0:
                st.success("⭐ Zaten yıldız!")
            else:
                st.info(f"~{years} yılda yıldız")

            st.subheader("📈 Gelişim")
            eff_pa = calculate_effective_pa(pa, hidden.get("Profesyonellik", 15), preset)
            growth = simulate_growth(age, ca, eff_pa, hidden.get("Profesyonellik", 15), position)

            if growth:
                ages = [y for y, _ in growth]
                vals = [v for _, v in growth]
                peak = max(vals)

                # Fiziksel ve zihinsel attr yaş eğrileri
                phys_keys = ["Hız", "Hızlanma", "Çeviklik", "Dayanıklılık"]
                ment_keys = ["Karar Alma", "Vizyon", "Konsantrasyon", "Önsezi"]
                phys_base = sum(all_attrs.get(k, 10) for k in phys_keys) / len(phys_keys)
                ment_base = sum(all_attrs.get(k, 10) for k in ment_keys) / len(ment_keys)


                def phys_mult(yr):
                    if yr >= 36: return 0.55
                    if yr >= 33: return 0.68
                    if yr >= 30: return 0.82
                    if yr >= 28: return 0.93
                    return 1.0


                def ment_mult(yr):
                    if yr >= 38: return 0.88
                    if yr >= 35: return 0.95
                    if yr >= 32: return 0.98
                    return 1.0


                phys_curve = [phys_base * phys_mult(yr) for yr in ages]
                ment_curve = [ment_base * ment_mult(yr) for yr in ages]

                fig, (ax, ax2) = plt.subplots(2, 1, figsize=(3.8, 4.0),
                                              gridspec_kw={"height_ratios": [2, 1.2]})
                fig.patch.set_facecolor("#0f0f1a")
                for a in (ax, ax2):
                    a.set_facecolor("#0f0f1a")
                    a.tick_params(colors="#aaa", labelsize=7)
                    a.spines[:].set_color("#333355")

                # Üst panel: CA gelişimi
                ax.plot(ages, vals, color="#2ecc71", linewidth=2, label="CA")
                ax.axhline(pa, color="#e67e22", ls="--", lw=1.2, label=f"PA {pa}")
                ax.axhline(eff_pa, color="#666", ls=":", lw=1.0, label=f"Eff {eff_pa}")
                ax.set_ylim(max(0, min(vals) - 10), min(200, pa + 15))
                ax.legend(fontsize=6, labelcolor="#aaa", facecolor="#0f0f1a", loc="upper right")
                ax.set_title("CA Gelişimi", fontsize=7, color="#8b949e", pad=3)

                # Alt panel: Fiziksel hızlı düşer, Zihinsel yavaş düşer
                ax2.plot(ages, phys_curve, color="#e74c3c", linewidth=1.8, label="Fiziksel")
                ax2.plot(ages, ment_curve, color="#9b59b6", linewidth=1.8, label="Zihinsel")
                if min(ages) <= 30 <= max(ages):
                    ax2.axvline(30, color="#444", ls=":", lw=1)
                    ax2.text(30.2, min(phys_curve) * 1.02, "30", fontsize=6, color="#666")
                ax2.legend(fontsize=6, labelcolor="#aaa", facecolor="#0f0f1a", loc="upper right")
                ax2.set_title("Özellik Yaşlanması", fontsize=7, color="#8b949e", pad=3)

                plt.tight_layout(pad=0.6)
                st.pyplot(fig, use_container_width=False)
                plt.close(fig)
                st.caption(f"PA kullanımı: %{int(peak / pa * 100)}")

        with col_right:
            # ── Attribute Kartı — DÜZENLENEBİLİR ──────────
            st.subheader("📋 Attribute Kartı")
            st.caption(
                "⚫ ≤4   ● 5-9   🟡 10-13   🟢 14-16   🔵 17+  |  ✏️ Değerleri düzenleyebilirsin — CA otomatik güncellenir")


            def attr_color(v):
                if v >= 17: return "#3498db"
                if v >= 14: return "#2ecc71"
                if v >= 10: return "#f1c40f"
                if v >= 5:  return "#e0e0e0"
                return "#9e9e9e"


            def editable_attr_col(title, attr_dict, prefix):
                st.markdown(f"**{title}**")
                result = {}
                for attr, default_val in attr_dict.items():
                    _key = f"edit_{prefix}_{attr}"
                    if _key not in st.session_state:
                        st.session_state[_key] = default_val
                    # Mevcut değeri oku (widget güncel değeri)
                    cur_val = st.session_state.get(_key, default_val)
                    # Renge göre stil
                    if cur_val >= 17:
                        vc = "#3498db";
                        dot = "🔵"
                    elif cur_val >= 14:
                        vc = "#2ecc71";
                        dot = "🟢"
                    elif cur_val >= 10:
                        vc = "#f1c40f";
                        dot = "🟡"
                    elif cur_val >= 5:
                        vc = "#e0e0e0";
                        dot = "●"
                    else:
                        vc = "#9e9e9e";
                        dot = "⚫"
                    col_l, col_r = st.columns([2, 1])
                    col_l.markdown(
                        f"<div style='font-size:0.73rem;color:#aaa;padding:2px 0;line-height:1.4'>"
                        f"<span style='color:{vc};font-weight:700;font-size:1.00rem'>{cur_val}</span>&nbsp;&nbsp;{attr} "
                        f"</div>",
                        unsafe_allow_html=True
                    )
                    new_val = col_r.number_input(
                        "", min_value=1, max_value=20,
                        key=_key, label_visibility="collapsed"
                    )
                    result[attr] = new_val
                return result


            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                editable_attr_col("⚙️ Teknik", tech, "tech")
            with c2:
                editable_attr_col("🧠 Zihinsel", mental, "ment")
            with c3:
                editable_attr_col("💪 Fiziksel", phys, "phys")
            with c4:
                editable_attr_col("🧤 Kaleci", gk, "gk")
            with c5:
                editable_attr_col("🔒 Gizli", hidden, "hid")

            # CA banner — tüm değerler zaten en üstte hesaplandı
            grade_e = scout_grade(ca)
            gc_e = {"A": "#2ecc71", "B": "#27ae60", "C": "#f1c40f", "D": "#e67e22", "E": "#e74c3c"}.get(grade_e, "#aaa")
            delta = ca - orig_ca
            delta_str = f"+{delta}" if delta > 0 else str(delta) if delta < 0 else "±0"
            delta_color = "#2ecc71" if delta > 0 else "#e74c3c" if delta < 0 else "#555"
            st.markdown(
                f"<div style='background:#161b22;border-radius:8px;padding:6px 14px;margin-top:8px;"
                f"font-size:14px;color:#ccc;text-align:center'>"
                f"CA = <span style='color:#2ecc71;font-weight:900;font-size:18px'>{ca}</span>"
                f"  <span style='color:{gc_e};font-weight:800'>({grade_e})</span>"
                f"  <span style='color:{delta_color};font-size:13px;font-weight:700'>{delta_str}</span>"
                f"  <span style='color:#555;font-size:11px'> vs orijinal {orig_ca}</span></div>",
                unsafe_allow_html=True
            )

        # ── YAŞ SİMÜLASYONU ──────────────────────────────────────────────────
        st.divider()
        st.subheader("⏳ Yaş Simülasyonu")
        st.caption("1'er yıllık periyotlarda gelişim ve yaşlanma — fiziksel erkenden düşer, zihinsel geç düşer.")

        # Her attribute'un zirve yaşı ve düşüş hızı
        ATTR_PROFILE = {
            # (peak_age, decline_speed): 1=çok hızlı, 2=hızlı, 3=orta, 4=yavaş, 5=çok yavaş
            "Hız"                    : (26, 1), "Hızlanma"         : (26, 1),
            "Çeviklik"               : (27, 1), "Güç"              : (28, 2),
            "Zıplama"                : (27, 2), "Dayanıklılık"     : (28, 2),
            "Denge"                  : (29, 3), "Vücut Zindeliği"  : (29, 3),
            "Bitiricilik"            : (29, 3), "Dripling"         : (28, 2),
            "Orta Yapma"             : (29, 3), "Top Kapma"        : (29, 3),
            "İlk Kontrol"            : (30, 4), "Pas"              : (30, 4),
            "Teknik"                 : (31, 4), "Markaj"           : (30, 3),
            "Uzaktan Şut"            : (30, 3), "Kafa Vuruşu"      : (30, 3),
            "Korner"                 : (31, 4), "Penaltı Kullanma" : (31, 4),
            "Serbest Vuruş Kullanma" : (31, 4), "Uzun Taç"         : (30, 3),
            "Karar Alma"             : (32, 5), "Önsezi"           : (32, 5),
            "Konsantrasyon"          : (32, 5), "Pozisyon Alma"    : (32, 5),
            "Vizyon"                 : (33, 5), "Soğukkanlılık"    : (33, 5),
            "Takım Oyunu"            : (33, 5), "Liderlik"         : (34, 5),
            "Topsuz Alan"            : (30, 4), "Çalışkanlık"      : (29, 3),
            "Agresiflik"             : (29, 3), "Cesaret"          : (31, 5),
            "Kararlılık"             : (32, 5), "Özel Yetenek"     : (30, 4),

            # Kaleci
            "Refleksler"        : (27, 1), "Elle Kontrol"          : (28, 2),
            "Birebir"           : (28, 2), "Bölge Hakimiyeti"      : (30, 3),
            "Hava Topları"      : (29, 3), "İletişim"              : (32, 4),
            "Degaj"             : (30, 4), "Elle Oyun Başlatma"    : (30, 4),
            "Pas (K)"           : (31, 4), "İlk Kontrol (K)"       : (30, 4),
            "Pozisyon Alma (K)" : (32, 5), "Ani Çıkış Eğilimi"     : (28, 3),
            "Eksantriklik"      : (28, 3), "Yumrukla Uzaklaştırma" : (28, 3),

        }

        # Decline hızı çarpanları (peak_age'den sonra her 5 yıl için)
        DECLINE_PER_5YR = {1: 0.13, 2: 0.09, 3: 0.06, 4: 0.035, 5: 0.018}

        # Gelişim: simulate_growth'tan CA eğrisi al (her yıl)
        prof         = hidden.get("Profesyonellik", 14)
        eff_pa_sim   = calculate_effective_pa(pa, prof, preset)
        growth_curve = simulate_growth(age, ca, eff_pa_sim, prof, position)
        ca_by_year   = {age: ca}
        for yr, val in growth_curve:
            ca_by_year[yr] = val
        # Eksik yıllar için interpolasyon
        for y in range(age, 41):
            if y not in ca_by_year:
                ca_by_year[y] = ca_by_year.get(y - 1, ca)


        def sim_attr(attr, base_val, current_age, target_age):
            """
            Gelişim fazı: CA büyümesiyle orantılı artış (peak_age'e kadar)
            Düşüş fazı: peak_age'den sonra hıza göre düşüş
            """
            peak_age, speed = ATTR_PROFILE.get(attr, (29, 3))
            decline_per5 = DECLINE_PER_5YR[speed]

            # Gelişim fazı — CA oranına göre attr büyür
            ca_now = ca_by_year.get(current_age, ca)
            ca_peak = ca_by_year.get(peak_age, ca_by_year.get(min(peak_age, 40), ca))
            ca_tgt = ca_by_year.get(min(target_age, 40), ca_by_year.get(40, ca))

            # Peak'teki değer
            if ca_now > 0 and ca_peak > ca_now:
                growth_ratio = min(ca_peak / ca_now, 1.0 + (peak_age - current_age) * 0.025)
            else:
                growth_ratio = 1.0
            peak_val = min(20, base_val * growth_ratio)

            # Target yaşında değer
            if target_age <= peak_age:
                # Henüz zirveye ulaşmamış — CA oranıyla büyü
                if ca_now > 0 and ca_tgt > ca_now:
                    ratio = min(ca_tgt / ca_now, growth_ratio)
                else:
                    ratio = 1.0
                result = base_val * ratio
            else:
                # Zirveyi geçti — düşüş
                years_past_peak = target_age - peak_age
                periods_past = years_past_peak / 5.0
                decline = 1.0 - (decline_per5 * periods_past)
                decline = max(0.40, decline)
                result = peak_val * decline

            return max(1, min(20, round(result)))


        # Periyotları belirle
        periods = list(range(age, 46))  # Her yıl, 45'e kadar

        # Tüm attr'ları simüle et
        all_sim = {}
        for yr in periods:
            all_sim[yr] = {}
            for attr in list(tech.keys()) + list(mental.keys()) + list(phys.keys()):
                src = tech if attr in tech else (mental if attr in mental else phys)
                all_sim[yr][attr] = sim_attr(attr, src[attr], age, yr)
            if POSITION_BASE[position] == "KL":
                for attr in gk:
                    all_sim[yr][attr] = sim_attr(attr, gk[attr], age, yr)
            # CA
            merged = {k: all_sim[yr][k] for k in all_sim[yr]}
            all_sim[yr]["__ca__"] = min(calculate_ca(merged, position, weak_foot), pa)

        # ── Grafik ────────────────────────────────────────────────────────────
        phys_keys = ["Hız", "Hızlanma", "Çeviklik", "Dayanıklılık", "Güç"]
        ment_keys = ["Karar Alma", "Vizyon", "Konsantrasyon", "Önsezi", "Soğukkanlılık"]
        tech_keys = ["Dripling", "Pas", "İlk Kontrol", "Teknik", "Bitiricilik"]


        def cat_avg(yr, keys):
            vals = [all_sim[yr].get(k, 10) for k in keys]
            return round(sum(vals) / len(vals), 1)


        phys_trend = [cat_avg(yr, phys_keys) for yr in periods]
        ment_trend = [cat_avg(yr, ment_keys) for yr in periods]
        tech_trend = [cat_avg(yr, tech_keys) for yr in periods]
        ca_trend = [all_sim[yr]["__ca__"] for yr in periods]

        fig_sim, (ax_top, ax_bot) = plt.subplots(1, 2, figsize=(10, 3.2))
        fig_sim.patch.set_facecolor("#0d1117")
        for ax in (ax_top, ax_bot):
            ax.set_facecolor("#0d1117")
            ax.tick_params(colors="#aaa", labelsize=8)
            ax.spines[:].set_color("#333355")

        ax_top.plot(periods, phys_trend, color="#e74c3c", lw=2.2, marker="o", ms=5, label="Fiziksel")
        ax_top.plot(periods, ment_trend, color="#9b59b6", lw=2.2, marker="o", ms=5, label="Zihinsel")
        ax_top.plot(periods, tech_trend, color="#3498db", lw=2.2, marker="o", ms=5, label="Teknik")
        ax_top.axvline(30, color="#444", ls=":", lw=1)
        ax_top.set_title("Kategori Ortalamaları", fontsize=9, color="#8b949e", pad=5)
        ax_top.legend(fontsize=7, labelcolor="#aaa", facecolor="#0d1117")
        ax_top.set_xticks(periods)
        ax_top.set_xlabel("Yaş", fontsize=8, color="#8b949e")
        ax_top.set_ylim(0, 20)
        ax_top.grid(color="#1e2235", lw=0.6)

        ax_bot.plot(periods, ca_trend, color="#2ecc71", lw=2.5, marker="o", ms=6)
        ax_bot.fill_between(periods, ca_trend, alpha=0.12, color="#2ecc71")
        ax_bot.axhline(pa, color="#e67e22", ls="--", lw=1.2, label=f"PA {pa}")
        ax_bot.axvline(30, color="#444", ls=":", lw=1)
        ax_bot.set_title("CA Değişimi", fontsize=9, color="#8b949e", pad=5)
        ax_bot.legend(fontsize=7, labelcolor="#aaa", facecolor="#0d1117")
        ax_bot.set_xticks(periods)
        ax_bot.set_xlabel("Yaş", fontsize=8, color="#8b949e")
        ax_bot.set_ylim(max(0, min(ca_trend) - 15), min(200, pa + 20))
        ax_bot.grid(color="#1e2235", lw=0.6)

        plt.tight_layout(pad=0.8)
        st.pyplot(fig_sim, use_container_width=True)
        plt.close(fig_sim)


        # ── Tablo — TÜM ATTR'LAR ─────────────────────────────────────────────
        def val_color(v):
            if v >= 17: return "#3498db"
            if v >= 14: return "#2ecc71"
            if v >= 10: return "#f1c40f"
            if v >= 5:  return "#e0e0e0"
            return "#9e9e9e"


        # Kategori sırası
        cat_sections = [
            ("💪 FİZİKSEL", "#e74c3c", [(a, phys) for a in PHYSICAL if a in phys]),
            ("🧠 ZİHİNSEL", "#9b59b6", [(a, mental) for a in MENTAL if a in mental]),
            ("⚙️ TEKNİK", "#3498db", [(a, tech) for a in TECHNICAL if a in tech]),
        ]
        if POSITION_BASE[position] == "KL":
            cat_sections.append(("🧤 KALECİ", "#e67e22", [(a, gk) for a in GOALKEEPER if a in gk]))

        tbl = "<div style='overflow-x:auto'><table style='border-collapse:collapse;font-size:1.00rem;white-space:nowrap'>"
        # Başlık
        tbl += "<tr style='color:#8b949e;border-bottom:2px solid #30363d;background:#0d1117;position:sticky;top:0'>"
        tbl += "<td style='padding:5px 10px;font-weight:700;position:sticky;left:0;background:#0d1117;z-index:2'>Özellik</td>"
        for yr in periods:
            bold = "font-weight:900;color:#e0e0e0;" if yr == age else ""
            bg = "background:#161b22;" if yr % 5 == 0 and yr != age else ""
            tbl += f"<td style='padding:5px 5px;text-align:center;min-width:32px;{bold}{bg}'>{yr}</td>"
        tbl += "</tr>"

        for sec_label, sec_color, attr_list in cat_sections:
            tbl += (
                f"<tr style='background:#111'>"
                f"<td colspan='{1 + len(periods)}' style='padding:4px 10px;"
                f"color:{sec_color};font-size:0.70rem;font-weight:800;"
                f"position:sticky;left:0;background:#111'>{sec_label}</td></tr>"
            )
            for attr, src_dict in attr_list:
                base_val = src_dict[attr]
                tbl += "<tr style='border-bottom:1px solid #1a1a2e'>"
                tbl += f"<td style='padding:2px 10px;color:#ccc;position:sticky;left:0;background:#0d1117;font-size:1.00rem'>{attr}</td>"
                for yr in periods:
                    v = all_sim[yr].get(attr, base_val)
                    c = val_color(v)
                    bg = "background:#0f1420;" if yr % 5 == 0 and yr != age else ""
                    bold = "font-weight:900;" if yr == age else ""
                    tbl += f"<td style='padding:1px 4px;text-align:center;{bold}{bg}color:{c}'>{v}</td>"
                tbl += "</tr>"

        # CA satırı
        tbl += "<tr style='border-top:2px solid #30363d;background:#0a0a14'>"
        tbl += "<td style='padding:4px 10px;color:#2ecc71;font-weight:800;position:sticky;left:0;background:#0a0a14'>CA</td>"
        base_ca = all_sim[age]["__ca__"]
        for yr in periods:
            v = all_sim[yr]["__ca__"]
            diff = v - base_ca
            diff_str = ""
            if yr != age and diff != 0:
                dc = "#e74c3c" if diff < 0 else "#2ecc71"
                diff_str = f"<br><span style='font-size:8px;color:{dc}'>({diff:+d})</span>"
            bold = "font-weight:900;" if yr == age else ""
            bg = "background:#0f1420;" if yr % 5 == 0 and yr != age else ""
            tbl += f"<td style='padding:3px 4px;text-align:center;{bold}{bg}color:#2ecc71;line-height:1.2'>{v}{diff_str}</td>"
        tbl += "</tr></table></div>"

        st.markdown(tbl, unsafe_allow_html=True)

    # =========================================================
    st.divider()

with tab3:
    # MANUEL CA HESAPLAYICI
    # =========================================================
    st.divider()
    st.header("🧮 Manuel CA Hesaplayıcı")
    st.caption("Kendi oyuncunun özellik değerlerini gir — CA gerçek FM ağırlıklarına göre anlık hesaplanır.")

    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        man_pos = st.selectbox("Mevki", ALL_POSITIONS, key="man_pos")
    with mc2:
        man_wf = st.slider("Zayıf Ayak (1-20)", 1, 20, 8, key="man_wf")
    with mc3:
        man_country = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="man_country")

    # Tek satır CA özeti — attribute'lar okunduktan sonra buraya yazılacak
    ca_summary_placeholder = st.empty()

    base_pos = POSITION_BASE[man_pos]
    is_gk = base_pos == "KL"

    st.markdown("---")

    # Mevkiye göre kaç kolon
    man_cols = st.columns(5 if not is_gk else 5)

    man_attrs = {}

    # Teknik
    with man_cols[0]:
        st.markdown("**⚙️ Teknik**")
        for attr in TECHNICAL:
            man_attrs[attr] = st.number_input(attr, 1, 20, 10, key=f"man_{attr}")

    # Zihinsel
    with man_cols[1]:
        st.markdown("**🧠 Zihinsel**")
        for attr in MENTAL:
            man_attrs[attr] = st.number_input(attr, 1, 20, 10, key=f"man_{attr}")

    # Fiziksel
    with man_cols[2]:
        st.markdown("**💪 Fiziksel**")
        for attr in PHYSICAL:
            man_attrs[attr] = st.number_input(attr, 1, 20, 10, key=f"man_{attr}")

    # Kaleci (her zaman göster ama GK değilse düşük default)
    with man_cols[3]:
        st.markdown("**🧤 Kaleci**")
        gk_default = {
            "Refleksler"      : 13, "Elle Kontrol"         : 12, "Birebir"          : 11, "Bölge Hakimiyeti"  : 11,
            "Hava Topları"    : 12, "İletişim"             : 10, "Degaj"            : 9, "Elle Oyun Başlatma" : 9,
            "İlk Kontrol (K)" : 10, "Pas (K)"              : 9, "Ani Çıkış Eğilimi" : 7,
            "Eksantriklik"    : 6, "Yumrukla Uzaklaştırma" : 8,

        }
        for attr in GOALKEEPER:
            default = gk_default.get(attr, 8) if is_gk else 3
            man_attrs[attr] = st.number_input(attr, 1, 20, default, key=f"man_{attr}")

        # Gizli (isteğe bağlı — opsiyonel kolon)
        with man_cols[4]:
            st.markdown("**🔒 Gizli**")
            for attr in HIDDEN:
                man_attrs[attr] = st.number_input(attr, 1, 20, 10, key=f"man_{attr}")

    # ── CA Hesapla ─────────────────────────────────────────
    man_ca = calculate_ca(man_attrs, man_pos, man_wf)
    man_grade = scout_grade(man_ca)
    man_grade_color = {"A": "#2ecc71", "B": "#27ae60", "C": "#f1c40f", "D": "#e67e22", "E": "#e74c3c"}.get(man_grade,
                                                                                                           "#aaa")

    # Tek satır özet — mevki seçiminin altındaki placeholder'a yaz
    ca_summary_placeholder.markdown(
        f"<div style='background:#161b22;border-radius:8px;padding:8px 16px;margin-bottom:4px;"
        f"font-size:15px;color:#ccc'>Hesaplanan CA = "
        f"<span style='color:#2ecc71;font-weight:900;font-size:18px'>{man_ca}</span>"
        f" &nbsp;·&nbsp; Scout: <span style='color:{man_grade_color};font-weight:800'>{man_grade}</span>"
        f" &nbsp;·&nbsp; Zayıf Ayak: <span style='color:#8b949e'>{man_wf} ({weak_foot_label(man_wf)})</span>"
        "</div>",
        unsafe_allow_html=True
    )

    # Radar
    man_radar_attrs = {k: v for k, v in man_attrs.items() if k not in HIDDEN_SET}
    man_fig = radar_chart(man_radar_attrs, man_pos)

    res_left, res_mid, res_right = st.columns([1.5, 1, 2])
    with res_left:
        st.markdown(f"""
        <div style='background:#0d1117;border:1px solid #30363d;border-radius:12px;padding:20px;text-align:center'>
          <div style='font-size:13px;color:#8b949e;margin-bottom:4px'>Hesaplanan CA</div>
          <div style='font-size:52px;font-weight:900;color:#2ecc71'>{man_ca}</div>
          <div style='font-size:28px;font-weight:900;color:{man_grade_color};margin-top:4px'>Scout: {man_grade}</div>
          <div style='font-size:12px;color:#8b949e;margin-top:12px'>Zayıf Ayak: {man_wf} · {weak_foot_label(man_wf)}</div>
          <div style='background:#1e1e2e;border-radius:6px;height:10px;margin-top:8px;overflow:hidden'>
            <div style='width:{man_ca / 200 * 100:.1f}%;height:100%;background:linear-gradient(90deg,#2ecc71,#27ae60);border-radius:6px'></div>
          </div>
          <div style='font-size:11px;color:#555;margin-top:4px'>{man_ca} / 200</div>
        </div>
        """, unsafe_allow_html=True)

    with res_mid:
        # Kategori barları
        tech_m = {k: man_attrs[k] for k in TECHNICAL}
        ment_m = {k: man_attrs[k] for k in MENTAL}
        phys_m2 = {k: man_attrs[k] for k in PHYSICAL}
        st.markdown("**Kategori Ortalamaları**")
        st.markdown(category_bar("⚙️ Teknik", tech_m, "#58a6ff"), unsafe_allow_html=True)
        st.markdown(category_bar("🧠 Zihinsel", ment_m, "#bc8cff"), unsafe_allow_html=True)
        st.markdown(category_bar("💪 Fiziksel", phys_m2, "#2ecc71"), unsafe_allow_html=True)
        if is_gk:
            gk_m = {k: man_attrs[k] for k in GOALKEEPER}
            st.markdown(category_bar("🧤 Kaleci", gk_m, "#e67e22"), unsafe_allow_html=True)

        # Ağırlıklı top attr'lar
        wts = ATTRIBUTE_WEIGHTS.get(base_pos, {})
        top_weighted = sorted(
            [(k, man_attrs.get(k, 1), w) for k, w in wts.items() if k not in HIDDEN_SET],
            key=lambda x: -x[1] * x[2]
        )[:5]
        st.markdown("**💡 En etkili özellikler**")
        for attr, val, w in top_weighted:
            c = "#3498db" if val >= 17 else "#2ecc71" if val >= 14 else "#f1c40f" if val >= 10 else "#e0e0e0"
            st.markdown(
                f"<span style='color:{c};font-size:12px'>● {attr}: <b>{val}</b> <span style='color:#555'>(×{w:.1f})</span></span>",
                unsafe_allow_html=True)

    with res_right:
        st.pyplot(man_fig, use_container_width=True)
        plt.close(man_fig)

    # ── TÜM MEVKİLERDE CA SIRALAMASI ──────────────────────────────────
    st.divider()
    st.subheader("📍 Tüm Mevkilerde CA")
    st.caption("Girilen attribute'lar bu oyuncunun hangi mevkide ne kadar değerli olduğunu gösterir.")

    # Her base mevki için CA hesapla
    pos_ca_list = []
    for pos in ALL_POSITIONS:
        pos_ca = calculate_ca(man_attrs, pos, man_wf)
        pos_ca_list.append((pos, pos_ca))

    # CA'ya göre sırala
    pos_ca_list.sort(key=lambda x: -x[1])

    max_ca_val = pos_ca_list[0][1]

    # Pozisyon grupları için renkler
    pos_colors = {
        "ST"       : "#e74c3c", "KF (Sol)"    : "#e74c3c", "KF (Sağ)"    : "#e74c3c",
        "OOS"      : "#e67e22", "KANAT (Sol)" : "#e67e22", "KANAT (Sağ)" : "#e67e22",
        "OS"       : "#f1c40f", "DM"          : "#f1c40f",
        "KB (Sol)" : "#2ecc71", "KB (Sağ)"    : "#2ecc71",
        "D (Sol)"  : "#3498db", "D (Sağ)"     : "#3498db", "DOS"         : "#3498db",
        "KL"       : "#9b59b6",

    }
    grade_colors = {"A": "#2ecc71", "B": "#27ae60", "C": "#f1c40f", "D": "#e67e22", "E": "#e74c3c"}

    rank_html = "<div style='display:flex;flex-direction:column;gap:5px'>"
    for i, (pos, pca) in enumerate(pos_ca_list):
        pct = pca / max(max_ca_val, 1) * 100
        pgrade = scout_grade(pca)
        gc = grade_colors.get(pgrade, "#aaa")
        pc = pos_colors.get(pos, "#aaa")
        medal = ["🥇", "🥈", "🥉"][i] if i < 3 else f"<span style='color:#555;font-size:11px'>#{i + 1}</span>"
        best = " ← En İyi" if i == 0 else ""

        rank_html += f"""
        <div style='display:flex;align-items:center;gap:8px;padding:4px 8px;
          background:{"#161b22" if i < 3 else "#0d1117"};border-radius:6px;
          border-left:3px solid {pc if i < 3 else "#222"}'>
          <span style='width:28px;text-align:center'>{medal}</span>
          <span style='width:90px;color:{pc};font-weight:{"700" if i < 3 else "400"};font-size:0.82rem'>{pos}</span>
          <div style='flex:1;background:#1e1e2e;border-radius:4px;height:8px;overflow:hidden'>
            <div style='width:{pct:.1f}%;height:100%;background:{pc};border-radius:4px;opacity:{"0.9" if i < 3 else "0.5"}'></div>
          </div>
          <span style='width:36px;text-align:right;font-weight:700;color:#f0f0f0;font-size:0.85rem'>{pca}</span>
          <span style='width:20px;text-align:center;font-weight:800;color:{gc};font-size:0.85rem'>{pgrade}</span>
          <span style='font-size:11px;color:#2ecc71;width:60px'>{best}</span>
        </div>"""

    rank_html += "</div>"
    st.markdown(rank_html, unsafe_allow_html=True)

with tab4:
    FORMATIONS = {
        "4-3-3"   : ["KL", "D (Sağ)", "DOS", "DOS", "D (Sol)", "OS", "DM", "OS", "KF (Sağ)", "ST", "KF (Sol)"],
        "4-4-2"   : ["KL", "D (Sağ)", "DOS", "DOS", "D (Sol)", "KANAT (Sağ)", "OS", "OS", "KANAT (Sol)", "ST", "ST"],
        "4-2-3-1" : ["KL", "D (Sağ)", "DOS", "DOS", "D (Sol)", "DM", "DM", "KANAT (Sağ)", "OOS", "KANAT (Sol)", "ST"],
        "3-5-2"   : ["KL", "DOS", "DOS", "DOS", "KB (Sağ)", "OS", "DM", "OS", "KB (Sol)", "ST", "ST"],
        "3-4-3"   : ["KL", "DOS", "DOS", "DOS", "KB (Sağ)", "OS", "OS", "KB (Sol)", "KF (Sağ)", "ST", "KF (Sol)"],
        "5-3-2"   : ["KL", "KB (Sağ)", "DOS", "DOS", "DOS", "KB (Sol)", "OS", "DM", "OS", "ST", "ST"],
        "4-1-4-1" : ["KL", "D (Sağ)", "DOS", "DOS", "D (Sol)", "DM", "KANAT (Sağ)", "OS", "OS", "KANAT (Sol)", "ST"],

    }

    FORMATION_PITCH = {
        "4-3-3": [(5, 0), (1, 1), (2.7, 1), (5, 1), (7.3, 1), (5, 2), (7, 3.2), (3, 3.2), (1, 3.5), (5, 4), (9, 3.5)],
        "4-4-2": [(5, 0), (1, 1), (2.7, 1), (5, 1), (7.3, 1), (1, 2.5), (3.3, 2.5), (6.7, 2.5), (9, 2.5), (3.5, 4),
                  (6.5, 4)],
        "4-2-3-1": [(5, 0), (1, 1), (2.7, 1), (5, 1), (7.3, 1), (3, 2), (7, 2), (1, 3.2), (5, 3.2), (9, 3.2), (5, 4.2)],
        "3-5-2": [(5, 0), (2.5, 1), (5, 1), (7.5, 1), (9, 2), (1, 2), (3.5, 2.5), (5, 2), (6.5, 2.5), (3.5, 4),
                  (6.5, 4)],
        "3-4-3": [(5, 0), (2.5, 1), (5, 1), (7.5, 1), (9, 2.2), (1, 2.2), (4, 3), (6, 3), (1, 4), (5, 4.2), (9, 4)],
        "5-3-2": [(5, 0), (1, 1), (2.5, 1), (5, 1), (7.5, 1), (9, 1), (2.5, 2.5), (5, 2.5), (7.5, 2.5), (3.5, 4),
                  (6.5, 4)],
        "4-1-4-1": [(5, 0), (1, 1), (2.7, 1), (5, 1), (7.3, 1), (5, 2), (1, 3), (3.3, 3), (6.7, 3), (9, 3), (5, 4.3)],
    }


    st.header("🏟️ Takım Kadro Üretici")
    
    sq_col1, sq_col2 = st.columns([1, 3])
    with sq_col1:
        formation = st.selectbox("Formasyon", list(FORMATIONS.keys()), key="sq_formation")
        sq_age = st.slider("Ortalama Yaş (±3)", 18, 36, 26, key="sq_age")
        sq_preset = st.selectbox("Kadro Seviyesi", ["Average", "Star", "Superstar", "Wonderkid"], key="sq_preset")
        sq_country = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="sq_country")
    
    with sq_col2:
        if st.button("⚽ Kadro Üret", type="primary", key="sq_btn"):
            positions = FORMATIONS[formation]
            squad = []
            total_value_num = 0
    
            for pos in positions:
                p_age = max(16, min(38, sq_age + random.randint(-3, 3)))
                t_ca, p_pa = get_ca_pa(sq_preset, p_age, sq_country)
                t, m, ph, gk_a, hid = generate_all_attributes(pos, p_age, sq_preset, sq_country, t_ca)
                bp = POSITION_BASE[pos]
                aa = {**t, **m, **ph, **gk_a} if bp == "KL" else {**t, **m, **ph}
                ft = generate_foot(pos)
                wf = generate_weak_foot(ft, pos)
                p_ca = min(calculate_ca(aa, pos, wf), p_pa)
                p_name = generate_name(sq_country)
                p_val = calculate_transfer_value(p_ca, p_pa, p_age)
                squad.append({"isim": p_name, "mevki": pos, "yaş": p_age, "CA": p_ca, "PA": p_pa, "değer": p_val})
    
    
            # Toplam değer
                def parse_value(v):
                    try:
                        v = v.replace("€", "").strip()
                        if "Milyon" in v: return float(v.replace("Milyon", "").strip()) * 1_000_000
                        if "M" in v: return float(v.replace("M", "").strip()) * 1_000_000
                        if "K" in v: return float(v.replace("K", "").strip()) * 1_000
                        return float(v)
                    except:
                        return 0
    
    
            total = sum(parse_value(p["değer"]) for p in squad)
            total_str = f"€{total / 1_000_000:.1f} M" if total >= 1_000_000 else f" €{total / 1_000:.0f}K"
            avg_ca = sum(p["CA"] for p in squad) // len(squad)
    
            # Pitch görselleştirme
            pitch_coords = FORMATION_PITCH.get(formation, [])
            fig_p, ax_p = plt.subplots(figsize=(6, 5.5))
            fig_p.patch.set_facecolor("#0d1117")
            ax_p.set_facecolor("#0d3e1a")
    
            # Saha çizgileri
            for lx in np.linspace(0, 10, 1):
                ax_p.axvline(lx, color="#1a5c2a", lw=0.5)
            ax_p.add_patch(plt.Rectangle((0.5, 0.1), 9, 4.6, fill=False, edgecolor="#2ecc71", lw=1.5))
            ax_p.add_patch(plt.Rectangle((2.5, 0.1), 5, 1.2, fill=False, edgecolor="#2ecc71", lw=1))
            ax_p.add_patch(plt.Rectangle((2.5, 3.5), 5, 1.2, fill=False, edgecolor="#2ecc71", lw=1))
            ax_p.add_patch(plt.Circle((5, 2.4), 0.8, fill=False, edgecolor="#2ecc71", lw=1))
            ax_p.axhline(2.4, color="#2ecc71", lw=0.7, alpha=0.6)
    
            # Oyuncular
            preset_colors = {"Average": "#3498db", "Wonderkid": "#f1c40f", "Star": "#e67e22", "Superstar": "#e74c3c"}
            dot_color = preset_colors.get(sq_preset, "#3498db")
    
            for i, (player, coord) in enumerate(zip(squad, pitch_coords)):
                cx, cy = coord
                ax_p.add_patch(plt.Circle((cx, cy + 0.1), 0.42, color=dot_color, zorder=5))
                short = player["isim"].split()[-1][:9]
                ax_p.text(cx, cy + 0.1, f"{player['CA']}", ha="center", va="center",
                          fontsize=6.5, fontweight="bold", color="#000", zorder=6)
                ax_p.text(cx, cy - 0.38, short, ha="center", va="top",
                          fontsize=5.5, color="#e0e0e0", zorder=6)
    
            ax_p.set_xlim(0, 10)
            ax_p.set_ylim(-0.3, 5.2)
            ax_p.set_aspect("equal")
            ax_p.axis("off")
            ax_p.set_title(f"{formation}  |  Ort. CA: {avg_ca}  |  Toplam Değer: {total_str}",
                           fontsize=9, color="#8b949e", pad=6)
            plt.tight_layout(pad=0.2)
            st.pyplot(fig_p, use_container_width=True)
            plt.close(fig_p)
    
            # Oyuncu listesi tablosu
            st.markdown("**📋 Kadro Listesi**")
            html = """<table style='width:100%;border-collapse:collapse;font-size:0.78rem'>
            <tr style='color:#8b949e;border-bottom:1px solid #30363d'>
              <th style='text-align:left;padding:4px 6px'>İsim</th>
              <th style='padding:4px 6px'>Mevki</th>
              <th style='padding:4px 6px'>Yaş</th>
              <th style='padding:4px 6px'>CA</th>
              <th style='padding:4px 6px'>PA</th>
              <th style='padding:4px 6px'>Değer</th>
            </tr>"""
            for p in squad:
                ca_color = "#3498db" if p["CA"] >= 170 else "#2ecc71" if p["CA"] >= 150 else "#f1c40f" if p[
                                                                                                              "CA"] >= 120 else "#e0e0e0"
                html += f"""<tr style='border-bottom:1px solid #1e2235'>
                  <td style='padding:3px 6px;color:#e0e0e0'>{p['isim']}</td>
                  <td style='padding:3px 6px;text-align:center;color:#58a6ff'>{p['mevki']}</td>
                  <td style='padding:3px 6px;text-align:center;color:#8b949e'>{p['yaş']}</td>
                  <td style='padding:3px 6px;text-align:center;font-weight:bold;color:{ca_color}'>{p['CA']}</td>
                  <td style='padding:3px 6px;text-align:center;color:#8b949e'>{p['PA']}</td>
                  <td style='padding:3px 6px;text-align:center;color:#2ecc71'>{p['değer']}</td>
                </tr>"""
            html += f"""<tr style='border-top:2px solid #30363d;background:#161b22'>
              <td colspan='5' style='padding:5px 6px;color:#aaa;font-weight:bold'>TOPLAM KADRO DEĞERİ</td>
              <td style='padding:5px 6px;text-align:center;font-weight:bold;color:#2ecc71;font-size:0.85rem'>{total_str}</td>
            </tr></table>"""
            st.markdown(html, unsafe_allow_html=True)
    
    # =========================================================

# =========================================================
# TAB 5 — SIMÜLASYON (Sezon Simülatörü + Kariyer Modu)
# Kaldırmak için: tabs tanımından "📅 Simülasyon" ve
# aşağıdaki "with tab5:" bloğunu silin.
# =========================================================
with tab5:
    sim_tab1, sim_tab2 = st.tabs(["⚽ Sezon Simülatörü", "🎯 Kariyer Modu"])

    # ──────────────────────────────────────────────────────
    # SEZON SİMÜLATÖRÜ
    # ──────────────────────────────────────────────────────
    with sim_tab1:
        st.subheader("⚽ Sezon Simülatörü")
        st.caption("Üretilen oyuncunun bir sezonluk performansını simüle et.")

        # ── Oyuncu Girişi ─────────────────────────────────
        ss_col1, ss_col2, ss_col3, ss_col4 = st.columns(4)
        with ss_col1:
            ss_pos = st.selectbox("Mevki", ALL_POSITIONS, key="ss_pos")
        with ss_col2:
            ss_ca = st.slider("CA", 40, 200, 140, key="ss_ca")
        with ss_col3:
            ss_age = st.slider("Yaş", 15, 44, 24, key="ss_age")
        with ss_col4:
            ss_inj = st.slider("Sakatlanma Eğilimi (1-20)", 1, 20, 8, key="ss_inj")

        ss_col5, ss_col6 = st.columns(2)
        with ss_col5:
            ss_league = st.selectbox("Lig Seviyesi", [
                "Elit (Premier League, La Liga)",
                "İyi (Bundesliga, Serie A)",
                "Orta (Süper Lig, Eredivisie)",
                "Alt (Championship, 2. Lig)",
            ], key="ss_league")
        with ss_col6:
            ss_role = st.selectbox("Takım İçi Rol", [
                "Vazgeçilmez (Hep 11'de)",
                "Önemli Oyuncu (Çoğunlukla 11'de)",
                "Rotasyon",
                "Yedek / Parça Parça",
            ], key="ss_role")

        if st.button("🎲 Sezonu Simüle Et", type="primary", key="ss_btn"):
            import math

            base_pos = POSITION_BASE[ss_pos]

            # ── Maç sayısı ────────────────────────────────
            role_games = {
                "Vazgeçilmez (Hep 11'de)": (55, 70),
                "Önemli Oyuncu (Çoğunlukla 11'de)": (40, 54),
                "Rotasyon": (20, 39),
                "Yedek / Parça Parça": (10, 19),
            }
            lo, hi = role_games[ss_role]
            total_games = random.randint(lo, hi)

            # ── Sakatlanma ────────────────────────────────
            inj_prob = ss_inj / 20 * 0.60  # max %60
            injured = random.random() < inj_prob
            inj_weeks = random.randint(2, 16) if injured else 0
            inj_games = min(total_games - 1, int(inj_weeks * 0.8))
            games_played = max(0, total_games - inj_games)

            # Sakatlanmanın sezon içindeki başlangıç maçı (deterministik seed)
            if injured and inj_games > 0:
                _rng_inj = random.Random(ss_ca * 31 + total_games)
                _max_start = max(1, total_games - inj_games)
                inj_start = _rng_inj.randint(1, _max_start)  # 1-indexed
                inj_end = inj_start + inj_games - 1
            else:
                inj_start = inj_end = -1

            # ── Temel istatistik katsayıları ──────────────
            ca_factor = ss_ca / 200  # 0-1 arası kalite çarpanı

            # Lig zorluğu — güçlü ligde daha az istatistik
            league_mult = {
                "Elit (Premier League, La Liga)": 0.75,
                "İyi (Bundesliga, Serie A)": 0.88,
                "Orta (Süper Lig, Eredivisie)": 1.00,
                "Alt (Championship, 2. Lig)": 1.18,
            }[ss_league]


            def rand_stat(per_game_mean, games, extra_mult=1.0):
                if games == 0: return 0
                per = per_game_mean * ca_factor * league_mult * extra_mult
                total = sum(max(0, random.gauss(per, per * 0.35)) for _ in range(games))
                return round(total, 1)


            # ── Tüm maç notlarını ÖNCE üret — her şey bundan türer ──────
            rng_form = random.Random(ss_ca * 13 + games_played)
            _cur = rng_form.gauss(6.5 + ca_factor * 1.5, 0.3)
            full_ratings = []
            for _gi in range(games_played):
                _cur += rng_form.gauss(0, 0.45)
                _cur = max(4.5, min(10.0, _cur))
                full_ratings.append(round(_cur, 2))

            # Form grafiği için ilk 20
            form_series = full_ratings[:20]


            # ── Maç başı ham istatistik ───────────────────────────────────────
            def _mstat(note, pos_base, idx):
                """Nota orantılı maç istatistikleri — sezon toplamları buradan türer."""
                r2 = random.Random(idx * 97 + int(note * 100))
                q  = max(0.0, min(1.0, (note - 4.5) / 5.5))

                def rgoal(lam_max):
                    """Poisson dağılımıyla gol sayısı.
                    lam = q * lam_max → beklenen gol/maç.
                    Not=10 ST için lam=0.75 → ~0.75 gol/maç, nadiren 2-3.
                    Not=7  ST için lam=0.34 → ~0.34 gol/maç.
                    Not=5  ST için lam=0.04 → çok nadir gol."""
                    lam = q * lam_max
                    if lam <= 0: return 0
                    p0 = math.exp(-lam)
                    p1 = p0 * lam
                    p2 = p1 * lam / 2
                    p3 = p2 * lam / 3
                    rv = r2.random()
                    if rv < p0:          return 0
                    if rv < p0 + p1:       return 1
                    if rv < p0 + p1 + p2:    return 2
                    if rv < p0 + p1 + p2 + p3: return 3
                    return 4

                def rq(hi, lo_frac=0.30):
                    """Alt sınır q*hi*lo_frac — iyi notta minimum da yüksek."""
                    top = int(q * hi)
                    bot = int(q * hi * lo_frac)
                    return r2.randint(bot, top) if top > bot else top

                def hit(p): return 1 if r2.random() < q * p else 0

                d = dict(gol=0, asist=0, sut=0, dribling=0, kp=0,
                         orta=0, topk=0, mud=0, blok=0, pas=0, kurtaris=0, cs=0)
                d = dict(gol=0, asist=0, sut=0, dribling=0, kp=0,
                         orta=0, topk=0, mud=0, blok=0, pas=0, kurtaris=0, cs=0)
                if pos_base == "ST":
                    g = rgoal(0.75); a = rgoal(0.25); s = rq(8)
                    d.update(gol=g, asist=a, sut=max(s, g))
                elif pos_base == "KF":
                    g = rgoal(0.55); a = rgoal(0.35); dr = rq(5)
                    d.update(gol=g, asist=a, dribling=max(dr, g))
                elif pos_base == "OOS":
                    g = rgoal(0.35); a = rgoal(0.50); kp = rq(4)
                    d.update(gol=g, asist=a, kp=max(kp, g))
                elif pos_base == "KANAT":
                    g = rgoal(0.38); a = rgoal(0.42)
                    d.update(gol=g, asist=a, orta=rq(8, 0.25))
                elif pos_base == "OS":
                    d.update(gol=rgoal(0.15), asist=rgoal(0.30), topk=rq(7, 0.25),
                             pas=int(20+q*50)+r2.randint(-3,3))
                elif pos_base == "DM":
                    d.update(gol=rgoal(0.08), asist=rgoal(0.18), topk=rq(8, 0.25),
                             pas=int(20+q*45)+r2.randint(-3,3))
                elif pos_base == "KB":
                    d.update(asist=rgoal(0.20), topk=rq(6, 0.25), orta=rq(5, 0.20))
                elif pos_base == "D":
                    d.update(topk=rq(6, 0.25), mud=rq(5, 0.20),
                             pas=int(15+q*35)+r2.randint(-2,2))
                elif pos_base == "DOS":
                    d.update(topk=rq(8, 0.25), mud=rq(6, 0.25), blok=rq(3, 0.20))
                elif pos_base == "KL":
                    d.update(kurtaris=rq(6, 0.25),
                             cs=(1 if r2.random() < q*0.65 else 0))
                return d


            all_mstats = [_mstat(rt, base_pos, i) for i, rt in enumerate(full_ratings)]


            # ── Sezon toplamları — maç istatistiklerinden türet ───────────────
            def _S(key):
                return sum(m[key] for m in all_mstats)


            real_avg_note = round(sum(full_ratings) / max(len(full_ratings), 1), 2)
            _pas_base = {"ST": 72, "KF": 75, "OOS": 78, "KANAT": 74,
                         "OS": 84, "DM": 83, "KB": 80, "D": 82, "DOS": 80, "KL": 72}
            pas_isabet = round(_pas_base.get(base_pos, 78) + (real_avg_note - 6.5) * 2.5, 1)

            if base_pos == "ST":
                xg = round(_S("gol") * 1.15 + random.uniform(-1, 2), 1)
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🎯 xG": max(0, xg),
                         "🎯 Şut": _S("sut"),
                         "📊 Şut/Maç": round(_S("sut") / max(games_played, 1), 1),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "KF":
                xg = round(_S("gol") * 1.1 + random.uniform(-0.5, 1.5), 1)
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🎯 xG": max(0, xg), "⚡ Dribling": _S("dribling"),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "OOS":
                xg = round(_S("gol") * 1.05 + random.uniform(-0.5, 1), 1)
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🎯 xG": max(0, xg), "🔑 Kilit Pas": _S("kp"),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "KANAT":
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🎯 xG": round(_S("gol") * 1.1 + 0.5, 1),
                         "🎯 Başarılı Çapraz": _S("orta"),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "OS":
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🛡️ Top Kapma": _S("topk"),
                         "📊 Ort. Pas": round(_S("pas") / max(games_played, 1), 1),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "DM":
                stats = {"⚽ Gol": _S("gol"), "🎯 Asist": _S("asist"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "🛡️ Top Kapma": _S("topk"),
                         "📊 Ort. Pas": round(_S("pas") / max(games_played, 1), 1),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "KB":
                stats = {"🎯 Asist": _S("asist"), "🛡️ Top Kapma": _S("topk"),
                         "⚽ Gol Katkısı": round(_S("gol") + _S("asist"), 1),
                         "✈️ Başarılı Orta": _S("orta"),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "D":
                stats = {"🛡️ Top Kapma": _S("topk"), "🧱 Müdahale": _S("mud"),
                         "📊 Ort. Pas": round(_S("pas") / max(games_played, 1), 1),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "DOS":
                stats = {"🛡️ Top Kapma": _S("topk"), "🧱 Müdahale": _S("mud"),
                         "🚫 Blok": _S("blok"),
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            elif base_pos == "KL":
                kur_pct = round(random.gauss(68 + ca_factor * 15, 4), 1)
                stats = {"🧤 Kurtarış": _S("kurtaris"), "🚫 Gol Yenmedi": _S("cs"),
                         "🎯 Kurtarış %": kur_pct,
                         "✅ Pas İsabeti %": pas_isabet, "🏆 Not Ort.": real_avg_note}
            else:
                stats = {"🏆 Not Ort.": real_avg_note}

            # ── Sonuç kartı ───────────────────────────────
            season_rating = (
                "🔥 Sezonun Oyuncusu" if real_avg_note >= 7.8 else
                "⭐ Harika Sezon" if real_avg_note >= 7.3 else
                "✅ İyi Sezon" if real_avg_note >= 6.8 else
                "😐 Ortalama Sezon" if real_avg_note >= 6.3 else
                "📉 Hayal Kırıklığı"
            )

            inj_text = (f"🏥 Sakatlandı — **{inj_weeks} hafta** ({inj_games} maç kaçırdı)" if injured
                        else "✅ Sezonu Sağlıklı Tamamladı")

            st.markdown(f"""
            <div style='background:linear-gradient(135deg,#0d1117,#161b22);border:1px solid #30363d;
            border-radius:12px;padding:18px 24px;margin-bottom:16px'>
              <div style='font-size:20px;font-weight:900;color:#f0f0f0;margin-bottom:4px'>{season_rating}</div>
              <div style='font-size:13px;color:#8b949e'>
                {ss_pos} · CA {ss_ca} · {ss_age} yaş · {ss_league.split("(")[0].strip()}
              </div>
              <div style='margin-top:12px;font-size:13px;color:#ccc'>{inj_text}</div>
              <div style='display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:14px'>
                <div style='background:#21262d;border-radius:8px;padding:10px;text-align:center'>
                  <div style='font-size:10px;color:#8b949e'>TOPLAM MAÇ</div>
                  <div style='font-size:24px;font-weight:800;color:#2ecc71'>{games_played}</div>
                </div>
                <div style='background:#21262d;border-radius:8px;padding:10px;text-align:center'>
                  <div style='font-size:10px;color:#8b949e'>KADRO RÖLÜ</div>
                  <div style='font-size:12px;font-weight:700;color:#58a6ff;margin-top:4px'>{ss_role.split("(")[0].strip()}</div>
                </div>
                <div style='background:#21262d;border-radius:8px;padding:10px;text-align:center'>
                  <div style='font-size:10px;color:#8b949e'>LİG</div>
                  <div style='font-size:11px;font-weight:700;color:#f1c40f;margin-top:4px'>{ss_league.split("(")[1].replace(")", "").split(",")[0].strip()}</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # İstatistik tablosu
            st.subheader("📊 Sezon İstatistikleri")
            stat_html = "<div style='display:grid;grid-template-columns:repeat(3,1fr);gap:8px'>"
            for stat_name, stat_val in stats.items():
                color = "#2ecc71" if "Not" in stat_name else "#58a6ff" if "%" in stat_name else "#f1c40f"
                stat_html += f"""
                <div style='background:#161b22;border-radius:8px;padding:10px;text-align:center;border:1px solid #21262d'>
                  <div style='font-size:11px;color:#8b949e;margin-bottom:4px'>{stat_name}</div>
                  <div style='font-size:22px;font-weight:800;color:{color}'>{stat_val}</div>
                </div>"""
            stat_html += "</div>"
            st.markdown(stat_html, unsafe_allow_html=True)

            # Form grafiği
            if form_series:
                st.subheader("📈 Form Grafiği")
                fig_f, ax_f = plt.subplots(figsize=(20, 2.5))
                fig_f.patch.set_facecolor("#0d1117")
                ax_f.set_facecolor("#0d1117")
                colors_form = ["#e74c3c" if v < 6.3 else "#f1c40f" if v < 7.0 else "#2ecc71"
                               for v in form_series]
                ax_f.bar(range(len(form_series)), form_series, color=colors_form, alpha=0.85, width=0.7)
                ax_f.axhline(7.0, color="#2ecc71", ls="--", lw=1, alpha=0.5, label="İyi form")
                ax_f.axhline(6.3, color="#e74c3c", ls="--", lw=1, alpha=0.5, label="Kötü form")
                ax_f.set_ylim(5, 10)
                ax_f.set_xlabel("Maç", fontsize=8, color="#8b949e")
                ax_f.set_ylabel("Not", fontsize=8, color="#8b949e")
                ax_f.tick_params(colors="#aaa", labelsize=7)
                ax_f.spines[:].set_color("#333355")
                ax_f.legend(fontsize=7, labelcolor="#aaa", facecolor="#0d1117")
                plt.tight_layout()
                st.pyplot(fig_f, use_container_width=True)
                plt.close(fig_f)

            # ── Maç Maç Değerlendirme Tablosu ─────────────────────────────
            if games_played > 0:
                st.subheader("📋 Maç Maç Değerlendirme")

                # Kaç maç göster — tüm sezon takvimi (sakatlık dahil)
                show_n = st.slider(
                    "Kaç maç göster?",
                    min_value=5, max_value=total_games,
                    value=min(total_games, 70),
                    key="ss_match_count"
                ) if total_games >= 5 else total_games


                # ── Değerlendirme etiketi ──────────────────────────────────
                def match_label(note):
                    if note >= 9.0: return "🌟 Efsane Performans", "#FFD700"
                    if note >= 8.5: return "🔥 Olağanüstü", "#2ecc71"
                    if note >= 8.0: return "⭐ Maçın Adamı", "#27ae60"
                    if note >= 7.5: return "👏 Harika Performans", "#58a6ff"
                    if note >= 7.0: return "✅ İyi Performans", "#3498db"
                    if note >= 6.5: return "👍 Ortalama Üstü", "#9b59b6"
                    if note >= 6.0: return "😐 Ortalama", "#8b949e"
                    if note >= 5.5: return "👎 Zayıf Performans", "#e67e22"
                    if note >= 5.0: return "😟 Kötü Maç", "#e74c3c"
                    return "💀 Berbat Performans", "#b71c1c"


                # ── Stat metni dict'ten formatla ──────────────────────────
                def match_stat_text(d, pb):
                    if pb == "ST":
                        return f"{d['gol']} Gol {d['asist']} Asist  {d['sut']} Şut"
                    elif pb == "KF":
                        return f"{d['gol']} Gol {d['asist']} Asist  {d['dribling']} Dribling"
                    elif pb == "OOS":
                        return f"{d['gol']} Gol {d['asist']} Asist  {d['kp']} KP"
                    elif pb == "KANAT":
                        return f"{d['gol']} Gol {d['asist']} Asist  {d['orta']} Orta"
                    elif pb == "OS":
                        return f"{d['asist']} Asist  {d['topk']} TopK  {d['pas']} Pas"
                    elif pb == "DM":
                        return f"{d['asist']} Asist  {d['topk']} TopK  {d['pas']} Pas"
                    elif pb == "KB":
                        return f"{d['asist']} Asist  {d['topk']} TopK  {d['orta']} Orta"
                    elif pb == "D":
                        return f"{d['topk']} TopK  {d['mud']} Müd  {d['pas']} Pas"
                    elif pb == "DOS":
                        return f"{d['topk']} TopK  {d['mud']} Müd  {d['blok']} Blok"
                    elif pb == "KL":
                        return f"{d['kurtaris']} Kur  {d['cs']} GoY"
                    return "—"


                # ── HTML tablosu — sakatlık satırları dahil ───────────────
                tbl_m = "<div style='overflow-x:auto'>"
                tbl_m += ("<table style='border-collapse:collapse;font-size:0.88rem;"
                          "width:100%;white-space:nowrap'>")
                tbl_m += ("<tr style='background:#0d1117;color:#8b949e;"
                          "border-bottom:2px solid #30363d'>"
                          "<td style='padding:5px 10px;font-weight:700'>Maç</td>"
                          "<td style='padding:5px 10px;font-weight:700;text-align:center'>Not</td>"
                          "<td style='padding:5px 10px;font-weight:700'>Değerlendirme</td>"
                          "<td style='padding:5px 10px;font-weight:700;text-align:center'>İstatistik</td>"
                          "<td style='padding:5px 10px;font-weight:700;text-align:center'>Form</td>"
                          "</tr>")

                played_ptr = 0  # full_ratings / all_mstats indeksi
                for game_num in range(1, show_n + 1):
                    is_inj = (inj_start != -1 and inj_start <= game_num <= inj_end)
                    row_bg = "#161b22" if game_num % 2 == 0 else "#0d1117"
                    if is_inj:
                        # Sakatlık satırı
                        tbl_m += (
                            f"<tr style='border-bottom:1px solid #1a1a2e;background:#1a0d0d'>"
                            f"<td style='padding:4px 10px;color:#8b949e'>{game_num}. Maç</td>"
                            f"<td style='padding:4px 10px;text-align:center;color:#e74c3c;font-weight:700'>—</td>"
                            f"<td style='padding:4px 10px;color:#e74c3c;font-weight:700'>"
                            f"🏥 Sakatlandı ({inj_weeks} hafta)</td>"
                            f"<td style='padding:4px 10px;text-align:center;color:#555'>—</td>"
                            f"<td style='padding:4px 14px;min-width:120px'>"
                            f"<div style='background:#1e1e2e;border-radius:4px;height:8px'></div></td>"
                            f"</tr>"
                        )
                    else:
                        if played_ptr >= len(full_ratings):
                            break
                        note = full_ratings[played_ptr]
                        msd = all_mstats[played_ptr]
                        played_ptr += 1
                        lbl, clr = match_label(note)
                        stat_txt = match_stat_text(msd, base_pos)
                        bar_pct = max(0, min(100, int((note - 4.5) / 5.5 * 100)))
                        note_sz = "1.05rem" if note >= 8.0 else "0.95rem"
                        tbl_m += (
                            f"<tr style='border-bottom:1px solid #1a1a2e;background:{row_bg}'>"
                            f"<td style='padding:4px 10px;color:#8b949e'>{game_num}. Maç</td>"
                            f"<td style='padding:4px 10px;text-align:center;font-weight:900;"
                            f"font-size:{note_sz};color:{clr}'>{note:.2f}</td>"
                            f"<td style='padding:4px 10px;color:{clr};font-weight:600'>{lbl}</td>"
                            f"<td style='padding:4px 10px;text-align:center;color:#aaa;"
                            f"font-size:0.82rem'>{stat_txt}</td>"
                            f"<td style='padding:4px 14px;min-width:120px'>"
                            f"<div style='background:#1e1e2e;border-radius:4px;height:8px;overflow:hidden'>"
                            f"<div style='width:{bar_pct}%;height:100%;background:{clr};"
                            f"border-radius:4px;opacity:0.85'></div></div></td>"
                            f"</tr>"
                        )

                # Özet alt satırı
                best_i = full_ratings.index(max(full_ratings))
                worst_i = full_ratings.index(min(full_ratings))
                b_lbl, b_c = match_label(full_ratings[best_i])
                w_lbl, w_c = match_label(full_ratings[worst_i])
                cnt_8plus = sum(1 for r in full_ratings if r >= 8.0)

                tbl_m += (
                    f"<tr style='border-top:2px solid #30363d;background:#0a0a14'>"
                    f"<td colspan='2' style='padding:5px 10px;color:#2ecc71;font-weight:800'>"
                    f"Sezon Ort. {real_avg_note:.2f}</td>"
                    f"<td colspan='3' style='padding:5px 10px;color:#8b949e;font-size:0.82rem'>"
                    f"En iyi: <b style='color:{b_c}'>{full_ratings[best_i]:.2f}</b>"
                    f" ({best_i + 1}. maç) &nbsp;·&nbsp; "
                    f"En kötü: <b style='color:{w_c}'>{full_ratings[worst_i]:.2f}</b>"
                    f" ({worst_i + 1}. maç) &nbsp;·&nbsp; "
                    f"8.0+ not: <b style='color:#58a6ff'>{cnt_8plus}</b> maç"
                    f"</td></tr>"
                )
                tbl_m += "</table></div>"
                st.markdown(tbl_m, unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────
    # KARİYER MODU
    # ──────────────────────────────────────────────────────
    with sim_tab2:
        st.subheader("🎯 Kariyer Modu")
        st.caption(
            "Oyuncunu 15 yaşından emekliliğe kadar yönet. Her yıl antrenman odağı seç, kulüp seç, gelişimi takip et.")

        # Session state başlat
        if "career_player" not in st.session_state:
            st.session_state.career_player = None
        if "career_log" not in st.session_state:
            st.session_state.career_log = []
        if "career_year" not in st.session_state:
            st.session_state.career_year = 0

        # ── Yeni Kariyer Başlat ───────────────────────────
        st.markdown("#### 👤 Yeni Oyuncu Oluştur")
        cr_col1, cr_col2, cr_col3 = st.columns(3)
        with cr_col1:
            cr_pos = st.selectbox("Mevki", ALL_POSITIONS, key="cr_pos")
        with cr_col2:
            cr_country = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="cr_country")
        with cr_col3:
            cr_preset = st.selectbox("Profil", ["Wonderkid", "Average", "Star"], key="cr_preset")

        if st.button("🌟 Kariyer Başlat", type="primary", key="cr_start"):
            start_age = 15
            t_ca, pa = get_ca_pa(cr_preset, start_age, cr_country)
            tech, mental, phys, gk, hidden = generate_all_attributes(
                cr_pos, start_age, cr_preset, cr_country, t_ca)
            base_pos_cr = POSITION_BASE[cr_pos]
            all_a = {**tech, **mental, **phys, **gk} if base_pos_cr == "KL" else {**tech, **mental, **phys}
            ft = generate_foot(cr_pos)
            wf = generate_weak_foot(ft, cr_pos)
            ca = min(calculate_ca(all_a, cr_pos, wf), pa)

            st.session_state.career_player = {
                "name"    : generate_name(cr_country),
                "pos"     : cr_pos,
                "country" : cr_country,
                "flag"    : COUNTRY_FLAG.get(cr_country, "🏳️"),
                "age"     : start_age,
                "ca"      : ca,
                "pa"      : pa,
                "preset"  : cr_preset,
                "tech"    : tech,
                "mental"  : mental,
                "phys"    : phys,
                "gk"      : gk,
                "hidden"  : hidden,
                "wf"      : wf,
                "club"    : "Kulüpsüz",
                "league"  : "—",
                "retired" : False,

            }
            st.session_state.career_log = []
            st.session_state.career_year = 0
            st.rerun()

        # ── Aktif Kariyer ─────────────────────────────────
        cp = st.session_state.career_player
        if cp:
            # Oyuncu başlık kartı
            grade = scout_grade(cp["pa"])
            g_color = {"A": "#2ecc71", "B": "#27ae60", "C": "#f1c40f", "D": "#e67e22", "E": "#e74c3c"}.get(grade,
                                                                                                           "#aaa")
            st.markdown(f"""
            <div style='background:#0d1117;border:1px solid #30363d;border-radius:12px;
            padding:14px 20px;margin:12px 0;display:flex;align-items:center;gap:20px'>
              <div>
                <div style='font-size:18px;font-weight:800;color:#f0f0f0'>{cp["flag"]} {cp["name"]}</div>
                <div style='font-size:12px;color:#8b949e'>{cp["pos"]} · {cp["country"]} · {cp["age"]} yaş</div>
                <div style='font-size:12px;color:#8b949e;margin-top:2px'>🏟️ {cp["club"]}</div>
              </div>
              <div style='margin-left:auto;text-align:right'>
                <div style='font-size:28px;font-weight:900;color:#2ecc71'>CA {cp["ca"]}</div>
                <div style='font-size:14px;color:#8b949e'>PA {cp["pa"]} · Scout <span style='color:{g_color}'>{grade}</span></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            if cp.get("retired"):
                st.success(f"🏁 {cp['name']} emekli oldu! Toplam {len(st.session_state.career_log)} sezon oynandı.")
            else:
                st.markdown(f"#### 📅 Sezon {st.session_state.career_year + 1} — {cp['age']} Yaş")

                # ── Yıllık Kararlar ───────────────────────
                dec_col1, dec_col2 = st.columns(2)

                with dec_col1:
                    st.markdown("**🏋️ Antrenman Odağı**")
                    focus = st.radio("Bu yıl ne üzerine çalışıyorsun?", [
                        "⚡ Hız & Çeviklik (Fiziksel)",
                        "🧠 Karar Alma & Vizyon (Zihinsel)",
                        "⚙️ Teknik & Pas (Teknik)",
                        "💪 Güç & Dayanıklılık (Kondisyon)",
                        "📋 Genel Gelişim",
                    ], key=f"cr_focus_{st.session_state.career_year}")

                with dec_col2:
                    st.markdown("**🏟️ Kulüp Seç**")
                    club_options = {
                        "Elit Kulüp (Top 5 Lig)": {"league": "Elit", "ca_req": 160, "dev_bonus": 0.8},
                        "İyi Kulüp (Orta Lig)": {"league": "İyi", "ca_req": 130, "dev_bonus": 1.0},
                        "Orta Kulüp (Alt Lig)": {"league": "Orta", "ca_req": 90, "dev_bonus": 1.2},
                        "Küçük Kulüp (Oyun Süresi!)": {"league": "Küçük", "ca_req": 0, "dev_bonus": 1.5},
                        "Mevcut Kulüpte Kal": {"league": cp.get("league", "—"), "ca_req": 0, "dev_bonus": 1.0},
                    }
                    club_choice = st.selectbox("Kulüp seç", list(club_options.keys()),
                                               key=f"cr_club_{st.session_state.career_year}")

                # ── Sezonu Oyna ───────────────────────────
                if st.button("▶️ Sezonu Oyna", key=f"cr_play_{st.session_state.career_year}"):
                    club_info = club_options[club_choice]
                    focus_label = focus.split("(")[0].strip()

                    # CA gelişimi
                    prof = cp["hidden"].get("Profesyonellik", 12)
                    eff_pa = calculate_effective_pa(cp["pa"], prof, cp["preset"])
                    base_pos_cr2 = POSITION_BASE[cp["pos"]]

                    # Antrenman bonusu
                    focus_bonus = {
                        "⚡ Hız & Çeviklik"       : 1.25,
                        "🧠 Karar Alma & Vizyon" : 1.25,
                        "⚙️ Teknik & Pas"        : 1.20,
                        "💪 Güç & Dayanıklılık"  : 1.20,
                        "📋 Genel Gelişim"       : 1.10,

                    }
                    f_mult = next((v for k, v in focus_bonus.items() if k in focus), 1.10)
                    dev_mult = club_info["dev_bonus"] * f_mult

                    # Yaşa göre gelişim faktörü
                    age = cp["age"]
                    af = (1.30 if age <= 18 else 1.00 if age <= 22 else
                    0.80 if age <= 26 else 0.50 if age <= 30 else 0.20)
                    gap = max(0, eff_pa - cp["ca"])
                    delta = int(af * dev_mult * min(gap, 12) * (prof / 15) * random.uniform(0.8, 1.2))
                    delta = max(0, min(delta, gap))

                    # Yaşlıysa düşüş
                    if age >= 32:
                        delta -= random.randint(1, max(1, int((age - 30) * 0.8)))
                    new_ca = max(1, min(cp["pa"], cp["ca"] + delta))

                    # Attribute güncellemesi (odak alanında +1-2)
                    focus_attrs = {
                        "⚡ Hız & Çeviklik"       : ["Hız", "Hızlanma", "Çeviklik", "Denge"],
                        "🧠 Karar Alma & Vizyon" : ["Karar Alma", "Vizyon", "Önsezi", "Konsantrasyon"],
                        "⚙️ Teknik & Pas"        : ["Pas", "Teknik", "İlk Kontrol", "Dripling"],
                        "💪 Güç & Dayanıklılık"  : ["Güç", "Dayanıklılık", "Zıplama", "Vücut Zindeliği"],

                    }
                    focused_attrs = next((v for k, v in focus_attrs.items() if k in focus), [])
                    for attr in focused_attrs:
                        for src in [cp["tech"], cp["mental"], cp["phys"]]:
                            if attr in src and delta > 0:
                                src[attr] = min(20, src[attr] + random.randint(0, 1))

                    # Sakatlanma
                    inj_prob = cp["hidden"].get("Sakatlanma Eğilimi", 8) / 20 * 0.45
                    injured = random.random() < inj_prob
                    inj_weeks = random.randint(2, 10) if injured else 0

                    # Maç sayısı
                    role_prob = min(1.0, new_ca / max(club_info["ca_req"], new_ca) * 0.9)
                    max_games = int(st.session_state.get('cfg_max_games', 70))
                    games_base = int(max_games * role_prob * random.uniform(0.7, 1.0))
                    games_played2 = max(2, games_base - int(inj_weeks * 0.8))

                    # Gol/Asist (mevkiye göre) — std mean ile orantılı, sabit değil
                    gol_per_game = {
                        "ST": 0.45, "KF": 0.30, "OOS": 0.22, "KANAT": 0.18,
                        "OS": 0.10, "DM": 0.06, "DOS": 0.03, "D": 0.02, "KB": 0.04, "KL": 0.00
                    }.get(base_pos_cr2, 0.08)
                    ast_per_game = {
                        "ST": 0.15, "KF": 0.22, "OOS": 0.28, "KANAT": 0.25,
                        "OS": 0.16, "DM": 0.12, "DOS": 0.03, "D": 0.07, "KB": 0.10, "KL": 0.00
                    }.get(base_pos_cr2, 0.10)


                    def safe_gauss_sum(mean, games, scale=0.5):
                        """std = mean*scale ile gauss, negatif sıfırlanır."""
                        if mean <= 0 or games == 0: return 0
                        std = max(0.01, mean * scale)
                        return int(sum(max(0, random.gauss(mean * (new_ca / 150), std))
                                       for _ in range(games)))


                    if base_pos_cr2 == "KL":
                        # Kaleci: gol/asist yok, kurtarış ve gol yemedi var
                        season_goals = 0
                        season_assists = 0
                        season_saves = safe_gauss_sum(3.2, games_played2, 0.3)
                        season_cs = max(0, int(games_played2 * random.gauss(0.28 + (new_ca / 200) * 0.15, 0.05)))
                    else:
                        season_goals = safe_gauss_sum(gol_per_game, games_played2)
                        season_assists = safe_gauss_sum(ast_per_game, games_played2)
                        season_saves = 0
                        season_cs = 0
                    avg_note = round(random.gauss(6.0 + (new_ca / 200) * 3, 0.4), 2)

                    # Kariyer logu
                    log_entry = {
                        "sezon"      : st.session_state.career_year + 1,
                        "yaş"        : age,
                        "kulüp"      : club_choice.split("(")[0].strip(),
                        "lig"        : club_info["league"],
                        "maç"        : games_played2,
                        "gol"        : season_goals,
                        "asist"      : season_assists,
                        "kurtarış"   : season_saves,
                        "gol_yemedi" : season_cs,
                        "not"        : avg_note,
                        "ca"         : new_ca,
                        "odak"       : focus_label,
                        "sakatlanma" : f"{inj_weeks}hf" if injured else "—",
                        "delta_ca"   : delta if delta >= 0 else delta,
                        "is_gk"      : base_pos_cr2 == "KL",

                    }
                    st.session_state.career_log.append(log_entry)

                    # Oyuncu güncelle
                    cp["ca"] = new_ca
                    cp["age"] += 1
                    cp["club"] = club_choice.split("(")[0].strip()
                    cp["league"] = club_info["league"]
                    _ret = int(st.session_state.get("cfg_ret_age", 46))
                    if cp["age"] >= _ret or (cp["age"] >= _ret - 3 and new_ca < 80):
                        cp["retired"] = True
                    st.session_state.career_year += 1
                    st.rerun()

            # ── Kariyer Logu ──────────────────────────────
            if st.session_state.career_log:
                st.markdown("---")
                st.subheader("📋 Kariyer Geçmişi")

                # CA Grafiği
                log = st.session_state.career_log
                ages_log = [e["yaş"] for e in log]
                ca_log = [e["ca"] for e in log]

                fig_cr, ax_cr = plt.subplots(figsize=(10, 2.8))
                fig_cr.patch.set_facecolor("#0d1117")
                ax_cr.set_facecolor("#0d1117")
                ax_cr.plot(ages_log, ca_log, color="#2ecc71", lw=2.5, marker="o", ms=6)
                ax_cr.fill_between(ages_log, ca_log, alpha=0.10, color="#2ecc71")
                ax_cr.axhline(cp["pa"], color="#e67e22", ls="--", lw=1.2, label=f"PA {cp['pa']}")
                ax_cr.set_title("CA Kariyer Eğrisi", fontsize=9, color="#8b949e")
                ax_cr.tick_params(colors="#aaa", labelsize=7)
                ax_cr.spines[:].set_color("#333355")
                ax_cr.legend(fontsize=7, labelcolor="#aaa", facecolor="#0d1117")
                ax_cr.grid(color="#1e2235", lw=0.5)
                plt.tight_layout()
                st.pyplot(fig_cr, use_container_width=True)
                plt.close(fig_cr)

                # Tablo — kaleci ve saha oyuncusu ayrı sütunlar
                is_gk_career = POSITION_BASE.get(cp.get("pos", "ST"), "") == "KL"
                tbl_cr = "<div style='overflow-x:auto'><table style='width:100%;border-collapse:collapse;font-size:1.00rem;white-space:nowrap'>"
                tbl_cr += "<tr style='color:#8b949e;border-bottom:2px solid #30363d;background:#0d1117'>"
                if is_gk_career:
                    headers = ["Sezon", "Yaş", "Kulüp", "Lig", "Maç", "Kurtarış", "Gol Yemedi", "Not", "CA", "ΔCA",
                               "Odak", "Sakatlık"]
                else:
                    headers = ["Sezon", "Yaş", "Kulüp", "Lig", "Maç", "Gol", "Asist", "⚡ Katkı", "Not", "CA", "ΔCA",
                               "Odak", "Sakatlık"]
                for h in headers:
                    tbl_cr += f"<td style='padding:5px 8px;font-weight:700'>{h}</td>"
                tbl_cr += "</tr>"
                for e in log:
                    dc = e["delta_ca"]
                    dc_color = "#2ecc71" if dc > 0 else "#e74c3c" if dc < 0 else "#aaa"
                    dc_str = f"+{dc}" if dc > 0 else str(dc)
                    note_c = "#2ecc71" if e["not"] >= 7 else "#f1c40f" if e["not"] >= 6.5 else "#e74c3c"
                    if is_gk_career:
                        col3 = f"<td style='padding:3px 8px;text-align:center;color:#58a6ff'>{e.get('kurtarış', 0)}</td>"
                        col4 = f"<td style='padding:3px 8px;text-align:center;color:#2ecc71'>{e.get('gol_yemedi', 0)}</td>"
                        col5 = ""
                    else:
                        katki = e['gol'] + e['asist']
                        col3 = f"<td style='padding:3px 8px;text-align:center;color:#f1c40f'>{e['gol']}</td>"
                        col4 = f"<td style='padding:3px 8px;text-align:center;color:#3498db'>{e['asist']}</td>"
                        col5 = f"<td style='padding:3px 8px;text-align:center;font-weight:700;color:#e67e22'>{katki}</td>"
                    tbl_cr += f"""<tr style='border-bottom:1px solid #1e2235'>
                      <td style='padding:3px 8px;color:#8b949e'>{e["sezon"]}</td>
                      <td style='padding:3px 8px;color:#ccc'>{e["yaş"]}</td>
                      <td style='padding:3px 8px;color:#58a6ff'>{e["kulüp"]}</td>
                      <td style='padding:3px 8px;color:#8b949e'>{e["lig"]}</td>
                      <td style='padding:3px 8px;text-align:center'>{e["maç"]}</td>
                      {col3}{col4}{col5}
                      <td style='padding:3px 8px;text-align:center;color:{note_c}'>{e["not"]}</td>
                      <td style='padding:3px 8px;text-align:center;color:#2ecc71;font-weight:700'>{e["ca"]}</td>
                      <td style='padding:3px 8px;text-align:center;color:{dc_color}'>{dc_str}</td>
                      <td style='padding:3px 8px;color:#8b949e;font-size:0.72rem'>{e["odak"]}</td>
                      <td style='padding:3px 8px;color:{"#e74c3c" if e["sakatlanma"] != "—" else "#444"}'>{e["sakatlanma"]}</td>
                    </tr>"""
                tbl_cr += "</table></div>"
                st.markdown(tbl_cr, unsafe_allow_html=True)

                # Özet istatistikler
                total_goals = sum(e["gol"] for e in log)
                total_assists = sum(e["asist"] for e in log)
                total_saves = sum(e.get("kurtarış", 0) for e in log)
                total_cs = sum(e.get("gol_yemedi", 0) for e in log)
                total_games = sum(e["maç"] for e in log)
                total_seasons = len(log)
                avg_note_career = round(sum(e["not"] for e in log) / max(len(log), 1), 2)

                st.markdown("**📊 Kariyer Toplamları**")
                if is_gk_career:
                    sum_cols = st.columns(5)
                    sum_cols[0].metric("Sezon", total_seasons)
                    sum_cols[1].metric("Maç", total_games)
                    sum_cols[2].metric("Kurtarış", total_saves)
                    sum_cols[3].metric("Gol Yemedi", total_cs)
                    sum_cols[4].metric("Not Ort.", avg_note_career)
                else:
                    sum_cols = st.columns(6)
                    sum_cols[0].metric("Sezon", total_seasons)
                    sum_cols[1].metric("Maç", total_games)
                    sum_cols[2].metric("⚽ Gol", total_goals)
                    sum_cols[3].metric("🎯 Asist", total_assists)
                    sum_cols[4].metric("⚡ Gol Katkısı", total_goals + total_assists)
                    sum_cols[5].metric("Not Ort.", avg_note_career)

            # Kariyer sıfırla
            if st.button("🔄 Kariyeri Sıfırla", key="cr_reset"):
                st.session_state.career_player = None
                st.session_state.career_log = []
                st.session_state.career_year = 0
                st.rerun()

# =========================================================
# TAB 6 — ANALİTİK
# =========================================================
with tab6:
    an_tab1, an_tab2, an_tab3, an_tab4 = st.tabs([
        "⚖️ Oyuncu Karşılaştırma",
        "📍 Pozisyon Fit Analizi",
        "🏟️ Kadro Denge Skoru",
        "📈 CA/PA Verimlilik",
    ])

    # ══════════════════════════════════════════════════════
    # AN-TAB 1 — OYUNCU KARŞILAŞTIRMA
    # ══════════════════════════════════════════════════════
    with an_tab1:
        st.subheader("⚖️ Oyuncu Karşılaştırma")
        st.caption("İki oyuncu üret — radar overlay, attribute fark tablosu ve skor karşılaştırması.")

        cmp_col1, cmp_col2 = st.columns(2)

        def _cmp_form(col, key_prefix, label, default_pos="ST"):
            with col:
                st.markdown(f"**{label}**")
                pos  = st.selectbox("Mevki", ALL_POSITIONS,
                                    index=ALL_POSITIONS.index(default_pos),
                                    key=f"{key_prefix}_pos")
                age  = st.slider("Yaş", 15, 40, 22, key=f"{key_prefix}_age")
                pre  = st.selectbox("Profil",
                                    ["Average","Wonderkid","Star","Superstar"],
                                    key=f"{key_prefix}_pre")
                cty  = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()),
                                    key=f"{key_prefix}_cty")
                return pos, age, pre, cty

        p1_pos, p1_age, p1_pre, p1_cty = _cmp_form(cmp_col1, "p1", "🔵 Oyuncu 1", "ST")
        p2_pos, p2_age, p2_pre, p2_cty = _cmp_form(cmp_col2, "p2", "🔴 Oyuncu 2", "KF (Sol)")

        if st.button("🎲 İki Oyuncuyu Üret & Karşılaştır", type="primary", key="cmp_gen"):
            def _gen(pos, age, pre, cty):
                tca, pa = get_ca_pa(pre, age, cty)
                tech, mental, phys, gk, hidden = generate_all_attributes(pos, age, pre, cty, tca)
                bp = POSITION_BASE[pos]
                all_a = {**tech,**mental,**phys,**gk} if bp=="KL" else {**tech,**mental,**phys}
                ft = generate_foot(pos)
                wf = generate_weak_foot(ft, pos)
                ca = min(calculate_ca(all_a, pos, wf), pa)
                name = generate_name(cty)
                personality = detect_personality(hidden)
                return {
                    "name": name, "pos": pos, "age": age, "pre": pre,
                    "cty": cty, "ca": ca, "pa": pa, "wf": wf,
                    "tech": tech, "mental": mental, "phys": phys,
                    "gk": gk, "hidden": hidden,
                    "all_a": all_a, "personality": personality,
                    "flag": COUNTRY_FLAG.get(cty, "🏳️"),
                }

            st.session_state["cmp_p1"] = _gen(p1_pos, p1_age, p1_pre, p1_cty)
            st.session_state["cmp_p2"] = _gen(p2_pos, p2_age, p2_pre, p2_cty)

        if "cmp_p1" in st.session_state and "cmp_p2" in st.session_state:
            P1 = st.session_state["cmp_p1"]
            P2 = st.session_state["cmp_p2"]

            # ── Başlık kartları ───────────────────────────
            h1, h2 = st.columns(2)
            for col, P, color in [(h1, P1, "#3498db"), (h2, P2, "#e74c3c")]:
                grade = scout_grade(P["pa"])
                gc = {"A":"#2ecc71","B":"#27ae60","C":"#f1c40f","D":"#e67e22","E":"#e74c3c"}.get(grade,"#aaa")
                col.markdown(
                    f"<div style='background:#161b22;border:1px solid {color};border-radius:10px;"
                    f"padding:12px 16px'>"
                    f"<div style='font-size:16px;font-weight:800;color:{color}'>{P['flag']} {P['name']}</div>"
                    f"<div style='font-size:12px;color:#8b949e'>{P['pos']} · {P['cty']} · {P['age']} yaş</div>"
                    f"<div style='margin-top:6px;font-size:18px;font-weight:900;color:#2ecc71'>CA {P['ca']}"
                    f"  <span style='font-size:13px;color:#8b949e'>/ PA {P['pa']}</span>"
                    f"  <span style='font-size:14px;color:{gc}'> {grade}</span></div>"
                    f"<div style='font-size:11px;color:#8b949e;margin-top:4px'>"
                    f"{P['personality']['name']}</div></div>",
                    unsafe_allow_html=True
                )

            # ── Radar Overlay ─────────────────────────────
            st.markdown("### 📡 Radar Karşılaştırma")
            bp1 = POSITION_BASE[P1["pos"]]
            bp2 = POSITION_BASE[P2["pos"]]
            cats = RADAR_CATEGORIES_GK if (bp1=="KL" or bp2=="KL") else RADAR_CATEGORIES

            cat_names = list(cats.keys())
            def _cat_avg(p_all, cat_attrs):
                vals = [p_all.get(a,1) for a in cat_attrs if a in p_all]
                return round(sum(vals)/len(vals)) if vals else 1

            v1 = [_cat_avg(P1["all_a"], cats[c]) for c in cat_names]
            v2 = [_cat_avg(P2["all_a"], cats[c]) for c in cat_names]
            N  = len(cat_names)
            angles = [np.pi/2 - i*2*np.pi/N for i in range(N)]
            n1 = [x/20 for x in v1]
            n2 = [x/20 for x in v2]

            fig_cmp, ax_c = plt.subplots(figsize=(5.5, 5.5))
            fig_cmp.patch.set_facecolor("#0d1117")
            ax_c.set_facecolor("#0d1117")
            ax_c.set_xlim(-1.5,1.5); ax_c.set_ylim(-1.5,1.5)
            ax_c.set_aspect("equal"); ax_c.axis("off")

            for r in [0.25,0.5,0.75,1.0]:
                cx = [r*np.cos(a) for a in np.linspace(0,2*np.pi,120)]
                cy = [r*np.sin(a) for a in np.linspace(0,2*np.pi,120)]
                ax_c.plot(cx, cy, color="#1e2235", lw=0.8)
            for a in angles:
                ax_c.plot([0,np.cos(a)],[0,np.sin(a)],color="#1e2235",lw=0.8)

            def _poly(vals_norm, color, label, alpha_fill=0.18):
                px = [v*np.cos(a) for v,a in zip(vals_norm,angles)]
                py = [v*np.sin(a) for v,a in zip(vals_norm,angles)]
                ax_c.fill(px+[px[0]], py+[py[0]], color=color, alpha=alpha_fill)
                ax_c.plot(px+[px[0]], py+[py[0]], color=color, lw=2.2, label=label)
                ax_c.scatter(px, py, color=color, s=24)

            _poly(n1, "#3498db", P1["name"].split()[-1])
            _poly(n2, "#e74c3c", P2["name"].split()[-1])

            for a, name, val1, val2 in zip(angles, cat_names, v1, v2):
                lx = 1.2*np.cos(a); ly = 1.2*np.sin(a)
                ha = "center" if abs(lx)<0.15 else ("left" if lx>0 else "right")
                va = "center" if abs(ly)<0.15 else ("bottom" if ly>0 else "top")
                ax_c.text(lx, ly, f"{val1}–{val2}", ha=ha, va=va,
                          fontsize=8, fontweight="bold", color="#e0e0e0")
                ax_c.text(lx*1.28, ly*1.28, name, ha=ha, va=va,
                          fontsize=6.5, color="#8b949e")

            ax_c.legend(fontsize=8, labelcolor="#ccc", facecolor="#0d1117",
                        loc="lower right")
            plt.tight_layout(pad=0.2)
            r_l, r_r = st.columns([1.2, 1])
            with r_l:
                st.pyplot(fig_cmp, use_container_width=True)
            plt.close(fig_cmp)

            # ── Attribute Fark Tablosu ─────────────────────
            with r_r:
                st.markdown("**Attribute Farkları (sadece ±2+)**")
                all_keys = list(P1["tech"]) + list(P1["mental"]) + list(P1["phys"])
                rows = []
                for attr in all_keys:
                    v1a = P1["all_a"].get(attr, 1)
                    v2a = P2["all_a"].get(attr, 1)
                    diff = v1a - v2a
                    if abs(diff) >= 2:
                        rows.append((attr, v1a, v2a, diff))
                rows.sort(key=lambda x: -abs(x[3]))

                diff_html = "<table style='width:100%;border-collapse:collapse;font-size:0.8rem'>"
                diff_html += ("<tr style='color:#8b949e;border-bottom:1px solid #30363d'>"
                              "<td style='padding:3px 6px'>Özellik</td>"
                              "<td style='padding:3px 6px;text-align:center;color:#3498db'>P1</td>"
                              "<td style='padding:3px 6px;text-align:center;color:#e74c3c'>P2</td>"
                              "<td style='padding:3px 6px;text-align:center'>Fark</td></tr>")
                for attr, v1a, v2a, diff in rows[:20]:
                    dc = "#3498db" if diff > 0 else "#e74c3c"
                    ds = f"+{diff}" if diff > 0 else str(diff)
                    diff_html += (f"<tr style='border-bottom:1px solid #1a1a2e'>"
                                  f"<td style='padding:2px 6px;color:#ccc'>{attr}</td>"
                                  f"<td style='padding:2px 6px;text-align:center;color:#3498db;font-weight:700'>{v1a}</td>"
                                  f"<td style='padding:2px 6px;text-align:center;color:#e74c3c;font-weight:700'>{v2a}</td>"
                                  f"<td style='padding:2px 6px;text-align:center;color:{dc};font-weight:800'>{ds}</td></tr>")
                diff_html += "</table>"
                st.markdown(diff_html, unsafe_allow_html=True)

            # ── Skor Karşılaştırma ─────────────────────────
            st.markdown("### 🏅 Kategori Skor Karşılaştırması")
            cat_keys = [
                ("⚙️ Teknik", list(P1["tech"].keys()), "#3498db"),
                ("🧠 Zihinsel", list(P1["mental"].keys()), "#9b59b6"),
                ("💪 Fiziksel", list(P1["phys"].keys()), "#2ecc71"),
            ]
            sc1, sc2, sc3 = st.columns(3)
            for col, (label, keys, color) in zip([sc1,sc2,sc3], cat_keys):
                avg1 = sum(P1["all_a"].get(k,1) for k in keys) / max(len(keys),1)
                avg2 = sum(P2["all_a"].get(k,1) for k in keys) / max(len(keys),1)
                winner = "🔵" if avg1 > avg2 else ("🔴" if avg2 > avg1 else "🟡")
                col.markdown(
                    f"<div style='background:#161b22;border-radius:8px;padding:10px;text-align:center'>"
                    f"<div style='font-size:11px;color:#8b949e'>{label}</div>"
                    f"<div style='font-size:12px;color:{color};margin:6px 0'>{winner}</div>"
                    f"<div style='display:flex;justify-content:space-between'>"
                    f"<span style='color:#3498db;font-weight:700;font-size:16px'>{avg1:.1f}</span>"
                    f"<span style='color:#555'>vs</span>"
                    f"<span style='color:#e74c3c;font-weight:700;font-size:16px'>{avg2:.1f}</span>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

    # ══════════════════════════════════════════════════════
    # AN-TAB 2 — POZİSYON FIT ANALİZİ
    # ══════════════════════════════════════════════════════
    with an_tab2:
        st.subheader("📍 Pozisyon Fit Analizi")
        st.caption("Bir oyuncu üret — hangi mevkide ne kadar değerli, neden orada iyi/kötü.")

        pf_col1, pf_col2, pf_col3, pf_col4 = st.columns(4)
        with pf_col1: pf_pos = st.selectbox("Ana Mevki", ALL_POSITIONS, key="pf_pos")
        with pf_col2: pf_age = st.slider("Yaş", 15, 40, 24, key="pf_age")
        with pf_col3: pf_pre = st.selectbox("Profil", ["Average","Wonderkid","Star","Superstar"], key="pf_pre")
        with pf_col4: pf_cty = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="pf_cty")

        if st.button("🔍 Analiz Et", type="primary", key="pf_btn"):
            tca, pa = get_ca_pa(pf_pre, pf_age, pf_cty)
            tech, mental, phys, gk, hidden = generate_all_attributes(pf_pos, pf_age, pf_pre, pf_cty, tca)
            bp = POSITION_BASE[pf_pos]
            all_a = {**tech,**mental,**phys,**gk} if bp=="KL" else {**tech,**mental,**phys}
            ft = generate_foot(pf_pos)
            wf = generate_weak_foot(ft, pf_pos)
            ca = min(calculate_ca(all_a, pf_pos, wf), pa)
            name = generate_name(pf_cty)
            flag = COUNTRY_FLAG.get(pf_cty, "🏳️")

            st.session_state["pf_player"] = {
                "name":name,"flag":flag,"pos":pf_pos,"age":pf_age,
                "pre":pf_pre,"cty":pf_cty,"ca":ca,"pa":pa,"wf":wf,
                "all_a":all_a,"tech":tech,"mental":mental,"phys":phys,
            }

        if "pf_player" in st.session_state:
            PF = st.session_state["pf_player"]
            grade = scout_grade(PF["pa"])
            gc = {"A":"#2ecc71","B":"#27ae60","C":"#f1c40f","D":"#e67e22","E":"#e74c3c"}.get(grade,"#aaa")
            st.markdown(
                f"<div style='background:#161b22;border:1px solid #30363d;border-radius:10px;"
                f"padding:10px 16px;margin-bottom:12px'>"
                f"<span style='font-size:16px;font-weight:800;color:#f0f0f0'>{PF['flag']} {PF['name']}</span>"
                f"<span style='color:#8b949e;margin-left:12px'>{PF['pos']} · {PF['age']} yaş</span>"
                f"<span style='color:#2ecc71;font-weight:900;font-size:16px;margin-left:12px'>CA {PF['ca']}</span>"
                f"<span style='color:{gc};margin-left:8px'>({grade})</span></div>",
                unsafe_allow_html=True
            )

            # Tüm mevkilerde CA hesapla
            pos_ca_all = [(pos, calculate_ca(PF["all_a"], pos, PF["wf"])) for pos in ALL_POSITIONS]
            pos_ca_all.sort(key=lambda x: -x[1])
            max_ca = pos_ca_all[0][1]

            # ── Yatay bar chart (matplotlib) ──────────────
            fig_pf, ax_pf = plt.subplots(figsize=(8, 5))
            fig_pf.patch.set_facecolor("#0d1117")
            ax_pf.set_facecolor("#0d1117")

            pos_colors_map = {
                "ST":"#e74c3c","KF (Sol)":"#e74c3c","KF (Sağ)":"#e74c3c",
                "OOS":"#e67e22","KANAT (Sol)":"#e67e22","KANAT (Sağ)":"#e67e22",
                "OS":"#f1c40f","DM":"#f1c40f",
                "KB (Sol)":"#2ecc71","KB (Sağ)":"#2ecc71",
                "D (Sol)":"#3498db","D (Sağ)":"#3498db","DOS":"#3498db",
                "KL":"#9b59b6",
            }
            labels = [p for p,_ in pos_ca_all]
            values = [c for _,c in pos_ca_all]
            colors_bar = [pos_colors_map.get(p,"#aaa") for p in labels]

            bars = ax_pf.barh(range(len(labels)), values,
                              color=colors_bar, alpha=0.8, height=0.6)
            ax_pf.set_yticks(range(len(labels)))
            ax_pf.set_yticklabels(labels, fontsize=8, color="#ccc")
            ax_pf.set_xlabel("CA", fontsize=8, color="#8b949e")
            ax_pf.set_xlim(0, max_ca * 1.15)
            ax_pf.tick_params(colors="#aaa", labelsize=7)
            ax_pf.spines[:].set_color("#1e2235")
            ax_pf.set_facecolor("#0d1117")
            ax_pf.invert_yaxis()

            # Değerleri barların üstüne yaz
            for i, (pos, ca_val) in enumerate(pos_ca_all):
                grade_p = scout_grade(ca_val)
                ax_pf.text(ca_val + max_ca*0.01, i, f"{ca_val} ({grade_p})",
                           va="center", fontsize=7.5, color="#e0e0e0", fontweight="bold")
                if pos == PF["pos"]:
                    ax_pf.get_yticklabels()[i].set_color("#2ecc71")
                    ax_pf.get_yticklabels()[i].set_fontweight("bold")

            ax_pf.set_title(f"{PF['name']} — Tüm Mevkilerde CA", fontsize=9, color="#8b949e", pad=8)
            plt.tight_layout(pad=0.5)

            pf_l, pf_r = st.columns([1.5, 1])
            with pf_l:
                st.pyplot(fig_pf, use_container_width=True)
            plt.close(fig_pf)

            # ── Güçlü/Zayıf Yönler Analizi ───────────────
            with pf_r:
                st.markdown("**🎯 En Uygun 3 Mevki**")
                for i, (pos, ca_val) in enumerate(pos_ca_all[:3]):
                    medal = ["🥇","🥈","🥉"][i]
                    w = ATTRIBUTE_WEIGHTS.get(POSITION_BASE[pos], {})
                    # Bu mevkide en kritik ve en düşük attr'lar
                    scored = sorted(
                        [(a, PF["all_a"].get(a,1)*wt)
                         for a,wt in w.items() if a not in HIDDEN_SET and a not in GK_ATTRS_SET],
                        key=lambda x: -x[1]
                    )
                    top3 = [a for a,_ in scored[:3]]
                    bot3 = [a for a,_ in scored[-3:] if PF["all_a"].get(a,1) < 10]
                    st.markdown(
                        f"<div style='background:#161b22;border-radius:8px;padding:8px 12px;margin-bottom:6px'>"
                        f"<div style='font-weight:700;color:#f0f0f0'>{medal} {pos} — CA {ca_val}</div>"
                        f"<div style='font-size:11px;color:#2ecc71;margin-top:3px'>✅ {', '.join(top3)}</div>"
                        + (f"<div style='font-size:11px;color:#e74c3c;margin-top:2px'>⚠️ {', '.join(bot3)}</div>" if bot3 else "")
                        + "</div>",
                        unsafe_allow_html=True
                    )

                st.markdown("**⚠️ En Zayıf 3 Mevki**")
                for pos, ca_val in pos_ca_all[-3:]:
                    grade_p = scout_grade(ca_val)
                    gc_p = {"A":"#2ecc71","B":"#27ae60","C":"#f1c40f","D":"#e67e22","E":"#e74c3c"}.get(grade_p,"#aaa")
                    st.markdown(
                        f"<span style='color:#e74c3c'>❌ {pos}</span>"
                        f"<span style='color:#8b949e'> — CA {ca_val}</span>"
                        f"<span style='color:{gc_p};font-weight:700'> ({grade_p})</span>",
                        unsafe_allow_html=True
                    )

    # ══════════════════════════════════════════════════════
    # AN-TAB 3 — KADRO DENGE SKORU
    # ══════════════════════════════════════════════════════
    with an_tab3:
        st.subheader("🏟️ Kadro Denge Skoru")
        st.caption("Bir kadro üret → yaş dağılımı, kategori dengesi, mevki derinliği, kişilik uyumu.")

        kb_col1, kb_col2, kb_col3, kb_col4 = st.columns(4)
        with kb_col1: kb_form = st.selectbox("Formasyon", ["4-3-3","4-4-2","4-2-3-1","3-5-2","3-4-3","5-3-2","4-1-4-1"], key="kb_form")
        with kb_col2: kb_pre  = st.selectbox("Kadro Seviyesi", ["Average","Star","Superstar","Wonderkid"], key="kb_pre")
        with kb_col3: kb_age  = st.slider("Ort. Yaş", 18, 36, 25, key="kb_age")
        with kb_col4: kb_cty  = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="kb_cty")

        FORMATIONS_KB = {
            "4-3-3"  :["KL","D (Sağ)","DOS","DOS","D (Sol)","OS","DM","OS","KF (Sağ)","ST","KF (Sol)"],
            "4-4-2"  :["KL","D (Sağ)","DOS","DOS","D (Sol)","KANAT (Sağ)","OS","OS","KANAT (Sol)","ST","ST"],
            "4-2-3-1":["KL","D (Sağ)","DOS","DOS","D (Sol)","DM","DM","KANAT (Sağ)","OOS","KANAT (Sol)","ST"],
            "3-5-2"  :["KL","DOS","DOS","DOS","KB (Sağ)","OS","DM","OS","KB (Sol)","ST","ST"],
            "3-4-3"  :["KL","DOS","DOS","DOS","KB (Sağ)","OS","OS","KB (Sol)","KF (Sağ)","ST","KF (Sol)"],
            "5-3-2"  :["KL","KB (Sağ)","DOS","DOS","DOS","KB (Sol)","OS","DM","OS","ST","ST"],
            "4-1-4-1":["KL","D (Sağ)","DOS","DOS","D (Sol)","DM","KANAT (Sağ)","OS","OS","KANAT (Sol)","ST"],
        }

        if st.button("📊 Kadro Üret & Analiz Et", type="primary", key="kb_btn"):
            positions_kb = FORMATIONS_KB[kb_form]
            squad_kb = []
            for pos in positions_kb:
                p_age = max(16, min(38, kb_age + random.randint(-4, 4)))
                t_ca, p_pa = get_ca_pa(kb_pre, p_age, kb_cty)
                t,m,ph,gk_a,hid = generate_all_attributes(pos, p_age, kb_pre, kb_cty, t_ca)
                bp = POSITION_BASE[pos]
                aa = {**t,**m,**ph,**gk_a} if bp=="KL" else {**t,**m,**ph}
                ft = generate_foot(pos)
                wf = generate_weak_foot(ft, pos)
                p_ca = min(calculate_ca(aa, pos, wf), p_pa)
                personality = detect_personality(hid)
                squad_kb.append({
                    "isim": generate_name(kb_cty),
                    "pos": pos, "yaş": p_age,
                    "CA": p_ca, "PA": p_pa,
                    "tech": t, "mental": m, "phys": ph,
                    "all_a": aa, "hidden": hid,
                    "personality": personality,
                })
            st.session_state["kb_squad"] = squad_kb

        if "kb_squad" in st.session_state:
            squad = st.session_state["kb_squad"]
            ages  = [p["yaş"] for p in squad]
            cas   = [p["CA"]  for p in squad]
            pas   = [p["PA"]  for p in squad]

            avg_ca  = sum(cas)/len(cas)
            avg_age = sum(ages)/len(ages)
            avg_pa  = sum(pas)/len(pas)
            pot_use = avg_ca / avg_pa * 100

            # Denge skoru hesapla
            age_balance = max(0, 100 - abs(avg_age - 26) * 8)
            ca_var = (sum((c-avg_ca)**2 for c in cas)/len(cas))**0.5
            ca_balance = max(0, 100 - ca_var * 2.5)
            pos_groups = {"Hücum":["ST","KF","OOS","KANAT"], "Orta":["OS","DM"], "Defans":["DOS","D","KB"], "Kale":["KL"]}
            group_cas = {}
            for grp, bases in pos_groups.items():
                grp_cas = [p["CA"] for p in squad if POSITION_BASE[p["pos"]] in bases]
                group_cas[grp] = sum(grp_cas)/max(len(grp_cas),1)
            balance_range = max(group_cas.values()) - min(group_cas.values())
            tactical_balance = max(0, 100 - balance_range * 1.5)

            # Kişilik dengesi
            negative_profiles = {"😤 Sorunlu Karakter","💣 Zehirli Unsur","💔 Motivasyon Yoksunu","💤 Tembel Deha"}
            neg_count = sum(1 for p in squad if p["personality"]["name"] in negative_profiles)
            personality_score = max(0, 100 - neg_count * 20)

            total_score = (age_balance*0.25 + ca_balance*0.25 + tactical_balance*0.30 + personality_score*0.20)

            # ── Metrik kartları ───────────────────────────
            m1,m2,m3,m4,m5 = st.columns(5)
            for col, label, val, unit in [
                (m1,"Ort. CA",f"{avg_ca:.0f}",""),
                (m2,"Ort. Yaş",f"{avg_age:.1f}",""),
                (m3,"PA Kullanım",f"{pot_use:.0f}","%"),
                (m4,"Sorunlu Oyuncu",neg_count,"kişi"),
                (m5,"Denge Skoru",f"{total_score:.0f}","/100"),
            ]:
                score_color = "#2ecc71" if total_score>=70 else "#f1c40f" if total_score>=50 else "#e74c3c"
                c = score_color if col==m5 else "#f0f0f0"
                col.markdown(
                    f"<div style='background:#161b22;border-radius:8px;padding:10px;text-align:center'>"
                    f"<div style='font-size:10px;color:#8b949e'>{label}</div>"
                    f"<div style='font-size:22px;font-weight:900;color:{c}'>{val}{unit}</div></div>",
                    unsafe_allow_html=True
                )

            # ── Alt grafikler ──────────────────────────────
            st.markdown("---")
            fig_kb, axes = plt.subplots(1, 3, figsize=(13, 3.5))
            fig_kb.patch.set_facecolor("#0d1117")

            # 1) Yaş dağılımı
            ax_age = axes[0]; ax_age.set_facecolor("#0d1117")
            age_colors = ["#2ecc71" if 21<=a<=28 else "#f1c40f" if a<=32 else "#e74c3c" for a in ages]
            ax_age.bar(range(len(ages)), ages, color=age_colors, alpha=0.85, width=0.6)
            ax_age.axhline(26, color="#f1c40f", ls="--", lw=1, label="Altın çağ (26)")
            ax_age.axhline(30, color="#e74c3c", ls=":", lw=1, label="30+")
            ax_age.set_title("Yaş Dağılımı", fontsize=9, color="#8b949e")
            ax_age.tick_params(colors="#aaa", labelsize=7)
            ax_age.spines[:].set_color("#1e2235")
            ax_age.legend(fontsize=6, labelcolor="#aaa", facecolor="#0d1117")
            ax_age.set_ylabel("Yaş", fontsize=7, color="#8b949e")

            # 2) CA/PA bar
            ax_ca = axes[1]; ax_ca.set_facecolor("#0d1117")
            x = range(len(squad))
            ax_ca.bar(x, [p["PA"] for p in squad], color="#e67e22", alpha=0.35, label="PA", width=0.6)
            ax_ca.bar(x, [p["CA"] for p in squad], color="#2ecc71", alpha=0.85, label="CA", width=0.6)
            ax_ca.set_title("CA / PA", fontsize=9, color="#8b949e")
            ax_ca.tick_params(colors="#aaa", labelsize=7)
            ax_ca.spines[:].set_color("#1e2235")
            ax_ca.legend(fontsize=6, labelcolor="#aaa", facecolor="#0d1117")
            ax_ca.set_ylabel("CA/PA", fontsize=7, color="#8b949e")

            # 3) Kategori dengesi (radar)
            ax_bal = axes[2]; ax_bal.set_facecolor("#0d1117")
            grp_labels = list(group_cas.keys())
            grp_vals   = [group_cas[g] for g in grp_labels]
            bar_colors = ["#e74c3c","#f1c40f","#3498db","#9b59b6"]
            ax_bal.bar(grp_labels, grp_vals, color=bar_colors, alpha=0.85, width=0.5)
            ax_bal.set_title("Grup CA Ortalaması", fontsize=9, color="#8b949e")
            ax_bal.tick_params(colors="#aaa", labelsize=7)
            ax_bal.spines[:].set_color("#1e2235")
            for i,(g,v) in enumerate(zip(grp_labels, grp_vals)):
                ax_bal.text(i, v+1, f"{v:.0f}", ha="center", fontsize=8, color="#e0e0e0", fontweight="bold")

            plt.tight_layout(pad=0.6)
            st.pyplot(fig_kb, use_container_width=True)
            plt.close(fig_kb)

            # ── Kişilik profil dağılımı ───────────────────
            st.markdown("**🧬 Kişilik Profil Dağılımı**")
            from collections import Counter
            pers_count = Counter(p["personality"]["name"] for p in squad)
            pers_html = "<div style='display:flex;flex-wrap:wrap;gap:8px;margin-top:4px'>"
            for name_p, cnt in pers_count.most_common():
                color_p = next((p["personality"]["color"] for p in squad if p["personality"]["name"]==name_p), "#888")
                neg = name_p in negative_profiles
                border = f"border:1px solid {color_p}" if neg else "border:1px solid #30363d"
                pers_html += (
                    f"<div style='background:#161b22;{border};border-radius:8px;"
                    f"padding:5px 10px;font-size:12px'>"
                    f"<span style='color:{color_p}'>{name_p}</span>"
                    f"<span style='color:#555;margin-left:6px'>×{cnt}</span></div>"
                )
            pers_html += "</div>"
            st.markdown(pers_html, unsafe_allow_html=True)

            # Denge skoru detayı
            st.markdown("**📊 Denge Skoru Detayı**")
            for label, val in [
                ("Yaş Dengesi", age_balance),
                ("CA Homojenliği", ca_balance),
                ("Taktiksel Denge", tactical_balance),
                ("Kişilik Uyumu", personality_score),
            ]:
                pct = val
                c = "#2ecc71" if pct>=70 else "#f1c40f" if pct>=50 else "#e74c3c"
                st.markdown(
                    f"<div style='margin:4px 0'>"
                    f"<div style='display:flex;justify-content:space-between;font-size:12px;color:#ccc'>"
                    f"<span>{label}</span><span style='color:{c};font-weight:700'>{pct:.0f}</span></div>"
                    f"<div style='background:#1e1e2e;border-radius:4px;height:7px'>"
                    f"<div style='width:{pct:.0f}%;height:100%;background:{c};border-radius:4px'></div>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

    # ══════════════════════════════════════════════════════
    # AN-TAB 4 — CA/PA VERİMLİLİK ANALİZİ
    # ══════════════════════════════════════════════════════
    with an_tab4:
        st.subheader("📈 CA/PA Verimlilik Analizi")
        st.caption("Profesyonellik, hırs ve preset — gelişim potansiyeline gerçekten ne kadar etki ediyor?")

        ef_col1, ef_col2, ef_col3 = st.columns(3)
        with ef_col1: ef_pos = st.selectbox("Mevki", ALL_POSITIONS, key="ef_pos")
        with ef_col2: ef_pre = st.selectbox("Profil", ["Average","Wonderkid","Star","Superstar"], key="ef_pre")
        with ef_col3: ef_pa  = st.slider("PA", 100, 200, 170, key="ef_pa")

        ef_start_age = st.slider("Başlangıç Yaşı", 15, 22, 17, key="ef_start")

        if st.button("📊 Verimlilik Analizi", type="primary", key="ef_btn"):
            # 5 farklı profesyonellik seviyesi
            prof_levels = [
                ("🔴 Düşük (5)",     5,  "#e74c3c"),
                ("🟠 Orta-Alt (9)",  9,  "#e67e22"),
                ("🟡 Orta (13)",     13, "#f1c40f"),
                ("🟢 Yüksek (17)",   17, "#2ecc71"),
                ("🔵 Maks (20)",     20, "#3498db"),
            ]
            start_ca = int(ef_pa * 0.40)

            fig_ef, (ax_top, ax_bot) = plt.subplots(1, 2, figsize=(12, 4))
            fig_ef.patch.set_facecolor("#0d1117")

            reach_data = []
            for label, prof, color in prof_levels:
                eff = calculate_effective_pa(ef_pa, prof, ef_pre)
                growth = simulate_growth(ef_start_age, start_ca, eff, prof, ef_pos)
                ages_g = [ef_start_age] + [y for y,_ in growth]
                vals_g = [start_ca]     + [v for _,v in growth]

                ax_top.plot(ages_g, vals_g, color=color, lw=2, label=f"{label} → {eff}")
                reach_pct = eff / ef_pa * 100
                reach_data.append((label, prof, eff, reach_pct, color))

            ax_top.axhline(ef_pa, color="#555", ls="--", lw=1, label=f"PA {ef_pa}")
            ax_top.set_facecolor("#0d1117")
            ax_top.tick_params(colors="#aaa", labelsize=7)
            ax_top.spines[:].set_color("#1e2235")
            ax_top.set_title("Profesyonelliğe Göre CA Gelişimi", fontsize=9, color="#8b949e")
            ax_top.set_xlabel("Yaş", fontsize=8, color="#8b949e")
            ax_top.set_ylabel("CA", fontsize=8, color="#8b949e")
            ax_top.legend(fontsize=6.5, labelcolor="#ccc", facecolor="#0d1117", loc="upper left")
            ax_top.grid(color="#1e2235", lw=0.5)

            # Sağ: PA erişim bar chart
            ax_bot.set_facecolor("#0d1117")
            labels_b = [r[0] for r in reach_data]
            reach_b  = [r[2] for r in reach_data]
            reach_pct_b = [r[3] for r in reach_data]
            colors_b = [r[4] for r in reach_data]
            bars_b = ax_bot.barh(labels_b, reach_b, color=colors_b, alpha=0.85, height=0.5)
            ax_bot.axvline(ef_pa, color="#555", ls="--", lw=1)
            for i, (val, pct) in enumerate(zip(reach_b, reach_pct_b)):
                ax_bot.text(val+1, i, f"{val} (%{pct:.0f})", va="center",
                            fontsize=8, color="#e0e0e0", fontweight="bold")
            ax_bot.set_title(f"Erişilen Efektif PA (toplam PA={ef_pa})", fontsize=9, color="#8b949e")
            ax_bot.set_xlabel("Efektif PA", fontsize=8, color="#8b949e")
            ax_bot.tick_params(colors="#aaa", labelsize=7)
            ax_bot.spines[:].set_color("#1e2235")
            ax_bot.set_xlim(0, ef_pa * 1.15)

            plt.tight_layout(pad=0.8)
            st.pyplot(fig_ef, use_container_width=True)
            plt.close(fig_ef)

            # ── Sayısal özet ──────────────────────────────
            st.markdown("**📊 Profesyonellik Etki Tablosu**")
            tbl_ef = ("<div style='overflow-x:auto'><table style='border-collapse:collapse;"
                      "font-size:0.85rem;width:100%'>"
                      "<tr style='color:#8b949e;border-bottom:2px solid #30363d'>"
                      "<td style='padding:5px 10px'>Profesyonellik</td>"
                      "<td style='padding:5px 10px;text-align:center'>Efektif PA</td>"
                      "<td style='padding:5px 10px;text-align:center'>PA Kullanım</td>"
                      "<td style='padding:5px 10px;text-align:center'>Kaybedilen PA</td></tr>")
            for label, prof, eff, reach_pct, color in reach_data:
                lost = ef_pa - eff
                tbl_ef += (
                    f"<tr style='border-bottom:1px solid #1a1a2e'>"
                    f"<td style='padding:4px 10px;color:{color};font-weight:700'>{label}</td>"
                    f"<td style='padding:4px 10px;text-align:center;color:#f0f0f0;font-weight:700'>{eff}</td>"
                    f"<td style='padding:4px 10px;text-align:center;color:"
                    f"{'#2ecc71' if reach_pct>=90 else '#f1c40f' if reach_pct>=75 else '#e74c3c'}'>"
                    f"%{reach_pct:.0f}</td>"
                    f"<td style='padding:4px 10px;text-align:center;color:#e74c3c'>{lost}</td></tr>"
                )
            tbl_ef += "</table></div>"
            st.markdown(tbl_ef, unsafe_allow_html=True)

            # ── Yorum ─────────────────────────────────────
            low_eff  = reach_data[0][2]
            high_eff = reach_data[-1][2]
            diff_eff = high_eff - low_eff
            st.info(
                f"**Profesyonellik etkisi:** Aynı PA {ef_pa}'lı oyuncuda "
                f"profesyonellik 5→20 arasında değiştiğinde erişilen CA farkı **{diff_eff}** puan. "
                f"Bu, bir mevki grubunu baştan sona değiştirebilecek bir fark."
            )



# =========================================================
# TAB 7 — GENÇLİK AKADEMİSİ MODU
# =========================================================
with tab7:
    st.subheader("🎓 Gençlik Akademisi")
    st.caption("Altyapı oyuncuları üret, yıllar içinde kimin yükseldiğini izle, sözleşme kararları ver.")

    ak_c1, ak_c2, ak_c3, ak_c4 = st.columns(4)
    with ak_c1:
        ak_club = st.selectbox("Kulüp Tipi", ["Elit (Top 5)", "İyi (Orta Lig)", "Küçük Kulüp"], key="ak_club")
    with ak_c2:
        ak_age_grp = st.selectbox("Yaş Grubu", ["U15 (13-15)", "U17 (15-17)", "U19 (17-19)", "U21 (19-21)"], key="ak_age_grp")
    with ak_c3:
        ak_cty = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="ak_cty")
    with ak_c4:
        ak_n = st.slider("Oyuncu Sayısı", 6, 16,
                         int(st.session_state.get("cfg_ak_count", 10)), key="ak_n")

    ak_years = int(st.session_state.get("cfg_ak_years", 5))
    st.info(f"Simülasyon: **{ak_years} yıl** (Ayarlar panelinden değiştirebilirsin)")

    # Kulüp tipine göre PA baskı faktörü
    club_pa_factor = {"Elit (Top 5)": 0.85, "İyi (Orta Lig)": 1.0, "Küçük Kulüp": 1.15}[ak_club]

    # Yaş grubu başlangıç yaşı
    age_map = {"U15 (13-15)": (13, 15), "U17 (15-17)": (15, 17),
               "U19 (17-19)": (17, 19), "U21 (19-21)": (19, 21)}
    ak_age_lo, ak_age_hi = age_map[ak_age_grp]

    if st.button("🎓 Akademiyi Üret & Simüle Et", type="primary", key="ak_btn"):
        random.seed()
        ak_players = []
        pos_pool = ["ST","KF (Sol)","KF (Sağ)","OOS","KANAT (Sol)","KANAT (Sağ)",
                    "OS","DM","DOS","KB (Sol)","KB (Sağ)","D (Sol)","D (Sağ)","KL"]

        for i in range(ak_n):
            pos   = random.choice(pos_pool)
            p_age = random.randint(ak_age_lo, ak_age_hi)
            # Akademi oyuncuları: çoğu Wonderkid ya da Average, yüksek PA ama düşük CA
            pre   = random.choices(["Wonderkid","Average","Star"], weights=[55, 35, 10])[0]
            tca, pa = get_ca_pa(pre, p_age, ak_cty)
            # Elit kulüpler daha iyi PA'lı oyuncular keşfeder
            pa = min(200, int(pa * club_pa_factor))
            tech, mental, phys, gk, hidden = generate_all_attributes(pos, p_age, pre, ak_cty, tca)
            bp   = POSITION_BASE[pos]
            aa   = {**tech,**mental,**phys,**gk} if bp=="KL" else {**tech,**mental,**phys}
            ft   = generate_foot(pos); wf = generate_weak_foot(ft, pos)
            ca   = min(calculate_ca(aa, pos, wf), pa)
            name = generate_name(ak_cty)
            prof = hidden.get("Profesyonellik", 12)
            pers = detect_personality(hidden)

            # ak_years yıl simülasyonu
            eff_pa    = calculate_effective_pa(pa, prof, pre)
            growth    = simulate_growth(p_age, ca, eff_pa, prof, pos)
            final_age = p_age + ak_years
            ca_at_end = ca
            trajectory = [ca]
            for yr, val in growth:
                if yr <= final_age:
                    ca_at_end = val
                    trajectory.append(val)

            # Keşif skoru: PA ne kadarına ulaşıldı + kişilik puanı
            pa_reach = ca_at_end / pa * 100
            pos_profiles = {"Wonderkid": 3, "Star": 2, "Average": 1}
            discovery_score = int(pa_reach * 0.6 + (pa / 200 * 100) * 0.4)
            prospect = ("🌟 Efsane Yetenek" if pa >= 185 else
                       "⭐ İyi Potansiyel" if pa >= 160 else
                       "📋 Gelişebilir" if pa >= 130 else "❌ Yetersiz")

            ak_players.append({
                "isim": name, "pos": pos, "yaş": p_age, "pre": pre,
                "ca_start": ca, "pa": pa, "eff_pa": eff_pa,
                "ca_end": ca_at_end, "pa_reach": pa_reach,
                "trajectory": trajectory, "prof": prof,
                "discovery_score": discovery_score, "prospect": prospect,
                "flag": COUNTRY_FLAG.get(ak_cty, "🏳️"),
                "personality": pers,
                "tech": tech, "mental": mental, "phys": phys,
                "hidden": hidden, "all_a": aa,
            })

        ak_players.sort(key=lambda x: -x["discovery_score"])
        st.session_state["ak_players"] = ak_players
        st.session_state["ak_age_lo"]  = ak_age_lo

    if "ak_players" in st.session_state:
        players = st.session_state["ak_players"]
        ak_years = int(st.session_state.get("cfg_ak_years", 5))

        # ── Genel İstatistik ──────────────────────────────
        total = len(players)
        signed = sum(1 for p in players if p["pa"] >= 150)
        avg_pa = sum(p["pa"] for p in players) / total
        best   = players[0]

        s1,s2,s3,s4 = st.columns(4)
        for col, label, val, color in [
            (s1, "Toplam Oyuncu", total, "#f0f0f0"),
            (s2, "Sözleşme (PA≥150)", signed, "#2ecc71"),
            (s3, "Ort. PA", f"{avg_pa:.0f}", "#f1c40f"),
            (s4, "En İyi Yetenek", best["isim"].split()[-1], "#3498db"),
        ]:
            col.markdown(
                f"<div style='background:#161b22;border-radius:8px;padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#8b949e'>{label}</div>"
                f"<div style='font-size:18px;font-weight:900;color:{color}'>{val}</div></div>",
                unsafe_allow_html=True
            )

        st.markdown("---")

        # ── Gelişim Grafiği ───────────────────────────────
        st.markdown("**📈 Tüm Oyuncuların CA Gelişimi**")
        fig_ak, ax_ak = plt.subplots(figsize=(11, 4))
        fig_ak.patch.set_facecolor("#0d1117")
        ax_ak.set_facecolor("#0d1117")
        colors_ak = ["#2ecc71","#3498db","#e67e22","#9b59b6","#e74c3c",
                     "#1abc9c","#f1c40f","#ff6b35","#00bcd4","#8e44ad",
                     "#27ae60","#2980b9","#d35400","#7d3c98","#c0392b","#16a085"]
        age_lo = st.session_state.get("ak_age_lo", 15)
        for i, p in enumerate(players):
            traj = p["trajectory"]
            ages_traj = list(range(p["yaş"], p["yaş"] + len(traj)))
            c = colors_ak[i % len(colors_ak)]
            ax_ak.plot(ages_traj, traj, color=c, lw=2,
                       label=p["isim"].split()[-1], marker="o", ms=5)
            ax_ak.text(ages_traj[-1]+0.1, traj[-1], f"{traj[-1]}",
                       fontsize=6.5, color=c, va="center")

        ax_ak.set_xlabel("Yaş", fontsize=8, color="#8b949e")
        ax_ak.set_ylabel("CA", fontsize=8, color="#8b949e")
        ax_ak.set_title(f"Akademi {ak_years} Yıl Gelişim Simülasyonu", fontsize=9, color="#8b949e")
        ax_ak.tick_params(colors="#aaa", labelsize=7)
        ax_ak.spines[:].set_color("#1e2235")
        ax_ak.legend(fontsize=5.5, labelcolor="#ccc", facecolor="#0d1117",
                     ncol=3, loc="upper left")
        ax_ak.grid(color="#1e2235", lw=0.5)
        plt.tight_layout(pad=0.5)
        st.pyplot(fig_ak, use_container_width=True)
        plt.close(fig_ak)

        # ── Oyuncu Tablosu & Kararlar ─────────────────────
        st.markdown("**📋 Keşif Raporu & Sözleşme Kararları**")
        st.caption("▶ Oyuncuya tıkla → mevcut özellikler + radar görünsün")

        def _attr_color_ak(v):
            if v >= 17: return "#3498db"
            if v >= 14: return "#2ecc71"
            if v >= 10: return "#f1c40f"
            if v >= 5:  return "#e0e0e0"
            return "#9e9e9e"

        def _attr_table_html(attrs, title, color):
            h = f"<div style='font-size:10px;font-weight:700;color:{color};margin-bottom:4px'>{title}</div>"
            h += "<table style='width:100%;border-collapse:collapse;font-size:0.78rem'>"
            for k, v in attrs.items():
                c = _attr_color_ak(v)
                h += (f"<tr><td style='padding:1px 4px;color:#aaa'>{k}</td>"
                      f"<td style='padding:1px 4px;text-align:right;font-weight:700;color:{c}'>{v}</td></tr>")
            h += "</table>"
            return h

        for p in players:
            pa_pct   = p["pa_reach"]
            pr_color = ("#2ecc71" if p["pa"] >= 185 else "#f1c40f" if p["pa"] >= 160
                        else "#e67e22" if p["pa"] >= 130 else "#e74c3c")
            rec_action = "✅ SÖZLEŞME" if p["pa"] >= 150 else "⚠️ İZLE" if p["pa"] >= 120 else "❌ BIRAK"
            rec_color  = "#2ecc71" if p["pa"] >= 150 else "#f1c40f" if p["pa"] >= 120 else "#e74c3c"

            exp_label = (
                f"{p['flag']} {p['isim']}  ·  {p['pos']}  ·  "
                f"CA {p['ca_start']}→{p['ca_end']}  ·  PA {p['pa']}  ·  {rec_action}"
            )
            with st.expander(exp_label, expanded=False):
                # Üst bilgi satırı
                st.markdown(
                    f"<div style='background:#161b22;border-left:3px solid {pr_color};"
                    f"border-radius:6px;padding:8px 14px;margin-bottom:8px;"
                    f"display:flex;align-items:center;gap:16px'>"
                    f"<div>"
                    f"<div style='font-weight:700;color:#f0f0f0;font-size:14px'>"
                    f"{p['flag']} {p['isim']}</div>"
                    f"<div style='font-size:11px;color:#8b949e'>"
                    f"{p['pos']} · {p['yaş']} yaş · {p['pre']} · "
                    f"{p['personality']['name']}</div>"
                    f"</div>"
                    f"<div style='margin-left:auto;text-align:right'>"
                    f"<div style='font-size:20px;font-weight:900;color:{pr_color}'>"
                    f"PA {p['pa']}</div>"
                    f"<div style='font-size:12px;color:#8b949e'>"
                    f"CA {p['ca_start']} → {p['ca_end']} · "
                    f"PA kullanım %{pa_pct:.0f}</div>"
                    f"<div style='font-weight:800;color:{rec_color};font-size:13px'>"
                    f"{rec_action} · {p['prospect']}</div>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

                # PA kullanım barı
                st.markdown(
                    f"<div style='background:#1e1e2e;border-radius:4px;height:8px;margin-bottom:10px'>"
                    f"<div style='width:{min(100,pa_pct):.0f}%;height:100%;"
                    f"background:{pr_color};border-radius:4px'></div></div>",
                    unsafe_allow_html=True
                )

                # Attribute kolonları + Radar
                ac1, ac2, ac3, ac4, ac5 = st.columns([1, 1, 1, 0.8, 1.6])

                with ac1:
                    st.markdown(
                        _attr_table_html(p["tech"], "⚙️ TEKNİK", "#58a6ff"),
                        unsafe_allow_html=True
                    )
                with ac2:
                    st.markdown(
                        _attr_table_html(p["mental"], "🧠 ZİHİNSEL", "#bc8cff"),
                        unsafe_allow_html=True
                    )
                with ac3:
                    st.markdown(
                        _attr_table_html(p["phys"], "💪 FİZİKSEL", "#2ecc71"),
                        unsafe_allow_html=True
                    )
                with ac4:
                    st.markdown(
                        _attr_table_html(p["hidden"], "🔒 GİZLİ", "#f1c40f"),
                        unsafe_allow_html=True
                    )
                with ac5:
                    # Radar
                    fig_ak_r = radar_chart(p["all_a"], p["pos"])
                    st.pyplot(fig_ak_r, use_container_width=True)
                    plt.close(fig_ak_r)

        # Özet
        st.markdown("---")
        sozlesme = [p for p in players if p["pa"] >= 150]
        izle     = [p for p in players if 120 <= p["pa"] < 150]
        birak    = [p for p in players if p["pa"] < 120]
        st.markdown(
            f"**Karar Özeti:** "
            f"<span style='color:#2ecc71'>✅ Sözleşme: {len(sozlesme)}</span> &nbsp;·&nbsp; "
            f"<span style='color:#f1c40f'>⚠️ İzle: {len(izle)}</span> &nbsp;·&nbsp; "
            f"<span style='color:#e74c3c'>❌ Bırak: {len(birak)}</span>",
            unsafe_allow_html=True
        )

# =========================================================
# TAB 8 — MİNİ LİG MOTORU
# =========================================================
with tab8:
    st.subheader("🏆 Mini Lig Motoru")
    st.caption("Takımlar üret, tam sezon oyna, puan tablosu & gol krallığı görüntüle.")

    # Takım sayısı sidebar'dan gelir (Mini Lig Varsayılanları)
    ml_n = int(st.session_state.get("cfg_lig_teams", 8))
    st.info(f"📍 Takım sayısı: **{ml_n}** — Sol paneldeki **Mini Lig Varsayılanları**'ndan ayarla", icon="ℹ️")

    ml_c2, ml_c3, ml_c4 = st.columns(3)
    with ml_c2:
        ml_pre = st.selectbox("Kadro Seviyesi", ["Average","Star","Superstar","Karışık"], key="ml_pre")
    with ml_c3:
        ml_cty = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="ml_cty")
    with ml_c4:
        ml_hg = st.slider("Ev Sahibi Avantajı", 0, 8, 3, key="ml_hg",
                          help="Ev sahibinin CA'sına eklenen sanal bonus")

    # Önceden tanımlı kulüp isimleri
    CLUB_NAMES = [
        "Kartalspor","Yıldız FK","Bozkurt United","Anadolu City",
        "Mavi Deniz","Kızıl Orman","Altın Hilal","Gümüş Aslan",
        "Dağ Kartalı","Yeşil Vadi","Demir Kale","Rüzgar FK",
        "Şimşek United","Fırtına City","Gece Yıldızı","Şahin FK",
        "Çelik Spor","Poyraz FK","Demirspor","Akıncılar",
        "Volkan United","Tunç FK","Kuzey Yıldızı","Ateş Spor",
    ]

    PRESET_WEIGHTS = {
        "Average"   : {"Average":1.0,"Star":0.0,"Superstar":0.0},
        "Star"      : {"Average":0.0,"Star":1.0,"Superstar":0.0},
        "Superstar" : {"Average":0.0,"Star":0.0,"Superstar":1.0},
        "Karışık"   : {"Average":0.5,"Star":0.35,"Superstar":0.15},
    }

    if st.button("⚽ Ligi Başlat", type="primary", key="ml_btn"):
        random.seed()
        pweights = PRESET_WEIGHTS[ml_pre]
        presets  = list(pweights.keys())
        pw_vals  = list(pweights.values())

        # Takımları üret
        teams = []
        for i in range(ml_n):
            tname  = CLUB_NAMES[i] if i < len(CLUB_NAMES) else f"Takım {i+1}"
            t_pre  = random.choices(presets, weights=pw_vals)[0]
            t_age  = random.randint(22, 28)
            # Her takım için 11 oyuncu CA ortalaması
            t_cas  = []
            for pos in ["KL","D (Sağ)","DOS","DOS","D (Sol)","OS","DM","OS","KF (Sağ)","ST","KF (Sol)"]:
                t_ca_p, t_pa_p = get_ca_pa(t_pre, t_age, ml_cty)
                t_cas.append(t_ca_p)
            avg_ca_team = sum(t_cas) / len(t_cas)
            teams.append({
                "isim": tname, "pre": t_pre,
                "avg_ca": avg_ca_team,
                "O":0,"G":0,"B":0,"M":0,"AG":0,"YG":0,"P":0,
            })
            
        # Puan tablosu başlangıç
        team_map = {t["isim"]: t for t in teams}

        # Goller için oyuncu havuzu (basit)
        scorers = {t["isim"]: {} for t in teams}

        def sim_match(home, away, hg_bonus):
            """CA bazlı maç simülasyonu."""
            h_ca = home["avg_ca"] + hg_bonus
            a_ca = away["avg_ca"]
            diff = (h_ca - a_ca) / 40  # normalize
            # Her takım için gol beklentisi
            h_lam = max(0.3, 1.4 + diff)
            a_lam = max(0.3, 1.4 - diff)
            # Poisson yaklaşımı
            h_goals = sum(1 for _ in range(10) if random.random() < h_lam/10)
            a_goals = sum(1 for _ in range(10) if random.random() < a_lam/10)
            return h_goals, a_goals

        # Round-robin
        match_log = []
        for i, h in enumerate(teams):
            for j, a in enumerate(teams):
                if i == j: continue
                hg, ag = sim_match(h, a, ml_hg)
                # Puan güncelle
                h["AG"] += hg; h["YG"] += ag; h["O"] += 1
                a["AG"] += ag; a["YG"] += hg; a["O"] += 1
                if hg > ag:
                    h["G"] += 1; h["P"] += 3; a["M"] += 1
                elif hg < ag:
                    a["G"] += 1; a["P"] += 3; h["M"] += 1
                else:
                    h["B"] += 1; h["P"] += 1
                    a["B"] += 1; a["P"] += 1
                # Gol atanları kaydet (basit)
                for _ in range(hg):
                    scorer = f"Oyuncu-{random.randint(1,11)}"
                    scorers[h["isim"]][scorer] = scorers[h["isim"]].get(scorer, 0) + 1
                for _ in range(ag):
                    scorer = f"Oyuncu-{random.randint(1,11)}"
                    scorers[a["isim"]][scorer] = scorers[a["isim"]].get(scorer, 0) + 1
                match_log.append((h["isim"], hg, ag, a["isim"]))

        # Gol farkı
        for t in teams: t["GD"] = t["AG"] - t["YG"]
        # Sırala
        teams.sort(key=lambda x: (-x["P"], -x["GD"], -x["AG"]))

        st.session_state["ml_teams"]  = teams
        st.session_state["ml_log"]    = match_log
        st.session_state["ml_scorers"] = scorers

    if "ml_teams" in st.session_state:
        teams     = st.session_state["ml_teams"]
        match_log = st.session_state["ml_log"]
        scorers   = st.session_state["ml_scorers"]

        ml_sub1, ml_sub2 = st.tabs(["📊 Puan Tablosu", "⚽ Gol Krallığı"])

        # ── Puan Tablosu ──────────────────────────────────
        with ml_sub1:
            st.markdown("### 📊 Puan Tablosu")
            tbl_html = ("<div style='overflow-x:auto'>"
                        "<table style='width:100%;border-collapse:collapse;font-size:0.88rem'>"
                        "<tr style='color:#8b949e;border-bottom:2px solid #30363d;background:#0d1117'>"
                        "<td style='padding:6px 10px'>#</td>"
                        "<td style='padding:6px 10px'>Takım</td>"
                        "<td style='padding:6px 10px;text-align:center'>Seviye</td>"
                        "<td style='padding:6px 10px;text-align:center'>O</td>"
                        "<td style='padding:6px 10px;text-align:center'>G</td>"
                        "<td style='padding:6px 10px;text-align:center'>B</td>"
                        "<td style='padding:6px 10px;text-align:center'>M</td>"
                        "<td style='padding:6px 10px;text-align:center'>AG</td>"
                        "<td style='padding:6px 10px;text-align:center'>YG</td>"
                        "<td style='padding:6px 10px;text-align:center'>GF</td>"
                        "<td style='padding:6px 10px;text-align:center;font-weight:800'>P</td>"
                        "<td style='padding:6px 10px;text-align:center'>CA Ort.</td>"
                        "</tr>")
            for rank, t in enumerate(teams, 1):
                if rank == 1:
                    row_bg = "background:#0a1f0a"; medal = "🥇"
                elif rank == 2:
                    row_bg = "background:#0f1a20"; medal = "🥈"
                elif rank == 3:
                    row_bg = "background:#1a140a"; medal = "🥉"
                elif rank <= len(teams) - 2:
                    row_bg = "background:#0d1117" if rank%2==0 else "background:#161b22"
                    medal  = str(rank)
                else:
                    row_bg = "background:#1a0d0d"; medal = f"⬇️{rank}"
                pre_color = {"Superstar":"#e74c3c","Star":"#e67e22",
                             "Average":"#3498db"}.get(t["pre"],"#aaa")
                gd_str = f"+{t['GD']}" if t["GD"] > 0 else str(t["GD"])
                gd_c   = "#2ecc71" if t["GD"] > 0 else "#e74c3c" if t["GD"] < 0 else "#aaa"
                tbl_html += (
                    f"<tr style='border-bottom:1px solid #1e2235;{row_bg}'>"
                    f"<td style='padding:5px 10px;color:#8b949e'>{medal}</td>"
                    f"<td style='padding:5px 10px;color:#f0f0f0;font-weight:700'>{t['isim']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:{pre_color};font-size:11px'>{t['pre']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#aaa'>{t['O']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#2ecc71;font-weight:700'>{t['G']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#f1c40f'>{t['B']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#e74c3c'>{t['M']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#aaa'>{t['AG']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#aaa'>{t['YG']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:{gd_c};font-weight:700'>{gd_str}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#2ecc71;font-weight:900;font-size:15px'>{t['P']}</td>"
                    f"<td style='padding:5px 10px;text-align:center;color:#8b949e;font-size:11px'>{t['avg_ca']:.0f}</td>"
                    f"</tr>"
                )
            tbl_html += "</table></div>"
            st.markdown(tbl_html, unsafe_allow_html=True)

            # ── CA vs Puan scatter ─────────────────────────
            st.markdown("### 📈 CA Ortalaması vs Puan")
            fig_ml, ax_ml = plt.subplots(figsize=(8, 3.5))
            fig_ml.patch.set_facecolor("#0d1117")
            ax_ml.set_facecolor("#0d1117")
            pre_colors = {"Superstar":"#e74c3c","Star":"#e67e22","Average":"#3498db"}
            for t in teams:
                c = pre_colors.get(t["pre"],"#aaa")
                ax_ml.scatter(t["avg_ca"], t["P"], color=c, s=80, zorder=3)
                ax_ml.annotate(t["isim"], (t["avg_ca"], t["P"]),
                               textcoords="offset points", xytext=(5,3),
                               fontsize=6.5, color="#ccc")
            ax_ml.set_xlabel("Kadro CA Ortalaması", fontsize=8, color="#8b949e")
            ax_ml.set_ylabel("Puan", fontsize=8, color="#8b949e")
            ax_ml.set_title("CA Kalitesi ↔ Lig Puanı Korelasyonu", fontsize=9, color="#8b949e")
            ax_ml.tick_params(colors="#aaa", labelsize=7)
            ax_ml.spines[:].set_color("#1e2235")
            ax_ml.grid(color="#1e2235", lw=0.5)
            # Trend çizgisi
            cas_x = [t["avg_ca"] for t in teams]
            pts_y = [t["P"] for t in teams]
            z = np.polyfit(cas_x, pts_y, 1)
            p_fit = np.poly1d(z)
            xs = np.linspace(min(cas_x), max(cas_x), 100)
            ax_ml.plot(xs, p_fit(xs), color="#555", lw=1.5, ls="--")
            # Korelasyon katsayısı
            corr = np.corrcoef(cas_x, pts_y)[0,1]
            ax_ml.text(0.02, 0.95, f"r = {corr:.2f}", transform=ax_ml.transAxes,
                       fontsize=8, color="#8b949e", va="top")
            plt.tight_layout(pad=0.5)
            st.pyplot(fig_ml, use_container_width=True)
            plt.close(fig_ml)

            # Şampiyon banner
            champ = teams[0]
            st.markdown(
                f"<div style='background:linear-gradient(135deg,#1a2f1a,#0d1117);"
                f"border:2px solid #f1c40f;border-radius:12px;padding:16px 24px;margin-top:12px;"
                f"text-align:center'>"
                f"<div style='font-size:28px'>🏆</div>"
                f"<div style='font-size:20px;font-weight:900;color:#f1c40f'>{champ['isim']}</div>"
                f"<div style='font-size:13px;color:#8b949e'>"
                f"{champ['P']} puan · {champ['G']} galibiyet · {champ['AG']} gol</div>"
                f"</div>",
                unsafe_allow_html=True
            )

        # ── Gol Krallığı ──────────────────────────────────
        with ml_sub2:
            st.markdown("### ⚽ Gol Krallığı")
            all_scorers = []
            for team_name, scorer_dict in scorers.items():
                for player, goals in scorer_dict.items():
                    team_pre = next(t["pre"] for t in teams if t["isim"] == team_name)
                    all_scorers.append({
                        "oyuncu": player,
                        "takım": team_name,
                        "gol": goals,
                        "pre": team_pre,
                    })
            all_scorers.sort(key=lambda x: -x["gol"])

            top_n = min(20, len(all_scorers))
            goal_html = ("<table style='width:100%;border-collapse:collapse;font-size:0.88rem'>"
                         "<tr style='color:#8b949e;border-bottom:2px solid #30363d'>"
                         "<td style='padding:5px 10px'>#</td>"
                         "<td style='padding:5px 10px'>Oyuncu</td>"
                         "<td style='padding:5px 10px'>Takım</td>"
                         "<td style='padding:5px 10px;text-align:center'>⚽ Gol</td></tr>")
            for i, s in enumerate(all_scorers[:top_n]):
                medal = "🥇" if i==0 else "🥈" if i==1 else "🥉" if i==2 else str(i+1)
                bg = "background:#0a1f0a" if i==0 else "background:#0f1a20" if i==1 else "background:#1a140a" if i==2 else ""
                goal_html += (
                    f"<tr style='border-bottom:1px solid #1e2235;{bg}'>"
                    f"<td style='padding:4px 10px;color:#8b949e'>{medal}</td>"
                    f"<td style='padding:4px 10px;color:#f0f0f0'>{s['oyuncu']}</td>"
                    f"<td style='padding:4px 10px;color:#58a6ff'>{s['takım']}</td>"
                    f"<td style='padding:4px 10px;text-align:center;"
                    f"font-weight:900;color:#2ecc71;font-size:15px'>{s['gol']}</td></tr>"
                )
            goal_html += "</table>"
            st.markdown(goal_html, unsafe_allow_html=True)

            # Gol dağılımı bar chart
            fig_g, ax_g = plt.subplots(figsize=(9, 3))
            fig_g.patch.set_facecolor("#0d1117")
            ax_g.set_facecolor("#0d1117")
            top15 = all_scorers[:15]
            xlabels = [f"{s['oyuncu'][:8]}\n({s['takım'][:6]})" for s in top15]
            ax_g.bar(range(len(top15)), [s["gol"] for s in top15],
                     color="#2ecc71", alpha=0.85, width=0.6)
            ax_g.set_xticks(range(len(top15)))
            ax_g.set_xticklabels(xlabels, fontsize=6, color="#ccc", rotation=30, ha="right")
            ax_g.set_ylabel("Gol", fontsize=8, color="#8b949e")
            ax_g.set_title("Top 15 Golcü", fontsize=9, color="#8b949e")
            ax_g.tick_params(colors="#aaa", labelsize=7)
            ax_g.spines[:].set_color("#1e2235")
            ax_g.grid(axis="y", color="#1e2235", lw=0.5)
            for i, s in enumerate(top15):
                ax_g.text(i, s["gol"]+0.1, str(s["gol"]),
                          ha="center", fontsize=7, color="#2ecc71", fontweight="bold")
            plt.tight_layout(pad=0.5)
            st.pyplot(fig_g, use_container_width=True)
            plt.close(fig_g)



# ── Maaş & Sözleşme Hesaplama ────────────────────────────────────────

def calculate_wage(ca, pa, age, personality_name="📋 Standart Profesyonel", country="Türkiye"):
    """CA/PA/yaş/kişilik bazlı haftalık maaş hesabı (€)."""
    base_weekly = (ca / 200) ** 2.1 * 500_000

    # Yaş faktörü — 26-28 peak, genç az ister, yaşlı çok ister (deneyim primi)
    if age <= 19:   age_f = 0.45
    elif age <= 22: age_f = 0.65
    elif age <= 25: age_f = 0.85
    elif age <= 28: age_f = 1.00
    elif age <= 31: age_f = 0.95
    elif age <= 34: age_f = 0.80
    else:           age_f = 0.60

    # PA bonusu — potansiyel yüksekse daha fazla ister
    pa_ratio = pa / max(ca, 1)
    pa_bonus = 1.0 + max(0, (pa_ratio - 1.0) * 0.25)

    # Kişilik faktörü
    pers_f = {
        "⭐ Lider": 1.20, "🎯 Mükemmeliyetçi": 1.18,
        "🔥 Hırslı Profesyonel": 1.15, "💎 Büyük Maç Oyuncusu": 1.15,
        "🧊 Buzdan Sinirler": 1.10, "🦁 Savaşçı": 1.08,
        "🎭 Medya Yıldızı": 1.12, "⚡ Karizmatik İsyankâr": 1.10,
        "💤 Tembel Deha": 0.85, "💔 Motivasyon Yoksunu": 0.75,
        "😤 Sorunlu Karakter": 0.80, "💣 Zehirli Unsur": 0.70,
        "📖 Sakin Profesyonel": 0.95, "🤝 Takım Oyuncusu": 0.92,
        "📋 Standart Profesyonel": 1.00,
    }.get(personality_name, 1.00)

    weekly = base_weekly * age_f * pa_bonus * pers_f
    return max(500, int(weekly))


def format_wage(weekly_eur):
    """Haftalık maaşı Türkçe formatta göster."""
    if weekly_eur >= 1_000_000:
        return f"{weekly_eur/1_000_000:.2f}M €/hafta"
    elif weekly_eur >= 1_000:
        return f"{weekly_eur/1_000:.0f}K €/hafta"
    return f"€{weekly_eur:,}/hafta"


def calculate_contract_years(age, ca, pa):
    """Yaş ve CA/PA'ya göre önerilen sözleşme süresi."""
    if age <= 19:   return 5
    elif age <= 22: return 4
    elif age <= 25: return 4 if pa >= 160 else 3
    elif age <= 28: return 3
    elif age <= 31: return 2
    elif age <= 33: return 1
    return 1


def calculate_signing_bonus(ca, pa, age):
    """İmzalama bonusu — PA/CA farkı ve yaşa göre."""
    potential_mult = 1.0 + (pa - ca) / 200 * 1.5
    age_mult = max(0.3, 1.0 - (age - 20) * 0.03)
    bonus = (ca / 200) ** 2.0 * 5_000_000 * potential_mult * age_mult
    return max(0, int(bonus))


def negotiate_transfer(player_value, budget, aggression=1.0):
    """
    Basit müzakere motoru.
    aggression: 0.8=tutumlu, 1.0=normal, 1.2=agresif
    Döndürür: (teklif, durum, mesaj)
    """
    asking = player_value * random.uniform(1.10, 1.40)  # satıcı marjı
    offer  = player_value * aggression * random.uniform(0.88, 1.05)
    offer  = min(offer, budget)

    if offer >= asking * 0.92:
        status = "✅ KABUL"
        msg = f"Teklif kabul edildi! Ödenen: {offer/1e6:.1f}M €"
        final = offer
    elif offer >= asking * 0.80:
        counter = asking * random.uniform(0.94, 0.98)
        status = "🔄 KARŞI TEKLİF"
        msg = f"Karşı teklif geldi: {counter/1e6:.1f}M €"
        final = counter
    else:
        status = "❌ RED"
        msg = "Teklif reddedildi. Çok düşük."
        final = None

    return offer, asking, final, status, msg


with tab9:
    tr_sub1, tr_sub2 = st.tabs(["🛒 Transfer Pazarı", "📋 Sözleşme Hesaplayıcı"])

    # ══════════════════════════════════════════════════════
    # TR-TAB 1 — TRANSFER PAZARI
    # ══════════════════════════════════════════════════════
    with tr_sub1:
        st.subheader("🛒 Transfer Pazarı")
        st.caption("Bütçeni belirle, mevki seç, piyasayı tara — satın al veya sat.")

        # ── Session state başlat ──────────────────────────
        if "tr_budget"  not in st.session_state: st.session_state.tr_budget  = 100_000_000
        if "tr_squad"   not in st.session_state: st.session_state.tr_squad   = []
        if "tr_market"  not in st.session_state: st.session_state.tr_market  = []
        if "tr_log"     not in st.session_state: st.session_state.tr_log     = []
        if "tr_revenue" not in st.session_state: st.session_state.tr_revenue = 0

        # ── Bütçe & Durum Paneli ──────────────────────────
        spent   = st.session_state.get("tr_spent", 0)
        revenue = st.session_state.tr_revenue
        budget  = st.session_state.tr_budget
        kalan   = budget - spent + revenue

        b1, b2, b3, b4 = st.columns(4)
        for col, label, val, color in [
            (b1, "Toplam Bütçe", f"{budget/1e6:.0f}M €", "#f0f0f0"),
            (b2, "Harcanan",     f"{spent/1e6:.1f}M €",  "#e74c3c"),
            (b3, "Satış Geliri", f"{revenue/1e6:.1f}M €","#2ecc71"),
            (b4, "Kalan",        f"{kalan/1e6:.1f}M €",
             "#2ecc71" if kalan >= 0 else "#e74c3c"),
        ]:
            col.markdown(
                f"<div style='background:#161b22;border-radius:8px;padding:10px;text-align:center'>"
                f"<div style='font-size:10px;color:#8b949e'>{label}</div>"
                f"<div style='font-size:18px;font-weight:900;color:{color}'>{val}</div></div>",
                unsafe_allow_html=True
            )

        # Bütçe bar
        used_pct = min(100, (spent - revenue) / max(budget, 1) * 100)
        bar_c = "#2ecc71" if used_pct < 70 else "#f1c40f" if used_pct < 90 else "#e74c3c"
        st.markdown(
            f"<div style='background:#1e1e2e;border-radius:4px;height:8px;margin:6px 0 12px'>"
            f"<div style='width:{used_pct:.0f}%;height:100%;background:{bar_c};border-radius:4px'></div></div>",
            unsafe_allow_html=True
        )

        st.divider()

        # ── Kontroller ────────────────────────────────────
        tr_c1, tr_c2, tr_c3, tr_c4, tr_c5 = st.columns([1.2, 1, 1, 1, 1])
        with tr_c1:
            new_budget = st.number_input("Bütçe (€)", 10_000_000, 1_000_000_000,
                                         st.session_state.tr_budget, 10_000_000,
                                         key="tr_budget_input",
                                         format="%d")
            if new_budget != st.session_state.tr_budget:
                st.session_state.tr_budget = new_budget
                st.session_state.tr_spent  = 0
                st.rerun()
        with tr_c2:
            tr_pos = st.selectbox("Hedef Mevki", ["Hepsi"] + ALL_POSITIONS, key="tr_pos")
        with tr_c3:
            tr_pre = st.selectbox("Oyuncu Kalitesi",
                                  ["Average","Wonderkid","Star","Superstar","Karışık"],
                                  key="tr_pre")
        with tr_c4:
            tr_cty = st.selectbox("Ülke", ["Karışık"] + list(COUNTRY_PROFILES.keys()),
                                  key="tr_cty")
        with tr_c5:
            tr_agg = st.select_slider("Müzakere Tarzı",
                                      options=["Tutumlu","Normal","Agresif"],
                                      value="Normal", key="tr_agg")
            agg_map = {"Tutumlu": 0.82, "Normal": 1.0, "Agresif": 1.18}

        if st.button("🔍 Piyasayı Tara (10 Oyuncu)", type="primary", key="tr_scan"):
            random.seed()
            market = []
            presets_pool = (["Average","Wonderkid","Star","Superstar"]
                            if tr_pre == "Karışık" else [tr_pre])
            countries_pool = (list(COUNTRY_PROFILES.keys())
                              if tr_cty == "Karışık" else [tr_cty])
            positions_pool = (ALL_POSITIONS if tr_pos == "Hepsi"
                              else [tr_pos])

            for _ in range(10):
                pos  = random.choice(positions_pool)
                pre  = random.choice(presets_pool)
                cty  = random.choice(countries_pool)
                age  = random.randint(17, 33)
                tca, pa = get_ca_pa(pre, age, cty)
                tech, mental, phys, gk, hidden = generate_all_attributes(
                    pos, age, pre, cty, tca)
                bp  = POSITION_BASE[pos]
                aa  = {**tech,**mental,**phys,**gk} if bp=="KL" else {**tech,**mental,**phys}
                ft  = generate_foot(pos)
                wf  = generate_weak_foot(ft, pos)
                ca  = min(calculate_ca(aa, pos, wf), pa)
                name = generate_name(cty)
                pers = detect_personality(hidden)
                value_str = calculate_transfer_value(ca, pa, age)
                # Sayısal değer — parse et
                v_num = 0
                try:
                    v_clean = value_str.replace("Milyon €","").replace("Bin €","").replace("€","").replace(",","").strip()
                    if "Milyon" in value_str:
                        v_num = float(v_clean) * 1_000_000
                    elif "Bin" in value_str:
                        v_num = float(v_clean) * 1_000
                    else:
                        v_num = float(v_clean)
                except:
                    v_num = ca * 500_000
                weekly = calculate_wage(ca, pa, age, pers["name"], cty)
                years  = calculate_contract_years(age, ca, pa)
                market.append({
                    "isim": name, "pos": pos, "pre": pre, "cty": cty,
                    "age": age, "ca": ca, "pa": pa, "wf": wf,
                    "value_str": value_str, "value": v_num,
                    "weekly": weekly, "years": years,
                    "personality": pers, "flag": COUNTRY_FLAG.get(cty,"🏳️"),
                    "tech": tech, "mental": mental, "phys": phys,
                    "gk": gk, "hidden": hidden,
                    "satildi": False,
                })
            st.session_state.tr_market = market

        # ── Piyasa Listesi ────────────────────────────────
        if st.session_state.tr_market:
            st.markdown("### 🏪 Piyasadaki Oyuncular")
            agg_val = agg_map[st.session_state.get("tr_agg","Normal")]

            for idx, p in enumerate(st.session_state.tr_market):
                if p.get("satildi"): continue
                pr_c = {"Superstar":"#e74c3c","Star":"#e67e22",
                        "Wonderkid":"#f1c40f","Average":"#3498db"}.get(p["pre"],"#aaa")
                pers_c = p["personality"]["color"]

                exp_lbl = (
                    f"{p['flag']} {p['isim']}  ·  {p['pos']}  ·  "
                    f"CA {p['ca']} / PA {p['pa']}  ·  {p['value_str']}"
                )
                with st.expander(exp_lbl, expanded=False):
                    ic1, ic2, ic3 = st.columns([1.2, 1, 1])
                    with ic1:
                        st.markdown(
                            f"<div style='background:#161b22;border-left:3px solid {pr_c};"
                            f"border-radius:8px;padding:10px 14px'>"
                            f"<div style='font-weight:800;font-size:14px;color:#f0f0f0'>"
                            f"{p['flag']} {p['isim']}</div>"
                            f"<div style='font-size:11px;color:#8b949e'>"
                            f"{p['pos']} · {p['cty']} · {p['age']} yaş</div>"
                            f"<div style='font-size:11px;color:{pers_c};margin-top:3px'>"
                            f"{p['personality']['name']}</div>"
                            f"<div style='margin-top:8px'>"
                            f"<span style='color:#2ecc71;font-weight:900;font-size:16px'>CA {p['ca']}</span>"
                            f"<span style='color:#8b949e'> / PA {p['pa']}</span></div>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                    with ic2:
                        st.markdown(
                            f"<div style='background:#161b22;border-radius:8px;padding:10px 14px'>"
                            f"<div style='font-size:10px;color:#8b949e'>Transfer Değeri</div>"
                            f"<div style='font-weight:800;color:#f1c40f;font-size:15px'>{p['value_str']}</div>"
                            f"<div style='font-size:10px;color:#8b949e;margin-top:6px'>Haftalık Maaş</div>"
                            f"<div style='font-weight:700;color:#e67e22'>{format_wage(p['weekly'])}</div>"
                            f"<div style='font-size:10px;color:#8b949e;margin-top:6px'>Sözleşme Önerisi</div>"
                            f"<div style='font-weight:700;color:#aaa'>{p['years']} yıl</div>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                    with ic3:
                        offer, asking, final, status, msg = negotiate_transfer(
                            p["value"], kalan, agg_val)
                        status_c = "#2ecc71" if "KABUL" in status else "#f1c40f" if "KARŞI" in status else "#e74c3c"
                        st.markdown(
                            f"<div style='background:#161b22;border-radius:8px;padding:10px 14px'>"
                            f"<div style='font-size:10px;color:#8b949e'>Teklifimiz</div>"
                            f"<div style='font-weight:700;color:#3498db'>{offer/1e6:.1f}M €</div>"
                            f"<div style='font-size:10px;color:#8b949e;margin-top:4px'>İstenen</div>"
                            f"<div style='font-weight:700;color:#e67e22'>{asking/1e6:.1f}M €</div>"
                            f"<div style='margin-top:8px;font-weight:800;color:{status_c}'>{status}</div>"
                            f"<div style='font-size:11px;color:#8b949e'>{msg}</div>"
                            f"</div>",
                            unsafe_allow_html=True
                        )

                    # ── Attribute Detayı ─────────────────
                    st.markdown("---")
                    def _tr_tbl(attrs, title, color):
                        h = (f"<div style='font-size:10px;font-weight:700;"
                             f"color:{color};margin-bottom:3px'>{title}</div>"
                             "<table style='width:100%;border-collapse:collapse;font-size:0.75rem'>")
                        for k, v in attrs.items():
                            c = ("#3498db" if v>=17 else "#2ecc71" if v>=14
                                 else "#f1c40f" if v>=10 else "#e0e0e0" if v>=5 else "#9e9e9e")
                            h += (f"<tr><td style='padding:1px 3px;color:#aaa'>{k}</td>"
                                  f"<td style='padding:1px 3px;text-align:right;"
                                  f"font-weight:700;color:{c}'>{v}</td></tr>")
                        return h + "</table>"

                    _is_gk_p = POSITION_BASE[p["pos"]] == "KL"
                    if _is_gk_p:
                        _col_t, _col_m, _col_p, _col_k, _col_g, _col_r = st.columns([1, 1, 1, 1, 0.85, 1.8])
                    else:
                        _col_t, _col_m, _col_p, _col_g, _col_r = st.columns([1, 1, 1, 0.85, 1.8])
                        _col_k = None
                    with _col_t:
                        st.markdown(_tr_tbl(p["tech"],   "⚙️ TEKNİK",   "#58a6ff"), unsafe_allow_html=True)
                    with _col_m:
                        st.markdown(_tr_tbl(p["mental"], "🧠 ZİHİNSEL", "#bc8cff"), unsafe_allow_html=True)
                    with _col_p:
                        st.markdown(_tr_tbl(p["phys"],   "💪 FİZİKSEL", "#2ecc71"), unsafe_allow_html=True)
                    if _is_gk_p and _col_k:
                        with _col_k:
                            _gk_attrs = {k: v for k, v in p.get("gk", {}).items()
                                         if k in GK_ATTRS_SET}
                            st.markdown(_tr_tbl(_gk_attrs, "🧤 KALECİ", "#e67e22"), unsafe_allow_html=True)
                    with _col_g:
                        _hid = p.get("hidden", {})
                        st.markdown(_tr_tbl(_hid, "🔒 GİZLİ", "#f1c40f"), unsafe_allow_html=True)
                    with _col_r:
                        _aa_tr = {**p["tech"], **p["mental"], **p["phys"]}
                        if _is_gk_p:
                            _aa_tr.update(p.get("gk", {}))
                        _fig_tr = radar_chart(_aa_tr, p["pos"])
                        st.pyplot(_fig_tr, use_container_width=True)
                        plt.close(_fig_tr)
                    st.markdown("---")

                    # Satın Al butonu
                    btn_cols = st.columns([1, 1, 2])
                    with btn_cols[0]:
                        if final and final <= kalan:
                            if st.button(f"✅ Satın Al ({final/1e6:.1f}M €)",
                                         key=f"buy_{idx}", type="primary"):
                                p["satildi"] = True
                                p["satin_alma_fiyati"] = final
                                st.session_state.tr_squad.append(p)
                                st.session_state.tr_spent = \
                                    st.session_state.get("tr_spent", 0) + final
                                st.session_state.tr_log.append(
                                    f"✅ Satın alındı: {p['isim']} — {final/1e6:.1f}M €")
                                st.rerun()
                        elif not final:
                            st.markdown(
                                "<span style='color:#e74c3c;font-size:12px'>Teklif reddedildi</span>",
                                unsafe_allow_html=True)
                        else:
                            st.markdown(
                                "<span style='color:#e74c3c;font-size:12px'>Bütçe yetersiz</span>",
                                unsafe_allow_html=True)

        st.divider()

        # ── Kadrom (satın alınanlar) & Satış ─────────────
        if st.session_state.tr_squad:
            st.markdown("### 👥 Kadrom")
            st.caption("▶ Oyuncuya tıkla → özellikler görünsün")
            for sidx, sp in enumerate(st.session_state.tr_squad):
                fiyat = sp.get("satin_alma_fiyati", sp["value"])
                _lbl = (f"{sp['flag']} {sp['isim']}  ·  {sp['pos']}  ·  "
                        f"CA {sp['ca']} / PA {sp['pa']}  ·  "
                        f"Ödenen: {fiyat/1e6:.1f}M €  ·  {format_wage(sp['weekly'])}")
                with st.expander(_lbl, expanded=False):
                    _is_gk_sq = POSITION_BASE[sp["pos"]] == "KL"
                    if _is_gk_sq:
                        _sq_t, _sq_m, _sq_p, _sq_k, _sq_g, _sq_r = st.columns([1, 1, 1, 1, 0.85, 1.8])
                    else:
                        _sq_t, _sq_m, _sq_p, _sq_g, _sq_r = st.columns([1, 1, 1, 0.85, 1.8])
                        _sq_k = None
                    with _sq_t:
                        st.markdown(_tr_tbl(sp["tech"],   "⚙️ TEKNİK",   "#58a6ff"), unsafe_allow_html=True)
                    with _sq_m:
                        st.markdown(_tr_tbl(sp["mental"], "🧠 ZİHİNSEL", "#bc8cff"), unsafe_allow_html=True)
                    with _sq_p:
                        st.markdown(_tr_tbl(sp["phys"],   "💪 FİZİKSEL", "#2ecc71"), unsafe_allow_html=True)
                    if _is_gk_sq and _sq_k:
                        with _sq_k:
                            _gk_attrs_sq = {k: v for k, v in sp.get("gk", {}).items()
                                            if k in GK_ATTRS_SET}
                            st.markdown(_tr_tbl(_gk_attrs_sq, "🧤 KALECİ", "#e67e22"), unsafe_allow_html=True)
                    with _sq_g:
                        _hid_sq = sp.get("hidden", {})
                        st.markdown(_tr_tbl(_hid_sq, "🔒 GİZLİ", "#f1c40f"), unsafe_allow_html=True)
                    with _sq_r:
                        _aa_sq = {**sp["tech"], **sp["mental"], **sp["phys"]}
                        if _is_gk_sq:
                            _aa_sq.update(sp.get("gk", {}))
                        _fig_sq = radar_chart(_aa_sq, sp["pos"])
                        st.pyplot(_fig_sq, use_container_width=True)
                        plt.close(_fig_sq)
                    st.markdown("---")
                    _sell_c1, _sell_c2 = st.columns([1, 3])
                    with _sell_c1:
                        # Satış fiyatı: PA potansiyeli + kişilik + pazar talebi
                        _pa_mult = 1.0 + (sp["pa"] - sp["ca"]) / 200 * 0.8
                        _age_mult = max(0.7, 1.15 - (sp["age"] - 22) * 0.03)
                        _base_sell = sp["value"] * _pa_mult * _age_mult
                        # Alış fiyatının altına düşme - minimum %95'i
                        _buy_floor = sp.get("satin_alma_fiyati", sp["value"]) * 0.95
                        sell_price = max(_buy_floor, _base_sell * random.uniform(1.05, 1.55))
                        profit = sell_price - sp.get("satin_alma_fiyati", sp["value"])
                        profit_str = (f"+{profit/1e6:.1f}M €" if profit >= 0
                                      else f"{profit/1e6:.1f}M €")
                        profit_c = "#2ecc71" if profit >= 0 else "#e74c3c"
                        st.markdown(
                            f"<div style='font-size:11px;color:#8b949e'>Tahmini satış</div>"
                            f"<div style='font-weight:700;color:#f1c40f'>{sell_price/1e6:.1f}M €</div>"
                            f"<div style='font-size:11px;color:{profit_c}'>"
                            f"Kâr/Zarar: {profit_str}</div>",
                            unsafe_allow_html=True
                        )
                    with _sell_c2:
                        # Manuel fiyat ayarı
                        custom_price = st.number_input(
                            "Satış fiyatını belirle (M €)",
                            min_value=0.1,
                            max_value=500.0,
                            value=round(sell_price/1e6, 1),
                            step=0.5,
                            key=f"sell_price_{sidx}",
                            format="%.1f"
                        )
                        if st.button(f"💰 Sat ({custom_price:.1f}M €)",
                                     key=f"sell_{sidx}", type="primary"):
                            final_sell = custom_price * 1e6
                            final_profit = final_sell - sp.get("satin_alma_fiyati", sp["value"])
                            fp_str = (f"+{final_profit/1e6:.1f}M €" if final_profit >= 0
                                      else f"{final_profit/1e6:.1f}M €")
                            st.session_state.tr_revenue += final_sell
                            st.session_state.tr_log.append(
                                f"💰 Satıldı: {sp['isim']} — {custom_price:.1f}M € ({fp_str})")
                            st.session_state.tr_squad.pop(sidx)
                            st.rerun()

        # ── İşlem Geçmişi ─────────────────────────────────
        if st.session_state.tr_log:
            st.markdown("### 📜 İşlem Geçmişi")
            for log_item in reversed(st.session_state.tr_log[-15:]):
                c = "#2ecc71" if "✅" in log_item else "#e67e22"
                st.markdown(
                    f"<div style='font-size:12px;color:{c};padding:2px 0'>{log_item}</div>",
                    unsafe_allow_html=True
                )
            if st.button("🗑️ Geçmişi Temizle", key="tr_clear_log"):
                st.session_state.tr_log = []
                st.rerun()

        # Sıfırla
        if st.button("🔄 Transferi Sıfırla", key="tr_reset"):
            for k in ["tr_budget","tr_squad","tr_market","tr_log",
                      "tr_revenue","tr_spent"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

    # ══════════════════════════════════════════════════════
    # TR-TAB 2 — SÖZLEŞME HESAPLAYICI
    # ══════════════════════════════════════════════════════
    with tr_sub2:
        st.subheader("📋 Sözleşme Hesaplayıcı")
        st.caption("Bir oyuncunun CA/PA/yaş/kişilik bilgilerini gir — maaş, sözleşme ve bonservis hesapla.")

        sc_c1, sc_c2, sc_c3, sc_c4 = st.columns(4)
        with sc_c1: sc_ca  = st.slider("CA", 40, 200, 140, key="sc_ca")
        with sc_c2: sc_pa  = st.slider("PA", 40, 200, 170, key="sc_pa")
        with sc_c3: sc_age = st.slider("Yaş", 15, 40, 24, key="sc_age")
        with sc_c4: sc_cty = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="sc_cty")

        sc_pers = st.selectbox("Kişilik",
            [p["name"] for p in PERSONALITY_PROFILES if not p.get("_fallback")],
            key="sc_pers")
        sc_pos = st.selectbox("Mevki", ALL_POSITIONS, key="sc_pos")

        # Hesapla
        weekly     = calculate_wage(sc_ca, sc_pa, sc_age, sc_pers, sc_cty)
        years      = calculate_contract_years(sc_age, sc_ca, sc_pa)
        signing    = calculate_signing_bonus(sc_ca, sc_pa, sc_age)
        value_str  = calculate_transfer_value(sc_ca, sc_pa, sc_age)
        bonservis  = weekly * 52 * years * random.uniform(1.5, 3.0)

        st.markdown("---")
        st.markdown("### 💰 Hesaplama Sonucu")

        r1, r2, r3, r4, r5 = st.columns(5)
        metrics = [
            (r1, "Haftalık Maaş",     format_wage(weekly),               "#f1c40f"),
            (r2, "Yıllık Maaş",       f"{weekly*52/1e6:.2f}M €/yıl",    "#e67e22"),
            (r3, "Sözleşme Süresi",   f"{years} yıl",                    "#3498db"),
            (r4, "İmzalama Bonusu",   f"{signing/1e6:.1f}M €",           "#9b59b6"),
            (r5, "Transfer Değeri",   value_str,                          "#2ecc71"),
        ]
        for col, label, val, color in metrics:
            col.markdown(
                f"<div style='background:#161b22;border-radius:8px;padding:12px;text-align:center'>"
                f"<div style='font-size:10px;color:#8b949e;margin-bottom:4px'>{label}</div>"
                f"<div style='font-size:15px;font-weight:900;color:{color}'>{val}</div></div>",
                unsafe_allow_html=True
            )

        # Kişilik etkisi açıklaması
        pers_effect = {
            "⭐ Lider": ("+20%", "Soyunma odası değeri için prim talep eder"),
            "🎯 Mükemmeliyetçi": ("+18%", "Kendine yüksek değer biçer"),
            "🔥 Hırslı Profesyonel": ("+15%", "Büyük kulüp için yüksek talep"),
            "💎 Büyük Maç Oyuncusu": ("+15%", "Final maçlarına katkı primi"),
            "🧊 Buzdan Sinirler": ("+10%", "Kritik anlardaki değeri yansır"),
            "🎭 Medya Yıldızı": ("+12%", "Sponsorluk değeri maaşa yansır"),
            "💤 Tembel Deha": ("-15%", "Tutarsız performans nedeniyle indirim"),
            "💔 Motivasyon Yoksunu": ("-25%", "Motivasyon sorunu yüksek risk"),
            "😤 Sorunlu Karakter": ("-20%", "Soyunma odası riski fiyata yansır"),
            "💣 Zehirli Unsur": ("-30%", "Ciddi karakter riski"),
        }.get(sc_pers, ("±0%", "Standart maaş profili"))

        eff_color = "#2ecc71" if "+" in pers_effect[0] else "#e74c3c" if "-" in pers_effect[0] else "#aaa"
        st.markdown(
            f"<div style='background:#161b22;border-left:3px solid {eff_color};"
            f"border-radius:8px;padding:10px 16px;margin-top:12px'>"
            f"<div style='font-size:12px;color:{eff_color};font-weight:700'>"
            f"🧬 Kişilik Etkisi: {sc_pers} → {pers_effect[0]}</div>"
            f"<div style='font-size:11px;color:#8b949e;margin-top:3px'>{pers_effect[1]}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

        # Karşılaştırmalı tablo — aynı mevkide farklı CA'lar
        st.markdown("### 📊 CA Bazlı Maaş Skalası")
        st.caption(f"Aynı yaş ({sc_age}) ve kişilik ({sc_pers}) ile farklı CA seviyeleri")

        scale_data = [(ca_v, calculate_wage(ca_v, max(ca_v, sc_pa), sc_age, sc_pers))
                      for ca_v in [60, 80, 100, 120, 140, 160, 180, 200]]
        fig_sc, ax_sc = plt.subplots(figsize=(8, 2.8))
        fig_sc.patch.set_facecolor("#0d1117")
        ax_sc.set_facecolor("#0d1117")
        ca_vals  = [d[0] for d in scale_data]
        wage_vals= [d[1]/1000 for d in scale_data]  # K €
        bars_sc  = ax_sc.bar(ca_vals, wage_vals,
                              color=["#2ecc71" if c==sc_ca else "#21262d" for c in ca_vals],
                              width=12, alpha=0.9)
        ax_sc.axvline(sc_ca, color="#2ecc71", lw=1.5, ls="--", alpha=0.6)
        for bar, w in zip(bars_sc, wage_vals):
            ax_sc.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                       f"{w:.0f}K", ha="center", fontsize=7,
                       color="#e0e0e0", fontweight="bold")
        ax_sc.set_xlabel("CA", fontsize=8, color="#8b949e")
        ax_sc.set_ylabel("Haftalık Maaş (K €)", fontsize=8, color="#8b949e")
        ax_sc.tick_params(colors="#aaa", labelsize=7)
        ax_sc.spines[:].set_color("#1e2235")
        ax_sc.grid(axis="y", color="#1e2235", lw=0.5)
        plt.tight_layout(pad=0.4)
        st.pyplot(fig_sc, use_container_width=True)
        plt.close(fig_sc)


# =========================================================
# TAB 10 — OYUNCU HAVUZU
# =========================================================
with tab10:
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

    st.subheader("📦 Oyuncu Havuzu")
    st.caption("Ürettiğin tüm oyuncular burada birikir. Filtrele, karşılaştır, Excel'e aktar.")

    pool = st.session_state.get("player_pool", [])

    if not pool:
        st.info("Henüz oyuncu yok. **Oyuncu Üret** sekmesinden oyuncu üretince otomatik buraya eklenir.")
    else:
        # ── Üst panel ────────────────────────────────────
        ph_col1, ph_col2, ph_col3, ph_col4 = st.columns(4)
        ph_col1.metric("Toplam Oyuncu", len(pool))
        ph_col2.metric("Favori", sum(1 for p in pool if p.get("favori")))
        ph_col3.metric("Ort. CA", f"{sum(p['ca'] for p in pool)//len(pool)}")
        ph_col4.metric("Ort. PA", f"{sum(p['pa'] for p in pool)//len(pool)}")

        # ── Filtreler ─────────────────────────────────────
        with st.expander("🔍 Filtrele & Sırala", expanded=False):
            f1, f2, f3, f4, f5 = st.columns(5)
            with f1:
                f_mevki = st.multiselect("Mevki", ALL_POSITIONS, key="ph_mevki")
            with f2:
                f_pre = st.multiselect("Profil",
                    ["Average","Wonderkid","Star","Superstar"], key="ph_pre")
            with f3:
                f_ca_min, f_ca_max = st.slider("CA Aralığı", 1, 200,
                    (1, 200), key="ph_ca")
            with f4:
                f_age_min, f_age_max = st.slider("Yaş Aralığı", 15, 45,
                    (15, 45), key="ph_age")
            with f5:
                f_sort = st.selectbox("Sırala",
                    ["CA ↓","CA ↑","PA ↓","Yaş ↓","Yaş ↑","Son Eklenen"],
                    key="ph_sort")
            f_favori = st.checkbox("Sadece favoriler", key="ph_fav")

        # Filtre uygula
        filtered = pool[:]
        if f_mevki:
            filtered = [p for p in filtered if p["position"] in f_mevki]
        if f_pre:
            filtered = [p for p in filtered if p["preset"] in f_pre]
        filtered = [p for p in filtered
                    if f_ca_min <= p["ca"] <= f_ca_max
                    and f_age_min <= p["age"] <= f_age_max]
        if f_favori:
            filtered = [p for p in filtered if p.get("favori")]

        sort_map = {
            "CA ↓": lambda x: -x["ca"],
            "CA ↑": lambda x:  x["ca"],
            "PA ↓": lambda x: -x["pa"],
            "Yaş ↓": lambda x: -x["age"],
            "Yaş ↑": lambda x:  x["age"],
            "Son Eklenen": lambda x: -x.get("_idx", 0),
        }
        filtered.sort(key=sort_map.get(f_sort, lambda x: -x["ca"]))

        st.caption(f"**{len(filtered)}** oyuncu gösteriliyor (toplam {len(pool)})")

        # ── Excel Export ─────────────────────────────────
        def build_excel(players):
            wb = Workbook()

            # ── Renk paleti ──────────────────────────────
            C_HEADER_BG  = "1E2D40"   # koyu lacivert
            C_HEADER_FG  = "FFFFFF"
            C_ALT_ROW    = "F0F4F8"   # açık gri-mavi
            C_GREEN      = "2ECC71"
            C_ORANGE     = "E67E22"
            C_YELLOW     = "F1C40F"
            C_BLUE       = "3498DB"
            C_RED        = "E74C3C"
            C_DARK       = "2C3E50"
            C_BORDER     = "BDC3C7"

            thin = Side(style="thin", color=C_BORDER)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def hdr_cell(ws, row, col, val, bg=C_HEADER_BG, fg=C_HEADER_FG,
                         bold=True, sz=10, align="center"):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=True)
                c.border = border
                return c

            def val_cell(ws, row, col, val, bold=False, color=None,
                         align="center", bg=None):
                c = ws.cell(row=row, column=col, value=val)
                fc = color or C_DARK
                c.font = Font(bold=bold, color=fc, size=9, name="Arial")
                if bg:
                    c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal=align, vertical="center")
                c.border = border
                return c

            def attr_color(v):
                if v >= 17: return C_BLUE
                if v >= 14: return C_GREEN
                if v >= 10: return "7F6000"   # sarı-kahve
                if v >= 5:  return C_DARK
                return "808080"

            # ── SAYFA 1: GENEL BİLGİ ─────────────────────
            ws1 = wb.active
            ws1.title = "Genel Bilgi"
            ws1.freeze_panes = "A2"
            ws1.sheet_view.showGridLines = False

            headers1 = ["#","İsim","Ülke","Mevki","Yaş","Preset","CA","PA",
                        "CA/PA %","Değer","Haftalık Maaş","Sözleşme (yıl)",
                        "Baskın Ayak","Zayıf Ayak","Boy","Kilo",
                        "Kişilik","Scout Notu","Favori"]
            col_widths1 = [4,22,14,12,6,12,7,7,9,16,15,13,12,10,7,7,22,10,7]

            for ci, (h, w) in enumerate(zip(headers1, col_widths1), 1):
                hdr_cell(ws1, 1, ci, h)
                ws1.column_dimensions[get_column_letter(ci)].width = w
            ws1.row_dimensions[1].height = 30

            for ri, p in enumerate(players, 2):
                ca_pa_pct = round(p["ca"] / max(p["pa"],1) * 100, 1)
                weekly    = calculate_wage(p["ca"], p["pa"], p["age"],
                                           p["personality"]["name"],
                                           p.get("country","Türkiye"))
                yrs       = calculate_contract_years(p["age"], p["ca"], p["pa"])
                grade     = scout_grade(p["pa"])
                is_alt    = (ri % 2 == 0)
                row_bg    = C_ALT_ROW if is_alt else "FFFFFF"
                fav       = "⭐" if p.get("favori") else ""

                grade_c = {"A":C_GREEN,"B":"27AE60","C":"7F6000",
                           "D":C_ORANGE,"E":C_RED}.get(grade, C_DARK)
                pre_c   = {"Superstar":C_RED,"Star":C_ORANGE,
                           "Wonderkid":"7F6000","Average":C_BLUE}.get(p["preset"], C_DARK)

                row_vals = [
                    (ri-1,       False, None,    "center", row_bg),
                    (p["name"],  True,  C_DARK,  "left",   row_bg),
                    (p.get("country","—"), False, None, "center", row_bg),
                    (p["position"], False, C_BLUE, "center", row_bg),
                    (p["age"],   False, None,    "center", row_bg),
                    (p["preset"],False, pre_c,   "center", row_bg),
                    (p["ca"],    True,  C_GREEN, "center", row_bg),
                    (p["pa"],    True,  C_ORANGE,"center", row_bg),
                    (f"{ca_pa_pct}%", False,
                     C_GREEN if ca_pa_pct>=80 else C_ORANGE if ca_pa_pct>=60 else C_RED,
                     "center", row_bg),
                    (calculate_transfer_value(p["ca"],p["pa"],p["age"]),
                     False, C_GREEN, "left", row_bg),
                    (format_wage(weekly),  False, None, "center", row_bg),
                    (yrs,        False, None,    "center", row_bg),
                    (p.get("foot","—"),  False, None, "center", row_bg),
                    (p.get("weak_foot",4), False,
                     C_GREEN if p.get("weak_foot",4)>=13 else
                     C_ORANGE if p.get("weak_foot",4)>=9 else C_RED,
                     "center", row_bg),
                    (p.get("height","—"), False, None, "center", row_bg),
                    (p.get("weight","—"), False, None, "center", row_bg),
                    (p["personality"]["name"], False, None, "left", row_bg),
                    (grade,      True,  grade_c, "center", row_bg),
                    (fav,        False, None,    "center", row_bg),
                ]
                for ci, (val, bold, col, align, bg) in enumerate(row_vals, 1):
                    val_cell(ws1, ri, ci, val, bold=bold,
                             color=col, align=align, bg=bg)
                ws1.row_dimensions[ri].height = 18

            # Otofilt
            ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}1"

            # ── SAYFA 2: TEKNİK ──────────────────────────
            ws2 = wb.create_sheet("Teknik")
            ws2.freeze_panes = "C2"
            ws2.sheet_view.showGridLines = False

            tech_cols = list(players[0]["tech"].keys()) if players else []
            hdr2 = ["#", "İsim"] + tech_cols
            ws2.column_dimensions["A"].width = 4
            ws2.column_dimensions["B"].width = 22
            for ci, h in enumerate(hdr2, 1):
                hdr_cell(ws2, 1, ci, h)
                if ci > 2:
                    ws2.column_dimensions[get_column_letter(ci)].width = 9
            ws2.row_dimensions[1].height = 30

            for ri, p in enumerate(players, 2):
                is_alt = (ri % 2 == 0)
                rb = C_ALT_ROW if is_alt else "FFFFFF"
                val_cell(ws2, ri, 1, ri-1, bg=rb)
                val_cell(ws2, ri, 2, p["name"], bold=True, align="left", bg=rb)
                for ci, attr in enumerate(tech_cols, 3):
                    v = p["tech"].get(attr, 1)
                    val_cell(ws2, ri, ci, v, bold=(v>=14),
                             color=attr_color(v), bg=rb)
                ws2.row_dimensions[ri].height = 16

            ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdr2))}1"

            # ── SAYFA 3: ZİHİNSEL ────────────────────────
            ws3 = wb.create_sheet("Zihinsel")
            ws3.freeze_panes = "C2"
            ws3.sheet_view.showGridLines = False

            ment_cols = list(players[0]["mental"].keys()) if players else []
            hdr3 = ["#", "İsim"] + ment_cols
            ws3.column_dimensions["A"].width = 4
            ws3.column_dimensions["B"].width = 22
            for ci, h in enumerate(hdr3, 1):
                hdr_cell(ws3, 1, ci, h)
                if ci > 2:
                    ws3.column_dimensions[get_column_letter(ci)].width = 9
            ws3.row_dimensions[1].height = 30

            for ri, p in enumerate(players, 2):
                is_alt = (ri % 2 == 0)
                rb = C_ALT_ROW if is_alt else "FFFFFF"
                val_cell(ws3, ri, 1, ri-1, bg=rb)
                val_cell(ws3, ri, 2, p["name"], bold=True, align="left", bg=rb)
                for ci, attr in enumerate(ment_cols, 3):
                    v = p["mental"].get(attr, 1)
                    val_cell(ws3, ri, ci, v, bold=(v>=14),
                             color=attr_color(v), bg=rb)
                ws3.row_dimensions[ri].height = 16

            ws3.auto_filter.ref = f"A1:{get_column_letter(len(hdr3))}1"

            # ── SAYFA 4: FİZİKSEL ────────────────────────
            ws4 = wb.create_sheet("Fiziksel")
            ws4.freeze_panes = "C2"
            ws4.sheet_view.showGridLines = False

            phys_cols = list(players[0]["phys"].keys()) if players else []
            hdr4 = ["#", "İsim"] + phys_cols
            ws4.column_dimensions["A"].width = 4
            ws4.column_dimensions["B"].width = 22
            for ci, h in enumerate(hdr4, 1):
                hdr_cell(ws4, 1, ci, h)
                if ci > 2:
                    ws4.column_dimensions[get_column_letter(ci)].width = 9
            ws4.row_dimensions[1].height = 30

            for ri, p in enumerate(players, 2):
                is_alt = (ri % 2 == 0)
                rb = C_ALT_ROW if is_alt else "FFFFFF"
                val_cell(ws4, ri, 1, ri-1, bg=rb)
                val_cell(ws4, ri, 2, p["name"], bold=True, align="left", bg=rb)
                for ci, attr in enumerate(phys_cols, 3):
                    v = p["phys"].get(attr, 1)
                    val_cell(ws4, ri, ci, v, bold=(v>=14),
                             color=attr_color(v), bg=rb)
                ws4.row_dimensions[ri].height = 16

            ws4.auto_filter.ref = f"A1:{get_column_letter(len(hdr4))}1"

            # ── SAYFA 5: GİZLİ ───────────────────────────
            ws5 = wb.create_sheet("Gizli")
            ws5.freeze_panes = "C2"
            ws5.sheet_view.showGridLines = False

            hid_cols = list(players[0]["hidden"].keys()) if players else []
            hdr5 = ["#", "İsim", "Kişilik"] + hid_cols
            ws5.column_dimensions["A"].width = 4
            ws5.column_dimensions["B"].width = 22
            ws5.column_dimensions["C"].width = 24
            for ci, h in enumerate(hdr5, 1):
                hdr_cell(ws5, 1, ci, h)
                if ci > 3:
                    ws5.column_dimensions[get_column_letter(ci)].width = 9
            ws5.row_dimensions[1].height = 30

            for ri, p in enumerate(players, 2):
                is_alt = (ri % 2 == 0)
                rb = C_ALT_ROW if is_alt else "FFFFFF"
                val_cell(ws5, ri, 1, ri-1, bg=rb)
                val_cell(ws5, ri, 2, p["name"], bold=True, align="left", bg=rb)
                val_cell(ws5, ri, 3, p["personality"]["name"],
                         align="left", bg=rb)
                for ci, attr in enumerate(hid_cols, 4):
                    v = p["hidden"].get(attr, 1)
                    val_cell(ws5, ri, ci, v, bold=(v>=14),
                             color=attr_color(v), bg=rb)
                ws5.row_dimensions[ri].height = 16

            ws5.auto_filter.ref = f"A1:{get_column_letter(len(hdr5))}1"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        # ── Export butonları ──────────────────────────────
        ex1, ex2, ex3 = st.columns([1, 1, 3])
        with ex1:
            if HAS_OPENPYXL and filtered:
                buf = build_excel(filtered)
                st.download_button(
                    "📥 Excel İndir (Filtrelenen)",
                    data=buf,
                    file_name="fm_oyuncular.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
            elif not HAS_OPENPYXL:
                st.warning("openpyxl kurulu değil: `pip install openpyxl`")
        with ex2:
            if HAS_OPENPYXL and pool:
                buf_all = build_excel(pool)
                st.download_button(
                    "📥 Excel İndir (Tümü)",
                    data=buf_all,
                    file_name="fm_oyuncular_tum.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        st.divider()

        # ── Oyuncu listesi ────────────────────────────────
        st.markdown(f"### 👥 Oyuncu Listesi ({len(filtered)})")
        for idx, p in enumerate(filtered):
            grade = scout_grade(p["pa"])
            gc = {"A":"#2ecc71","B":"#27ae60","C":"#f1c40f",
                  "D":"#e67e22","E":"#e74c3c"}.get(grade,"#aaa")
            pr_c = {"Superstar":"#e74c3c","Star":"#e67e22",
                    "Wonderkid":"#f1c40f","Average":"#3498db"}.get(p["preset"],"#aaa")
            fav_icon = "⭐" if p.get("favori") else "☆"
            ca_pa_pct = int(p["ca"]/max(p["pa"],1)*100)
            bar_c = "#2ecc71" if ca_pa_pct>=75 else "#f1c40f" if ca_pa_pct>=50 else "#e74c3c"

            exp_lbl = (f"{fav_icon}  {p.get('flag','🏳️')} {p['name']}  ·  "
                       f"{p['position']}  ·  CA {p['ca']} / PA {p['pa']}  ·  "
                       f"{p.get('age',0)} yaş  ·  {p['preset']}")
            with st.expander(exp_lbl, expanded=False):
                # Üst kart
                rc1, rc2, rc3 = st.columns([2, 1.2, 1])
                with rc1:
                    st.markdown(
                        f"<div style='background:#161b22;border-left:3px solid {pr_c};"
                        f"border-radius:8px;padding:10px 14px'>"
                        f"<div style='font-size:15px;font-weight:800;color:#f0f0f0'>"
                        f"{p.get('flag','🏳️')} {p['name']}</div>"
                        f"<div style='font-size:11px;color:#8b949e'>"
                        f"{p['position']} · {p.get('country','?')} · {p['age']} yaş · {p['preset']}</div>"
                        f"<div style='font-size:11px;color:{p['personality']['color']};margin-top:3px'>"
                        f"{p['personality']['name']}</div>"
                        f"<div style='margin-top:8px'>"
                        f"<span style='font-size:18px;font-weight:900;color:#2ecc71'>CA {p['ca']}</span>"
                        f"<span style='color:#8b949e'> / PA {p['pa']}</span>"
                        f"<span style='color:{gc};font-weight:700;margin-left:8px'>{grade}</span></div>"
                        f"<div style='background:#1e1e2e;border-radius:4px;height:7px;margin-top:6px'>"
                        f"<div style='width:{ca_pa_pct}%;height:100%;background:{bar_c};"
                        f"border-radius:4px'></div></div></div>",
                        unsafe_allow_html=True
                    )
                with rc2:
                    st.markdown(
                        f"<div style='background:#161b22;border-radius:8px;padding:10px 14px'>"
                        f"<div style='font-size:10px;color:#8b949e'>Değer</div>"
                        f"<div style='font-weight:700;color:#f1c40f'>"
                        f"{calculate_transfer_value(p['ca'],p['pa'],p['age'])}</div>"
                        f"<div style='font-size:10px;color:#8b949e;margin-top:6px'>Haftalık Maaş</div>"
                        f"<div style='font-weight:700;color:#e67e22'>"
                        f"{format_wage(calculate_wage(p['ca'],p['pa'],p['age'],p['personality']['name']))}</div>"
                        f"<div style='font-size:10px;color:#8b949e;margin-top:6px'>Sözleşme</div>"
                        f"<div style='color:#aaa'>{calculate_contract_years(p['age'],p['ca'],p['pa'])} yıl</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                with rc3:
                    # Favori & Sil butonları
                    fav_lbl = "⭐ Favoriden Çıkar" if p.get("favori") else "☆ Favoriye Ekle"
                    pool_idx = next((i for i,x in enumerate(pool) if x.get("_id")==p.get("_id")), None)
                    if pool_idx is not None:
                        if st.button(fav_lbl, key=f"ph_fav_{idx}"):
                            st.session_state.player_pool[pool_idx]["favori"] =                                 not pool[pool_idx].get("favori", False)
                            st.rerun()
                        if st.button("🗑️ Sil", key=f"ph_del_{idx}"):
                            st.session_state.player_pool.pop(pool_idx)
                            st.rerun()

                # Attribute tabloları
                st.markdown("---")
                pa1, pa2, pa3, pa4 = st.columns([1,1,1,0.9])
                def _ph_tbl(attrs, title, color):
                    h = (f"<div style='font-size:10px;font-weight:700;color:{color};"
                         f"margin-bottom:3px'>{title}</div>"
                         f"<table style='width:100%;border-collapse:collapse;font-size:0.75rem'>")
                    for k, v in attrs.items():
                        c = ("#3498db" if v>=17 else "#2ecc71" if v>=14
                             else "#f1c40f" if v>=10 else "#e0e0e0" if v>=5 else "#9e9e9e")
                        h += (f"<tr><td style='padding:1px 3px;color:#aaa'>{k}</td>"
                              f"<td style='padding:1px 3px;text-align:right;"
                              f"font-weight:700;color:{c}'>{v}</td></tr>")
                    return h + "</table>"
                with pa1:
                    st.markdown(_ph_tbl(p["tech"],   "⚙️ TEKNİK",   "#58a6ff"), unsafe_allow_html=True)
                with pa2:
                    st.markdown(_ph_tbl(p["mental"], "🧠 ZİHİNSEL", "#bc8cff"), unsafe_allow_html=True)
                with pa3:
                    st.markdown(_ph_tbl(p["phys"],   "💪 FİZİKSEL", "#2ecc71"), unsafe_allow_html=True)
                with pa4:
                    st.markdown(_ph_tbl(p["hidden"], "🔒 GİZLİ",    "#f1c40f"), unsafe_allow_html=True)

        st.divider()
        hav_c1, hav_c2 = st.columns([1, 4])
        with hav_c1:
            if st.button("🗑️ Havuzu Temizle", key="ph_clear"):
                st.session_state.player_pool = []
                st.rerun()

st.caption("© Football Manager Oyuncu Oluşturma | Streamlit + Python | 2026 Enes Özkan")

# =========================================================
# TAB 9 — TRANSFER & SÖZLEŞME
# =========================================================

# =========================================================
# TAB 11 — FM OYUNCU ANALİZİ
# =========================================================
with tab11:
    st.subheader("🔍 FM Oyuncu Analizi")
    st.caption(
        "FM'deki oyuncunun değerlerini gir → CA doğrulama, radar, "
        "tüm mevki fit analizi, kişilik, kariyer projeksiyonu."
    )
    st.info(
        "**Nasıl kullanılır:** FM'de oyuncunun profiline gir, "
        "attribute değerlerini buraya gir, **Analiz Et**'e bas.",
        icon="ℹ️"
    )

    # ── Temel Bilgiler ─────────────────────────────────────
    st.markdown("#### 👤 Temel Bilgiler")
    bi1, bi2, bi3, bi4, bi5 = st.columns(5)
    with bi1: fm_name  = st.text_input("Oyuncu Adı", "Oyuncu", key="fm_name")
    with bi2: fm_pos   = st.selectbox("Mevki", ALL_POSITIONS, key="fm_pos")
    with bi3: fm_age   = st.number_input("Yaş", 15, 45, 24, key="fm_age")
    with bi4: fm_cty   = st.selectbox("Ülke", list(COUNTRY_PROFILES.keys()), key="fm_cty")
    with bi5: fm_ca_fm = st.number_input("FM'deki CA", 1, 200, 130, key="fm_ca_fm",
                                          help="FM'in gösterdiği CA değeri")

    bi6, bi7, bi8, bi9, bi10 = st.columns(5)
    with bi6:  fm_pa     = st.number_input("FM'deki PA", 1, 200, 160, key="fm_pa")
    with bi7:  fm_foot   = st.selectbox("Baskın Ayak",
                               ["Sağ","Sol","Her İkisi"], key="fm_foot")
    with bi8:  fm_wf     = st.number_input("Zayıf Ayak (1-20)", 1, 20, 8, key="fm_wf")
    with bi9:  fm_height = st.number_input("Boy (cm)", 155, 210, 180, key="fm_height")
    with bi10: fm_weight = st.number_input("Kilo (kg)", 55, 110, 75, key="fm_weight")

    st.divider()

    # ── Attribute Girişi ───────────────────────────────────
    st.markdown("#### 📋 Attribute Değerleri")
    st.caption("FM'de gördüğün değerleri gir — renk otomatik güncellenir")

    fm_is_gk = POSITION_BASE[fm_pos] == "KL"

    # Her satırda 3 attribute (tab sırası için)
    def _fm_attr_section(attrs_list, prefix, title, color):
        st.markdown(
            f"<div style='font-size:11px;font-weight:700;color:{color};"
            f"margin:8px 0 4px'>{title}</div>",
            unsafe_allow_html=True
        )
        result = {}
        for i in range(0, len(attrs_list), 3):
            row_attrs = attrs_list[i:i+3]
            cols = st.columns(3)
            for j, attr in enumerate(row_attrs):
                with cols[j]:
                    val = st.number_input(
                        attr, 1, 20, 10, step=1,
                        key=f"fm_{prefix}_{attr}"
                    )
                    result[attr] = val
        return result

    col_left, col_right = st.columns([1.6, 1])
    with col_left:
        t1, t2, t3 = st.tabs(["⚙️ Teknik", "🧠 Zihinsel", "💪 Fiziksel"])
        with t1:
            fm_tech = _fm_attr_section(TECHNICAL, "tech", "⚙️ Teknik", "#58a6ff")
        with t2:
            fm_mental = _fm_attr_section(MENTAL, "ment", "🧠 Zihinsel", "#bc8cff")
        with t3:
            fm_phys = _fm_attr_section(PHYSICAL, "phys", "💪 Fiziksel", "#2ecc71")

    with col_right:
        if fm_is_gk:
            st.markdown(
                "<div style='font-size:11px;font-weight:700;color:#e67e22;"
                "margin:8px 0 4px'>🧤 Kaleci</div>",
                unsafe_allow_html=True
            )
            fm_gk = {}
            for attr in GOALKEEPER:
                fm_gk[attr] = st.number_input(
                    attr, 1, 20,
                    GK_ATTR_MEANS.get(attr, 10),
                    key=f"fm_gk_{attr}"
                )
        else:
            fm_gk = {attr: 1 for attr in GOALKEEPER}

        st.markdown("---")
        st.markdown(
            "<div style='font-size:11px;font-weight:700;color:#f1c40f;"
            "margin:8px 0 4px'>🔒 Gizli Özellikler</div>",
            unsafe_allow_html=True
        )
        fm_hidden = {}
        for attr in HIDDEN:
            fm_hidden[attr] = st.number_input(
                attr, 1, 20, 10, key=f"fm_hid_{attr}"
            )

    st.divider()

    # ── Analiz Butonu ──────────────────────────────────────
    if st.button("🔍 Analiz Et", type="primary", key="fm_analyze"):

        fm_all_a = ({**fm_tech, **fm_mental, **fm_phys, **fm_gk}
                    if fm_is_gk else {**fm_tech, **fm_mental, **fm_phys})

        calc_ca   = calculate_ca(fm_all_a, fm_pos, fm_wf)
        ca_delta  = calc_ca - fm_ca_fm
        pers      = detect_personality(fm_hidden)
        value_str = calculate_transfer_value(fm_ca_fm, fm_pa, fm_age)
        weekly    = calculate_wage(fm_ca_fm, fm_pa, fm_age,
                                   pers["name"], fm_cty)
        yrs       = calculate_contract_years(fm_age, fm_ca_fm, fm_pa)
        sign_bon  = calculate_signing_bonus(fm_ca_fm, fm_pa, fm_age)
        grade     = scout_grade(fm_pa)
        flag      = COUNTRY_FLAG.get(fm_cty, "🏳️")

        st.session_state["fm_result"] = {
            "name": fm_name, "pos": fm_pos, "age": fm_age,
            "cty": fm_cty, "ca_fm": fm_ca_fm, "pa": fm_pa,
            "calc_ca": calc_ca, "ca_delta": ca_delta,
            "foot": fm_foot, "wf": fm_wf,
            "height": fm_height, "weight": fm_weight,
            "pers": pers, "value_str": value_str,
            "weekly": weekly, "yrs": yrs, "sign_bon": sign_bon,
            "grade": grade, "flag": flag,
            "tech": fm_tech, "mental": fm_mental,
            "phys": fm_phys, "gk": fm_gk,
            "hidden": fm_hidden, "all_a": fm_all_a,
        }

    # ── Sonuç ─────────────────────────────────────────────
    if "fm_result" in st.session_state:
        R = st.session_state["fm_result"]
        gc = {"A":"#2ecc71","B":"#27ae60","C":"#f1c40f",
              "D":"#e67e22","E":"#e74c3c"}.get(R["grade"],"#aaa")
        delta_c = "#2ecc71" if R["ca_delta"] >= 0 else "#e74c3c"
        delta_s = f"+{R['ca_delta']}" if R["ca_delta"] >= 0 else str(R["ca_delta"])

        st.markdown("---")
        st.markdown("## 📊 Analiz Sonucu")

        # Oyuncu başlık kartı
        st.markdown(
            f"<div style='background:linear-gradient(135deg,#0d1117,#161b22);"
            f"border:1px solid #30363d;border-radius:12px;padding:16px 22px;"
            f"margin-bottom:16px'>"
            f"<div style='font-size:22px;font-weight:900;color:#f0f0f0'>"
            f"{R['flag']} {R['name']}</div>"
            f"<div style='color:#8b949e;font-size:13px'>"
            f"{R['pos']} · {R['cty']} · {R['age']} yaş · "
            f"{R['foot']} · {R['height']} cm / {R['weight']} kg</div>"
            f"<div style='margin-top:10px;display:flex;gap:24px;align-items:center'>"
            f"<div><div style='font-size:10px;color:#8b949e'>FM CA</div>"
            f"<div style='font-size:24px;font-weight:900;color:#2ecc71'>{R['ca_fm']}</div></div>"
            f"<div><div style='font-size:10px;color:#8b949e'>Hesaplanan CA</div>"
            f"<div style='font-size:24px;font-weight:900;color:#3498db'>{R['calc_ca']}</div></div>"
            f"<div><div style='font-size:10px;color:#8b949e'>Fark</div>"
            f"<div style='font-size:20px;font-weight:900;color:{delta_c}'>{delta_s}</div></div>"
            f"<div><div style='font-size:10px;color:#8b949e'>PA</div>"
            f"<div style='font-size:20px;font-weight:700;color:#e67e22'>{R['pa']}</div></div>"
            f"<div><div style='font-size:10px;color:#8b949e'>Scout</div>"
            f"<div style='font-size:20px;font-weight:900;color:{gc}'>{R['grade']}</div></div>"
            f"</div></div>",
            unsafe_allow_html=True
        )

        # CA fark uyarısı
        if abs(R["ca_delta"]) > 10:
            st.warning(
                f"⚠️ FM CA ({R['ca_fm']}) ile hesaplanan CA ({R['calc_ca']}) arasında "
                f"**{abs(R['ca_delta'])} puanlık** fark var. "
                f"{'Girilen attribute değerlerini kontrol et.' if R['ca_delta'] < 0 else 'FM bu oyuncuyu daha yüksek değerlendiriyor olabilir.'}"
            )

        # ── Ana Layout ────────────────────────────────────
        res_l, res_m, res_r = st.columns([1.1, 1, 1.2])

        with res_l:
            # Radar
            st.markdown("**📡 Radar**")
            fig_fm = radar_chart(R["all_a"], R["pos"])
            st.pyplot(fig_fm, use_container_width=True)
            plt.close(fig_fm)

            # Kişilik
            st.markdown("**🧬 Kişilik**")
            pers = R["pers"]
            st.markdown(
                f"<div style='background:#161b22;border-left:3px solid "
                f"{pers['color']};border-radius:8px;padding:10px 14px'>"
                f"<div style='font-weight:800;color:{pers['color']}'>"
                f"{pers['name']}</div>"
                f"<div style='font-size:11px;color:#8b949e;margin-top:3px'>"
                f"{pers['desc']}</div></div>",
                unsafe_allow_html=True
            )

            # Öne çıkan gizli değerler
            highlights = sorted(
                [(k,v) for k,v in R["hidden"].items() if v >= 14],
                key=lambda x: -x[1]
            )
            if highlights:
                st.markdown("**🔒 Öne Çıkan Gizli**")
                bdg = ""
                for attr, val in highlights[:8]:
                    c = "#3498db" if val>=17 else "#2ecc71"
                    bdg += (f"<span style='background:#21262d;border-radius:8px;"
                            f"padding:2px 8px;font-size:11px;color:{c};"
                            f"margin:2px;display:inline-block'>{attr} {val}</span>")
                st.markdown(bdg, unsafe_allow_html=True)

        with res_m:
            # Finansal Özet
            st.markdown("**💰 Finansal**")
            for label, val, color in [
                ("Transfer Değeri", R["value_str"], "#f1c40f"),
                ("Haftalık Maaş",  format_wage(R["weekly"]), "#e67e22"),
                (f"Sözleşme",       f"{R['yrs']} yıl", "#3498db"),
                ("İmzalama Bonusu",f"{R['sign_bon']/1e6:.1f}M €", "#9b59b6"),
            ]:
                st.markdown(
                    f"<div style='background:#161b22;border-radius:6px;"
                    f"padding:8px 12px;margin-bottom:4px'>"
                    f"<div style='font-size:10px;color:#8b949e'>{label}</div>"
                    f"<div style='font-weight:700;color:{color}'>{val}</div></div>",
                    unsafe_allow_html=True
                )

            # Kariyer projeksiyonu
            st.markdown("**📈 Kariyer Projeksiyonu**")
            prof = R["hidden"].get("Profesyonellik", 12)
            eff_pa = calculate_effective_pa(R["pa"], prof, "Star")
            growth = simulate_growth(R["age"], R["ca_fm"], eff_pa, prof, R["pos"])
            if growth:
                ages_g = [R["age"]] + [y for y,_ in growth]
                vals_g = [R["ca_fm"]] + [v for _,v in growth]
                fig_g, ax_g = plt.subplots(figsize=(4, 2.5))
                fig_g.patch.set_facecolor("#0d1117")
                ax_g.set_facecolor("#0d1117")
                ax_g.plot(ages_g, vals_g, color="#2ecc71", lw=2, marker="o", ms=4)
                ax_g.fill_between(ages_g, vals_g, alpha=0.1, color="#2ecc71")
                ax_g.axhline(R["pa"], color="#e67e22", ls="--", lw=1,
                             label=f"PA {R['pa']}")
                ax_g.axhline(eff_pa, color="#555", ls=":", lw=1,
                             label=f"Eff {eff_pa}")
                ax_g.set_ylim(max(0, min(vals_g)-10), min(200, R["pa"]+15))
                ax_g.tick_params(colors="#aaa", labelsize=6)
                ax_g.spines[:].set_color("#1e2235")
                ax_g.legend(fontsize=6, labelcolor="#aaa", facecolor="#0d1117")
                ax_g.grid(color="#1e2235", lw=0.5)
                plt.tight_layout(pad=0.3)
                st.pyplot(fig_g, use_container_width=True)
                plt.close(fig_g)

            # Zirve CA ve kaç yıl
            peak = max(vals_g) if growth else R["ca_fm"]
            st.caption(f"Zirve CA: **{peak}** · PA kullanım: **%{int(peak/R['pa']*100)}**")

        with res_r:
            # Tüm mevkilerde CA
            st.markdown("**📍 Mevki Fit Analizi**")
            pos_cas = sorted(
                [(pos, calculate_ca(R["all_a"], pos, R["wf"])) for pos in ALL_POSITIONS],
                key=lambda x: -x[1]
            )
            max_pc = pos_cas[0][1]
            pos_c_map = {
                "ST":"#e74c3c","KF (Sol)":"#e74c3c","KF (Sağ)":"#e74c3c",
                "OOS":"#e67e22","KANAT (Sol)":"#e67e22","KANAT (Sağ)":"#e67e22",
                "OS":"#f1c40f","DM":"#f1c40f",
                "KB (Sol)":"#2ecc71","KB (Sağ)":"#2ecc71",
                "D (Sol)":"#3498db","D (Sağ)":"#3498db","DOS":"#3498db",
                "KL":"#9b59b6",
            }
            fit_html = "<div style='display:flex;flex-direction:column;gap:3px'>"
            for i,(pos,ca_v) in enumerate(pos_cas):
                pct  = ca_v / max(max_pc,1) * 100
                pc   = pos_c_map.get(pos,"#aaa")
                bold = pos == R["pos"]
                medal = ["🥇","🥈","🥉"][i] if i<3 else ""
                is_main = "★ " if pos == R["pos"] else ""
                gr = scout_grade(ca_v)
                fit_html += (
                    f"<div style='display:flex;align-items:center;gap:6px;"
                    f"padding:2px 4px;border-radius:4px;"
                    f"background:{'#161b22' if bold else 'transparent'}'>"
                    f"<span style='width:20px;font-size:10px'>{medal}</span>"
                    f"<span style='width:80px;font-size:{'12px' if bold else '10px'};"
                    f"color:{pc};font-weight:{'800' if bold else '400'}'>"
                    f"{is_main}{pos}</span>"
                    f"<div style='flex:1;background:#1e1e2e;border-radius:3px;height:6px'>"
                    f"<div style='width:{pct:.0f}%;height:100%;background:{pc};"
                    f"opacity:{'1' if i<3 else '0.5'};border-radius:3px'></div></div>"
                    f"<span style='width:28px;font-size:10px;color:#f0f0f0;"
                    f"font-weight:{'700' if bold else '400'}'>{ca_v}</span>"
                    f"<span style='width:14px;font-size:9px;color:{gc}'>{gr}</span>"
                    f"</div>"
                )
            fit_html += "</div>"
            st.markdown(fit_html, unsafe_allow_html=True)

        # ── Havuza Kaydet ──────────────────────────────────
        st.divider()
        sv1, sv2 = st.columns([1, 4])
        with sv1:
            _fm_id = f"FM_{R['name']}_{R['age']}_{R['ca_fm']}"
            _in_pool = any(p.get("_id") == _fm_id
                          for p in st.session_state.get("player_pool", []))
            if _in_pool:
                st.success("✅ Havuzda", icon="📦")
            else:
                if st.button("📦 Havuza Ekle", key="fm_to_pool", type="primary"):
                    entry = {
                        "_id"       : _fm_id,
                        "_idx"      : len(st.session_state.get("player_pool",[])),
                        "name"      : R["name"],
                        "flag"      : R["flag"],
                        "position"  : R["pos"],
                        "age"       : R["age"],
                        "country"   : R["cty"],
                        "preset"    : "Star",
                        "ca"        : R["ca_fm"],
                        "pa"        : R["pa"],
                        "personality": R["pers"],
                        "height"    : R["height"],
                        "weight"    : R["weight"],
                        "foot"      : R["foot"],
                        "weak_foot" : R["wf"],
                        "favori"    : False,
                        "tech"      : R["tech"],
                        "mental"    : R["mental"],
                        "phys"      : R["phys"],
                        "gk"        : R["gk"],
                        "hidden"    : R["hidden"],
                    }
                    if "player_pool" not in st.session_state:
                        st.session_state.player_pool = []
                    st.session_state.player_pool.append(entry)
                    st.rerun()
        with sv2:
            st.caption("📦 Oyuncu Havuzu'na eklenirse Excel'e aktarabilirsin.")

st.caption("© Football Manager Oyuncu Oluşturma | Streamlit + Python | 2026 Enes Özkan")
