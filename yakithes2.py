import streamlit as st
import pandas as pd
import math
import io
import os
import plotly.graph_objects as go
import folium
from streamlit_folium import st_folium
from math import radians, sin, cos, sqrt, atan2
from datetime import time, timedelta, datetime
import requests
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Font kayıt: Arial (Windows) → Calibri (Windows) → DejaVu (Linux/Mac)
def _font_yukle():
    import os, glob
    from reportlab.pdfbase.pdfmetrics import registerFontFamily

    _here = os.path.dirname(os.path.abspath(__file__))

    def _dene(f_n, f_b, path_n, path_b):
        if os.path.exists(path_n) and os.path.exists(path_b):
            try:
                pdfmetrics.registerFont(TTFont(f_n, path_n))
                pdfmetrics.registerFont(TTFont(f_b, path_b))
                registerFontFamily(f_n, normal=f_n, bold=f_b)
                return f_n, f_b
            except Exception:
                pass
        return None

    # 1. Yerel dizin (py ile aynı klasör) — Arial
    r = _dene("PDF_Arial", "PDF_Arial_Bold",
              os.path.join(_here, "arial.ttf"),
              os.path.join(_here, "arialbd.ttf"))
    if r: return r

    # 2. Windows sistem — Arial
    for root in [r"C:\Windows\Fonts", os.path.expandvars(r"%SystemRoot%\Fonts")]:
        r = _dene("PDF_Arial", "PDF_Arial_Bold",
                  os.path.join(root, "arial.ttf"),
                  os.path.join(root, "arialbd.ttf"))
        if r: return r

    # 3. Yerel dizin — Calibri
    r = _dene("PDF_Calibri", "PDF_Calibri_Bold",
              os.path.join(_here, "calibri.ttf"),
              os.path.join(_here, "calibrib.ttf"))
    if r: return r

    # 4. Windows sistem — Calibri
    for root in [r"C:\Windows\Fonts", os.path.expandvars(r"%SystemRoot%\Fonts")]:
        r = _dene("PDF_Calibri", "PDF_Calibri_Bold",
                  os.path.join(root, "calibri.ttf"),
                  os.path.join(root, "calibrib.ttf"))
        if r: return r

    # 5. DejaVu — bilinen yollar
    for dn, db in [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
    ]:
        r = _dene("PDF_DejaVu", "PDF_DejaVu_Bold", dn, db)
        if r: return r

    # 6. DejaVu — glob ile otomatik ara
    hits_n = glob.glob("/usr/share/fonts/**/DejaVuSans.ttf", recursive=True)
    hits_b = glob.glob("/usr/share/fonts/**/DejaVuSans-Bold.ttf", recursive=True)
    if hits_n and hits_b:
        r = _dene("PDF_DejaVu", "PDF_DejaVu_Bold", hits_n[0], hits_b[0])
        if r: return r

    # 7. Son çare — Helvetica (Türkçe karakter olmadan ama app çökmez)
    return "Helvetica", "Helvetica-Bold"

PDF_FONT, PDF_FONT_BOLD = _font_yukle()
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Araç Yol Bilgisayarı", page_icon="🚗", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.page-title { font-size:2rem; font-weight:700; color:#1F4E78; margin-bottom:4px; }
.page-sub   { color:#888; font-size:0.9rem; margin-bottom:20px; }
.status-ok   { background:linear-gradient(135deg,#375623,#70AD47); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(55,86,35,.3); }
.status-fail { background:linear-gradient(135deg,#7B0000,#C00000); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(192,0,0,.3); }
.status-ev-ok   { background:linear-gradient(135deg,#4a1a7a,#7030A0); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(112,48,160,.3); margin-top:6px; }
.status-ev-multi { background:linear-gradient(135deg,#1a237e,#1976D2); color:white; padding:16px 22px;
               border-radius:12px; display:flex; align-items:center; gap:14px;
               font-size:1.1rem; font-weight:700; box-shadow:0 4px 12px rgba(25,118,210,.3); margin-top:6px; }
.status-icon { font-size:2rem; }
.status-detail { font-weight:400; font-size:0.85rem; opacity:.85; margin-top:2px; }
.section-header { font-size:1rem; font-weight:700; color:#1F4E78;
    border-left:4px solid #1F4E78; padding-left:10px; margin:24px 0 10px; }
.route-strip { background:#EBF3FB; border-radius:10px; padding:12px 16px; margin:8px 0 10px;
               border-left:4px solid #1F4E78; }
.route-strip-title { font-size:0.72rem; font-weight:700; color:#1F4E78;
    text-transform:uppercase; letter-spacing:.6px; margin-bottom:8px; }
.route-seg { display:flex; align-items:center; gap:6px; font-size:0.85rem; color:#333; margin:4px 0; }
.route-seg-km { margin-left:auto; font-weight:700; color:#1F4E78; background:white;
    padding:2px 8px; border-radius:20px; font-size:0.78rem; border:1px solid #C8DDF0; }
.route-total { border-top:1px dashed #C8DDF0; margin-top:8px; padding-top:8px;
    font-weight:700; font-size:0.9rem; color:#1F4E78; display:flex; justify-content:space-between; }
.etki-box { background:#FFF8E7; border:1px solid #FFE082; border-radius:8px;
    padding:10px 14px; font-size:0.82rem; color:#5D4037; margin-top:6px; }
[data-testid="metric-container"] { background:white; border:1px solid #E8EDF2;
    border-radius:12px; padding:14px 18px; box-shadow:0 2px 8px rgba(0,0,0,.05); }
[data-testid="metric-container"] > label { font-size:0.72rem !important; color:#888 !important;
    font-weight:600 !important; text-transform:uppercase; letter-spacing:.5px; }
[data-testid="stMetricValue"] > div { font-size:1.4rem !important; font-weight:700 !important; color:#1F4E78 !important; }
[data-testid="stMetricDelta"] > div { font-size:0.75rem !important; color:#999 !important; }
[data-testid="stMetricDeltaIcon"] { display:none !important; }
[data-testid="stSidebar"] { background:#F8FAFC; border-right:1px solid #E2E8F0; }
[data-testid="stSidebar"] label { font-size:0.8rem !important; font-weight:600 !important; color:#555 !important; }
</style>
""", unsafe_allow_html=True)

# ── Hava durumu (Open-Meteo — API key gerektirmez) ───────────────────────────
# WMO 4677 hava kodu → (emoji, Türkçe açıklama)
WMO_KODU = {
    0:  ("☀️",  "Açık gökyüzü"),
    1:  ("🌤",  "Çoğunlukla açık"),
    2:  ("⛅",  "Parçalı bulutlu"),
    3:  ("☁️",  "Kapalı"),
    45: ("🌫",  "Sisli"),
    48: ("🌫",  "Kırağılı sis"),
    51: ("🌦",  "Hafif çiseleyen"),
    53: ("🌦",  "Orta çiseleyen"),
    55: ("🌧",  "Yoğun çiseleyen"),
    56: ("🌦",  "Dondurucu hafif çiseleyen"),
    57: ("🌧",  "Dondurucu yoğun çiseleyen"),
    61: ("🌧",  "Hafif yağmur"),
    63: ("🌧",  "Orta yağmur"),
    65: ("🌧",  "Yoğun yağmur"),
    66: ("🌧",  "Dondurucu hafif yağmur"),
    67: ("🌧",  "Dondurucu yoğun yağmur"),
    71: ("❄️",  "Hafif kar"),
    73: ("❄️",  "Orta kar yağışı"),
    75: ("❄️",  "Yoğun kar yağışı"),
    77: ("❄️",  "Kar taneleri"),
    80: ("🌦",  "Hafif sağanak"),
    81: ("🌧",  "Orta sağanak"),
    82: ("⛈",  "Şiddetli sağanak"),
    85: ("❄️",  "Hafif kar sağanağı"),
    86: ("❄️",  "Yoğun kar sağanağı"),
    95: ("⛈",  "Gök gürültülü fırtına"),
    96: ("⛈",  "Dolu fırtınası (küçük)"),
    99: ("⛈",  "Dolu fırtınası (büyük)"),
}

def hava_getir(lat, lon, hedef_dt):
    """Open-Meteo saatlik + günlük tahmin → hedef saat + gün verisi (API key yok, 16 gün)"""
    cache_key = (round(lat, 2), round(lon, 2), hedef_dt.strftime("%Y%m%d%H"))
    if cache_key in st.session_state.hava_cache:
        return st.session_state.hava_cache[cache_key]

    try:
        params = {
            "latitude":        lat,
            "longitude":       lon,
            "hourly": ",".join([
                "temperature_2m", "apparent_temperature", "dewpoint_2m",
                "precipitation_probability", "precipitation", "rain",
                "snowfall", "weathercode", "cloudcover",
                "windspeed_10m", "windgusts_10m", "winddirection_10m",
                "visibility", "relativehumidity_2m", "surface_pressure",
                "uv_index", "is_day", "freezing_level_height",
            ]),
            "daily":           "sunrise,sunset",
            "timezone":        "Europe/Istanbul",
            "forecast_days":   16,
            "wind_speed_unit": "kmh",
        }
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params=params, timeout=10
        )
        r.raise_for_status()
        data  = r.json()
        times = data["hourly"]["time"]

        # Hedef zamana en yakın saati bul
        hedef_ts  = hedef_dt.timestamp()
        best_idx  = 0
        best_diff = float("inf")
        for i, t_str in enumerate(times):
            diff = abs(datetime.fromisoformat(t_str).timestamp() - hedef_ts)
            if diff < best_diff:
                best_diff = diff
                best_idx  = i

        h = data["hourly"]

        # ── Güneş doğuş / batış (daily) ──────────────────────────────────────
        hedef_tarih = hedef_dt.strftime("%Y-%m-%d")
        daily_dates = data.get("daily", {}).get("time", [])
        sunrise_str = sunset_str = "—"
        try:
            d_idx = daily_dates.index(hedef_tarih)
            raw_sr = data["daily"]["sunrise"][d_idx]  # "2024-01-01T07:15"
            raw_ss = data["daily"]["sunset"][d_idx]
            sunrise_str = datetime.fromisoformat(raw_sr).strftime("%H:%M")
            sunset_str  = datetime.fromisoformat(raw_ss).strftime("%H:%M")
        except (ValueError, IndexError, KeyError):
            pass

        # ── Günün saatlik dizileri (sparkline için) ───────────────────────────
        prefix = hedef_tarih + "T"
        gun_idxler = [i for i, t in enumerate(times) if t.startswith(prefix)]
        day_temps        = [h["temperature_2m"][i]           for i in gun_idxler]
        day_precip_prob  = [h["precipitation_probability"][i] for i in gun_idxler]
        day_precip       = [h["precipitation"][i]             for i in gun_idxler]
        day_weathercodes = [h["weathercode"][i]               for i in gun_idxler]
        day_hours        = [times[i][11:16]                   for i in gun_idxler]  # "HH:MM"

        result = {
            "temp":           h["temperature_2m"][best_idx],
            "feels_like":     h["apparent_temperature"][best_idx],
            "dewpoint":       h["dewpoint_2m"][best_idx],
            "humidity":       h["relativehumidity_2m"][best_idx],
            "pressure":       h["surface_pressure"][best_idx],
            "cloudcover":     h["cloudcover"][best_idx],
            "weathercode":    h["weathercode"][best_idx],
            "precip_prob":    h["precipitation_probability"][best_idx],
            "precip":         h["precipitation"][best_idx],
            "rain":           h["rain"][best_idx],
            "snow_cm":        h["snowfall"][best_idx],
            "windspeed":      h["windspeed_10m"][best_idx],
            "windgusts":      h["windgusts_10m"][best_idx],
            "winddirection":  h["winddirection_10m"][best_idx],
            "visibility":     h["visibility"][best_idx],
            "uv_index":       h["uv_index"][best_idx],
            "is_day":         h["is_day"][best_idx],
            "freeze_lvl_m":   h["freezing_level_height"][best_idx],
            "time_str":       times[best_idx],
            # yeni alanlar
            "sunrise":        sunrise_str,
            "sunset":         sunset_str,
            "day_hours":      day_hours,
            "day_temps":      day_temps,
            "day_precip_prob":day_precip_prob,
            "day_precip":     day_precip,
            "day_wcodes":     day_weathercodes,
        }
        st.session_state.hava_cache[cache_key] = result
        return result
    except Exception as e:
        return {"hata": str(e)}


def hava_kalite_getir(lat, lon, hedef_dt):
    """Open-Meteo Air Quality → AQI (EU), PM2.5, PM10 (API key yok, 5 gün)"""
    cache_key = ("aqi", round(lat, 2), round(lon, 2), hedef_dt.strftime("%Y%m%d%H"))
    if cache_key in st.session_state.hava_cache:
        return st.session_state.hava_cache[cache_key]

    fark_gun = (hedef_dt - datetime.now()).total_seconds() / 86400
    if fark_gun > 5:
        return {"hata": "5 günden uzak, AQI tahmini yok"}

    try:
        params = {
            "latitude":     lat,
            "longitude":    lon,
            "hourly":       "pm2_5,pm10,european_aqi",
            "timezone":     "Europe/Istanbul",
            "forecast_days": 5,
        }
        r = requests.get(
            "https://air-quality-api.open-meteo.com/v1/air-quality",
            params=params, timeout=10
        )
        r.raise_for_status()
        data  = r.json()
        times = data["hourly"]["time"]

        hedef_ts  = hedef_dt.timestamp()
        best_idx  = 0
        best_diff = float("inf")
        for i, t_str in enumerate(times):
            diff = abs(datetime.fromisoformat(t_str).timestamp() - hedef_ts)
            if diff < best_diff:
                best_diff = diff
                best_idx  = i

        h = data["hourly"]
        result = {
            "pm25":   h["pm2_5"][best_idx],
            "pm10":   h["pm10"][best_idx],
            "aqi_eu": h["european_aqi"][best_idx],
        }
        st.session_state.hava_cache[cache_key] = result
        return result
    except Exception as e:
        return {"hata": str(e)}


def sparkline_svg(day_hours, day_temps, day_precip_prob, gecis_saat_str,
                  width=400, height=56):
    """24 saatlik sıcaklık + yağış ihtimali sparkline SVG'si üretir."""
    if not day_temps or len(day_temps) < 2:
        return ""

    n     = len(day_temps)
    t_min = min(day_temps)
    t_max = max(day_temps)
    t_rng = max(t_max - t_min, 1.0)

    pad_l = 30; pad_r = 6; pad_t = 8; pad_b = 18
    iw = width - pad_l - pad_r    # iç genişlik
    ih = height - pad_t - pad_b   # iç yükseklik

    def tx(i): return pad_l + i * iw / max(n - 1, 1)
    def ty(t): return pad_t + ih - (t - t_min) / t_rng * ih

    # Yağış ihtimali çubukları (arka plan)
    bars = ""
    bar_w = iw / n
    for i, pp in enumerate(day_precip_prob or [0] * n):
        if pp > 0:
            bh    = (pp / 100) * ih
            bx    = pad_l + i * bar_w
            by    = pad_t + ih - bh
            alpha = min(pp / 100 * 0.45, 0.45)
            bars += (f'<rect x="{bx:.1f}" y="{by:.1f}" '
                     f'width="{bar_w:.1f}" height="{bh:.1f}" '
                     f'fill="rgba(25,118,210,{alpha:.2f})" rx="1"/>')

    # Sıcaklık eğrisi
    pts = " ".join(f"{tx(i):.1f},{ty(t):.1f}" for i, t in enumerate(day_temps))

    # Geçiş saati marker'ı
    cur_idx = 0
    if gecis_saat_str and day_hours:
        hedef_h = gecis_saat_str[:2]
        for i, h in enumerate(day_hours):
            if h[:2] == hedef_h:
                cur_idx = i
                break
    cx = tx(cur_idx)
    cy = ty(day_temps[min(cur_idx, n - 1)])

    # Saat etiketleri (0, 6, 12, 18)
    saat_lbls = ""
    for target_h in [0, 6, 12, 18]:
        for i, h in enumerate(day_hours):
            if int(h[:2]) == target_h:
                lx = tx(i)
                saat_lbls += (f'<text x="{lx:.0f}" y="{height - 2}" '
                              f'text-anchor="middle" font-size="8" fill="#AAA">'
                              f'{target_h:02d}:00</text>')
                break

    # Min/Max sıcaklık etiketleri (sol kenar)
    temp_lbls = (
        f'<text x="{pad_l - 3}" y="{pad_t + 5}" text-anchor="end" '
        f'font-size="8" fill="#1F4E78">{t_max:.0f}°</text>'
        f'<text x="{pad_l - 3}" y="{pad_t + ih}" text-anchor="end" '
        f'font-size="8" fill="#1F4E78">{t_min:.0f}°</text>'
    )

    return (
        f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg" '
        f'style="display:block;border-radius:6px">'
        f'<rect width="{width}" height="{height}" fill="rgba(255,255,255,0.6)" rx="6"/>'
        f'{bars}'
        f'<polyline points="{pts}" fill="none" stroke="#1F4E78" '
        f'stroke-width="1.8" stroke-linejoin="round"/>'
        f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="4" '
        f'fill="#E65100" stroke="white" stroke-width="1.5"/>'
        f'{temp_lbls}{saat_lbls}'
        f'</svg>'
    )


def hesapla_yol_sicakligi(air_temp, cloudcover, is_day, humidity, uv_idx):
    """Hava sıcaklığından tahmini yol yüzeyi sıcaklığını hesaplar."""
    cloud_fac = 1.0 - cloudcover / 100.0
    if is_day:
        solar_gain = cloud_fac * min(uv_idx * 1.8, 14)   # güneşli günde maks +14°C
        return air_temp + solar_gain
    else:
        rad_loss = cloud_fac * (1.0 - humidity / 100.0) * 6   # açık gecede maks -6°C
        return air_temp - rad_loss


def hesapla_thi(temp, humidity):
    """Sıcaklık-Nem Konforu İndeksi (THI). Döner: (thi_degeri, etiket, renk)"""
    # THI = T - (0.55 - 0.0055 * RH) * (T - 14.5)
    thi = temp - (0.55 - 0.0055 * humidity) * (temp - 14.5)
    if temp < 10:
        return thi, "Soğuk", "#1976D2"
    elif thi < 21:
        return thi, "Konforlu", "#2E7D32"
    elif thi < 24:
        return thi, "Hafif Sıcak", "#F9A825"
    elif thi < 27:
        return thi, "Bunaltıcı", "#E65100"
    else:
        return thi, "Çok Bunaltıcı", "#B71C1C"

def hava_karti(durak_adi, varis_str, veri, gecis_dt=None, kalite_veri=None):
    """Tam hava durumu kartı (Open-Meteo) — tüm göstergeler + sparkline."""
    if "hata" in veri:
        return (f'<div style="background:#FFF3CD;border-radius:10px;padding:12px;'
                f'margin:6px 0;border-left:4px solid #FFC107">'
                f'<b>{durak_adi}</b> ({varis_str}) — ⚠️ {veri["hata"]}</div>')

    # ── Temel veriler ─────────────────────────────────────────────────────────
    sicaklik    = veri["temp"]
    hissedilen  = veri["feels_like"]
    ciy_nkt     = veri["dewpoint"]
    nem         = veri["humidity"]
    basinc      = veri["pressure"]
    bulut_pct   = veri["cloudcover"]
    wcode       = veri["weathercode"]
    yagis_pct   = veri["precip_prob"]
    yagmur_mm   = veri["rain"]
    kar_cm      = veri["snow_cm"]
    kar_mm      = kar_cm * 10
    toplam_yag  = yagmur_mm + kar_mm
    ruzgar_kmh  = veri["windspeed"]
    hamle_kmh   = veri["windgusts"]
    ruzgar_deg  = veri["winddirection"]
    gorunum_m   = veri["visibility"]
    gorunum_km  = gorunum_m / 1000
    uv_idx      = veri["uv_index"]
    freeze_m    = veri["freeze_lvl_m"]
    is_day_flag = veri["is_day"]
    sunrise     = veri.get("sunrise", "—")
    sunset      = veri.get("sunset",  "—")

    emoji, durum = WMO_KODU.get(wcode, ("🌡", f"WMO-{wcode}"))

    yon_list   = ["K","KKD","KD","DKD","D","DGD","GD","GGD","G","GGB","GB","BGB","B","KBK","KB","KKB"]
    ruzgar_yon = yon_list[round(ruzgar_deg / 22.5) % 16]

    if gecis_dt is not None:
        saat = gecis_dt.hour
        gece_mi = saat < 6 or saat >= 21
    else:
        gece_mi = (is_day_flag == 0)
    zaman_ikon = "🌙 Gece" if gece_mi else "☀️ Gündüz"

    is_storm  = wcode in (95, 96, 99)
    is_snow   = wcode in (71, 73, 75, 77, 85, 86) or kar_mm > 0
    is_freeze = wcode in (56, 57, 66, 67)

    # ── Sürüş Güvenlik Skoru ─────────────────────────────────────────────────
    skor = 10.0
    skor_detay = []
    if gorunum_km < 0.2:
        skor -= 4.0; skor_detay.append("Yoğun sis")
    elif gorunum_km < 1:
        skor -= 2.5; skor_detay.append("Sis")
    elif gorunum_km < 4:
        skor -= 1.0; skor_detay.append("Düşük görüş")
    if kar_mm > 0:
        skor -= min(3.0, kar_mm * 0.08); skor_detay.append(f"Kar ({kar_cm:.1f}cm)")
    if yagmur_mm > 5:
        skor -= 1.5; skor_detay.append("Yoğun yağmur")
    elif yagmur_mm > 1:
        skor -= 0.8; skor_detay.append("Yağmur")
    if hamle_kmh > 90:
        skor -= 2.5; skor_detay.append(f"Fırtına hamlesi ({hamle_kmh:.0f}km/h)")
    elif hamle_kmh > 60:
        skor -= 1.5; skor_detay.append(f"Kuvvetli rüzgar ({hamle_kmh:.0f}km/h)")
    elif hamle_kmh > 40:
        skor -= 0.5; skor_detay.append(f"Rüzgarlı ({hamle_kmh:.0f}km/h)")
    if is_storm:
        skor -= 2.0; skor_detay.append("Gök gürültülü fırtına")
    if is_freeze:
        skor -= 1.5; skor_detay.append("Dondurucu yağış")
    if sicaklik < -5:
        skor -= 1.5; skor_detay.append("Şiddetli soğuk")
    elif sicaklik < 0:
        skor -= 0.8; skor_detay.append("Buzlanma riski")
    skor = max(0.0, min(10.0, skor))

    if skor >= 8:   skor_renk = "#2E7D32"; skor_bg = "#E8F5E9"; skor_lbl = "İyi"
    elif skor >= 6: skor_renk = "#F57F17"; skor_bg = "#FFF8E1"; skor_lbl = "Dikkatli"
    elif skor >= 4: skor_renk = "#E65100"; skor_bg = "#FFF3E0"; skor_lbl = "Zor"
    else:           skor_renk = "#B71C1C"; skor_bg = "#FFEBEE"; skor_lbl = "Tehlikeli"

    # ── Yakıt Etkisi ─────────────────────────────────────────────────────────
    if    sicaklik <  0: pct_sicak = 15
    elif  sicaklik < 10: pct_sicak = 8
    elif  sicaklik < 25: pct_sicak = 0
    elif  sicaklik < 35: pct_sicak = 5
    else:                pct_sicak = 8
    pct_yagis  = 8 if toplam_yag > 5 else (4 if toplam_yag > 1 else (2 if yagis_pct > 60 else 0))
    pct_ruzgar = 8 if hamle_kmh > 80 else (4 if hamle_kmh > 50 else (2 if hamle_kmh > 30 else 0))
    pct_toplam = pct_sicak + pct_yagis + pct_ruzgar
    yakit_etki = f"+{pct_toplam}%" if pct_toplam > 0 else "±0%"
    detay_parcalar = []
    if pct_sicak  > 0: detay_parcalar.append(f"sıcaklık +{pct_sicak}%")
    if pct_yagis  > 0: detay_parcalar.append(f"yağış +{pct_yagis}%")
    if pct_ruzgar > 0: detay_parcalar.append(f"rüzgar +{pct_ruzgar}%")
    yakit_detay = " · ".join(detay_parcalar) if detay_parcalar else "etkisiz"
    if   pct_toplam >= 15: yakit_renk = "#B71C1C"
    elif pct_toplam >= 8:  yakit_renk = "#E65100"
    elif pct_toplam >= 4:  yakit_renk = "#F57F17"
    elif pct_toplam > 0:   yakit_renk = "#1976D2"
    else:                   yakit_renk = "#2E7D32"

    # ── UV ───────────────────────────────────────────────────────────────────
    if   uv_idx >= 11: uv_renk = "#B71C1C"; uv_lbl = "Aşırı"
    elif uv_idx >= 8:  uv_renk = "#E65100"; uv_lbl = "Çok yüksek"
    elif uv_idx >= 6:  uv_renk = "#F57F17"; uv_lbl = "Yüksek"
    elif uv_idx >= 3:  uv_renk = "#F9A825"; uv_lbl = "Orta"
    else:               uv_renk = "#2E7D32"; uv_lbl = "Düşük"

    # ── Bulutluluk rengi ─────────────────────────────────────────────────────
    if   bulut_pct < 20: bulut_renk = "#2E7D32"
    elif bulut_pct < 60: bulut_renk = "#F57F17"
    else:                bulut_renk = "#546E7A"

    # ── Yol yüzeyi sıcaklığı ─────────────────────────────────────────────────
    yol_temp = hesapla_yol_sicakligi(sicaklik, bulut_pct, is_day_flag, nem, uv_idx)
    yol_buzlanma = yol_temp <= 1.0
    if yol_buzlanma:
        yol_temp_renk = "#B71C1C"
        yol_temp_ikon = "🧊"
    elif yol_temp >= 50:
        yol_temp_renk = "#E65100"
        yol_temp_ikon = "🔥"
    else:
        yol_temp_renk = "#546E7A"
        yol_temp_ikon = "🛣"

    # ── Nem konforu (THI) ─────────────────────────────────────────────────────
    thi_val, thi_lbl, thi_renk = hesapla_thi(sicaklik, nem)

    # ── Hava kalitesi (AQI) ───────────────────────────────────────────────────
    aqi_html = ""
    if kalite_veri and "hata" not in kalite_veri:
        aqi  = kalite_veri["aqi_eu"]
        pm25 = kalite_veri["pm25"]
        pm10 = kalite_veri["pm10"]
        if   aqi <= 20:  aqi_renk = "#2E7D32"; aqi_lbl = "Çok İyi"
        elif aqi <= 40:  aqi_renk = "#66BB6A"; aqi_lbl = "İyi"
        elif aqi <= 60:  aqi_renk = "#F9A825"; aqi_lbl = "Orta"
        elif aqi <= 80:  aqi_renk = "#E65100"; aqi_lbl = "Kötü"
        elif aqi <= 100: aqi_renk = "#B71C1C"; aqi_lbl = "Çok Kötü"
        else:             aqi_renk = "#6A0080"; aqi_lbl = "Tehlikeli"
        aqi_html = (
            f'<span>🌬 AQI: <b style="color:{aqi_renk}">{aqi:.0f} — {aqi_lbl}</b>'
            f' <span style="color:#888;font-size:0.8rem">'
            f'PM2.5:{pm25:.1f} PM10:{pm10:.1f} µg/m³</span></span>'
        )

    # ── Uyarı bandı ──────────────────────────────────────────────────────────
    uyarilar = []
    if is_snow:                           uyarilar.append("❄️ KAR")
    if is_storm:                          uyarilar.append("⛈ FIRTINA")
    if hamle_kmh > 60:                    uyarilar.append("💨 KUVVETLI RÜZGAR")
    if sicaklik < 0 or is_freeze:         uyarilar.append("🧊 BUZLANMA")
    if yol_buzlanma and not (sicaklik<0): uyarilar.append("🛣 YOL BUZLANMASI")
    if gorunum_km < 1:                    uyarilar.append("🌫 SIS")
    if yagis_pct > 70:                    uyarilar.append("🌧 YOĞUN YAĞIŞ")
    if freeze_m < 1000 and freeze_m > 0:  uyarilar.append("🏔 DÜŞÜK DON YÜKSEKLİĞİ")
    if kalite_veri and "aqi_eu" in kalite_veri and kalite_veri["aqi_eu"] > 60:
        uyarilar.append("🌬 HAVA KALİTESİ KÖTÜ")
    uyari_html = ""
    if uyarilar:
        uyari_html = " &nbsp;".join(f"<b style='color:#C62828'>{u}</b>" for u in uyarilar)
        uyari_html = f"<div style='margin:6px 0 2px;font-size:0.8rem'>{uyari_html}</div>"

    kenar = skor_renk
    arka  = skor_bg if skor < 6 else "#EBF3FB"

    # ── Don yüksekliği ────────────────────────────────────────────────────────
    freeze_html = ""
    if freeze_m > 0:
        if freeze_m < 1500:
            freeze_html = (f"<span style='color:#C62828'>🏔 Don yüksekliği: "
                           f"<b>{freeze_m:.0f}m</b> ⚠️ yol buzlanma riski</span>")
        else:
            freeze_html = (f"<span style='color:#666'>🏔 Don yüksekliği: "
                           f"<b>{freeze_m:.0f}m</b></span>")

    kar_html = f'<span>❄️ Kar: <b>{kar_cm:.1f}cm</b></span>' if kar_cm > 0 else ""

    # ── Sparkline ─────────────────────────────────────────────────────────────
    gecis_saat = gecis_dt.strftime("%H:%M") if gecis_dt else ""
    spark = sparkline_svg(
        veri.get("day_hours", []),
        veri.get("day_temps", []),
        veri.get("day_precip_prob", []),
        gecis_saat,
    )
    spark_html = (
        f'<div style="margin-top:10px">'
        f'<div style="font-size:0.72rem;color:#888;margin-bottom:3px">'
        f'📈 Günlük sıcaklık (mavi çubuk = yağış ihtimali · turuncu nokta = geçiş saati)</div>'
        f'{spark}</div>'
    ) if spark else ""

    return f"""<div style="background:{arka};border-radius:12px;padding:16px 20px;margin:8px 0;border-left:6px solid {kenar};box-shadow:0 2px 8px rgba(0,0,0,.06)">
<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px">
<div style="font-weight:700;font-size:1rem;color:#1F4E78">{emoji} {durak_adi} <span style="font-weight:400;font-size:0.82rem;color:#666">— {varis_str} {zaman_ikon}</span></div>
<div style="background:{skor_renk};color:white;border-radius:20px;padding:4px 14px;font-weight:700;font-size:0.85rem">{skor:.0f}/10 — {skor_lbl}</div>
</div>
{uyari_html}
<div style="display:flex;gap:20px;margin-top:10px;font-size:0.87rem;flex-wrap:wrap">
<span>🌡 <b>{sicaklik:.1f}°C</b> <span style="color:#888">(hissedilen {hissedilen:.1f}°C)</span></span>
<span>🌫 {durum}</span>
<span>☁️ Bulut: <b style="color:{bulut_renk}">%{bulut_pct}</b></span>
<span>💧 Nem: %{nem} <span style="color:#888">(çiy {ciy_nkt:.1f}°C)</span></span>
<span>📊 {basinc:.0f} hPa</span>
</div>
<div style="display:flex;gap:20px;margin-top:6px;font-size:0.87rem;flex-wrap:wrap">
<span>💨 Rüzgar: <b>{ruzgar_kmh:.0f} km/s</b> ({ruzgar_yon})</span>
<span>💨 Hamle: <b style="color:{'#C62828' if hamle_kmh>60 else 'inherit'}">{hamle_kmh:.0f} km/s</b></span>
<span>☀️ UV: <b style="color:{uv_renk}">{uv_idx:.0f} — {uv_lbl}</b></span>
<span>🌅 Doğuş/Batış: <b>{sunrise} / {sunset}</b></span>
</div>
<div style="display:flex;gap:20px;margin-top:6px;font-size:0.87rem;flex-wrap:wrap">
<span>👁 Görüş: <b style="color:{'#C62828' if gorunum_km < 2 else 'inherit'}">{gorunum_km:.1f} km{'  ⚠️' if gorunum_km < 2 else ''}</b></span>
<span>🌧 Yağış: <b>%{yagis_pct}</b>{f' ({yagmur_mm:.1f}mm)' if yagmur_mm > 0 else ''}</span>
{kar_html}
<span>⛽ <b style="color:{yakit_renk}">{yakit_etki}</b> <span style="color:#888;font-size:0.8rem">({yakit_detay})</span></span>
</div>
<div style="display:flex;gap:20px;margin-top:6px;font-size:0.87rem;flex-wrap:wrap">
<span>{yol_temp_ikon} Yol yüzeyi: <b style="color:{yol_temp_renk}">{yol_temp:.1f}°C{'  ⚠️' if yol_buzlanma else ''}</b></span>
<span>🧘 Nem konforu: <b style="color:{thi_renk}">{thi_lbl}</b> <span style="color:#888;font-size:0.8rem">(THI {thi_val:.0f})</span></span>
{aqi_html}
</div>
{f'<div style="display:flex;gap:16px;margin-top:6px;font-size:0.87rem;flex-wrap:wrap">{freeze_html}</div>' if freeze_html else ''}
{spark_html}
{f'<div style="margin-top:6px;font-size:0.78rem;color:#888">Skor: {", ".join(skor_detay)}</div>' if skor_detay else ''}
</div>"""

# ── Veri ─────────────────────────────────────────────────────────────────────
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ilce_data import ILCELER
IL_LIST = sorted(ILCELER.keys())

def ilce_listesi(il): return sorted(ILCELER.get(il, {}).keys())
def koordinat(il, ilce): return ILCELER[il][ilce]

ROAD_FACTOR = 1.30

def haversine_km(c1, c2):
    lat1,lon1 = c1; lat2,lon2 = c2
    R = 6371
    dlat,dlon = radians(lat2-lat1), radians(lon2-lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    return round(R*2*atan2(sqrt(a),sqrt(1-a))*ROAD_FACTOR)

def fmt_sure(h_float):
    h = int(h_float); m = int(round((h_float-h)*60))
    if m == 60: h+=1; m=0
    return f"{h}s {m:02d}dk" if h else f"{m}dk"

# ── Session state ─────────────────────────────────────────────────────────────
if "duraks"  not in st.session_state:
    st.session_state.duraks = [("İstanbul","Kadıköy"),("Ankara","Çankaya")]
if "km_ov"   not in st.session_state:
    st.session_state.km_ov = [0]
if "hava_cache" not in st.session_state:
    st.session_state.hava_cache = {}   # key: (lat,lon,dt_str) → veri

if "fiyatlar" not in st.session_state:
    st.session_state.fiyatlar = {"benzin":64.88,"dizel":55.0,"lpg":25.0,"elektrik":12.0,"kaynak":"","ts":None,"guncelleme":""}

def _sync():
    n = len(st.session_state.duraks)-1
    while len(st.session_state.km_ov) < n: st.session_state.km_ov.append(0)
    while len(st.session_state.km_ov) > n: st.session_state.km_ov.pop()
_sync()

# ── Yakıt fiyatlarını JSON'dan oku ───────────────────────────────────────────
def fiyat_json_oku():
    """fiyatlar.json dosyasından fiyatları okur."""
    try:
        import json as _json
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fiyatlar.json")
        with open(json_path, encoding="utf-8") as f:
            data = _json.load(f)
        st.session_state.fiyatlar["benzin"]    = float(data.get("benzin",   64.88))
        st.session_state.fiyatlar["dizel"]     = float(data.get("dizel",    55.00))
        st.session_state.fiyatlar["lpg"]       = float(data.get("lpg",      25.00))
        st.session_state.fiyatlar["elektrik"]  = float(data.get("elektrik", 12.00))
        st.session_state.fiyatlar["kaynak"]    = data.get("kaynak", "")
        st.session_state.fiyatlar["guncelleme"]= data.get("guncelleme", "")
        st.session_state.fiyatlar["ts"]        = data.get("guncelleme", "")
        return True
    except FileNotFoundError:
        return False
    except Exception as e:
        st.sidebar.warning(f"fiyatlar.json okunamadı: {e}")
        return False

# Uygulama ilk açıldığında JSON'dan otomatik yükle
if "json_yuklendi" not in st.session_state:
    fiyat_json_oku()
    st.session_state.json_yuklendi = True

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Güzergah ──────────────────────────────────────────────────────────────
    st.markdown("### 📍 Güzergah")
    to_remove = None
    n_d = len(st.session_state.duraks)
    for i in range(n_d):
        il, ilce = st.session_state.duraks[i]
        lbl = "🚩 Kalkış" if i==0 else ("🏁 Varış" if i==n_d-1 else f"📍 {i}. Durak")
        c1,c2 = st.columns([5,1])
        with c1:
            new_il   = st.selectbox(lbl, IL_LIST,
                index=IL_LIST.index(il) if il in IL_LIST else 0, key=f"il_{i}")
            ilceler  = ilce_listesi(new_il)
            safe_ilce= ilce if (new_il==il and ilce in ilceler) else ilceler[0]
            new_ilce = st.selectbox("", ilceler,
                index=ilceler.index(safe_ilce), key=f"ilce_{i}", label_visibility="collapsed")
            st.session_state.duraks[i] = (new_il, new_ilce)
        with c2:
            st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)
            if n_d > 2 and st.button("✕", key=f"del_{i}"):
                to_remove = i
        if i < n_d-1:
            a_il,a_ilce = st.session_state.duraks[i]
            b_il,b_ilce = st.session_state.duraks[i+1]
            auto_km = haversine_km(koordinat(a_il,a_ilce), koordinat(b_il,b_ilce))
            ov = st.number_input(f"↕ Segment km (0 = tahmini {auto_km} km)",
                value=int(st.session_state.km_ov[i]), min_value=0, step=1,
                key=f"km_ov_{i}", help="Google Maps'teki gerçek km'yi gir, 0 = otomatik")
            st.session_state.km_ov[i] = ov

    if to_remove is not None:
        st.session_state.duraks.pop(to_remove); _sync(); st.rerun()

    ca, cb = st.columns(2)
    with ca:
        if st.button("＋ Durak", use_container_width=True):
            st.session_state.duraks.append(("Ankara", ilce_listesi("Ankara")[0]))
            _sync(); st.rerun()
    with cb:
        if st.button("↺ Sıfırla", use_container_width=True):
            st.session_state.duraks=[("İstanbul","Kadıköy"),("Ankara","Çankaya")]
            st.session_state.km_ov=[0]; st.rerun()

    segments = []
    for i in range(len(st.session_state.duraks)-1):
        a_il,a_ilce = st.session_state.duraks[i]
        b_il,b_ilce = st.session_state.duraks[i+1]
        ov = st.session_state.km_ov[i]
        km = ov if ov>0 else haversine_km(koordinat(a_il,a_ilce), koordinat(b_il,b_ilce))
        src = "📍" if ov>0 else "📐"
        segments.append((f"{a_il}/{a_ilce}", f"{b_il}/{b_ilce}", km, src))
    total_yol = sum(s[2] for s in segments)

    if segments:
        segs_html = "".join(
            f'<div class="route-seg">{src} {a} → {b}<span class="route-seg-km">{km} km</span></div>'
            for a,b,km,src in segments)
        st.markdown(f"""<div class="route-strip">
            <div class="route-strip-title">Güzergah Özeti</div>
            {segs_html}
            <div class="route-total"><span>Toplam</span><span>{total_yol} km</span></div>
        </div>""", unsafe_allow_html=True)
        st.caption("📍 Manuel  ·  📐 Tahmini (haversine ×1.30)")
    yol = max(total_yol, 1)

    # ── Mola ──────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### ⏱ Mola Ayarları")
    mola_araligi = st.number_input("Kaç saatte bir mola", value=2, min_value=1, max_value=6, step=1)
    mola_suresi  = st.number_input("Mola süresi (dk)",    value=15, min_value=5, max_value=120, step=5)

    # ── Araç ──────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🚗 Araç Bilgileri (Benzin)")
    hiz      = st.number_input("Ortalama Hız (km/s)", value=90, min_value=1, max_value=300)
    depo     = st.number_input("Depo Kapasitesi (L)",  value=53.0, min_value=1.0, step=0.5, format="%.1f")
    tuketim  = st.number_input("100km Tüketim (L)",    value=7.4,  min_value=0.1, step=0.1, format="%.1f")
    cikis    = st.time_input("Çıkış Saati", value=time(8,0))
    kisi     = st.number_input("Kişi Sayısı", value=1, min_value=1, max_value=20)
    aylik_km = st.number_input("Aylık Ortalama km", value=2000, min_value=0, step=100)

    # ── Hava & Yük Etkisi ─────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🌡 Hava & Yük Etkisi")
    sicaklik = st.slider("Hava Sıcaklığı (°C)", -20, 45, 20)
    yuk_kg   = st.slider("Araç Yükü / Bagaj (kg)", 0, 500, 0, step=25)

    if    sicaklik <  0: hava_kat = 1.15; hava_acik = "+15% (çok soğuk)"
    elif  sicaklik < 10: hava_kat = 1.08; hava_acik = "+8% (soğuk)"
    elif  sicaklik < 25: hava_kat = 1.00; hava_acik = "Etki yok (ideal)"
    elif  sicaklik < 35: hava_kat = 1.05; hava_acik = "+5% (sıcak)"
    else:                hava_kat = 1.08; hava_acik = "+8% (çok sıcak)"

    yuk_ek       = yuk_kg / 100 * 0.3        # her 100 kg → +0.3 L/100km
    tuketim_d    = tuketim * hava_kat + yuk_ek
    duzeltme_pct = (tuketim_d / tuketim - 1) * 100

    st.markdown(f"""<div class="etki-box">
    🌡 Hava: {hava_acik}<br>
    ⚖ Yük: +{yuk_ek:.1f} L/100km<br>
    <b>Düzeltilmiş tüketim: {tuketim:.1f} → {tuketim_d:.1f} L/100km ({duzeltme_pct:+.0f}%)</b>
    </div>""", unsafe_allow_html=True)

    # ── Yakıt Fiyatları ───────────────────────────────────────────────────────
    st.divider()
    st.markdown("### ⛽ Yakıt Fiyatları")

    col_j1, col_j2 = st.columns(2)
    with col_j1:
        if st.button("💲 Fiyatları Yükle", use_container_width=True,
                     help="fiyatlar.json dosyasındaki fiyatları uygular"):
            if fiyat_json_oku():
                st.success("Yüklendi!")
                st.rerun()
            else:
                st.error("fiyatlar.json bulunamadı")
    with col_j2:
        if st.session_state.fiyatlar.get("guncelleme"):
            kaynak = st.session_state.fiyatlar.get("kaynak","")
            st.caption(f"📅 {st.session_state.fiyatlar['guncelleme']}" +
                       (f"\n📌 {kaynak}" if kaynak else ""))

    fiyat  = st.number_input("Benzin (TL/L)",  value=st.session_state.fiyatlar["benzin"],
                              min_value=0.01, step=0.01, format="%.2f", key="inp_benzin")
    st.session_state.fiyatlar["benzin"] = fiyat

    dizel_fiyat   = st.number_input("Dizel (TL/L)",    value=st.session_state.fiyatlar["dizel"],
                                     min_value=0.01, step=0.01, format="%.2f", key="inp_dizel")
    st.session_state.fiyatlar["dizel"] = dizel_fiyat

    lpg_fiyat     = st.number_input("LPG (TL/L)",      value=st.session_state.fiyatlar["lpg"],
                                     min_value=0.01, step=0.01, format="%.2f", key="inp_lpg")
    st.session_state.fiyatlar["lpg"] = lpg_fiyat

    dizel_tuketim = st.number_input("Dizel 100km (L)", value=5.5,  min_value=0.1, step=0.1, format="%.1f")
    lpg_tuketim   = st.number_input("LPG 100km (L)",   value=10.0, min_value=0.1, step=0.1, format="%.1f")

    st.divider()
    st.markdown("### 🌤 Hava Durumu (Open-Meteo)")
    st.caption("✅ API key gerekmez — ücretsiz, 16 güne kadar tahmin.")

    st.divider()
    st.markdown("### ⚡ Elektrikli Araç (EV)")
    ev_fiyat   = st.number_input("Elektrik (TL/kWh)", value=st.session_state.fiyatlar.get("elektrik", 12.0), min_value=0.01, step=0.5, format="%.2f", key="inp_elektrik")
    st.session_state.fiyatlar["elektrik"] = ev_fiyat
    ev_tuketim = st.number_input("EV 100km (kWh)",    value=17.0,  min_value=0.1,  step=0.1, format="%.1f")
    ev_batarya = st.number_input("Batarya (kWh)",     value=64.8,  min_value=1.0,  step=1.0, format="%.1f")

# ── Hesaplamalar ─────────────────────────────────────────────────────────────
# Düzeltilmiş tüketim kullan
tuk = tuketim_d  # hava + yük etkili değer

km_yakit         = tuk / 100
km_maliyet       = km_yakit * fiyat
yol_yakiti       = yol * km_yakit
yol_maliyeti     = yol_yakiti * fiyat
tam_depo_menzil  = (depo / tuk) * 100
tam_depo_maliyet = depo * fiyat
km100_maliyet    = tuk * fiyat
gerekli_depo     = math.ceil(yol / tam_depo_menzil)
depoda_kalan     = depo - yol_yakiti
gidis_donus      = yol_maliyeti * 2
kisi_basi        = gidis_donus / kisi

sure_saat   = yol / hiz

# Mola hesabı: her segmentin sürüşü ayrı ayrı hesaplanır
# (kısa segmentlerde mola olmaz, tablo ile tutarlı olur)
if segments:
    mola_sayisi = sum(int((km / hiz) / mola_araligi) for _, _, km, _ in segments)
    mola_dakika = mola_sayisi * mola_suresi
else:
    mola_sayisi = int(sure_saat / mola_araligi)
    mola_dakika = mola_sayisi * mola_suresi

molali_sure = sure_saat + mola_dakika / 60
varis_dt    = datetime.combine(datetime.today(), cikis) + timedelta(hours=molali_sure)
varis_str   = varis_dt.strftime("%H:%M")
sure_str    = fmt_sure(molali_sure)
surus_str   = fmt_sure(sure_saat)
mola_str    = f"{mola_dakika}dk" + (f" ({mola_sayisi}×{mola_suresi}dk)" if mola_sayisi>0 else " (mola yok)")

aylik_gider   = aylik_km * km_maliyet
yillik_gider  = aylik_gider * 12
yakit_yeterli = yol <= tam_depo_menzil

def hesapla(km, fp, tuk_):
    m = km * tuk_ / 100; return m * fp, m

bnz_m, bnz_y = hesapla(yol, fiyat, tuk)
diz_m, diz_y = hesapla(yol, dizel_fiyat, dizel_tuketim * hava_kat + yuk_ek)
lpg_m, lpg_y = hesapla(yol, lpg_fiyat,   lpg_tuketim   * hava_kat + yuk_ek)
ev_enerji    = yol * ev_tuketim / 100
ev_m         = ev_enerji * ev_fiyat
ev_menzil    = (ev_batarya / ev_tuketim) * 100 if ev_tuketim > 0 else 0
ev_tas       = bnz_m - ev_m
ev_aylik_tas = (aylik_km * ev_tas / yol) if yol > 0 else 0
ev_yillik_tas= ev_aylik_tas * 12
ev_oran      = bnz_m / ev_m if ev_m > 0 else 0

def seg_sure_h(km):
    h = km / hiz
    return h + int(h / mola_araligi) * mola_suresi / 60

# Segment saatleri
def build_seg_rows():
    cur = datetime.combine(datetime.today(), cikis)
    rows = []
    for a,b,km,src in segments:
        dep = cur.strftime("%H:%M")
        sh  = seg_sure_h(km)
        arr_dt = cur + timedelta(hours=sh)
        arr = arr_dt.strftime("%H:%M")
        mol_dk = int((km/hiz)/mola_araligi) * mola_suresi
        rows.append({"Segment":f"{a} → {b}", "Mesafe":f"{km} km ({src})",
                     "Çıkış":dep, "Varış":arr,
                     "Sürüş":fmt_sure(km/hiz), "Mola":f"{mol_dk}dk" if mol_dk else "—",
                     "Toplam":fmt_sure(sh), "Yakıt Mal.":f"{km*km_maliyet:,.0f} ₺"})
        cur = arr_dt
    return rows

# Yakıt dolum noktaları (hangi duraklar)
def yakit_durak_indeksleri():
    idxs = []; kum = 0; sonraki = tam_depo_menzil
    for i,(a,b,km,src) in enumerate(segments):
        if kum + km > sonraki:
            idxs.append(i); sonraki += tam_depo_menzil
        kum += km
    return idxs

# ── Sayfa ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="page-title">🚗 Araç Yol Bilgisayarı</div>', unsafe_allow_html=True)
rl = " → ".join(f"{il}/{ilce}" for il,ilce in st.session_state.duraks)
st.markdown(f'<div class="page-sub">📍 {rl} &nbsp;·&nbsp; {total_yol} km</div>', unsafe_allow_html=True)

# Durum
if yakit_yeterli:
    st.markdown(f"""<div class="status-ok"><div class="status-icon">✅</div>
      <div><div>YAKIT YETERLİ</div>
      <div class="status-detail">Menzil {tam_depo_menzil:.0f} km · Yol {yol} km · Depoda {depoda_kalan:.1f} L kalır · Tüketim {tuk:.1f} L/100km</div>
      </div></div>""", unsafe_allow_html=True)
else:
    st.markdown(f"""<div class="status-fail"><div class="status-icon">⛽</div>
      <div><div>YAKIT YETMEZ — {gerekli_depo} depo gerekli</div>
      <div class="status-detail">Menzil {tam_depo_menzil:.0f} km · Yol {yol} km · {abs(depoda_kalan):.1f} L eksik · Tüketim {tuk:.1f} L/100km</div>
      </div></div>""", unsafe_allow_html=True)

# EV şarj durumu
_ev_sarj_sayisi = math.ceil(yol / ev_menzil) if ev_menzil > 0 else 0
_ev_sarj_kalan  = (_ev_sarj_sayisi * ev_menzil) - yol if ev_menzil > 0 else 0
if _ev_sarj_sayisi <= 1:
    st.markdown(f"""<div class="status-ev-ok"><div class="status-icon">⚡</div>
      <div><div>ELEKTRİKLİ ARAÇ — TEK ŞARJLA GİDİLİR</div>
      <div class="status-detail">Tam şarj menzili {ev_menzil:.0f} km · Yol {yol} km · Varışta {_ev_sarj_kalan:.0f} km menzil kalır</div>
      </div></div>""", unsafe_allow_html=True)
else:
    st.markdown(f"""<div class="status-ev-multi"><div class="status-icon">⚡</div>
      <div><div>ELEKTRİKLİ ARAÇ — {_ev_sarj_sayisi} ŞARJ GEREKLİ</div>
      <div class="status-detail">Tam şarj menzili {ev_menzil:.0f} km · Yol {yol} km · Her {ev_menzil:.0f} km'de bir şarj</div>
      </div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Ana metrikler (tablo dışında — her tab'da görünür)
r1c1,r1c2,r1c3,r1c4,r1c5,r1c6 = st.columns(6)
r1c1.metric("Yol Maliyeti",    f"{yol_maliyeti:,.0f} ₺",  f"{km_maliyet:.2f} ₺/km",  delta_color="off")
r1c2.metric("Gidiş-Dönüş",    f"{gidis_donus:,.0f} ₺",   f"{kisi} kişi, {kisi_basi:,.0f} ₺/kişi", delta_color="off")
r1c3.metric("km Maliyeti",     f"{km_maliyet:.2f} ₺/km",  f"100km = {km100_maliyet:.0f} ₺",         delta_color="off")
r1c4.metric("Yakıt Tüketimi",  f"{km_yakit:.3f} L/km",     f"Yolda {yol_yakiti:.1f} L",               delta_color="off")
r1c5.metric("Kişi Başı (G/D)", f"{kisi_basi:,.0f} ₺",    f"{kisi} kişi",                            delta_color="off")
r1c6.metric("Tahmini Varış",   varis_str, f"Çıkış {cikis.strftime('%H:%M')} · {sure_str}",          delta_color="off")
st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# Tabs
tab_ozet, tab_hava, tab_harita, tab_rapor = st.tabs(["📊 Özet", "🌤 Hava Durumu", "🗺 Harita", "📄 Rapor"])

# ═══════════════════════ ÖZET ═════════════════════════════════════════════════
with tab_ozet:
    r2c1,r2c2,r2c3,r2c4 = st.columns(4)
    r2c1.metric("Aylık Yakıt Gideri", f"{aylik_gider:,.0f} ₺", f"Yıllık {yillik_gider:,.0f} ₺", delta_color="off")
    r2c2.metric("100km Maliyeti",     f"{km100_maliyet:,.0f} ₺", f"Tüketim {tuk:.1f} L",         delta_color="off")
    r2c3.metric("Tam Depo Menzili",   f"{tam_depo_menzil:.0f} km", "Benzin",                      delta_color="off")
    r2c4.metric("Tam Depo Maliyeti",  f"{tam_depo_maliyet:,.0f} ₺", "Benzin",                     delta_color="off")

    # Tüm yakıt tipleri menzil & tam depo/şarj karşılaştırması
    st.markdown('<div class="section-header">🔋 Menzil & Tam Depo / Şarj Karşılaştırması</div>', unsafe_allow_html=True)

    dizel_tuk_d_pre = dizel_tuketim * hava_kat + yuk_ek
    lpg_tuk_d_pre   = lpg_tuketim   * hava_kat + yuk_ek

    diz_menzil  = (depo / dizel_tuk_d_pre) * 100  if dizel_tuk_d_pre > 0 else 0
    lpg_menzil  = (depo / lpg_tuk_d_pre)   * 100  if lpg_tuk_d_pre   > 0 else 0
    ev_menzil2  = (ev_batarya / ev_tuketim) * 100  if ev_tuketim      > 0 else 0

    diz_depo_mal = depo * dizel_fiyat
    lpg_depo_mal = depo * lpg_fiyat
    ev_sarj_mal  = ev_batarya * ev_fiyat

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Benzin Menzili",  f"{tam_depo_menzil:.0f} km", f"Tam depo: {tam_depo_maliyet:,.0f} ₺", delta_color="off")
    m2.metric("Dizel Menzili",   f"{diz_menzil:.0f} km",      f"Tam depo: {diz_depo_mal:,.0f} ₺",     delta_color="off")
    m3.metric("LPG Menzili",     f"{lpg_menzil:.0f} km",      f"Tam depo: {lpg_depo_mal:,.0f} ₺",     delta_color="off")
    m4.metric("EV Menzili",      f"{ev_menzil2:.0f} km",      f"Tam şarj: {ev_sarj_mal:,.0f} ₺",      delta_color="off")

    st.markdown('<div class="section-header">⏱ Süre Bilgileri</div>', unsafe_allow_html=True)
    t1,t2,t3,t4,t5 = st.columns(5)
    t1.metric("Sürüş Süresi", surus_str,  f"{yol} km / {hiz} km/s",                    delta_color="off")
    t2.metric("Mola Ayarı",   f"Her {mola_araligi}s → {mola_suresi}dk", f"{mola_sayisi} mola yapılacak", delta_color="off")
    t3.metric("Toplam Mola",  f"{mola_dakika}dk",  f"{mola_sayisi} × {mola_suresi}dk",  delta_color="off")
    t4.metric("Toplam Süre",  sure_str,   "Sürüş + mola",                               delta_color="off")
    t5.metric("Varış",        varis_str,  f"Çıkış: {cikis.strftime('%H:%M')}",          delta_color="off")

    if len(segments) > 1:
        st.markdown('<div class="section-header">📍 Güzergah Detayı</div>', unsafe_allow_html=True)
        seg_df = pd.DataFrame(build_seg_rows())
        st.dataframe(seg_df, hide_index=True, use_container_width=True)

    st.markdown('<div class="section-header">⛽ Yakıt Tipi Karşılaştırması</div>', unsafe_allow_html=True)

    def fmt_tas(v):
        if v == 0: return "—"
        return f"{'+'if v>0 else ''}{v:,.0f} ₺"

    dizel_tuk_d = dizel_tuketim * hava_kat + yuk_ek
    lpg_tuk_d   = lpg_tuketim   * hava_kat + yuk_ek

    # km başı tüketim ve maliyet
    bnz_lkm = tuk          / 100;  bnz_tkm = bnz_lkm * fiyat
    diz_lkm = dizel_tuk_d  / 100;  diz_tkm = diz_lkm * dizel_fiyat
    lpg_lkm = lpg_tuk_d    / 100;  lpg_tkm = lpg_lkm * lpg_fiyat
    ev_lkm  = ev_tuketim   / 100;  ev_tkm  = ev_lkm  * ev_fiyat

    comp = pd.DataFrame([
        {"Yakıt":"Benzin",  "Fiyat":f"{fiyat:.2f} ₺/L",      "L(kWh)/km":f"{bnz_lkm:.3f}", "₺/km":f"{bnz_tkm:.3f}", "100km":f"{tuk:.1f} L",          "Yol Yakıtı":f"{bnz_y:.1f} L",    "Yol Mal.":f"{bnz_m:,.0f} ₺",  "G/D":f"{bnz_m*2:,.0f} ₺",  "Tasarruf":"—"},
        {"Yakıt":"Dizel",   "Fiyat":f"{dizel_fiyat:.2f} ₺/L", "L(kWh)/km":f"{diz_lkm:.3f}", "₺/km":f"{diz_tkm:.3f}", "100km":f"{dizel_tuk_d:.1f} L",  "Yol Yakıtı":f"{diz_y:.1f} L",    "Yol Mal.":f"{diz_m:,.0f} ₺",  "G/D":f"{diz_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-diz_m)},
        {"Yakıt":"LPG",     "Fiyat":f"{lpg_fiyat:.2f} ₺/L",   "L(kWh)/km":f"{lpg_lkm:.3f}", "₺/km":f"{lpg_tkm:.3f}", "100km":f"{lpg_tuk_d:.1f} L",    "Yol Yakıtı":f"{lpg_y:.1f} L",    "Yol Mal.":f"{lpg_m:,.0f} ₺",  "G/D":f"{lpg_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-lpg_m)},
        {"Yakıt":"⚡ EV",   "Fiyat":f"{ev_fiyat:.2f} ₺/kWh",  "L(kWh)/km":f"{ev_lkm:.3f}",  "₺/km":f"{ev_tkm:.3f}",  "100km":f"{ev_tuketim:.1f} kWh", "Yol Yakıtı":f"{ev_enerji:.1f} kWh","Yol Mal.":f"{ev_m:,.0f} ₺",  "G/D":f"{ev_m*2:,.0f} ₺",  "Tasarruf":fmt_tas(bnz_m-ev_m)},
    ])
    st.dataframe(comp, hide_index=True, use_container_width=True)

    st.markdown('<div class="section-header">📊 Maliyet Grafiği</div>', unsafe_allow_html=True)
    fig = go.Figure(go.Bar(
        x=["Benzin","Dizel","LPG","⚡ EV"], y=[bnz_m,diz_m,lpg_m,ev_m],
        marker_color=["#ED7D31","#2E75B6","#70AD47","#7030A0"],
        text=[f"{v:,.0f} ₺" for v in [bnz_m,diz_m,lpg_m,ev_m]],
        textposition="outside", textfont=dict(size=13,family="Inter"),
    ))
    fig.add_hline(y=bnz_m, line_dash="dot", line_color="#ED7D31",
                  annotation_text=f"Benzin: {bnz_m:,.0f} ₺",
                  annotation_position="top right",
                  annotation_font=dict(color="#ED7D31",size=11))
    fig.update_layout(plot_bgcolor="white", paper_bgcolor="white",
        yaxis=dict(title=dict(text="Yol Maliyeti (₺)",font=dict(size=12)), gridcolor="#F0F0F0"),
        xaxis=dict(tickfont=dict(size=13)), margin=dict(t=20,b=20,l=20,r=20), height=300, showlegend=False)
    st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="section-header">⚡ EV Tasarruf Analizi</div>', unsafe_allow_html=True)
    e1,e2,e3,e4 = st.columns(4)
    e1.metric("Bu Seyahat",    f"{'+'if ev_tas>=0 else ''}{ev_tas:,.0f} ₺",         "Benzin'e kıyasla",  delta_color="off")
    e2.metric("Aylık Tasarruf",f"{'+'if ev_aylik_tas>=0 else ''}{ev_aylik_tas:,.0f} ₺", f"{aylik_km:,} km/ay",delta_color="off")
    e3.metric("Yıllık Tasarruf",f"{'+'if ev_yillik_tas>=0 else ''}{ev_yillik_tas:,.0f} ₺")
    e4.metric("Benzin/EV Oran",f"{ev_oran:.2f}×","EV ucuz" if ev_oran>1 else "EV pahalı",delta_color="off")

# ═══════════════════════ HARİTA ═══════════════════════════════════════════════
# ═══════════════════════ HAVA DURUMU ═════════════════════════════════════════
with tab_hava:
    st.markdown('<div class="section-header">🌤 Güzergah Boyunca Geçilen İllerin Hava Durumu</div>',
                unsafe_allow_html=True)
    st.caption("🌐 Open-Meteo kullanılıyor — API key gerekmez · 16 güne kadar saatlik tahmin")

    # ── İl merkezlerini hesapla (tüm ilçelerin ortalaması) ────────────────
    @st.cache_data
    def il_merkezleri():
        merkezler = {}
        for il, ilceler in ILCELER.items():
            lats = [v[0] for v in ilceler.values()]
            lons = [v[1] for v in ilceler.values()]
            merkezler[il] = (sum(lats)/len(lats), sum(lons)/len(lons))
        return merkezler

    IL_MERKEZ = il_merkezleri()

    def en_yakin_il(lat, lon):
        """Koordinata en yakın ili bul"""
        en_az = float('inf')
        bulunan = None
        for il, (ilat, ilon) in IL_MERKEZ.items():
            d = (lat - ilat)**2 + (lon - ilon)**2
            if d < en_az:
                en_az = d
                bulunan = il
        return bulunan

    def guzergah_illeri(segments, stop_times):
        """
        Her segment için düz hat üzerinde her ~30 km'de bir ara nokta üret,
        o noktaya en yakın ili bul. Zaman orantılı hesaplanır.
        Döner: [(il_adi, lat, lon, tahmini_dt), ...]  — tekrarsız sıralı liste
        """
        gecilen = []   # (il, lat, lon, dt)
        goruldü = set()

        for seg_idx, (_, _, km, _) in enumerate(segments):
            c1 = koordinat(*st.session_state.duraks[seg_idx])
            c2 = koordinat(*st.session_state.duraks[seg_idx + 1])
            dep_dt = stop_times[seg_idx]
            arr_dt = stop_times[seg_idx + 1]
            sure_sn = (arr_dt - dep_dt).total_seconds()

            # Her ~30 km'de bir nokta (en az 2 nokta: %0 ve %100)
            n_nokta = max(2, round(km / 30) + 1)
            for j in range(n_nokta):
                t = j / (n_nokta - 1)  # 0.0 → 1.0
                lat = c1[0] + (c2[0] - c1[0]) * t
                lon = c1[1] + (c2[1] - c1[1]) * t
                zaman = dep_dt + timedelta(seconds=sure_sn * t)
                il = en_yakin_il(lat, lon)
                if il not in goruldü:
                    goruldü.add(il)
                    il_lat, il_lon = IL_MERKEZ[il]
                    gecilen.append((il, il_lat, il_lon, zaman))

        return gecilen

    # Segment zamanlarını hesapla
    _cur_dt = datetime.combine(datetime.today(), cikis)
    _stop_times = [_cur_dt]
    for _, _, km, _ in segments:
        _cur_dt = _cur_dt + timedelta(hours=seg_sure_h(km))
        _stop_times.append(_cur_dt)

    # Geçilen illeri hesapla
    rota_illeri = guzergah_illeri(segments, _stop_times)

    col_btn, col_bilgi = st.columns([2, 5])
    with col_btn:
        if st.button("🔄 Hava Durumunu Getir / Yenile"):
            st.session_state.hava_cache = {}
    with col_bilgi:
        st.caption(f"Güzergahta **{len(rota_illeri)} il** tespit edildi "
                   f"(her ~30 km'de bir nokta, il sınırı tahmini).")

    st.markdown("<br>", unsafe_allow_html=True)

    html_kartlar = ""
    sicakliklar  = []
    simdi        = datetime.now()

    for il, lat, lon, gecis_dt in rota_illeri:
        gecis_str = gecis_dt.strftime("%d.%m %H:%M")
        fark_gun  = (gecis_dt - simdi).total_seconds() / 86400

        if fark_gun > 16:
            html_kartlar += (
                f'<div style="background:#F5F5F5;border-radius:10px;padding:10px 16px;'
                f'margin:6px 0;border-left:5px solid #9E9E9E;color:#666;font-size:0.85rem">'
                f'📅 <b>{il}</b> ({gecis_str}) — 16 günden uzak, tahmin mevcut değil</div>'
            )
            continue

        with st.spinner(f"{il} hava durumu alınıyor..."):
            veri        = hava_getir(lat, lon, gecis_dt)
            kalite_veri = hava_kalite_getir(lat, lon, gecis_dt)

        if "hata" not in veri:
            sicakliklar.append(veri["temp"])

        html_kartlar += hava_karti(il, gecis_str, veri, gecis_dt, kalite_veri)

    st.markdown(html_kartlar, unsafe_allow_html=True)

    # Güzergah özeti
    if sicakliklar:
        ort = sum(sicakliklar) / len(sicakliklar)
        en_dusuk = min(sicakliklar)
        en_yuksek = max(sicakliklar)
        if    ort <  0: etki = "+15% (çok soğuk)"
        elif  ort < 10: etki = "+8% (soğuk)"
        elif  ort < 25: etki = "±0% (ideal)"
        elif  ort < 35: etki = "+5% (sıcak)"
        else:            etki = "+8% (çok sıcak)"
        st.info(
            f"🌡 Güzergah sıcaklık — Ort: **{ort:.0f}°C** | "
            f"En düşük: **{en_dusuk:.0f}°C** | En yüksek: **{en_yuksek:.0f}°C**  \n"
            f"⛽ Tahmini yakıt etkisi: **{etki}** "
            f"(Sol menü → Hava & Yük ayarını buna göre güncelleyebilirsin)"
        )

with tab_harita:
    stop_coords = [koordinat(il,ilce) for il,ilce in st.session_state.duraks]
    center_lat  = sum(c[0] for c in stop_coords) / len(stop_coords)
    center_lon  = sum(c[1] for c in stop_coords) / len(stop_coords)

    # ── Online harita linkleri ────────────────────────────────────────────────
    st.markdown('<div class="section-header">🌐 Online Harita Uygulamalarında Aç</div>',
                unsafe_allow_html=True)

    # Google Maps: /dir/lat1,lon1/lat2,lon2/.../latN,lonN  ✅ Çoklu durak
    gm_noktalar = "/".join(f"{lat},{lon}" for lat, lon in stop_coords)
    gm_url      = f"https://www.google.com/maps/dir/{gm_noktalar}"

    # Yandex Maps: rtext=lat1,lon1~lat2,lon2~...  ✅ Çoklu durak
    yx_noktalar = "~".join(f"{lat},{lon}" for lat, lon in stop_coords)
    yx_url      = f"https://yandex.com.tr/maps/?rtext={yx_noktalar}&rtt=auto&lang=tr_TR"

    # Bing Maps: rtp=pos.lat1_lon1~pos.lat2_lon2~...  ✅ Çoklu durak
    bing_noktalar = "~".join(f"pos.{lat}_{lon}" for lat, lon in stop_coords)
    bing_url      = f"https://www.bing.com/maps?rtp={bing_noktalar}&mode=D"

    # Waze: sadece kalkış → varış ❌ Ara durak desteği yok
    wz_kalkis = stop_coords[0]
    wz_varis  = stop_coords[-1]
    wz_url    = (f"https://waze.com/ul?ll={wz_varis[0]},{wz_varis[1]}"
                 f"&navigate=yes&from={wz_kalkis[0]},{wz_kalkis[1]}")

    coklu = len(stop_coords) > 2

    lc1, lc2, lc3, lc4 = st.columns(4)
    lc1.link_button("🗺 Google Maps",  gm_url,   use_container_width=True)
    lc2.link_button("🗺 Yandex Maps",  yx_url,   use_container_width=True)
    lc3.link_button("🗺 Bing Maps",    bing_url, use_container_width=True)
    lc4.link_button("🚗 Waze",         wz_url,   use_container_width=True)

    if coklu:
        st.caption("✅ Google Maps · Yandex · Bing → tüm duraklar desteklenir  "
                   "⚠️ Waze → sadece kalkış→varış açılır, ara duraklar desteklenmiyor")
    else:
        st.caption("✅ Tüm uygulamalar tek güzergahı destekler")

    # Durak listesi
    with st.expander("📋 Koordinat Listesi (kopyala-yapıştır için)"):
        koord_listesi = "\n".join(
            f"{i+1}. {il}/{ilce} → {lat}, {lon}"
            for i, ((il, ilce), (lat, lon))
            in enumerate(zip(st.session_state.duraks, stop_coords))
        )
        st.code(koord_listesi, language=None)

    st.markdown('<div class="section-header">🗺 Rota Önizlemesi</div>',
                unsafe_allow_html=True)
    st.caption("Aşağıdaki harita tahmini güzergahı gösterir. "
               "Gerçek yol için yukarıdaki bağlantıları kullanın.")

    m = folium.Map(location=[center_lat,center_lon], zoom_start=6,
                   tiles="CartoDB positron")

    # Güzergah çizgisi
    folium.PolyLine(stop_coords, weight=3, color="#1F4E78", opacity=0.8,
                    dash_array=None).add_to(m)

    # Durak marker'ları
    seg_rows = build_seg_rows() if len(segments) > 0 else []
    for i,(il,ilce) in enumerate(st.session_state.duraks):
        lat, lon = koordinat(il, ilce)
        if i == 0:
            color, icon_name, label = "green",  "play",      "Kalkış"
        elif i == len(st.session_state.duraks)-1:
            color, icon_name, label = "red",    "stop",      "Varış"
        else:
            color, icon_name, label = "blue",   "info-sign", f"{i}. Durak"

        # Çıkış/varış saatini bul
        if i < len(seg_rows):
            zaman = f"Çıkış: {seg_rows[i]['Çıkış']}"
        elif seg_rows:
            zaman = f"Varış: {seg_rows[-1]['Varış']}"
        else:
            zaman = ""

        popup_html = f"""
        <b>{label}</b><br>
        {il} / {ilce}<br>
        {zaman}
        """
        folium.Marker([lat,lon],
            popup=folium.Popup(popup_html, max_width=180),
            tooltip=f"{label}: {il}/{ilce}",
            icon=folium.Icon(color=color, icon=icon_name)
        ).add_to(m)

    # Yakıt dolum noktaları
    yakit_idxler = yakit_durak_indeksleri()
    for idx in yakit_idxler:
        durak_idx = idx
        if durak_idx < len(st.session_state.duraks):
            il, ilce = st.session_state.duraks[durak_idx]
            lat, lon = koordinat(il, ilce)
            kum_km   = sum(s[2] for s in segments[:idx])
            folium.Marker([lat,lon],
                popup=folium.Popup(
                    f"<b>⛽ Yakıt Durağı</b><br>{il}/{ilce}<br>Yaklaşık {kum_km} km'de", max_width=160),
                tooltip=f"⛽ Yakıt al: {il}/{ilce}",
                icon=folium.Icon(color="orange", icon="tint", prefix="fa")
            ).add_to(m)

    col_map, col_info = st.columns([3, 1])
    with col_map:
        st_folium(m, height=520, use_container_width=True)
    with col_info:
        st.markdown("**Harita Göstergesi**")
        st.markdown("🟢 Kalkış noktası")
        st.markdown("🔵 Ara duraklar")
        st.markdown("🔴 Varış noktası")
        st.markdown("🟠 Yakıt dolum noktası")
        st.divider()
        st.markdown(f"**Toplam:** {total_yol} km")
        st.markdown(f"**Süre:** {sure_str}")
        st.markdown(f"**Varış:** {varis_str}")
        if yakit_idxler:
            st.divider()
            st.markdown(f"**Yakıt durağı:** {len(yakit_idxler)} kez")
            st.markdown(f"Menzil: {tam_depo_menzil:.0f} km")
        else:
            st.divider()
            st.markdown("✅ Tek depoda gidilir")
            st.markdown(f"Kalacak: {depoda_kalan:.1f} L")

# ═══════════════════════ RAPOR ════════════════════════════════════════════════
with tab_rapor:
    st.markdown('<div class="section-header">📄 Rapor Oluştur</div>', unsafe_allow_html=True)
    st.markdown("Seyahat planını PDF veya Excel olarak indir.")

    tarih = datetime.now().strftime("%d.%m.%Y")
    rota  = " → ".join(f"{il}/{ilce}" for il,ilce in st.session_state.duraks)

    # ── PDF ───────────────────────────────────────────────────────────────────
    def olustur_pdf():
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        stiller = getSampleStyleSheet()
        baslik_stili  = ParagraphStyle("B", parent=stiller["Heading1"],
            fontName=PDF_FONT_BOLD, fontSize=16, textColor=colors.HexColor("#1F4E78"), spaceAfter=4)
        alt_stili     = ParagraphStyle("A", parent=stiller["Normal"],
            fontName=PDF_FONT, fontSize=9, textColor=colors.grey, spaceAfter=12)
        bolum_stili   = ParagraphStyle("S", parent=stiller["Heading2"],
            fontName=PDF_FONT_BOLD, fontSize=11, textColor=colors.HexColor("#1F4E78"), spaceBefore=14, spaceAfter=6)

        def tablo(veriler, col_w=None):
            t = Table(veriler, colWidths=col_w)
            t.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
                ("FONTNAME",     (0,0), (-1,0), PDF_FONT_BOLD),
                ("FONTNAME",     (0,1), (-1,-1), PDF_FONT),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F0F4F8")]),
                ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
                ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",   (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ])); return t

        story = [
            Paragraph("🚗 Araç Yol Bilgisayarı — Seyahat Raporu", baslik_stili),
            Paragraph(f"{rota}  ·  {total_yol} km  ·  {tarih}", alt_stili),
            Paragraph("Özet Bilgiler", bolum_stili),
            tablo([
                ["Yol","Gidiş-Dönüş","Kişi Başı","Toplam Süre","Varış"],
                [f"{yol} km", f"{gidis_donus:,.0f} ₺", f"{kisi_basi:,.0f} ₺", sure_str, varis_str],
            ], [3*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]),
            Spacer(1,0.3*cm),
            tablo([
                ["Menzil","Tüketim (düz.)","Yakıt Fiyatı","Aylık Gider","Yıllık Gider"],
                [f"{tam_depo_menzil:.0f} km", f"{tuk:.1f} L/100km",
                 f"{fiyat:.2f} ₺/L", f"{aylik_gider:,.0f} ₺", f"{yillik_gider:,.0f} ₺"],
            ], [3.5*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]),
        ]

        if len(segments) > 1:
            story.append(Paragraph("Güzergah Detayı", bolum_stili))
            rows = build_seg_rows()
            tablo_v = [["Segment","Mesafe","Çıkış","Varış","Süre","Yakıt"]]
            for r in rows:
                tablo_v.append([r["Segment"], r["Mesafe"], r["Çıkış"], r["Varış"],
                                r["Toplam"], r["Yakıt Mal."]])
            story.append(tablo([tablo_v[0]] + tablo_v[1:],
                [6.5*cm,2.5*cm,1.8*cm,1.8*cm,2*cm,2.4*cm]))

        story.append(Paragraph("Yakıt Karşılaştırması", bolum_stili))
        story.append(tablo([
            ["Yakıt","Fiyat","Yol Maliyeti","G/D Toplam","Tasarruf"],
            ["Benzin",  f"{fiyat:.2f} ₺/L",      f"{bnz_m:,.0f} ₺", f"{bnz_m*2:,.0f} ₺", "—"],
            ["Dizel",   f"{dizel_fiyat:.2f} ₺/L", f"{diz_m:,.0f} ₺", f"{diz_m*2:,.0f} ₺", fmt_tas(bnz_m-diz_m)],
            ["LPG",     f"{lpg_fiyat:.2f} ₺/L",   f"{lpg_m:,.0f} ₺", f"{lpg_m*2:,.0f} ₺", fmt_tas(bnz_m-lpg_m)],
            ["EV",      f"{ev_fiyat:.2f} ₺/kWh",  f"{ev_m:,.0f} ₺",  f"{ev_m*2:,.0f} ₺",  fmt_tas(bnz_m-ev_m)],
        ], [3.5*cm,3.5*cm,3.5*cm,3.5*cm,3*cm]))

        if duzeltme_pct != 0:
            story.append(Spacer(1,0.3*cm))
            story.append(Paragraph(
                f"ℹ Tüketim {sicaklik}°C hava ve {yuk_kg}kg yük için düzeltilmiştir "
                f"({tuketim:.1f} → {tuk:.1f} L/100km, {duzeltme_pct:+.0f}%).", alt_stili))

        doc.build(story)
        buf.seek(0); return buf

    # ── Excel ─────────────────────────────────────────────────────────────────
    def olustur_excel():
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Seyahat Raporu"
        DARK  = "1F4E78"; WHT = "FFFFFF"; LBLU = "DEEAF1"; YELL = "FFF2CC"
        def hd(cell, val, bg=DARK, fc=WHT, bold=True, sz=11):
            cell.value = val
            cell.font  = Font(bold=bold, color=fc, size=sz, name="Calibri")
            cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=s,right=s,top=s,bottom=s)
        def vl(cell, val, bg=LBLU, bold=False):
            cell.value = val
            cell.font  = Font(bold=bold, color="000000", size=10, name="Calibri")
            cell.fill  = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=s,right=s,top=s,bottom=s)

        r = 1
        ws.merge_cells(f"A{r}:G{r}")
        hd(ws[f"A{r}"], f"ARAÇ YOL BİLGİSAYARI — {tarih}", sz=13)
        ws.row_dimensions[r].height = 24; r += 1
        ws.merge_cells(f"A{r}:G{r}")
        hd(ws[f"A{r}"], rota, bg="2E4057", sz=10)
        ws.row_dimensions[r].height = 18; r += 2

        hdrs = ["Yol","Gidiş-Dönüş","Kişi Başı","Toplam Süre","Menzil","Tüketim","Varış"]
        vals = [f"{yol} km", f"{gidis_donus:,.0f} ₺", f"{kisi_basi:,.0f} ₺",
                sure_str, f"{tam_depo_menzil:.0f} km", f"{tuk:.1f} L/100km", varis_str]
        for j,(h,v) in enumerate(zip(hdrs,vals),1):
            hd(ws.cell(r,j), h); vl(ws.cell(r+1,j), v)
        ws.row_dimensions[r].height=18; ws.row_dimensions[r+1].height=18; r+=3

        if len(segments)>1:
            seg_hdrs=["Segment","Mesafe","Çıkış","Varış","Sürüş","Mola","Toplam","Yakıt"]
            for j,h in enumerate(seg_hdrs,1): hd(ws.cell(r,j),h)
            ws.row_dimensions[r].height=18; r+=1
            for row in build_seg_rows():
                for j,k in enumerate(["Segment","Mesafe","Çıkış","Varış","Sürüş","Mola","Toplam","Yakıt Mal."],1):
                    bg="FFFFFF" if r%2==0 else "F5F8FC"
                    vl(ws.cell(r,j), row.get(k,""), bg=bg)
                ws.row_dimensions[r].height=16; r+=1
            r+=1

        yakit_hdrs=["Yakıt","Fiyat","Yol Maliyeti","G/D Toplam","Tasarruf"]
        for j,h in enumerate(yakit_hdrs,1): hd(ws.cell(r,j),h)
        ws.row_dimensions[r].height=18; r+=1
        for yakdata in [
            ("Benzin",  f"{fiyat:.2f}₺/L",      f"{bnz_m:,.0f}₺", f"{bnz_m*2:,.0f}₺","—"),
            ("Dizel",   f"{dizel_fiyat:.2f}₺/L", f"{diz_m:,.0f}₺", f"{diz_m*2:,.0f}₺",fmt_tas(bnz_m-diz_m)),
            ("LPG",     f"{lpg_fiyat:.2f}₺/L",   f"{lpg_m:,.0f}₺", f"{lpg_m*2:,.0f}₺",fmt_tas(bnz_m-lpg_m)),
            ("⚡ EV",   f"{ev_fiyat:.2f}₺/kWh",  f"{ev_m:,.0f}₺",  f"{ev_m*2:,.0f}₺", fmt_tas(bnz_m-ev_m)),
        ]:
            bg = "FFFFFF" if r%2==0 else "F5F8FC"
            for j,v in enumerate(yakdata,1): vl(ws.cell(r,j),v,bg=bg)
            ws.row_dimensions[r].height=16; r+=1

        for col in ws.columns:
            col_letter = None
            for cell in col:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
            mx = max((len(str(c.value or "")) for c in col if hasattr(c, "value")), default=10)
            ws.column_dimensions[col_letter].width = min(mx + 3, 30)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    # İndirme butonları
    dc1, dc2, dc3 = st.columns([1,1,2])
    with dc1:
        try:
            pdf_buf = olustur_pdf()
            st.download_button("📥 PDF İndir", data=pdf_buf,
                file_name=f"seyahat_{tarih.replace('.','')}.pdf",
                mime="application/pdf", use_container_width=True)
        except Exception as e:
            st.error(f"PDF hatası: {e}")
    with dc2:
        try:
            xl_buf = olustur_excel()
            st.download_button("📥 Excel İndir", data=xl_buf,
                file_name=f"seyahat_{tarih.replace('.','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.error(f"Excel hatası: {e}")

    st.divider()
    st.markdown("**Rapor içeriği:**")
    st.markdown("- Özet bilgiler (yol, maliyet, süre, varış)\n"
                "- Güzergah detay tablosu (çıkış/varış saatleri)\n"
                "- Yakıt tipi karşılaştırması (benzin, dizel, LPG, EV)\n"
                "- Hava/yük düzeltme notu")
    st.divider()
    st.caption("⛽ Yakıt fiyatlarını güncellemek için fiyatlar.json dosyasını düzenleyin.")
    st.caption(f"🔤 PDF fontu: {PDF_FONT}")
    st.caption("📐 Mesafeler koordinat tabanlı tahminidir (haversine × 1.30 yol katsayısı). Gerçek mesafe için navigasyon uygulaması kullanın.")
    st.caption("© Yakıt Hesabı | Streamlit + Python | 2026 Enes Özkan")
